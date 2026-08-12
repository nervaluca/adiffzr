import os
import sys

_this_dir = os.path.dirname(os.path.abspath(__file__))
_target_pkg = r'C:\Users\nerva\Desktop\printlog\innosetup3.2\ADIF_FZR_Modular'
if _target_pkg not in sys.path:
    sys.path.insert(0, _target_pkg)
if _this_dir not in sys.path:
    sys.path.insert(0, _this_dir)

import os
import sys
import re
import csv
import json
import openpyxl
import time
import math
from datetime import datetime, timedelta
import customtkinter as ctk
from tkinter import filedialog, messagebox, colorchooser
import tkinter.ttk as _ttk
import adif_io
from reportlab.lib.pagesizes import A4, landscape, LETTER
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, String, Rect
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend

try:
    from PIL import Image
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

from config import T, imposta_lingua, LINGUA, TRADUZIONI, MAPPA_CONTINENTI_DXCC, CONTINENTS_ORDER
from config import VERSIONE, BUILD_DATE, APP_TITOLO, PROGRAMID_ADIF
from utils.dxcc import dxcc_da_nominativo
from utils.maidenhead import locator_to_latlon, distanza_bearing, bearing_to_compass, estrai_locator_da_testo
from utils.formatting import chiedi_cartella_output
from utils.tooltip import _tip
from radio.omnirig import OmniRigControl
from radio.sdrconsole import SDRConsoleControl
from radio.bandplan import modo_da_bandplan
from radio import sat_db as SATDB
import theme as TH
from pdf.canvas import ElegantNumberedCanvas
from net.wsjtx import WSJTXListener
from net.dxcluster import DXClusterWindow
from gui.widgets import _WrapToolbar, SplashScreen, CalendarPopup
from gui.dialogs.filters import FiltriDialog
from gui.dialogs.preferences import ColoriDialog, ColoriHtmlDialog, OpzioniRegistroPDFDialog
from gui.dialogs.merge import UnisciDialog
from gui.dialogs.dupe_check import DuplicatiDialog
from gui.dialogs.charts import GraficiDialog
from gui.dialogs.satellite import SatellitiDialog
from gui.dialogs.qsl_designer import QSLCardDesignerDialog, SelezioneQSODialog, QSLMasterDialog, QSLCardDialog
from gui.dialogs.online_dialogs import (
    CloudlogUploadDialog, ClublogUploadDialog, LotwUploadDialog,
    EqslUploadDialog, QO100UploadDialog, HamQTHDialog,
    LotwDownloadDialog, EqslDownloadDialog, EqslUnconfirmedDialog
)

class ADIFtoPDFApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title(f"{APP_TITOLO}  ·  build {BUILD_DATE}  —  IW1FZR")
        self.geometry("1100x850")

        self.minsize(1000, 750)

        self._imposta_icona()
        # Gestione chiusura pulita: chiede di salvare e chiude OmniRig se avviato
        self.protocol("WM_DELETE_WINDOW", self._chiudi_app)

        # ── Splash screen ──────────────────────────
        self.withdraw()  # nasconde la finestra principale durante il caricamento
        splash = self._mostra_splash()

        self.filepath = ""
        self.qsos_caricati = []
        self._undo_stack = []   # stack per Ctrl+Z nella finestra principale
        self.qsos_filtrati = []
        self._log_modificato = False   # True se ci sono modifiche non salvate

        self.colori_pdf = {
            'primario': '#1A365D',
            'secondario': '#2B6CB0',
            'riga_pari': '#F7FAFC',
        }

        self.colori_html = {
            'primario':   '#1A365D',
            'secondario': '#2B6CB0',
            'bg_scuro':   '#0D1117',
            'bg_chiaro':  '#F7FAFC',
        }

        self.campi_disponibili = {
            'QSO_DATE':       {'nome': 'Data',       'width_base': 65},
            'TIME_ON':        {'nome': 'UTC',         'width_base': 50},
            'CALL':           {'nome': 'Callsign',    'width_base': 80},
            'NAME':           {'nome': 'Name',        'width_base': 85},
            'BAND':           {'nome': 'Banda',       'width_base': 55},
            'FREQ':           {'nome': 'Frequenza',   'width_base': 65},
            'MODE':           {'nome': 'Mode',        'width_base': 68},
            'RST_SENT':       {'nome': 'RST TX',      'width_base': 50},
            'RST_RCVD':       {'nome': 'RST RX',      'width_base': 50},
            'GRIDSQUARE':     {'nome': 'Locator',     'width_base': 65},
            'COUNTRY':        {'nome': 'Country',     'width_base': 152},
            'LOTW_QSL_RCVD':  {'nome': 'LoTW',        'width_base': 45},
            'EQSL_QSL_RCVD':  {'nome': 'eQSL',        'width_base': 45},
        }
        # Ordine colonne PDF (lista di tag) — personalizzabile dall'utente
        self.ordine_campi_pdf = list(self.campi_disponibili.keys())
        # Larghezze personalizzate (sovrascrivono width_base se impostate)
        self.width_pdf = {}
        # Titolo personalizzato e font size celle
        self.titolo_pdf_custom = ""
        self.font_size_pdf = 7

        self.checkboxes = {}
        self._widget_refs = {}
        self._coda_export = []  # lista di QSO (riferimenti) in coda per export, persistente tra filtri
        self._ultima_riga_toggle = None  # ultima riga marcata con Ctrl+Click (per range Shift+Click)
        self._checkbox_widgets = {}
        self.profilo_path   = os.path.join(os.path.expanduser("~"), ".adif_converter_profilo.json")
        self.profili_path   = os.path.join(os.path.expanduser("~"), ".adif_fzr_profili.json")
        self.profilo_attivo = None
        self.storico_path   = os.path.join(os.path.expanduser("~"), ".adif_converter_storico.json")
        self.storico_files  = []
        self.formati_qsl_path = os.path.join(os.path.expanduser("~"), ".adif_fzr_formati_qsl.json")
        self.impostazioni_apertura_path = os.path.join(
            os.path.expanduser("~"), ".adif_fzr_apertura.json")
        self.var_controllo_post_apertura = ctk.BooleanVar(value=False)
        self.var_colora_righe = ctk.BooleanVar(value=True)
        self._carica_impostazioni_apertura()
        self._campi_presenti_nel_log = set()  # popolato da _aggiorna_tree()
        self.var_lingua     = ctk.StringVar(value="IT")
        # Allinea var_tema al tema GIÀ impostato all'avvio (evita flash light→dark)
        self.var_tema       = ctk.StringVar(value=ctk.get_appearance_mode())
        self.var_dxcc_page  = ctk.BooleanVar(value=True)
        self.var_formato_pdf= ctk.StringVar(value="A4")

        # Variabili BooleanVar campi (inizializzate prima di crea_interfaccia)
        for tag in self.campi_disponibili:
            self.checkboxes[tag] = ctk.BooleanVar(value=(tag not in ('FREQ', 'NAME')))

        self.crea_interfaccia()
        self.carica_profilo()
        # Il tema è già stato applicato all'avvio da _tema_iniziale(); qui solo i colori
        self._aggiorna_colori_tree()
        self.state('zoomed')

        # ── Chiude lo splash e mostra la finestra principale ──
        self._chiudi_splash(splash)

    def _trova_immagine_splash(self):
        """Cerca un file immagine splash nelle posizioni note. Restituisce il path o None."""
        import os, sys
        basi = []
        if getattr(sys, 'frozen', False):
            mei = getattr(sys, '_MEIPASS', '')
            if mei:
                basi.append(mei)
            basi.append(os.path.dirname(sys.executable))
        try:
            basi.append(os.path.dirname(os.path.abspath(__file__)))
        except NameError:
            pass
        basi.append(os.getcwd())
        candidati = []
        for base in basi:
            if not base:
                continue
            for nome in ("splash.png", "splash.bmp", "splash.jpg", "splash.jpeg",
                          "banner.png", "banner.bmp", "logo.png"):
                candidati.append(os.path.join(base, nome))
                candidati.append(os.path.join(base, "assets", nome))
                candidati.append(os.path.join(base, "resources", nome))
        for path in candidati:
            if os.path.isfile(path):
                return path

        return None

    def _imposta_icona(self):
        """Imposta l'icona dell'app nella barra del titolo e nella taskbar di Windows.

        Cerca 'adif_fzr.ico' accanto allo script/eseguibile. Su Windows imposta
        anche un AppUserModelID esplicito: senza di esso la taskbar mostra
        l'icona di python.exe invece di quella dell'app.
        """
        import sys, os

        # 1) AppUserModelID (solo Windows): fa sì che la taskbar usi la nostra
        #    icona invece di raggruppare sotto python.exe.
        if sys.platform.startswith("win"):
            try:
                import ctypes
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                    "IW1FZR.IW1FZR.ADIFFZR.Logbook.2.5")
            except Exception:
                pass

        # 2) Trova il file .ico. Gestisce sia l'esecuzione da sorgente sia il
        #    bundle PyInstaller (sys._MEIPASS) sia la cartella dell'eseguibile.
        #    Cerca sia nella cartella base sia in assets/ e resources/.
        candidati = []
        nomi = ("adif_fzr.ico", "icon.ico", "logo.ico", "app.ico", "adiffzr.ico")
        basi = []
        if getattr(sys, "frozen", False):
            # Eseguibile PyInstaller: risorse in _MEIPASS + cartella dell'exe
            basi.append(getattr(sys, "_MEIPASS", ""))
            basi.append(os.path.dirname(sys.executable))
        try:
            basi.append(os.path.dirname(os.path.abspath(__file__)))
        except NameError:
            pass
        basi.append(os.getcwd())
        for base in basi:
            if not base:
                continue
            for sub in ("", "assets", "resources"):
                for nome in nomi:
                    candidati.append(os.path.join(base, sub, nome) if sub
                                     else os.path.join(base, nome))

        ico = next((p for p in candidati if os.path.isfile(p)), None)
        if not ico:
            return  # nessuna icona trovata: resta quella di default

        # 3) Applica l'icona. iconbitmap gestisce la barra del titolo e (con
        #    l'AppUserModelID sopra) anche la taskbar su Windows.
        try:
            self.iconbitmap(ico)
        except Exception:
            pass
        # Ritenta dopo che la finestra è pronta: alcuni temi/OS ignorano la
        # prima chiamata se fatta troppo presto.
        try:
            self.after(300, lambda: self._riapplica_icona(ico))
        except Exception:
            pass
        self._ico_path = ico

    def _riapplica_icona(self, ico):
        """Riapplica l'icona dopo l'avvio (workaround per alcune configurazioni)."""
        try:
            self.iconbitmap(ico)
        except Exception:
            pass

    def _mostra_splash(self):
        """Mostra una finestra splash durante l'avvio.
        Se trova un file splash.png/bmp/jpg (nella cartella dell'app, in
        assets/ o resources/) lo mostra a schermo intero centrato.
        Altrimenti mostra il banner testuale con logo e nome app."""
        import tkinter as _tk

        splash = ctk.CTkToplevel(self)
        splash.overrideredirect(True)
        splash.attributes("-topmost", True)

        img_path = self._trova_immagine_splash() if _PIL_OK else None
        sw = splash.winfo_screenwidth()
        sh = splash.winfo_screenheight()

        if img_path:
            try:
                pil_img = Image.open(img_path)
                iw, ih = pil_img.size
                # Scala l'immagine mantenendo le proporzioni, max 70% schermo
                max_w, max_h = int(sw * 0.6), int(sh * 0.6)
                scale = min(max_w / iw, max_h / ih, 1.0)
                w, h = int(iw * scale), int(ih * scale)

                ctk_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(w, h))

                x = (sw - w) // 2
                y = (sh - h) // 2
                splash.geometry(f"{w}x{h}+{x}+{y}")
                splash.configure(fg_color="#0f1c2e")

                lbl = ctk.CTkLabel(splash, text="", image=ctk_img)
                lbl.pack(fill="both", expand=True)
                splash._ctk_img_ref = ctk_img  # evita garbage collection
            except Exception:
                img_path = None  # fallback al testo se l'immagine non si carica

        if not img_path:
            w, h = 420, 260
            x = (sw - w) // 2
            y = (sh - h) // 2
            splash.geometry(f"{w}x{h}+{x}+{y}")

            bg = "#0f1c2e"
            splash.configure(fg_color=bg)

            frame = ctk.CTkFrame(splash, fg_color=bg, corner_radius=0,
                                  border_width=1, border_color=TH.PRIMARY)
            frame.pack(fill="both", expand=True)

            ctk.CTkLabel(frame, text="📻", font=ctk.CTkFont(size=52)).pack(pady=(36,6))
            ctk.CTkLabel(frame, text=f"ADIF FZR {VERSIONE}",
                          font=ctk.CTkFont(size=24, weight="bold"),
                          text_color=TH.LINK).pack()
            ctk.CTkLabel(frame, text=T("dv_app_sottotit"),
                          font=ctk.CTkFont(size=12),
                          text_color="#63B3ED").pack(pady=(2,18))
            ctk.CTkLabel(frame, text=T("dv_caricamento"),
                          font=ctk.CTkFont(size=11),
                          text_color="#4A6FA5").pack()

        splash.update_idletasks()
        splash.update()
        # Fade in
        splash.attributes('-alpha', 0.0)
        def _fade_in(a=0.0):
            a = min(a + 0.08, 1.0)
            try: splash.attributes('-alpha', a)
            except Exception: return
            if a < 1.0:
                self.after(20, lambda: _fade_in(a))
        _fade_in()
        self._splash_start = datetime.now()
        return splash

    def _chiudi_splash(self, splash):
        """Chiude lo splash con fade out, poi mostra la finestra principale."""
        elapsed_ms = int((datetime.now() - self._splash_start).total_seconds() * 1000)
        min_display_ms = 2500
        remaining = max(0, min_display_ms - elapsed_ms)

        def finish():
            # Fade out
            def _fade_out(a=1.0):
                a = max(a - 0.08, 0.0)
                try: splash.attributes('-alpha', a)
                except Exception:
                    _show_main(); return
                if a > 0:
                    self.after(20, lambda: _fade_out(a))
                else:
                    _show_main()
            def _show_main():
                try: splash.destroy()
                except Exception: pass
                self.deiconify()
                self.state('zoomed')
                self.lift()
                self.focus_force()
                # Riapplica l'icona ora che la finestra è visibile: è il momento
                # in cui Windows ridisegna la taskbar.
                if getattr(self, "_ico_path", None):
                    self.after(200, lambda: self._riapplica_icona(self._ico_path))
                self.after(150, self._controlla_primo_avvio)
            _fade_out()

        self.after(remaining, finish)

    # ══════════════════════════════════════════
    #  NUOVA INTERFACCIA EXCEL STYLE
    # ══════════════════════════════════════════
    def crea_interfaccia(self):
        import tkinter as _tk
        import tkinter.ttk as _ttk

        # Ripristina i colori del treeview quando la finestra principale
        # riceve il focus (es. dopo aver chiuso o messo in background un
        # dialog etichette che ha contaminato il tema TTK globale)
        self.bind("<FocusIn>", self._on_main_focus)

        # ── 1. MENUBAR nativa ─────────────────────────────────
        self._crea_menubar()

        # Riferimenti toolbar per cambio lingua
        self._tb1_refs = []
        self._tb2_refs = []
        self._sidebar_section_refs = []
        self._sidebar_lang_refs = []
        # Controllo radio via OmniRig (lazy: si connette solo quando serve)
        self._omnirig_reale = OmniRigControl()   # istanza OmniRig persistente
        self._sdrconsole = None                  # istanza SDR Console (se usata)
        self._radio_backend = "omnirig"          # "omnirig" o "sdrconsole"
        # self._omnirig è il riferimento che il display USA: punta al backend
        # attivo. Di default OmniRig; diventa SDRConsoleControl se selezionato.
        self._omnirig = self._omnirig_reale
        self._display_radio_win = None   # finestra display radio in tempo reale
        # NOTA: le impostazioni OmniRig (percorsi, flag avvio) vengono applicate
        # in _applica_profilo_avvio, quando il profilo attivo è realmente noto.
        # Qui teniamo avvio_auto = False per sicurezza, così il polling della
        # barra radio non fa partire OmniRig prima di conoscere il flag reale.
        self._omnirig.avvio_auto = False
        # NOTA: l'avvio automatico di OmniRig NON parte qui (nell'__init__ il
        # profilo attivo non è ancora noto, quindi il flag sarebbe sbagliato).
        # Parte in _applica_profilo_avvio, dopo che il profilo è stato caricato.

        # ── TOOLBAR UNIFICATA (avvolge su più righe se la finestra è stretta) ──
        is_dark = ctk.get_appearance_mode().lower() == "dark"
        _TB_BG      = ("#EAEFF4","#131313")   # sfondo toolbar
        _BTN_NEU    = ("#D4DCE6","#253043")    # neutro — unico colore per tutti
        _BTN_NEU_HV = ("#B8C5D4","#2A2A2A")
        _BTN_POS    = _BTN_NEU
        _BTN_POS_HV = _BTN_NEU_HV
        _BTN_ACT    = _BTN_NEU
        _BTN_ACT_HV = _BTN_NEU_HV
        _BTN_WRN    = _BTN_NEU
        _BTN_WRN_HV = _BTN_NEU_HV
        _TXT        = ("#1A202C","#E2E8F0")
        _TXT_NEU    = ("#2D3748","#CBD5E0")
        _FONT_TB    = ctk.CTkFont(size=12)

        self._tb_wrap = _WrapToolbar(self, fg_color=_TB_BG)
        self._tb_wrap.pack(fill="x", side="top")
        def _tbtn(text, cmd, tooltip=None,
                  fc=_BTN_NEU, hc=_BTN_NEU_HV,
                  tc=_TXT_NEU, lang_key=None, emoji="", key=None):
            b = ctk.CTkButton(self._tb_wrap, text=text, command=cmd,
                              height=32, width=0,
                              fg_color=fc, hover_color=hc, text_color=tc,
                              font=_FONT_TB, corner_radius=5)
            if tooltip: _tip(b, tooltip)
            if lang_key:
                self._tb1_refs.append((b, lang_key, emoji))
            self._tb_wrap.add(b)
            return b

        def _tsep():
            s = ctk.CTkLabel(self._tb_wrap, text="",
                             width=1, height=32,
                             fg_color=("#B0BBC8","#2A2A2A"))
            self._tb_wrap.add(s, is_sep=True)

        # ── Definizioni disponibili riga 1 (per personalizzazione) ──
        self._tb1_disponibili = {
            "apri_adif":      ("📂 Apri",         self.sfoglia_file,    "Apri file ADIF  Ctrl+O",      _BTN_NEU, "📂 "),
            "unisci":         ("🔗 Unisci",        self.apri_unisci,     "Unisci più file ADIF",        _BTN_NEU, "🔗 "),
            "importa_cbr":    ("📥 Da CBR",        self.importa_cbr,     "Importa Cabrillo",            _BTN_NEU, "📥 "),
            "salva_adif":     ("💾 Salva",         self.salva_adif,      "Salva ADIF  Ctrl+S",          _BTN_POS, "💾 "),
            "aggiungi_qso":   ("➕ Aggiungi QSO",  self.apri_aggiungi_qso,"Inserisci QSO  Ctrl+N",      _BTN_POS, "➕ "),
            "filtri_qso":     ("⚗ Filtri",        self.apri_filtri,     "Filtra QSO",                  _BTN_NEU, "⚗ "),
            "duplicati":      ("🔍 Dupe Check",    self.apri_duplicati,  "Dupe Check",                  _BTN_WRN, "🔍 "),
            "deduci_country": ("🌍 Country",       self.deduci_country_da_nominativo, "Deduci Country dal prefisso", _BTN_NEU, "🌍 "),
        }
        # ── Definizioni disponibili riga 2 ──
        self._tb2_disponibili = {
            "genera_pdf":    ("📄 Genera PDF",     self.processa_e_salva,       "Genera PDF  Ctrl+P",  _BTN_ACT, "📄 "),
            "esporta_csv":   ("📊 CSV",            self.esporta_csv,            "Esporta CSV",          _BTN_NEU, "📊 "),
            "esporta_html":  ("🌐 HTML",           self.esporta_html,           "Esporta HTML Web",     _BTN_NEU, "🌐 "),
            "esporta_excel": ("📗 Excel",          self.esporta_excel,          "Esporta Excel",        _BTN_NEU, "📗 "),
            "qsl_card":      ("📮 Stampa Etichette", self.apri_qsl_card,        "Stampa etichette QSL", _BTN_NEU, "📮 "),
            "grafici":       ("📊 Grafici",        self.apri_grafici,           "Grafici attività",     _BTN_NEU, "📊 "),
            "qsl_designer":  ("🎨 Card Designer",  self.apri_qsl_card_designer, "QSL Card Designer",   _BTN_NEU, "🎨 "),
            "storico":       ("🕐 Recenti",        self.apri_storico,           "File recenti",         _BTN_NEU, "🕐 "),
            "dx_cluster":    ("📡 DX Cluster",     self.apri_dx_cluster,        "Spot DX in tempo reale", _BTN_NEU, "📡 "),
        }

        # Frame interni referenziati da _ricostruisci_toolbar1/2
        # Ora _tb_wrap è l'unica toolbar — i frame interni puntano ad essa
        self._tb1_frame = self._tb_wrap
        self._tb2_frame = self._tb_wrap

        def tb_btn1(parent, text, cmd, tooltip=None, color=None, lang_key=None, emoji=""):
            fc = color or _BTN_NEU
            hc = _BTN_NEU_HV if fc == _BTN_NEU else (
                _BTN_POS_HV if fc == _BTN_POS else
                _BTN_ACT_HV if fc == _BTN_ACT else _BTN_WRN_HV)
            return _tbtn(text, cmd, tooltip, fc, hc, _TXT_NEU, lang_key, emoji)

        def tb_btn2(parent, text, cmd, tooltip=None, color=None, lang_key=None, emoji=""):
            return tb_btn1(parent, text, cmd, tooltip, color, lang_key, emoji)

        self._tb_btn1_factory = tb_btn1
        self._tb_btn2_factory = tb_btn2

        # Costruisce toolbar riga 1
        self._ricostruisci_toolbar1_wrap()

        # Separatore visivo tra gruppo 1 e gruppo 2
        _tsep()

        # Costruisce toolbar riga 2
        self._ricostruisci_toolbar2_wrap()

        # Profilo — etichetta + bottone a destra, fuori dal wrap (fisso)
        frame_profilo = ctk.CTkFrame(self, fg_color=_TB_BG, corner_radius=0)
        frame_profilo.pack(fill="x", side="top")
        ctk.CTkLabel(frame_profilo, text=T("profili")+":",
                     font=ctk.CTkFont(size=11),
                     text_color=("#4A5568","#94A3B8")).pack(side="right", padx=(0,4), pady=4)
        self.btn_profili = ctk.CTkButton(
            frame_profilo, text="—", command=self.apri_gestione_profili,
            width=130, height=28,
            fg_color=_BTN_ACT, hover_color=_BTN_ACT_HV,
            text_color=("#FFFFFF","#FFFFFF"),
            font=ctk.CTkFont(size=11))
        self.btn_profili.pack(side="right", padx=6, pady=4)

        # ── 3. AREA PRINCIPALE (due colonne) ──────────────────
        # ── 3. AREA PRINCIPALE (due colonne) ──────────────────
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=0, pady=0)
        main.columnconfigure(0, weight=0, minsize=260)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(0, weight=1)

        # ── 3a. PANNELLO SINISTRO ──────────────────────────────
        left = ctk.CTkScrollableFrame(main, width=260, corner_radius=0,
                                      fg_color=("#F0F4F8","#161616"),
                                      border_width=0)
        left.grid(row=0, column=0, sticky="nsew")

        # Sezione operatore
        s1 = self._lbl_section(left, T("sez_operatore"))
        self._sidebar_section_refs.append((s1, "sez_operatore"))
        self._widget_refs['indicativo'] = ctk.CTkLabel(left, text=T("indicativo"), anchor="w")
        self._widget_refs['indicativo'].pack(fill="x", padx=10, pady=(0,2))
        self.entry_owner = ctk.CTkEntry(left, placeholder_text=T("sb_owner_ph"))
        self.entry_owner.pack(fill="x", padx=10, pady=(0,6))
        self._widget_refs['nome_op'] = ctk.CTkLabel(left, text=T("nome_op"), anchor="w")
        self._widget_refs['nome_op'].pack(fill="x", padx=10, pady=(0,2))
        self.entry_details = ctk.CTkEntry(left, placeholder_text=T("sb_details_ph"))
        self.entry_details.pack(fill="x", padx=10, pady=(0,10))

        # ── DISPLAY RADIO (frequenza dal vivo via OmniRig) ──────────
        _is_dark = ctk.get_appearance_mode().lower() == "dark"
        radio_box = ctk.CTkFrame(left, fg_color=("#0E2038", "#0A1A2E"),
                                 corner_radius=8)
        radio_box.pack(fill="x", padx=10, pady=(0, 10))
        # Riga stato: pallino + label RADIO + badge SPLIT
        _rtop = ctk.CTkFrame(radio_box, fg_color="transparent")
        _rtop.pack(fill="x", padx=10, pady=(6, 0))
        self.sb_radio_dot = ctk.CTkLabel(_rtop, text="●", font=ctk.CTkFont(size=12),
                                         text_color="#64748B")
        self.sb_radio_dot.pack(side="left")
        # Nome radio (es. "IC-7600"); se non nota mostra "RADIO"
        self.sb_radio_nome = ctk.CTkLabel(_rtop, text="RADIO",
                                          font=ctk.CTkFont(size=10, weight="bold"),
                                          text_color="#64748B")
        self.sb_radio_nome.pack(side="left", padx=(4, 0))
        # Badge SPLIT (a destra, visibile solo quando lo split è attivo)
        self.sb_radio_split = ctk.CTkLabel(_rtop, text="", font=ctk.CTkFont(size=10, weight="bold"),
                                           text_color="#FFFFFF", fg_color="transparent",
                                           corner_radius=4, width=48)
        self.sb_radio_split.pack(side="right")
        # Etichetta VFO-A
        self.sb_radio_lbl_a = ctk.CTkLabel(radio_box, text="VFO-A",
                                           font=ctk.CTkFont(size=9, weight="bold"),
                                           text_color="#64748B")
        self.sb_radio_lbl_a.pack(padx=10, pady=(4, 0))
        # Frequenza grande (VFO-A)
        self.sb_radio_freq = ctk.CTkLabel(radio_box, text="—.—————",
                                          font=ctk.CTkFont(size=26, weight="bold"),
                                          text_color=TH.OK_TEXT, cursor="hand2")
        self.sb_radio_freq.pack(padx=10, pady=(0, 0))
        self.sb_radio_freq.bind("<Button-1>", lambda e: self._logga_da_barra_radio())
        # Modo e banda piccoli
        _rbot = ctk.CTkFrame(radio_box, fg_color="transparent")
        _rbot.pack(pady=(0, 6))
        self.sb_radio_mode = ctk.CTkLabel(_rbot, text="—",
                                          font=ctk.CTkFont(size=13, weight="bold"),
                                          text_color=TH.PRIMARY)
        self.sb_radio_mode.pack(side="left", padx=(0, 10))
        self.sb_radio_band = ctk.CTkLabel(_rbot, text="—",
                                          font=ctk.CTkFont(size=13, weight="bold"),
                                          text_color="#FBBF24")
        self.sb_radio_band.pack(side="left")
        # Separatore + VFO-B (visibili solo in split)
        self.sb_radio_sep = ctk.CTkFrame(radio_box, height=1, fg_color="#1E3A5F")
        self.sb_radio_lbl_b = ctk.CTkLabel(radio_box, text="VFO-B  ·  TX",
                                           font=ctk.CTkFont(size=9, weight="bold"),
                                           text_color="#64748B")
        self.sb_radio_vfob = ctk.CTkLabel(radio_box, text="",
                                          font=ctk.CTkFont(size=18, weight="bold"),
                                          text_color="#F472B6")
        # (sep, lbl_b e vfob vengono packati/nascosti dinamicamente in _aggiorna_barra_radio)
        # Pulsanti controllo VFO: swap A<->B e toggle split
        _rctrl = ctk.CTkFrame(radio_box, fg_color="transparent")
        _rctrl.pack(fill="x", padx=8, pady=(0, 8))
        self.sb_btn_swap = ctk.CTkButton(_rctrl, text="⇄ A/B",
                                         font=ctk.CTkFont(size=11, weight="bold"),
                                         height=26, fg_color="#2D3748",
                                         hover_color="#4A5568",
                                         command=self._radio_vfo_swap)
        self.sb_btn_swap.pack(side="left", expand=True, fill="x", padx=(0, 3))
        self.sb_btn_split = ctk.CTkButton(_rctrl, text="SPLIT",
                                          font=ctk.CTkFont(size=11, weight="bold"),
                                          height=26, fg_color="#2D3748",
                                          hover_color="#4A5568",
                                          command=self._radio_toggle_split)
        self.sb_btn_split.pack(side="left", expand=True, fill="x", padx=(3, 0))

        # Sezione campi PDF — NASCOSTA dalla sidebar, visibile nel dialog Genera PDF
        grid_cb = ctk.CTkFrame(left, fg_color="transparent")
        # (non packato — i checkbox colonne appaiono solo nel dialog Genera PDF)
        for idx, (tag, info) in enumerate(self.campi_disponibili.items()):
            row, col = idx // 2, idx % 2
            cb = ctk.CTkCheckBox(grid_cb, text=info['nome'], variable=self.checkboxes[tag],
                                 font=ctk.CTkFont(size=11))
            cb.grid(row=row, column=col, padx=4, pady=2, sticky="w")
            self._checkbox_widgets[tag] = cb

        # FILTRI espansi direttamente nella sidebar
        s2 = self._lbl_section(left, T("sez_filtri"))
        self._sidebar_section_refs.append((s2, "sez_filtri"))

        frame_fil = ctk.CTkFrame(left, fg_color="transparent")
        frame_fil.pack(fill="x", padx=8, pady=(0,4))

        def _lbl(chiave):
            w = ctk.CTkLabel(frame_fil, text=T(chiave), font=ctk.CTkFont(size=10),
                         text_color="gray", anchor="w")
            w.pack(fill="x", pady=(3,0))
            self._sidebar_lang_refs.append((w, chiave, None))

        _lbl("sb_fil_data_dal")
        fr_da = ctk.CTkFrame(frame_fil, fg_color="transparent")
        fr_da.pack(fill="x")
        self._fil_data_da = ctk.CTkEntry(fr_da, placeholder_text="dd/mm/yyyy",
                                          font=ctk.CTkFont(size=10), height=26)
        self._fil_data_da.pack(side="left", expand=True, fill="x")
        ctk.CTkButton(fr_da, text="📅", width=28, height=26, fg_color=TH.PRIMARY,
                      command=lambda: CalendarPopup(self, self._fil_data_da)
                      ).pack(side="left", padx=(3,0))

        _lbl("sb_fil_data_al")
        fr_a = ctk.CTkFrame(frame_fil, fg_color="transparent")
        fr_a.pack(fill="x")
        self._fil_data_al = ctk.CTkEntry(fr_a, placeholder_text="dd/mm/yyyy",
                                          font=ctk.CTkFont(size=10), height=26)
        self._fil_data_al.pack(side="left", expand=True, fill="x")
        ctk.CTkButton(fr_a, text="📅", width=28, height=26, fg_color=TH.PRIMARY,
                      command=lambda: CalendarPopup(self, self._fil_data_al)
                      ).pack(side="left", padx=(3,0))

        _lbl("sb_fil_callsign")
        self._fil_call = ctk.CTkEntry(frame_fil, font=ctk.CTkFont(size=10), height=26)
        self._fil_call.pack(fill="x")

        _lbl("sb_fil_banda")
        self._fil_banda_var = ctk.StringVar(value=T("filtri_tutte"))
        self._fil_banda_om = ctk.CTkOptionMenu(frame_fil, variable=self._fil_banda_var,
                                                values=[T("filtri_tutte")], height=26,
                                                font=ctk.CTkFont(size=10))
        self._fil_banda_om.pack(fill="x")

        _lbl("sb_fil_modo")
        self._fil_modo_var = ctk.StringVar(value=T("filtri_tutti"))
        self._fil_modo_om = ctk.CTkOptionMenu(frame_fil, variable=self._fil_modo_var,
                                               values=[T("filtri_tutti")], height=26,
                                               font=ctk.CTkFont(size=10))
        self._fil_modo_om.pack(fill="x")

        _lbl("sb_fil_satellite")
        self._fil_sat_var = ctk.StringVar(value=T("filtri_tutti"))
        _sat_base = [T("filtri_tutti"),"QO-100","RS-44","AO-91","AO-92","AO-7","FO-29",
                     "SO-50","IO-117","CAS-4A","XW-2C","ISS"]
        self._fil_sat_om = ctk.CTkOptionMenu(frame_fil, variable=self._fil_sat_var,
                                              values=_sat_base, height=26,
                                              font=ctk.CTkFont(size=10))
        self._fil_sat_om.pack(fill="x")

        fr_btns = ctk.CTkFrame(frame_fil, fg_color="transparent")
        fr_btns.pack(fill="x", pady=(6,0))
        _btn_applica = ctk.CTkButton(fr_btns, text=T("sb_fil_applica"),
                      command=self._applica_filtri_sidebar,
                      height=28, fg_color=TH.PRIMARY, font=ctk.CTkFont(size=10)
                      )
        _btn_applica.pack(side="left", expand=True, fill="x", padx=(0,3))
        self._sidebar_lang_refs.append((_btn_applica, "sb_fil_applica", None))
        ctk.CTkButton(fr_btns, text="✕",
                      command=self._reset_filtri_sidebar,
                      height=28, width=32, fg_color=TH.WARNING_H,
                      font=ctk.CTkFont(size=10)).pack(side="left")

        self.lbl_filtri_attivi = ctk.CTkLabel(left, text="",
                                               font=ctk.CTkFont(size=9),
                                               text_color=TH.WARN_TEXT, anchor="w")
        self.lbl_filtri_attivi.pack(fill="x", padx=10, pady=(2,6))

        # ── 3b. PANNELLO DESTRO (griglia QSO) ─────────────────
        right = ctk.CTkFrame(main, corner_radius=0, fg_color="transparent")
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)

        # Mini toolbar sopra la griglia
        tg = ctk.CTkFrame(right, height=36, corner_radius=0,
                          fg_color=("#D1DBE8","#131313"))
        tg.grid(row=0, column=0, sticky="ew")
        tg.pack_propagate(False)

        _lbl_logqso = ctk.CTkLabel(tg, text=T("sb_log_qso"), font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=("#1A365D","#90CDF4"))
        _lbl_logqso.pack(side="left", padx=12)
        self._sidebar_lang_refs.append((_lbl_logqso, "sb_log_qso", None))
        self.lbl_status = ctk.CTkLabel(tg, text=T("nessun_file"),
                                       text_color="gray", font=ctk.CTkFont(size=11))
        self.lbl_status.pack(side="left", padx=8)

        # ── Indicatore selezione QSO + esporta selezione ──
        self.btn_sel_tutti = ctk.CTkButton(
            tg, text=T("sb_seleziona_tutti"), command=self._seleziona_tutti_visibili,
            height=24, width=120, font=ctk.CTkFont(size=11),
            fg_color="#4A5568", hover_color="#2D3748",
            corner_radius=4)
        self.btn_sel_tutti.pack(side="left", padx=(8,4))

        self.lbl_selezione = ctk.CTkLabel(tg, text="", text_color=TH.OK_TEXT,
                                          font=ctk.CTkFont(size=11, weight="bold"))
        self.lbl_selezione.pack(side="left", padx=8)

        self.btn_vedi_coda = ctk.CTkButton(
            tg, text=T("sb_vedi_coda"), command=self.apri_coda_export,
            height=24, width=90, font=ctk.CTkFont(size=11),
            fg_color="#4A5568", hover_color="#2D3748",
            corner_radius=4, state="disabled")
        self.btn_vedi_coda.pack(side="left", padx=(0,4))

        self.btn_esporta_sel = ctk.CTkButton(
            tg, text=T("sb_esporta_sel"), command=self._menu_esporta_selezione,
            height=24, width=150, font=ctk.CTkFont(size=11),
            fg_color=("#2F855A","#276749"), hover_color=("#276749","#22543D"),
            corner_radius=4, state="disabled")
        self.btn_esporta_sel.pack(side="left", padx=4)

        self.btn_elimina_sel = ctk.CTkButton(
            tg, text=T("sb_elimina_sel"), command=self._elimina_selezione,
            height=24, width=140, font=ctk.CTkFont(size=11),
            fg_color=("#C53030","#9B2C2C"), hover_color=("#9B2C2C","#822727"),
            corner_radius=4, state="disabled")
        self.btn_elimina_sel.pack(side="left", padx=(0,4))

        self.btn_undo_main = ctk.CTkButton(
            tg, text=T("sb_annulla"), command=self._undo_main,
            height=24, width=90, font=ctk.CTkFont(size=11),
            fg_color="#4A5568", hover_color="#2D3748",
            corner_radius=4, state="disabled")
        self.btn_undo_main.pack(side="left", padx=(0,4))
        self.bind("<Control-z>", lambda e: self._undo_main())
        self.bind("<Control-Z>", lambda e: self._undo_main())

        self.btn_desel = ctk.CTkButton(
            tg, text="✕", command=self._deseleziona_qso,
            height=24, width=28, font=ctk.CTkFont(size=11),
            fg_color="#4A5568", hover_color="#2D3748",
            corner_radius=4, state="disabled")
        self.btn_desel.pack(side="left", padx=(0,4))

        # Registra i pulsanti toolbar con testo tradotto (aggiornamento lingua)
        for _b, _k in ((self.btn_sel_tutti, "sb_seleziona_tutti"),
                       (self.btn_vedi_coda, "sb_vedi_coda"),
                       (self.btn_esporta_sel, "sb_esporta_sel"),
                       (self.btn_elimina_sel, "sb_elimina_sel"),
                       (self.btn_undo_main, "sb_annulla")):
            self._sidebar_lang_refs.append((_b, _k, None))

        # ── Ricerca testuale ──
        self.entry_search = ctk.CTkEntry(tg, placeholder_text=T("sb_cerca_ph"),
                                          width=240, height=24, font=ctk.CTkFont(size=11))
        self.entry_search.pack(side="right", padx=(6,4), pady=5)
        self.entry_search.bind("<KeyRelease>", lambda e: self._esegui_ricerca())
        self.btn_search_clear = ctk.CTkButton(
            tg, text="✕", command=self._pulisci_ricerca,
            height=24, width=28, font=ctk.CTkFont(size=11),
            fg_color="#4A5568", hover_color="#2D3748",
            corner_radius=4)
        self.btn_search_clear.pack(side="right", padx=(0,4), pady=5)

        ctk.CTkButton(tg, text="✕ Reset filtro", command=self._rimuovi_filtro_rapido,
                      height=24, width=100, font=ctk.CTkFont(size=11),
                      fg_color=("#C05621","#9C4221"), hover_color=("#9C4221","#7B3618"),
                      corner_radius=4).pack(side="right", padx=6, pady=5)
        self.lbl_filtri = ctk.CTkLabel(tg, text=T("nessun_filtro"),
                                       text_color=TH.PRIMARY, font=ctk.CTkFont(size=11))
        self.lbl_filtri.pack(side="right", padx=4)

        # Treeview Excel-style — adatta colori a tema
        import tkinter.ttk as _ttv
        import tkinter as _tk_root
        is_dark = ctk.get_appearance_mode().lower() == "dark"
        # Palette griglia curata (tema chiaro/scuro)
        if is_dark:
            bg_tree   = "#12181F"   # sfondo righe
            bg_even   = "#1A222B"   # riga alternata (sottile)
            fg_tree   = "#DCE3EC"   # testo
            bg_field  = "#12181F"
            head_bg   = "#1E2A38"   # header tenue (non blu pieno)
            head_fg   = "#9FB3C8"   # testo header morbido
            head_act  = "#2B3B4E"   # header hover
            sel_bg    = "#2B6CB0"   # selezione
            grid_line = "#232D38"
        else:
            bg_tree   = "#FFFFFF"
            bg_even   = "#F1F5FA"   # alternata tenue (era #E4EAF2 troppo forte)
            fg_tree   = "#1F2A37"
            bg_field  = "#FFFFFF"
            head_bg   = "#EDF2F7"   # header chiaro elegante (non blu pieno)
            head_fg   = "#3A4A61"   # testo header grigio-blu
            head_act  = "#DCE6F0"
            sel_bg    = "#2B6CB0"
            grid_line = "#E2E8F0"

        style = _ttv.Style()
        style.theme_use("clam")
        style.configure("QSO.Treeview",
                        background=bg_tree, foreground=fg_tree,
                        rowheight=28, fieldbackground=bg_field,
                        borderwidth=0, font=("Segoe UI", 10))
        style.configure("QSO.Treeview.Heading",
                        background=head_bg, foreground=head_fg,
                        font=("Segoe UI Semibold", 10), relief="flat",
                        borderwidth=0, padding=(6, 8))
        style.map("QSO.Treeview",
                  background=[("selected", sel_bg)],
                  foreground=[("selected", "#FFFFFF")])
        style.map("QSO.Treeview.Heading",
                  background=[("active", head_act)])

        cols = ("n","data","utc","call","nome","banda","sat","prop","modo","contest","freq","rst_s","rst_r","country","state","locator","lotw","eqsl","banda_rx","freq_rx","sat_mode")
        self.tree = _ttv.Treeview(right, columns=cols, show="headings",
                                  style="QSO.Treeview", selectmode="extended")

        headers = {"n":"#","data":"Data","utc":"UTC","call":"Callsign","nome":"Nome Op.",
                   "banda":"Banda","sat":"Satellite","prop":"Prop.",
                   "modo":"Mode","contest":"Contest","freq":"Freq.",
                   "rst_s":"RST TX","rst_r":"RST RX","country":"Country","state":"Stato",
                   "locator":"Locator","lotw":"LoTW","eqsl":"eQSL",
                   "banda_rx":"Banda RX","freq_rx":"Freq. RX","sat_mode":"SAT Mode"}
        widths =  {"n":38,"data":75,"utc":52,"call":90,"nome":100,"banda":55,"sat":80,"prop":55,
                   "modo":55,"contest":80,"freq":65,"rst_s":52,"rst_r":52,"country":150,"state":55,
                   "locator":65,"lotw":42,"eqsl":42,
                   "banda_rx":60,"freq_rx":70,"sat_mode":60}
        for c in cols:
            self.tree.heading(c, text=headers[c],
                              command=lambda _c=c: self._sort_tree(_c))
            self.tree.column(c, width=widths[c], anchor="center", minwidth=30)
        self.tree.column("call",    anchor="w")
        self.tree.column("nome",    anchor="w")
        self.tree.column("country", anchor="w")
        self.tree.column("sat",     anchor="w")
        self.tree.column("contest", anchor="w")
        # Nascondi colonne dinamiche finché non servono
        self.tree.column("state",   width=0, minwidth=0, stretch=False)
        self.tree.column("sat",     width=0, minwidth=0, stretch=False)
        self.tree.column("prop",    width=0, minwidth=0, stretch=False)
        self.tree.column("contest", width=0, minwidth=0, stretch=False)
        # Applica ordine/visibilità colonne salvato nel profilo
        self.after(100, self._applica_colonne)

        # Tag colorazione per modo
        is_dark = ctk.get_appearance_mode().lower() == "dark"
        self.tree.tag_configure("even",    background=bg_even)
        self.tree.tag_configure("odd",     background=bg_tree)
        self.tree.tag_configure("sel",     background="#2B6CB0")
        self.tree.tag_configure("ft8",     background="#1A3A2A" if is_dark else "#E6F4EA", foreground="#4ADE80" if is_dark else "#166534")
        self.tree.tag_configure("ft4",     background="#1A3A2A" if is_dark else "#E6F4EA", foreground="#4ADE80" if is_dark else "#166534")
        self.tree.tag_configure("cw",      background="#1A2A3A" if is_dark else "#EBF4FF", foreground="#90CDF4" if is_dark else "#1A4480")
        self.tree.tag_configure("ssb",     background=bg_even)
        self.tree.tag_configure("sat_row", background="#2A1A3A" if is_dark else "#FAF0FF", foreground="#D8B4FE" if is_dark else "#6B21A8")
        self.tree.tag_configure("eme",     background="#3A1A1A" if is_dark else "#FFF0F0", foreground="#FCA5A5" if is_dark else "#991B1B")

        # Ctrl+Click → filtro rapido (non interferisce con selezione normale)
        vsb = _ttv.Scrollbar(right, orient="vertical",   command=self.tree.yview)
        hsb = _ttv.Scrollbar(right, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        hsb.grid(row=2, column=0, sticky="ew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        # ── PANNELLO MODIFICA INLINE ───────────────────────────
        self._edit_panel = ctk.CTkFrame(right, corner_radius=0,
                                         fg_color=("#EBF4FF","#101010"),
                                         border_width=1,
                                         border_color=("#2B6CB0","#242424"))
        self._edit_panel.grid(row=3, column=0, columnspan=2, sticky="ew")
        right.rowconfigure(3, weight=0)

        # Intestazione pannello
        ep_head = ctk.CTkFrame(self._edit_panel, height=28, corner_radius=0,
                               fg_color=("#2B6CB0","#242424"))
        ep_head.pack(fill="x")
        ep_head.pack_propagate(False)
        self._widget_refs['ep_modifica'] = ctk.CTkLabel(ep_head, text="✏  " + T("ep_modifica"),
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="white")
        self._widget_refs['ep_modifica'].pack(side="left", padx=10)
        self._ep_modifica_lbl = ctk.CTkLabel(ep_head, text=T("ep_seleziona"),
                                     font=ctk.CTkFont(size=10),
                                     text_color=TH.LINK)
        self._ep_modifica_lbl.pack(side="left", padx=6)
        self._ep_info = self._ep_modifica_lbl

        # Campi di modifica — scrollabile orizzontalmente, costruiti dinamicamente
        ep_scroll = ctk.CTkScrollableFrame(self._edit_panel, height=68,
                                            orientation="horizontal",
                                            fg_color="transparent")
        ep_scroll.pack(fill="x", padx=8, pady=(4,0))
        self._ep_fields_frame = ep_scroll
        self._ep_entries = {}
        self._ep_field_labels = {}

        # Pulsanti azione
        ep_btns = ctk.CTkFrame(self._edit_panel, fg_color="transparent")
        ep_btns.pack(fill="x", padx=8, pady=(0,6))

        self._widget_refs['ep_applica'] = ctk.CTkButton(ep_btns, text=T("ep_applica"),
                      command=self._ep_applica, width=90, height=28,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                      font=ctk.CTkFont(size=11, weight="bold"))
        self._widget_refs['ep_applica'].pack(side="left", padx=(0,4))
        self._widget_refs['ep_duplica'] = ctk.CTkButton(ep_btns, text=T("ep_duplica"),
                      command=self._ep_duplica, width=80, height=28,
                      fg_color=TH.PRIMARY, hover_color=TH.PRIMARY,
                      font=ctk.CTkFont(size=11))
        self._widget_refs['ep_duplica'].pack(side="left", padx=4)
        self._widget_refs['ep_su'] = ctk.CTkButton(ep_btns, text=T("ep_su"),
                      command=self._ep_su, width=60, height=28,
                      fg_color="#4A5568", hover_color="#2D3748",
                      font=ctk.CTkFont(size=11))
        self._widget_refs['ep_su'].pack(side="left", padx=4)
        self._widget_refs['ep_giu'] = ctk.CTkButton(ep_btns, text=T("ep_giu"),
                      command=self._ep_giu, width=60, height=28,
                      fg_color="#4A5568", hover_color="#2D3748",
                      font=ctk.CTkFont(size=11))
        self._widget_refs['ep_giu'].pack(side="left", padx=4)
        self._widget_refs['ep_elimina'] = ctk.CTkButton(ep_btns, text=T("ep_elimina"),
                      command=self._ep_elimina, width=80, height=28,
                      fg_color=TH.WARNING_H, hover_color=TH.WARNING_H,
                      font=ctk.CTkFont(size=11))
        self._widget_refs['ep_elimina'].pack(side="left", padx=4)
        self._widget_refs['ep_colonna'] = ctk.CTkButton(ep_btns, text=T("ep_colonna"),
                      command=self._ep_colonna, width=90, height=28,
                      fg_color="#4A5568", hover_color="#2D3748",
                      font=ctk.CTkFont(size=11))
        self._widget_refs['ep_colonna'].pack(side="left", padx=4)
        self._widget_refs['ep_normalizza'] = ctk.CTkButton(ep_btns, text=T("ep_normalizza"),
                      command=self._ep_normalizza, width=100, height=28,
                      fg_color="#4A5568", hover_color="#2D3748",
                      font=ctk.CTkFont(size=11))
        self._widget_refs['ep_normalizza'].pack(side="left", padx=4)
        self._widget_refs['dist_calc_from_qso'] = ctk.CTkButton(ep_btns, text=T("dist_calc_from_qso"),
                      command=self._ep_calcola_distanza, width=140, height=28,
                      fg_color=TH.PRIMARY, hover_color=TH.PRIMARY,
                      font=ctk.CTkFont(size=11))
        self._widget_refs['dist_calc_from_qso'].pack(side="left", padx=4)

        # Selezione nel tree → popola pannello edit
        self._last_tree_sel = []   # selezione nativa salvata per _elimina_selezione
        self.tree.bind("<<TreeviewSelect>>", lambda e: [
            self._ep_on_select(e),
            setattr(self, '_last_tree_sel', list(self.tree.selection())),
            self.btn_elimina_sel.configure(
                state="normal" if (self.tree.selection() or self._coda_export) else "disabled")
        ])
        # Ctrl+Click → filtro rapido su cella dati, toggle selezione export su colonna "#"
        self.tree.bind("<Control-Button-1>", lambda e: self._on_tree_click(e))
        # Shift+Click sulla colonna "#" → seleziona un range (dall'ultimo Ctrl+Click)
        self.tree.bind("<Shift-Button-1>", lambda e: self._on_tree_shift_click(e))
        # Click destro → menu contestuale
        self.tree.bind("<Button-3>", lambda e: self._on_tree_right_click(e))
        self._ep_idx = None  # indice QSO corrente

        # ── 4. STATUS BAR ─────────────────────────────────────
        sb = ctk.CTkFrame(self, height=26, corner_radius=0,
                          fg_color=("#1A365D","#0D0D0D"))
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)

        self.sb_qso  = ctk.CTkLabel(sb, text=T("dv_qso_dash"),
                                    font=ctk.CTkFont(size=10), text_color=TH.LINK)
        self.sb_qso.pack(side="left", padx=12)
        self.sb_dxcc = ctk.CTkLabel(sb, text=T("dv_dxcc_dash"),
                                    font=ctk.CTkFont(size=10), text_color=TH.LINK)
        self.sb_dxcc.pack(side="left", padx=8)
        self.sb_band = ctk.CTkLabel(sb, text=T("dv_bande_dash"),
                                    font=ctk.CTkFont(size=10), text_color=TH.LINK)
        self.sb_band.pack(side="left", padx=8)
        self.sb_filt = ctk.CTkLabel(sb, text="",
                                    font=ctk.CTkFont(size=10), text_color=TH.WARN_TEXT)
        self.sb_filt.pack(side="left", padx=8)
        ctk.CTkLabel(sb, text=f"{APP_TITOLO} · build {BUILD_DATE}  |  iw1fzr.it",
                     font=ctk.CTkFont(size=10),
                     text_color="#4A6FA5").pack(side="right", padx=12)
        # Avvia il polling della radio per il display nella sidebar
        self._radiobar_job = None
        self._radiobar_ultimo = {}
        self.after(1500, self._aggiorna_barra_radio)

        # Sort state
        self._sort_col = None
        self._sort_rev = False

    # ── Helper sezione sidebar ─────────────────
    def _lbl_section(self, parent, text):
        f = ctk.CTkFrame(parent, height=24, corner_radius=0,
                         fg_color=("#CBD5E0","#222222"))
        f.pack(fill="x", pady=(8,4))
        f.pack_propagate(False)
        lbl = ctk.CTkLabel(f, text=text, font=ctk.CTkFont(size=10, weight="bold"),
                     text_color=("#4A5568","#90CDF4"),
                     anchor="w")
        lbl.pack(side="left", padx=8)
        return lbl

    # ── Riempie il Treeview con i QSO caricati ─
    def _aggiorna_tree(self):
        self.tree.delete(*self.tree.get_children())
        qsos = self._qsos_attivi()

        # Campi presenti in almeno un QSO dell'intero log (non solo quelli
        # filtrati/visibili): usato dal pannello di modifica rapida per
        # mostrare sempre editabile un campo, anche se vuoto per il QSO
        # selezionato, purché compaia in almeno un altro QSO del log.
        self._campi_presenti_nel_log = set()
        for q in self.qsos_caricati:
            for k, v in q.items():
                if str(v).strip():
                    self._campi_presenti_nel_log.add(k.lower())

        # Controlla colonne dinamiche
        has_sat     = any(str(q.get('sat_name','')).strip()    for q in qsos)
        has_prop    = any(str(q.get('prop_mode','')).strip()   for q in qsos)
        has_contest = any(str(q.get('contest_id','')).strip()  for q in qsos)
        has_state   = any(str(q.get('state','')).strip()       for q in qsos)

        self.tree.column("sat",     width=80 if has_sat     else 0, minwidth=0 if not has_sat     else 30, stretch=False)
        self.tree.column("prop",    width=55 if has_prop    else 0, minwidth=0 if not has_prop    else 30, stretch=False)
        self.tree.column("contest", width=90 if has_contest else 0, minwidth=0 if not has_contest else 30, stretch=False)
        self.tree.column("state",   width=55 if has_state   else 0, minwidth=0 if not has_state   else 30, stretch=False)

        # Aggiorna displaycolumns: le colonne dinamiche (sat/prop/contest/state)
        # appaiono automaticamente se ci sono dati, sparendo altrimenti.
        # Rispetta l'ordine e le altre scelte di visibilità del profilo.
        try:
            ordine, nascoste = self._ordine_colonne()
            presenza = {"sat": has_sat, "prop": has_prop,
                        "contest": has_contest, "state": has_state}
            visibili = []
            for c in ordine:
                if c in presenza:
                    # colonna dinamica: visibile solo se ci sono dati
                    if presenza[c]:
                        visibili.append(c)
                elif c not in nascoste:
                    visibili.append(c)
            if visibili:
                self.tree.configure(displaycolumns=visibili)
        except Exception:
            pass

        is_dark = ctk.get_appearance_mode().lower() == "dark"

        for i, qso in enumerate(qsos):
            data = str(qso.get('qso_date',''))
            if len(data)==8:
                data = f"{data[6:8]}/{data[4:6]}/{data[0:4]}"
            ora = str(qso.get('time_on',''))
            if len(ora)>=4:
                ora = ora[:2]+":"+ora[2:4]
            lotw   = str(qso.get('lotw_qsl_rcvd','')).upper().strip()
            eqsl   = str(qso.get('eqsl_qsl_rcvd','')).upper().strip()
            lotw_d = str(qso.get('lotw_qslrdate','')).strip()
            eqsl_d = str(qso.get('eqsl_qslrdate','')).strip()
            # La data fa da fallback SOLO se il campo è vuoto (non se è "N")
            lotw_ok = (lotw in ('Y', 'V')) or (not lotw and lotw_d and lotw_d != '00000000')
            eqsl_ok = (eqsl == 'Y') or (not eqsl and eqsl_d and eqsl_d != '00000000')
            sat     = str(qso.get('sat_name','')).strip()
            prop    = str(qso.get('prop_mode','')).strip().upper()
            contest = str(qso.get('contest_id','')).strip()
            modo    = str(qso.get('mode','')).upper().strip()
            submode = str(qso.get('submode','')).upper().strip()
            # Per i log con MODE=MFSK e SUBMODE=FT4/FT8/FT2 ecc., mostra il
            # submode accanto al modo (es. "MFSK (FT4)"), come già avviene
            # per Banda/Satellite. Se submode è vuoto o coincide col modo,
            # non aggiunge nulla.
            modo_display = f"{modo} ({submode})" if submode and submode != modo else modo

            # Determina tag colore riga: riconosce FT8/FT4 sia quando sono
            # il MODE diretto, sia quando sono il SUBMODE di un MODE=MFSK
            # (caso comune nei log che usano ADIF con submode esplicito).
            modo_per_colore = submode if (modo == "MFSK" and submode) else modo
            if not self.var_colora_righe.get():
                tag = "even" if i%2==0 else "odd"
            elif sat:
                tag = "sat_row"
            elif prop in ('EME','MS','METEOR','AURORA','FAI','SPORADIC-E','ES'):
                tag = "eme"
            elif modo_per_colore in ('FT8',):
                tag = "ft8"
            elif modo_per_colore in ('FT4','FT2'):
                tag = "ft4"
            elif modo_per_colore in ('CW',):
                tag = "cw"
            else:
                tag = "even" if i%2==0 else "odd"

            # "#" mostra ✓ se il QSO è già in coda export (persiste tra filtri/ricerche)
            numero = "✓" if self._trova_in_coda(qso) != -1 else i+1

            self.tree.insert("", "end", iid=str(i), tags=(tag,), values=(
                numero, data, ora,
                str(qso.get('call','')).upper(),
                str(qso.get('name','')),
                str(qso.get('band','')).upper(),
                sat, prop, modo_display, contest,
                str(qso.get('freq','')),
                str(qso.get('rst_sent','')),
                str(qso.get('rst_rcvd','')),
                str(qso.get('country','')),
                str(qso.get('state','')).upper(),
                str(qso.get('gridsquare','')).upper(),
                "Y" if lotw_ok else "—",
                "Y" if eqsl_ok else "—",
                str(qso.get('band_rx','')).upper(),
                str(qso.get('freq_rx','')),
                str(qso.get('sat_mode','')).upper(),
            ))
        self._aggiorna_statusbar()
        self._ultima_riga_toggle = None
        if hasattr(self, 'lbl_selezione'):
            self._aggiorna_indicatore_selezione()
        self._aggiorna_sidebar_filtri()

    def _aggiorna_statusbar(self):
        qsos = self._qsos_attivi()
        tot = len(self.qsos_caricati)
        vis = len(qsos)
        stats = self.calcola_statistiche(qsos)
        self.sb_qso.configure(text=f"QSO: {vis}" + (f"/{tot}" if vis!=tot else ""))
        self.sb_dxcc.configure(text=f"DXCC: {stats['dxcc']}")
        self.sb_band.configure(text=f"Bande: {stats['bande']}")
        if vis != tot:
            self.sb_filt.configure(text=f"[Filtro attivo: {vis}/{tot}]")
        else:
            self.sb_filt.configure(text="")

    def _aggiorna_barra_radio(self):
        """Aggiorna il display radio nella sidebar leggendo da OmniRig.
        Un solo loop, sempre attivo, a intervalli di 1s. Non invasivo se la
        radio non è disponibile."""
        try:
            if not self.winfo_exists():
                return
        except Exception:
            return
        if not hasattr(self, "sb_radio_freq"):
            # widget non ancora creati: riprova più tardi
            try:
                self._radiobar_job = self.after(1000, self._aggiorna_barra_radio)
            except Exception:
                pass
            return
        rig = getattr(self, "_omnirig", None)
        freq_txt = "—.—————"
        mode_txt = "—"
        band_txt = "—"
        vfob_txt = ""
        split_attivo = False
        dot_col = "#64748B"
        freq_col = "#4ADE80"
        if rig is not None and rig.disponibile():
            try:
                fa, fb = rig.get_freq_ab()
                ma = rig.get_modo()
                if fa:
                    # Nome radio: leggilo una volta e memorizzalo (non cambia)
                    if not getattr(self, "_radio_nome_cache", None):
                        try:
                            nome = rig.nome_radio()
                            if nome:
                                self._radio_nome_cache = nome
                        except Exception:
                            pass
                    banda = self._banda_da_freq(fa/1_000_000) or ""
                    freq_txt = f"{fa/1_000_000:.5f}"
                    mode_txt = ma if ma else "—"
                    band_txt = banda if banda else "—"
                    dot_col = "#4ADE80"
                    # Stato split: usa la proprietà .Split (affidabile), con
                    # fallback sullo scostamento VFO.
                    try:
                        st = rig.is_split()
                    except Exception:
                        st = None
                    if st is None:
                        st = bool(fb and abs(fb - fa) > 100)
                    split_attivo = bool(st)
                    # VFO-B: mostralo SEMPRE se la radio espone FreqB
                    # (acceso in split, grigio spento in simplex).
                    is_sdrconsole = (getattr(self, "_radio_backend", "omnirig")
                                     == "sdrconsole")
                    if is_sdrconsole and fa and (2_400_000_000 <= fa <= 2_410_000_000):
                        # SDR Console su QO-100: la freq letta è la TX (uplink,
                        # banda 2400). Calcolo la RX (downlink) = TX + 8089.5 MHz
                        # (offset fisso del transponder QO-100, non invertente).
                        rx_hz = fa + 8_089_500_000
                        vfob_txt = f"{rx_hz/1_000_000:.5f}"
                        self._radiobar_rx_calc = True
                    elif fb:
                        vfob_txt = f"{fb/1_000_000:.5f}"
                        self._radiobar_rx_calc = False
                    else:
                        self._radiobar_rx_calc = False
                    self._radiobar_ultimo = {"hz": fa, "modo": ma, "banda": banda,
                                             "hz_b": fb, "split": split_attivo}
                else:
                    self._radiobar_ultimo = {}
            except Exception:
                self._radiobar_ultimo = {}
        else:
            dot_col = "#64748B"
            freq_col = "#64748B"
            self._radio_nome_cache = None   # reset: radio non connessa
        try:
            self.sb_radio_dot.configure(text_color=dot_col)
            # Nome radio: mostra quello memorizzato, o "RADIO" se sconosciuto
            _nome = getattr(self, "_radio_nome_cache", None)
            self.sb_radio_nome.configure(text=_nome if _nome else "RADIO",
                                         text_color=dot_col)
            self.sb_radio_freq.configure(text=freq_txt, text_color=freq_col)
            self.sb_radio_mode.configure(text=mode_txt)
            self.sb_radio_band.configure(text=band_txt)
            # Badge SPLIT
            if split_attivo:
                self.sb_radio_split.configure(text="SPLIT", fg_color="#DD6B20")
            else:
                self.sb_radio_split.configure(text="", fg_color="transparent")
            # Colore del pulsante SPLIT: acceso se attivo
            try:
                if split_attivo:
                    self.sb_btn_split.configure(fg_color="#DD6B20", hover_color=TH.WARNING_H)
                else:
                    self.sb_btn_split.configure(fg_color="#2D3748", hover_color="#4A5568")
            except Exception:
                pass
            # VFO-B: mostralo sempre se la radio espone FreqB.
            # Acceso (rosa + etichetta TX) quando split attivo,
            # grigio spento quando simplex.
            if vfob_txt:
                self.sb_radio_vfob.configure(text=vfob_txt)
                if getattr(self, "_radiobar_rx_calc", False):
                    # SDR Console / QO-100: la riga mostra la RX (downlink)
                    # calcolata dalla TX. Sempre accesa, in azzurro.
                    self.sb_radio_vfob.configure(text_color=TH.PRIMARY)
                    self.sb_radio_lbl_b.configure(text="RX · downlink",
                                                  text_color=TH.PRIMARY)
                    # E l'etichetta della freq principale diventa "TX · uplink"
                    self.sb_radio_lbl_a.configure(text="TX · uplink")
                elif split_attivo:
                    self.sb_radio_vfob.configure(text_color="#F472B6")
                    self.sb_radio_lbl_b.configure(text="VFO-B  ·  TX",
                                                  text_color="#F472B6")
                    self.sb_radio_lbl_a.configure(text="VFO-A")
                else:
                    # Grigio spento: VFO-B presente ma non in uso
                    self.sb_radio_vfob.configure(text_color="#4A5568")
                    self.sb_radio_lbl_b.configure(text="VFO-B",
                                                  text_color="#3A4A5F")
                    self.sb_radio_lbl_a.configure(text="VFO-A")
                if not self.sb_radio_sep.winfo_ismapped():
                    self.sb_radio_sep.pack(fill="x", padx=14, pady=(2, 4))
                    self.sb_radio_lbl_b.pack(padx=10, pady=(0, 0))
                    self.sb_radio_vfob.pack(padx=10, pady=(0, 8))
            else:
                # Nessun FreqB esposto: nascondi la sezione
                if self.sb_radio_sep.winfo_ismapped():
                    self.sb_radio_sep.pack_forget()
                    self.sb_radio_lbl_b.pack_forget()
                    self.sb_radio_vfob.pack_forget()
        except Exception:
            pass
        # Riprogramma il prossimo aggiornamento
        try:
            self._radiobar_job = self.after(1000, self._aggiorna_barra_radio)
        except Exception:
            pass

    def _radio_vfo_swap(self):
        """Scambia VFO-A e VFO-B sulla radio."""
        rig = getattr(self, "_omnirig", None)
        if rig is None or not rig.disponibile():
            return
        try:
            rig.vfo_swap()
        except Exception:
            pass
        # Aggiorna subito il display (senza aspettare il prossimo ciclo)
        try:
            self.after(150, self._aggiorna_barra_radio)
        except Exception:
            pass

    def _imposta_backend_radio(self, backend, porta_sdr=None):
        """Seleziona la sorgente del display radio: 'omnirig' o 'sdrconsole'.
        Il display usa self._omnirig, che qui viene fatto puntare al backend
        scelto. Grazie all'interfaccia comune, il resto dell'app non cambia."""
        self._radio_backend = backend
        if backend == "sdrconsole":
            if self._sdrconsole is None:
                self._sdrconsole = SDRConsoleControl(porta=porta_sdr or "COM11")
            elif porta_sdr:
                self._sdrconsole.imposta_porta(porta_sdr)
            self._omnirig = self._sdrconsole
            # Il nome radio si aggiorna al prossimo giro di polling
            self._radio_nome_cache = None
        else:
            self._omnirig = self._omnirig_reale
            self._radio_nome_cache = None

    def _radio_toggle_split(self):
        """Attiva/disattiva lo split sulla radio."""
        rig = getattr(self, "_omnirig", None)
        if rig is None or not rig.disponibile():
            return
        try:
            rig.toggle_split()
        except Exception:
            pass
        try:
            self.after(150, self._aggiorna_barra_radio)
        except Exception:
            pass

    def _logga_da_barra_radio(self):
        """Click sulla barra radio: apre Aggiungi QSO precompilando la
        frequenza/banda/modo correnti letti dalla radio."""
        d = getattr(self, "_radiobar_ultimo", None)
        if not d or not d.get("hz"):
            return
        # Riusa la logica del display: imposta _display_ultimo e chiama _logga
        self._display_ultimo = dict(d)
        self._logga_da_display()

    def _sort_tree(self, col):
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        rev = (self._sort_col == col and not self._sort_rev)
        try:
            data.sort(key=lambda x: (float(x[0]) if x[0].replace('.','',1).isdigit() else x[0].lower()), reverse=rev)
        except Exception:
            data.sort(reverse=rev)
        for i,(_, k) in enumerate(data):
            self.tree.move(k, "", i)
            tag = "even" if i%2==0 else "odd"
            self.tree.item(k, tags=(tag,))
        self._sort_col = col
        self._sort_rev = rev

    # ── Pannello modifica inline ───────────────
    def _on_tree_click(self, event):
        """Ctrl+Click su cella → filtro rapido per quel valore.
        Ctrl+Click sulla colonna '#' → toggle QSO nella coda export (persistente tra filtri)."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id or not col_id:
            return
        # Ricava il NOME REALE della colonna a questa posizione visiva.
        # Non usare una tupla statica: le colonne possono essere riordinate
        # (profilo) o nascoste (width=0), sfasando gli indici. tree.column(id)
        # con l'identificatore #N ritorna il nome vero della colonna lì.
        try:
            col_name = self.tree.column(col_id, "id")
        except Exception:
            col_name = ""
        if not col_name:
            # Fallback: mappa posizionale sulle colonne di visualizzazione reali
            try:
                dispcols = self.tree["displaycolumns"]
                if dispcols == ("#all",) or not dispcols:
                    dispcols = self.tree["columns"]
                idx = int(col_id.replace('#', '')) - 1
                if 0 <= idx < len(dispcols):
                    col_name = dispcols[idx]
            except Exception:
                return
        if not col_name:
            return

        # ── Colonna "#" → toggle QSO nella coda export, indipendente da tree.selection() ──
        if col_name == "n":
            qsos = self._qsos_attivi()
            idx = int(row_id)
            if 0 <= idx < len(qsos):
                qso = qsos[idx]
                added = self._toggle_coda_export(qso)
                self.tree.set(row_id, "n", "✓" if added else str(idx + 1))
            self._ultima_riga_toggle = row_id
            self._aggiorna_indicatore_selezione()
            return "break"

        col_to_adif = {
            "banda":"band","modo":"mode","country":"country","state":"state",
            "sat":"sat_name","prop":"prop_mode","contest":"contest_id",
            "locator":"gridsquare","call":"call","lotw":"lotw_qsl_rcvd","eqsl":"eqsl_qsl_rcvd"
        }
        if col_name not in col_to_adif:
            return

        val = self.tree.set(row_id, col_name)
        if not val or val == "—":
            return

        adif_field = col_to_adif[col_name]
        self.qsos_filtrati = [
            q for q in self.qsos_caricati
            if str(q.get(adif_field,'')).upper().strip() == val.upper().strip()
        ]
        n = len(self.qsos_filtrati)
        tot = len(self.qsos_caricati)
        self.lbl_filtri.configure(
            text=f"[{col_name.upper()}={val}  {n}/{tot} QSO — Ctrl+Click per rimuovere]",
            text_color=TH.WARN_TEXT
        )
        self.sb_filt.configure(text=f"Filtro rapido: {col_name.upper()}={val}")
        self._aggiorna_tree()
        return "break"

    # ── Shift+Click sulla colonna "#" → aggiunge un range alla coda export ──
    # ── Menu contestuale treeview (click destro) ───────────────

    def _on_tree_right_click(self, event):
        """Mostra menu contestuale con opzioni per il QSO sotto il cursore."""
        import tkinter as _tk
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        self.tree.selection_set(row_id)
        try:
            idx = int(row_id)
            qso = self._qsos_attivi()[idx]
        except (ValueError, IndexError):
            return

        call = str(qso.get('call', '')).upper().strip()
        menu = _tk.Menu(self, tearoff=0)
        menu.add_command(
            label=f"🔍 Cerca QSL Manager per {call or '?'}  [IK3QAR]",
            command=lambda: self._cerca_qsl_manager(call, idx),
            state="normal" if call else "disabled")
        menu.add_command(
            label=f"📡 Info stazione  [HamQTH]",
            command=lambda: self._cerca_hamqth(call, idx),
            state="normal" if call else "disabled")
        menu.add_separator()
        menu.add_command(
            label=T("menu_copia_call"),
            command=lambda: (self.clipboard_clear(), self.clipboard_append(call)))
        menu.add_command(
            label="🌐 Apri su QRZ.com",
            command=lambda: self._apri_url(f"https://www.qrz.com/db/{call}"),
            state="normal" if call else "disabled")
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _apri_url(self, url):
        import webbrowser
        webbrowser.open(url)

    def _cerca_hamqth(self, call, qso_idx=None):
        """Lookup callsign su HamQTH e mostra la finestra info."""
        if not call:
            return
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        hq_user = dati.get("hamqth_username", "").strip()
        hq_pass = dati.get("hamqth_password", "").strip()
        if not (hq_user and hq_pass):
            messagebox.showwarning(T("attenzione"), T("hqth_no_config"))
            return

        client = HamQTHClient(hq_user, hq_pass)
        info, err = client.lookup(call)

        if err == "NOT_FOUND":
            messagebox.showinfo(T("hqth_titolo"), T("hqth_non_trovato"), parent=self)
            return
        if err:
            messagebox.showerror("HamQTH", T("hqth_login_err", msg=err), parent=self)
            return
        HamQTHDialog(self, info, call, qso_idx)


    def _hint_hamqth(self):
        """Apre HamQTH lookup per il QSO selezionato, o mostra suggerimento."""
        sel = self.tree.selection()
        if sel:
            try:
                idx = int(sel[0])
                qso = self._qsos_attivi()[idx]
                call = str(qso.get('call', '')).upper().strip()
                if call:
                    self._cerca_hamqth(call, idx)
                    return
            except Exception:
                pass
        messagebox.showinfo("📡 HamQTH",
            "Seleziona un QSO nella griglia, poi:\n\n"
            "  • Click DESTRO sul QSO → 'Info stazione [HamQTH]'\n\n"
            "Configura le credenziali HamQTH in:\n"
            "  Strumenti → Profili operatore → Modifica",
            parent=self)

    def _hint_cerca_manager(self):
        """Apre il lookup QSL Manager per il QSO attualmente selezionato,
        oppure mostra un suggerimento se nessun QSO è selezionato."""
        sel = self.tree.selection()
        if sel:
            try:
                idx = int(sel[0])
                qso = self._qsos_attivi()[idx]
                call = str(qso.get('call', '')).upper().strip()
                if call:
                    self._cerca_qsl_manager(call, idx)
                    return
            except Exception:
                pass
        messagebox.showinfo(
            "🔍 Cerca QSL Manager",
            "Seleziona un QSO nella griglia, poi:\n\n"
            "  • Click DESTRO sul QSO → 'Cerca QSL Manager'\n\n"
            "  • Oppure usa 'Stampa QSL' → 'Cerca manager IK3QAR'\n"
            "    per cercare il manager di tutti i QSO selezionati in batch.",
            parent=self)
        if sel:
            try:
                idx = int(sel[0])
                qso = self._qsos_attivi()[idx]
                call = str(qso.get('call', '')).upper().strip()
                if call:
                    self._cerca_qsl_manager(call, idx)
                    return
            except Exception:
                pass
        messagebox.showinfo(
            "🔍 Cerca QSL Manager",
            "Seleziona un QSO nella griglia, poi:\n\n"
            "  • Click DESTRO sul QSO → 'Cerca QSL Manager'\n\n"
            "  • Oppure usa 'Stampa QSL' → 'Cerca manager IK3QAR'\n"
            "    per cercare il manager di tutti i QSO selezionati in batch.",
            parent=self)


    def _cerca_qsl_manager(self, call, qso_idx):
        """Interroga IK3QAR per il callsign e mostra il QSL Manager."""
        if not call:
            return
        url = f"https://www.ik3qar.it/manager/man_result.php?call={urllib.parse.quote(call)}"
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "ADIF-FZR/2.3 (+https://iw1fzr.it)"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                html = resp.read().decode("iso-8859-1", errors="replace")
            righe = re.findall(
                r'<td[^>]*>\s*(?:<[^>]+>)*([A-Z0-9/]+)(?:</[^>]+>)*\s*</td>'
                r'\s*<td[^>]*>\s*(?:<[^>]+>)*([A-Z0-9/]+)(?:</[^>]+>)*\s*</td>'
                r'\s*<td[^>]*>\s*(\d{4})\s*</td>'
                r'\s*<td[^>]*>\s*(.*?)\s*</td>',
                html, re.IGNORECASE | re.DOTALL)
            risultati = [(m, y, re.sub('<[^>]+>', '', i).strip())
                         for c, m, y, i in righe if c.upper() == call]
        except Exception as ex:
            messagebox.showerror("Errore lookup",
                f"Impossibile contattare IK3QAR:\n{ex}", parent=self)
            return
        self._mostra_qsl_manager(call, risultati, qso_idx, url)

    def _mostra_qsl_manager(self, call, risultati, qso_idx, url_sorgente, nota_fonte="IK3QAR"):
        """Finestra risultati QSL Manager lookup."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(f"QSL Manager — {call}")
        dlg.geometry("480x340")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text=f"🔍 QSL Manager per  {call}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(14,4), padx=20)
        ctk.CTkLabel(dlg, text=f"Fonte: {nota_fonte}",
                     font=ctk.CTkFont(size=9), text_color="gray").pack()

        if not risultati:
            ctk.CTkLabel(dlg, text=T("dv_no_mgr"),
                         font=ctk.CTkFont(size=11), text_color=TH.WARN_TEXT).pack(pady=20)
            ctk.CTkButton(dlg, text="🌐 Apri pagina web", height=30,
                          command=lambda: self._apri_url(url_sorgente)
                          ).pack(pady=4)
            ctk.CTkButton(dlg, text=T("cm_chiudi"), fg_color="#718096",
                          height=30, command=dlg.destroy).pack(pady=4)
            return

        # Mostra tutti i risultati (più recenti prima)
        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=16, pady=8)

        manager_piu_recente = risultati[-1][0] if risultati else ""

        for manager, anno, info in reversed(risultati):
            row = ctk.CTkFrame(scroll, fg_color="#141414" if manager == manager_piu_recente else "transparent",
                               corner_radius=6)
            row.pack(fill="x", pady=2, padx=2)
            testo = f"  {anno}  —  Via: {manager}"
            if info:
                testo += f"   ({info})"
            ctk.CTkLabel(row, text=testo,
                         font=ctk.CTkFont(size=11, weight="bold" if manager == manager_piu_recente else "normal"),
                         text_color=TH.OK_TEXT if manager == manager_piu_recente else "gray",
                         anchor="w").pack(side="left", padx=8, pady=6, fill="x", expand=True)
            ctk.CTkButton(row, text=T("pref_usa"), width=50, height=26,
                          fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                          font=ctk.CTkFont(size=10),
                          command=lambda m=manager, d=dlg: self._applica_qsl_manager(m, qso_idx, d)
                          ).pack(side="right", padx=8, pady=4)

        # Pulsante applica il più recente
        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=16, pady=(0,12))
        if manager_piu_recente:
            ctk.CTkButton(frame_btn,
                text=f"✔ Applica più recente: {manager_piu_recente}",
                command=lambda: self._applica_qsl_manager(manager_piu_recente, qso_idx, dlg),
                fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS, height=34,
                font=ctk.CTkFont(size=11, weight="bold")).pack(
                side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(frame_btn, text="🌐 Web", width=60, height=34,
                      fg_color=TH.PRIMARY,
                      command=lambda: self._apri_url(url_sorgente)).pack(side="left", padx=(0,6))
        ctk.CTkButton(frame_btn, text=T("cm_chiudi"), width=70, height=34,
                      fg_color="#718096", command=dlg.destroy).pack(side="left")

    def _applica_qsl_manager(self, manager, qso_idx, dlg):
        """Scrive il QSL Manager nel campo QSL_VIA del QSO."""
        try:
            qso = self._qsos_attivi()[qso_idx]
            qso['qsl_via'] = manager.upper()
            self._aggiorna_tree()
            messagebox.showinfo("Applicato",
                f"QSL_VIA = {manager.upper()} applicato al QSO.",
                parent=dlg)
            dlg.destroy()
        except Exception as ex:
            messagebox.showerror(T("dxc_errore"), str(ex), parent=dlg)

    def _on_tree_shift_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        row_id = self.tree.identify_row(event.y)
        if not row_id or not col_id:
            return
        # Nome reale della colonna (robusto a riordino/nascondimento)
        try:
            col_name = self.tree.column(col_id, "id")
        except Exception:
            col_name = ""
        if col_name != "n":
            return  # Shift+Click solo sulla colonna "#"

        qsos = self._qsos_attivi()

        if self._ultima_riga_toggle is None or not self.tree.exists(self._ultima_riga_toggle):
            # Nessun punto di partenza: comportati come Ctrl+Click singolo
            idx = int(row_id)
            if 0 <= idx < len(qsos):
                qso = qsos[idx]
                added = self._toggle_coda_export(qso)
                self.tree.set(row_id, "n", "✓" if added else str(idx + 1))
        else:
            start = int(self._ultima_riga_toggle)
            end = int(row_id)
            lo, hi = sorted((start, end))
            for i in range(lo, min(hi, len(qsos) - 1) + 1):
                iid = str(i)
                if self.tree.exists(iid):
                    self._aggiungi_coda_export(qsos[i])
                    self.tree.set(iid, "n", "✓")

        self._ultima_riga_toggle = row_id
        self._aggiorna_indicatore_selezione()
        return "break"

    # ── Coda export: lista di QSO (riferimenti agli stessi dict di qsos_caricati) ──
    def _trova_in_coda(self, qso):
        """Restituisce l'indice di qso nella coda export (per identità), o -1."""
        for i, q in enumerate(self._coda_export):
            if q is qso:
                return i
        return -1

    def _aggiungi_coda_export(self, qso):
        """Aggiunge un QSO alla coda export se non già presente. Restituisce True se aggiunto."""
        if self._trova_in_coda(qso) == -1:
            self._coda_export.append(qso)
            return True
        return False

    def _toggle_coda_export(self, qso):
        """Aggiunge/rimuove un QSO dalla coda export. Restituisce True se ora presente (aggiunto)."""
        i = self._trova_in_coda(qso)
        if i == -1:
            self._coda_export.append(qso)
            return True
        else:
            del self._coda_export[i]
            return False

    def _seleziona_tutti_visibili(self):
        """Aggiunge alla coda export tutti i QSO attualmente visibili in griglia
        (rispetta filtro/ricerca attivi). Non duplica QSO già in coda."""
        qsos = self._qsos_attivi()
        for i, qso in enumerate(qsos):
            iid = str(i)
            self._aggiungi_coda_export(qso)
            if self.tree.exists(iid):
                self.tree.set(iid, "n", "✓")
        if qsos:
            self._ultima_riga_toggle = str(len(qsos) - 1)
        self._aggiorna_indicatore_selezione()

    def _deseleziona_qso(self):
        """Svuota completamente la coda export."""
        qsos = self._qsos_attivi()
        for i, qso in enumerate(qsos):
            if self._trova_in_coda(qso) != -1:
                iid = str(i)
                if self.tree.exists(iid):
                    self.tree.set(iid, "n", str(i + 1))
        self._coda_export.clear()
        self._ultima_riga_toggle = None
        self._aggiorna_indicatore_selezione()

    # ── Indicatore coda export / esporta selezione ──
    def _aggiorna_indicatore_selezione(self):
        n = len(self._coda_export)
        if n >= 1:
            self.lbl_selezione.configure(text=f"📋 {n} in coda")
            self.btn_esporta_sel.configure(state="normal")
            self.btn_elimina_sel.configure(state="normal")
            self.btn_desel.configure(state="normal")
            self.btn_vedi_coda.configure(state="normal")
        else:
            self.lbl_selezione.configure(text="")
            self.btn_esporta_sel.configure(state="disabled")
            self.btn_elimina_sel.configure(state="disabled")
            self.btn_desel.configure(state="disabled")
            self.btn_vedi_coda.configure(state="disabled")

    def _qsos_da_selezione(self):
        """Restituisce la lista di QSO attualmente in coda export."""
        return list(self._coda_export)

    def _rimuovi_da_coda(self, qso):
        """Rimuove uno specifico QSO dalla coda export (usato dalla finestra di revisione)."""
        i = self._trova_in_coda(qso)
        if i != -1:
            del self._coda_export[i]
        # Aggiorna il segno ✓ nella griglia se il QSO è attualmente visibile
        qsos = self._qsos_attivi()
        for idx, q in enumerate(qsos):
            if q is qso:
                iid = str(idx)
                if self.tree.exists(iid):
                    self.tree.set(iid, "n", str(idx + 1))
                break
        self._aggiorna_indicatore_selezione()

    def apri_coda_export(self):
        """Apre una finestra che mostra tutti i QSO attualmente in coda per l'export,
        permettendo di rimuoverli singolarmente o di esportarli."""
        if not self._coda_export:
            return

        dlg = ctk.CTkToplevel(self)
        dlg.title("📋 Coda export")
        dlg.geometry("600x480")
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text=f"📋 Coda export — {len(self._coda_export)} QSO",
                      font=ctk.CTkFont(size=15, weight="bold")).pack(anchor="w", padx=16, pady=(14,4))
        ctk.CTkLabel(dlg,
            text=T("dv_coda_resta"),
            font=ctk.CTkFont(size=11), text_color=("#4A5568","#94A3B8")).pack(anchor="w", padx=16, pady=(0,8))

        scroll = ctk.CTkScrollableFrame(dlg)
        scroll.pack(fill="both", expand=True, padx=16, pady=4)

        def ridisegna():
            for w in scroll.winfo_children():
                w.destroy()
            if not self._coda_export:
                ctk.CTkLabel(scroll, text=T("dv_coda_vuota"), text_color="gray").pack(pady=20)
                dlg.destroy()
                return
            for qso in list(self._coda_export):
                call = str(qso.get('call','')).upper()
                data = str(qso.get('qso_date',''))
                if len(data) == 8:
                    data = f"{data[6:8]}/{data[4:6]}/{data[0:4]}"
                ora = str(qso.get('time_on',''))[:4]
                banda = str(qso.get('band','')).upper()
                modo = str(qso.get('mode','')).upper()
                sat = str(qso.get('sat_name','')).strip()
                desc = f"{call}  —  {data} {ora}  —  {banda}/{modo}"
                if sat:
                    desc += f"  —  🛰 {sat}"

                row = ctk.CTkFrame(scroll, fg_color=("#EBF4FF","#161616"))
                row.pack(fill="x", pady=2)
                ctk.CTkLabel(row, text=desc, font=ctk.CTkFont(size=12),
                              anchor="w").pack(side="left", padx=10, pady=6, fill="x", expand=True)
                ctk.CTkButton(row, text="✕", width=28, height=24,
                              fg_color="#4A5568", hover_color=TH.WARNING_H,
                              command=lambda q=qso: (self._rimuovi_da_coda(q), ridisegna())
                              ).pack(side="right", padx=8)

        ridisegna()

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=(8,14))
        btn_esp = ctk.CTkButton(btns, text="📤 Esporta selezione ▾",
                      height=34, fg_color=("#2F855A","#276749"), hover_color=("#276749","#22543D"),
                      font=ctk.CTkFont(size=12, weight="bold"))
        btn_esp.configure(command=lambda: self._menu_esporta_selezione(anchor_widget=btn_esp))
        btn_esp.pack(side="left", padx=(0,8))
        ctk.CTkButton(btns, text="🗑 Svuota coda", command=lambda: (self._deseleziona_qso(), dlg.destroy()),
                      height=34, width=120, fg_color=TH.WARNING_H,
                      hover_color=TH.WARNING_H).pack(side="left", padx=(0,8))
        ctk.CTkButton(btns, text=T("cm_chiudi"), command=dlg.destroy,
                      height=34, width=90, fg_color="#4A5568",
                      hover_color="#2D3748").pack(side="right")

    def _menu_esporta_selezione(self, anchor_widget=None):
        """Apre un menu per esportare solo i QSO in coda nel formato scelto."""
        import tkinter as _tk
        qsos_sel = self._qsos_da_selezione()
        if not qsos_sel:
            return
        if anchor_widget is None:
            anchor_widget = self.btn_esporta_sel
        menu = _tk.Menu(self, tearoff=0)
        menu.add_command(label=f"📄 PDF  ({len(qsos_sel)} QSO)",
                          command=lambda: self._esporta_con_subset(qsos_sel, self.processa_e_salva))
        menu.add_command(label=f"📊 CSV  ({len(qsos_sel)} QSO)",
                          command=lambda: self._esporta_con_subset(qsos_sel, self.esporta_csv))
        menu.add_command(label=f"📗 Excel  ({len(qsos_sel)} QSO)",
                          command=lambda: self._esporta_con_subset(qsos_sel, self.esporta_excel))
        menu.add_command(label=f"🌐 HTML  ({len(qsos_sel)} QSO)",
                          command=lambda: self._esporta_con_subset(qsos_sel, self.esporta_html))
        menu.add_separator()
        menu.add_command(label=f"💾 ADIF  ({len(qsos_sel)} QSO)",
                          command=lambda: self._esporta_con_subset(qsos_sel, self.salva_adif))

        try:
            x = anchor_widget.winfo_rootx()
            y = anchor_widget.winfo_rooty() + anchor_widget.winfo_height()
            menu.tk_popup(x, y)
        finally:
            menu.grab_release()

    def _esporta_con_subset(self, qsos_subset, export_func):
        """Esegue una funzione di esportazione limitandola temporaneamente
        ai QSO in coda, poi ripristina la vista precedente e svuota la coda."""
        old_filtrati = self.qsos_filtrati
        old_lbl = self.lbl_filtri.cget("text")
        old_color = self.lbl_filtri.cget("text_color")
        self.qsos_filtrati = list(qsos_subset)
        try:
            export_func()
        finally:
            self.qsos_filtrati = old_filtrati
            self.lbl_filtri.configure(text=old_lbl, text_color=old_color)
            self._coda_export.clear()
            self._ultima_riga_toggle = None
            self._aggiorna_tree()

    def _push_undo(self):
        """Salva snapshot del log corrente nello stack undo (max 10 livelli)."""
        self._undo_stack.append(list(self.qsos_caricati))
        if len(self._undo_stack) > 10:
            self._undo_stack.pop(0)
        if hasattr(self, 'btn_undo_main'):
            self.btn_undo_main.configure(state="normal")

    def _undo_main(self):
        """Ripristina l'ultimo snapshot del log (Ctrl+Z)."""
        if not self._undo_stack:
            messagebox.showinfo("Annulla", "Nessuna operazione da annullare.")
            return
        self.qsos_caricati = self._undo_stack.pop()
        self.qsos_filtrati = list(self.qsos_caricati)
        self._coda_export.clear()
        self._ultima_riga_toggle = None
        self._aggiorna_tree()
        if not self._undo_stack and hasattr(self, 'btn_undo_main'):
            self.btn_undo_main.configure(state="disabled")

    def _elimina_selezione(self):
        """Elimina dal log i QSO selezionati — sia dalla coda export
        (Ctrl+Click su colonna #) sia dalla selezione nativa del treeview
        (click/Shift+Click sulle righe)."""
        qsos_sel = list(self._coda_export)

        # Se la coda export è vuota, usa la selezione nativa salvata
        if not qsos_sel:
            qsos_attivi = self._qsos_attivi()
            for iid in getattr(self, '_last_tree_sel', []) or list(self.tree.selection()):
                try:
                    idx = int(iid)
                    if 0 <= idx < len(qsos_attivi):
                        qsos_sel.append(qsos_attivi[idx])
                except (ValueError, IndexError):
                    pass

        if not qsos_sel:
            return
        n = len(qsos_sel)
        if not messagebox.askyesno(
            "Conferma eliminazione",
            f"Eliminare {n} QSO selezionati dal log?\n\n"
            f"Puoi annullare con ↩ Annulla o Ctrl+Z."
        ):
            return
        self._push_undo()
        ids_da_rimuovere = {id(q) for q in qsos_sel}
        self.qsos_caricati = [q for q in self.qsos_caricati if id(q) not in ids_da_rimuovere]
        self._log_modificato = True
        if self.qsos_filtrati:
            self.qsos_filtrati = [q for q in self.qsos_filtrati if id(q) not in ids_da_rimuovere]
        self._coda_export.clear()
        self._ultima_riga_toggle = None
        self._aggiorna_tree()
        messagebox.showinfo("Eliminazione completata", f"{n} QSO eliminati dal log.")

    # Mappa completa campi ADIF → label display
    _EP_ALL_FIELDS = [
        ("call",           "Callsign",    12),
        ("qso_date",       "Data",        10),
        ("time_on",        "UTC",          8),
        ("band",           "Banda",        8),
        ("mode",           "Mode",         8),
        ("freq",           "Freq.",        9),
        ("rst_sent",       "RST TX",       7),
        ("rst_rcvd",       "RST RX",       7),
        ("gridsquare",     "Locator",      9),
        ("country",        "Country",     14),
        ("name",           "Nome",        10),
        ("qsl_via",        "QSL via",     10),
        ("sat_name",       "Satellite",   10),
        ("prop_mode",      "Prop.",         8),
        ("band_rx",        "Banda RX",      8),
        ("freq_rx",        "Freq. RX",      9),
        ("sat_mode",       "SAT Mode",      8),
        ("contest_id",     "Contest",     10),
        ("stx",            "STX",          6),
        ("srx",            "SRX",          6),
        ("lotw_qsl_rcvd",  "LoTW",         6),
        ("eqsl_qsl_rcvd",  "eQSL",         6),
        ("qsl_sent",       "QSL Sent",     7),
        ("qsl_rcvd",       "QSL Rcvd",     7),
        ("comment",        "Comment",     14),
        ("notes",          "Notes",       14),
        ("station_callsign","Station",    10),
        ("operator",       "Operator",    10),
        ("tx_pwr",         "Power",        7),
        ("antenna",        "Antenna",     10),
    ]

    # Campi sempre mostrati anche se vuoti
    _EP_ALWAYS = {"call","qso_date","time_on","band","mode","sat_name","rst_sent","rst_rcvd"}

    def _ep_on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        try:
            idx = int(iid)
        except ValueError:
            return
        qsos = self._qsos_attivi()
        if idx >= len(qsos):
            return
        self._ep_idx = idx
        qso = qsos[idx]
        call = str(qso.get('call','')).upper()
        data = str(qso.get('qso_date',''))
        if len(data)==8:
            data = f"{data[6:8]}/{data[4:6]}/{data[0:4]}"
        self._ep_info.configure(text=f"QSO #{idx+1}  —  {call}  {data}")

        # Svuota frame dinamico
        for w in self._ep_fields_frame.winfo_children():
            w.destroy()
        self._ep_entries = {}
        self._ep_field_labels = {}

        # Costruisci i campi con valore nel QSO selezionato, più quelli che
        # compaiono in almeno un altro QSO dell'intero log (così un campo
        # come RST o Country, anche se vuoto qui, resta sempre editabile
        # se è usato altrove nel log) e i pochi campi base sempre garantiti.
        campi_da_mostrare = getattr(self, '_campi_presenti_nel_log', set()) | self._EP_ALWAYS
        col = 0
        for tag, lbl, w in self._EP_ALL_FIELDS:
            val = str(qso.get(tag, '')).strip()
            if not val and tag not in campi_da_mostrare:
                continue
            f = ctk.CTkFrame(self._ep_fields_frame, fg_color="transparent")
            f.grid(row=0, column=col, padx=3, pady=2, sticky="nw")
            fl = ctk.CTkLabel(f, text=lbl, font=ctk.CTkFont(size=10),
                              text_color=("#4A5568","#94A3B8"))
            fl.pack(anchor="w")
            e = ctk.CTkEntry(f, width=w*7, height=26,
                             font=ctk.CTkFont(size=11, family="Consolas"))
            e.pack()
            e.insert(0, val)
            # Evidenzia campi satellite vuoti in arancione
            if tag == "sat_name" and not val:
                e.configure(border_color=TH.WARN_TEXT, border_width=2,
                            placeholder_text=T("dv_ph_so50"))
            self._ep_entries[tag] = e
            self._ep_field_labels[tag] = fl
            col += 1

    # ── Ricerca testuale multi-campo ───────────
    _CAMPI_RICERCA = ('call', 'name', 'qth', 'country', 'gridsquare',
                       'comment', 'notes', 'sat_name', 'contest_id',
                       'state', 'cnty', 'operator', 'station_callsign')

    def _esegui_ricerca(self):
        """Filtra il log in base al testo inserito nel box di ricerca,
        cercando in più campi (call, nome, QTH, country, locator, commento...)."""
        testo = self.entry_search.get().strip().lower()
        if not testo:
            self._azzera_filtro()
            return

        self.qsos_filtrati = [
            q for q in self.qsos_caricati
            if any(testo in str(q.get(campo, '')).lower() for campo in self._CAMPI_RICERCA)
        ]
        n = len(self.qsos_filtrati)
        tot = len(self.qsos_caricati)
        self.lbl_filtri.configure(
            text=f"🔍 '{self.entry_search.get().strip()}'  {n}/{tot} QSO",
            text_color=TH.PRIMARY
        )
        self.sb_filt.configure(text=f"Ricerca: '{self.entry_search.get().strip()}'")
        self._aggiorna_tree()

    def _pulisci_ricerca(self):
        self.entry_search.delete(0, 'end')
        self._azzera_filtro()

    def _reset_filtro(self):
        """Reset filtri dal pulsante sidebar."""
        self._azzera_filtro()

    def _aggiorna_sidebar_filtri(self):
        """Aggiorna il label stato filtri e popola banda/modo/sat dal log corrente."""
        if not hasattr(self, 'lbl_filtri_attivi'):
            return
        n_tot  = len(self.qsos_caricati)
        n_filt = len(self.qsos_filtrati) if hasattr(self,'qsos_filtrati') else n_tot
        if n_filt == n_tot or n_tot == 0:
            self.lbl_filtri_attivi.configure(text="", text_color="gray")
        else:
            self.lbl_filtri_attivi.configure(
                text=f"▶ {n_filt} / {n_tot} QSO visibili",
                text_color=TH.WARN_TEXT)
        # Popola OptionMenu con i valori del log
        if hasattr(self, '_fil_banda_om') and self.qsos_caricati:
            bande = [T("filtri_tutte")] + sorted({str(q.get('band','')).upper().strip()
                                        for q in self.qsos_caricati if q.get('band','')})
            self._fil_banda_om.configure(values=bande)
            modi = [T("filtri_tutti")] + sorted({str(q.get('mode','')).upper().strip()
                                       for q in self.qsos_caricati if q.get('mode','')})
            self._fil_modo_om.configure(values=modi)
            sat_log = sorted({str(q.get('sat_name','')).upper().strip()
                              for q in self.qsos_caricati
                              if q.get('sat_name','') and str(q.get('prop_mode','')).upper()=='SAT'})
            base = [T("filtri_tutti"),"QO-100","RS-44","AO-91","AO-92","AO-7","FO-29",
                    "SO-50","IO-117","CAS-4A","XW-2C","ISS"]
            extra = [s for s in sat_log if s not in base]
            self._fil_sat_om.configure(values=base + extra)

    def _azzera_filtro(self):
        """Rimuove qualsiasi filtro/ricerca attivi, mostrando tutto il log."""
        self.qsos_filtrati = list(self.qsos_caricati)
        self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
        self.sb_filt.configure(text="")
        self._aggiorna_tree()
        self._aggiorna_sidebar_filtri()

    def _rimuovi_filtro_rapido(self):
        self.entry_search.delete(0, 'end')
        self._azzera_filtro()

    def _ep_get_qso_originale(self):
        """Restituisce il QSO originale in qsos_caricati corrispondente all'indice filtrato."""
        if self._ep_idx is None:
            return None, None
        qsos = self._qsos_attivi()
        if self._ep_idx >= len(qsos):
            return None, None
        qso_filt = qsos[self._ep_idx]
        # Trova il corrispondente in qsos_caricati
        for i, q in enumerate(self.qsos_caricati):
            if q is qso_filt:
                return i, q
        return None, None

    def _ep_applica(self):
        orig_idx, qso = self._ep_get_qso_originale()
        if qso is None:
            return
        for tag, entry in self._ep_entries.items():
            val = entry.get().strip()
            if val:
                qso[tag.lower()] = val
            elif tag.lower() in qso:
                qso[tag.lower()] = val
        self._aggiorna_tree()
        # Riseleziona la stessa riga
        if self._ep_idx is not None:
            iid = str(self._ep_idx)
            if self.tree.exists(iid):
                self.tree.selection_set(iid)
                self.tree.see(iid)
        from tkinter import messagebox
        messagebox.showinfo("OK", "Modifiche applicate al log in memoria.\nRicorda di salvare il file ADIF.")

    def _ep_duplica(self):
        orig_idx, qso = self._ep_get_qso_originale()
        if qso is None:
            return
        import copy
        nuovo = copy.deepcopy(qso)
        self.qsos_caricati.insert(orig_idx + 1, nuovo)
        if self.qsos_filtrati:
            self.qsos_filtrati.insert(self._ep_idx + 1, nuovo)
        self._aggiorna_tree()

    def _ep_su(self):
        orig_idx, qso = self._ep_get_qso_originale()
        if orig_idx is None or orig_idx == 0:
            return
        self.qsos_caricati.insert(orig_idx - 1, self.qsos_caricati.pop(orig_idx))
        if self.qsos_filtrati and self._ep_idx > 0:
            self.qsos_filtrati.insert(self._ep_idx - 1, self.qsos_filtrati.pop(self._ep_idx))
            self._ep_idx -= 1
        self._aggiorna_tree()

    def _ep_giu(self):
        orig_idx, qso = self._ep_get_qso_originale()
        if orig_idx is None or orig_idx >= len(self.qsos_caricati) - 1:
            return
        self.qsos_caricati.insert(orig_idx + 1, self.qsos_caricati.pop(orig_idx))
        if self.qsos_filtrati and self._ep_idx < len(self.qsos_filtrati) - 1:
            self.qsos_filtrati.insert(self._ep_idx + 1, self.qsos_filtrati.pop(self._ep_idx))
            self._ep_idx += 1
        self._aggiorna_tree()

    def _ep_elimina(self):
        from tkinter import messagebox
        orig_idx, qso = self._ep_get_qso_originale()
        if qso is None:
            return
        call = str(qso.get('call','')).upper()
        if not messagebox.askyesno("Elimina QSO", f"Eliminare il QSO con {call}?"):
            return
        self._push_undo()
        self.qsos_caricati.remove(qso)
        if qso in self.qsos_filtrati:
            self.qsos_filtrati.remove(qso)
        self._ep_idx = None
        self._ep_info.configure(text=T("ep_seleziona"))
        for e in self._ep_entries.values():
            e.delete(0, 'end')
        self._aggiorna_tree()

    # Campi proposti da "Imposta colonna": elenco fisso e completo dei campi
    # ADIF più comuni, indipendente dal QSO selezionato nel pannello rapido
    # (prima usava solo i campi non vuoti del QSO corrente, troppo limitato
    # per un'operazione pensata per agire sull'intero log).
    _COLONNA_CAMPI_DISPONIBILI = [
        "call", "qso_date", "time_on", "band", "freq", "mode", "submode",
        "rst_sent", "rst_rcvd", "gridsquare", "country", "dxcc", "cont",
        "name", "qth", "sat_name", "prop_mode", "contest_id",
        "station_callsign", "operator", "my_gridsquare", "my_country",
        "my_cq_zone", "my_itu_zone", "tx_pwr", "antenna",
        "qsl_sent", "qsl_rcvd", "qsl_sent_via", "qsl_rcvd_via",
        "eqsl_qsl_sent", "eqsl_qsl_rcvd", "lotw_qsl_sent", "lotw_qsl_rcvd",
        "comment", "notes",
    ]

    def _ep_colonna(self):
        from tkinter import messagebox, simpledialog
        import tkinter as _tk
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        campi = self._COLONNA_CAMPI_DISPONIBILI
        dlg = ctk.CTkToplevel(self)
        dlg.title("Imposta colonna")
        dlg.geometry("320x200")
        dlg.grab_set()
        ctk.CTkLabel(dlg, text=T("dv_campo_imp"),
                     font=ctk.CTkFont(size=12)).pack(padx=15, pady=(15,5))
        var_campo = ctk.StringVar(value=campi[0])
        ctk.CTkOptionMenu(dlg, variable=var_campo, values=campi).pack(padx=15, fill="x")
        ctk.CTkLabel(dlg, text=T("dv_valore_tutti"),
                     font=ctk.CTkFont(size=12)).pack(padx=15, pady=(10,5))
        e_val = ctk.CTkEntry(dlg)
        e_val.pack(padx=15, fill="x")
        def applica():
            tag = var_campo.get()
            val = e_val.get().strip()
            n = len(self.qsos_caricati)
            if not messagebox.askyesno(
                "Conferma",
                f"Impostare {tag.upper()} = '{val}' su tutti i {n} QSO del log?\n"
                f"L'operazione sovrascrive eventuali valori già presenti in quel campo.",
                parent=dlg
            ):
                return
            for q in self.qsos_caricati:
                q[tag.lower()] = val
            if self.qsos_filtrati:
                for q in self.qsos_filtrati:
                    q[tag.lower()] = val
            self._aggiorna_tree()
            dlg.destroy()
        ctk.CTkButton(dlg, text=T("dv_applica_tutti"),
                      command=applica, fg_color="#4A5568").pack(padx=15, pady=12, fill="x")

    def _ep_normalizza(self):
        from tkinter import messagebox
        if not self.qsos_caricati:
            return
        # Correzione preliminare: bande non valide (es. 'infm') su QSO
        # satellitari con satellite noto -> banda corretta + FREQ di default.
        n_sat_corretti = self._correggi_banda_satellite(self.qsos_caricati)
        # Normalizzazione di base: uppercase + formato data/ora (comportamento storico)
        for q in self.qsos_caricati:
            for k in ('call','band','mode','country','gridsquare'):
                if k in q:
                    q[k] = str(q[k]).upper().strip()
            d = str(q.get('qso_date','')).replace('/','').replace('-','')
            if len(d)==8:
                q['qso_date'] = d
            t = str(q.get('time_on','')).replace(':','')
            if len(t)>=4:
                q['time_on'] = t[:4]
        # Fase 1 completa: omogenizzazione campi + calcolo FREQ di default per banda
        n_omogenea, tutti_campi, campi_critici_assenti = self._normalizza_fase1_omogenizza(self.qsos_caricati)

        # IMPORTANTE: dopo la normalizzazione azzera sempre filtri e ricerca.
        # La normalizzazione opera sull'intero log (qsos_caricati) ma se rimane
        # un filtro attivo la griglia mostra solo il sottoinsieme filtrato,
        # facendo sembrare il log incompleto. Mostrare sempre tutto il log.
        self.qsos_filtrati = list(self.qsos_caricati)
        if hasattr(self, 'entry_search'):
            self.entry_search.delete(0, 'end')
        self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
        if hasattr(self, 'sb_filt'):
            self.sb_filt.configure(text="")

        self._aggiorna_tree()
        def _dopo():
            self._aggiorna_tree()
        # Fase 2: dialogo valori di default per i campi assenti dallo schema
        self._normalizza_fase2_dialogo(self.qsos_caricati, tutti_campi, n_omogenea,
                                        parent=self, on_done=_dopo,
                                        n_sat_corretti=n_sat_corretti,
                                        campi_critici_assenti=campi_critici_assenti)

    # ── Distanza/Bearing dal QSO selezionato ───
    def _ep_calcola_distanza(self):
        from tkinter import messagebox
        if self._ep_idx is None:
            messagebox.showinfo(T("dist_calc_title"), T("dist_calc_no_locator"))
            return
        qsos = self._qsos_attivi()
        if self._ep_idx >= len(qsos):
            return
        qso = qsos[self._ep_idx]
        locator_qso = estrai_locator_da_testo(str(qso.get('gridsquare','')))
        if not locator_qso:
            messagebox.showinfo(T("dist_calc_title"), T("dist_calc_no_locator"))
            return
        my_locator = estrai_locator_da_testo(self.entry_details.get())
        if not my_locator:
            messagebox.showinfo(T("dist_calc_title"), T("dist_calc_no_my_locator"))
        self.apri_calcolatore_distanza(locator_other_prefill=locator_qso)

    # ── Esporta Cabrillo standalone ────────────
    def esporta_cabrillo_standalone(self):
        from tkinter import messagebox, filedialog
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_nessun_file"))
            return
        qsos_exp = list(self._qsos_attivi())

        dlg = ctk.CTkToplevel(self)
        dlg.title("Esporta Cabrillo Contest")
        dlg.geometry("480x460")
        dlg.resizable(False, False)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text=T("dv_esporta_cbr"),
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=10, padx=20)

        CONTEST_LIST = ["CQ-WW-CW","CQ-WW-SSB","CQ-WPX-CW","CQ-WPX-SSB",
            "IARU-HF","CQ-VHF","ARRL-DX-CW","ARRL-DX-SSB","ARRL-10","ARRL-160",
            "ARRL-SS-CW","ARRL-SS-SSB","EU-HF","ARI-HF","ARI-DX","ITALIA-40-80",
            "GENERIC","OTHER"]

        scroll_c = ctk.CTkScrollableFrame(dlg, height=300)
        scroll_c.pack(fill="x", padx=15, pady=4)

        def add_field(lbl, default="", ph=""):
            row = ctk.CTkFrame(scroll_c, fg_color="transparent")
            row.pack(fill="x", pady=3)
            ctk.CTkLabel(row, text=lbl, width=140, anchor="e",
                         font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,8))
            e = ctk.CTkEntry(row, width=260, placeholder_text=ph)
            e.insert(0, default)
            e.pack(side="left")
            return e

        profilo_call = self.entry_owner.get().strip().upper()
        f_call    = add_field("Callsign:",        profilo_call)
        ctk.CTkLabel(scroll_c, text=T("cm_contest"), anchor="e", width=140,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", padx=(15,0), pady=(4,0))
        contest_var = ctk.StringVar(value="CQ-WW-SSB")
        ctk.CTkOptionMenu(scroll_c, variable=contest_var,
                          values=CONTEST_LIST, width=260).pack(padx=15, pady=2)
        f_cat_op   = add_field("Category-Op:",    "SINGLE-OP")
        f_cat_band = add_field("Category-Band:",  "ALL")
        f_cat_mode = add_field("Category-Mode:",  "SSB")
        f_cat_pwr  = add_field("Category-Power:", "HIGH")
        f_ops      = add_field("Operators:",      profilo_call)
        f_club     = add_field("Club:",           "", "Nome club (opz.)")
        f_loc      = add_field("Location:",       "", "es. ITA / EU")
        f_score    = add_field("Claimed-Score:",  "0")

        def genera_cbr():
            call = f_call.get().strip().upper()
            if not call:
                messagebox.showwarning("Attenzione", "Inserire il callsign.")
                return
            nome_base = os.path.splitext(os.path.basename(self.filepath))[0] if self.filepath else "contest"
            save_path = filedialog.asksaveasfilename(
                parent=dlg, title=T("dv_salva_cbr"),
                defaultextension=".cbr",
                filetypes=[("Cabrillo","*.cbr *.log"),("All","*.*")],
                initialfile=f"{nome_base}_{contest_var.get()}.cbr")
            if not save_path:
                return
            nl = chr(10)
            try:
                with open(save_path, "w", encoding="ascii", errors="replace") as fw:
                    fw.write(f"START-OF-LOG: 3.0{nl}")
                    fw.write(f"CONTEST: {contest_var.get()}{nl}")
                    fw.write(f"CALLSIGN: {call}{nl}")
                    fw.write(f"CATEGORY-OPERATOR: {f_cat_op.get().strip()}{nl}")
                    fw.write(f"CATEGORY-BAND: {f_cat_band.get().strip()}{nl}")
                    fw.write(f"CATEGORY-MODE: {f_cat_mode.get().strip()}{nl}")
                    fw.write(f"CATEGORY-POWER: {f_cat_pwr.get().strip()}{nl}")
                    if f_ops.get().strip():
                        fw.write(f"OPERATORS: {f_ops.get().strip()}{nl}")
                    if f_club.get().strip():
                        fw.write(f"CLUB: {f_club.get().strip()}{nl}")
                    if f_loc.get().strip():
                        fw.write(f"LOCATION: {f_loc.get().strip()}{nl}")
                    fw.write(f"CLAIMED-SCORE: {f_score.get().strip() or '0'}{nl}")
                    fw.write(f"CREATED-BY: ADIF FZR 2.5{nl}{nl}")

                    FREQ_MAP = {"160m":"1800","80m":"3500","60m":"5357","40m":"7000",
                        "30m":"10100","20m":"14000","17m":"18068","15m":"21000",
                        "12m":"24890","10m":"28000","6m":"50000","2m":"144000",
                        "70cm":"432000","23cm":"1296000"}

                    for qso in qsos_exp:
                        q = {k.lower():v for k,v in qso.items()}
                        try:
                            freq_khz = str(int(float(q.get("freq","0"))*1000))
                        except Exception:
                            freq_khz = FREQ_MAP.get(q.get("band","").lower(),"14000")
                        modo = q.get("mode","SSB").upper()
                        if modo in ("FT8","FT4","WSPR","JT65","JT9","MFSK"):
                            modo = "DG"
                        elif modo in ("PSK31","PSK63","RTTY"):
                            modo = "RY"
                        data = q.get("qso_date","")
                        if len(data)==8:
                            data = f"{data[:4]}-{data[4:6]}-{data[6:]}"
                        ora = q.get("time_on","0000")[:4]
                        fw.write(
                            f"QSO: {freq_khz:>6} {modo:<2} {data} {ora} "
                            f"{q.get('station_callsign',call):<13} "
                            f"{q.get('rst_sent','59')[:3]} {q.get('stx','001'):<6} "
                            f"{q.get('call','').upper():<13} "
                            f"{q.get('rst_rcvd','59')[:3]} {q.get('srx','001')}{nl}"
                        )
                    fw.write(f"END-OF-LOG:{nl}")

                dlg.destroy()
                messagebox.showinfo("Esportato",
                    f"{len(qsos_exp)} QSO esportati in Cabrillo:{nl}{os.path.basename(save_path)}")
                try:
                    os.startfile(os.path.abspath(save_path))
                except Exception:
                    pass
            except Exception as ex:
                messagebox.showerror("Errore", f"Impossibile salvare:{nl}{ex}")

        ctk.CTkButton(dlg, text="📄 Genera Cabrillo", command=genera_cbr,
                      fg_color="#4A5568", hover_color="#2D3748",
                      font=ctk.CTkFont(size=13, weight="bold"),
                      height=38).pack(padx=15, pady=10, fill="x")

    # ── Dialog colonne PDF (era inline) ────────
    def _apri_dialog_colonne(self):
        import tkinter as _tk
        dlg = ctk.CTkToplevel(self)
        dlg.title("Colonne PDF")
        dlg.geometry("360x320")
        dlg.grab_set()
        ctk.CTkLabel(dlg, text=T("dv_sel_colonne_pdf"),
                     font=ctk.CTkFont(size=12, weight="bold")).pack(padx=15, pady=10)
        f = ctk.CTkScrollableFrame(dlg)
        f.pack(fill="both", expand=True, padx=15)
        for tag, info in self.campi_disponibili.items():
            cb = ctk.CTkCheckBox(f, text=info['nome'], variable=self.checkboxes[tag])
            cb.pack(anchor="w", pady=2)
        ctk.CTkButton(dlg, text=T("cm_chiudi"), command=dlg.destroy).pack(pady=10)

    # ── Dialog formato PDF ─────────────────────
    def _apri_dialog_formato(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Formato pagina PDF")
        dlg.geometry("280x160")
        dlg.grab_set()
        ctk.CTkLabel(dlg, text=T("dv_formato_pagina"), font=ctk.CTkFont(size=12)).pack(pady=20)
        ctk.CTkOptionMenu(dlg, variable=self.var_formato_pdf,
                          values=["A4","US Letter"], width=160).pack()
        ctk.CTkButton(dlg, text="OK", command=dlg.destroy, width=80).pack(pady=15)

    # ── Calcolatore Distanza / Bearing ─────────
    def apri_aggiungi_qso(self):
        """Finestra di inserimento rapido per QSO fatti manualmente (non
        loggati a computer). Data/Ora sono un orologio UTC live aggiornato
        finché non si preme Inserisci; Banda/Frequenza/Modo/Satellite
        restano gli ultimi valori usati tra un inserimento e l'altro;
        RST TX/RX si auto-compila in base al Modo (59/599/+01) mantenendo
        comunque la possibilità di modificarli a mano."""
        import tkinter as _tk
        import datetime as _dt

        BANDE = ['160m','80m','60m','40m','30m','20m','17m','15m','12m','10m',
                 '6m','4m','2m','1.25m','70cm','33cm','23cm','13cm','9cm',
                 '6cm','3cm','1.25cm','6mm']
        MODI = ['SSB','CW','FT8','FT4','FT2','RTTY','PSK31','MFSK','JT65','JT9',
                'WSPR','AM','FM','SSTV','PKT','OLIVIA','DOMINO','THOR']
        RST_DEFAULT = {
            'SSB':('59','59'), 'AM':('59','59'), 'FM':('59','59'),
            'CW':('599','599'),
            'FT8':('+01','+01'), 'FT4':('+01','+01'), 'FT2':('+01','+01'),
            'RTTY':('599','599'), 'PSK31':('599','599'), 'MFSK':('+01','+01'),
            'JT65':('+01','+01'), 'JT9':('+01','+01'), 'WSPR':('-10','-10'),
            'SSTV':('59','59'), 'PKT':('599','599'),
            'OLIVIA':('599','599'), 'DOMINO':('599','599'), 'THOR':('599','599'),
        }

        # Valori persistenti tra un'apertura e l'altra della finestra (e tra
        # un inserimento e l'altro nella stessa sessione).
        if not hasattr(self, '_addqso_ultimi'):
            self._addqso_ultimi = {'banda': '20m', 'freq': '', 'modo': 'SSB', 'sat': ''}

        dlg = ctk.CTkToplevel(self)
        dlg.title(T("addqso_titolo"))
        dlg.geometry("500x720")
        dlg.resizable(False, True)
        dlg.minsize(460, 520)
        dlg.lift(); dlg.focus_force()
        self._aggiungi_qso_dlg = dlg   # riferimento per precompilazione da DX Cluster
        dlg.attributes('-topmost', True)
        # Niente grab_set(): la finestra resta utilizzabile insieme alla
        # principale (puoi comunque scorrere/cliccare la griglia), ma
        # 'topmost' la mantiene sempre visibile sopra le altre finestre,
        # così non sparisce dietro cliccando altrove nel programma.

        ctk.CTkLabel(dlg, text="➕ " + T("addqso_titolo"),
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(16, 4))

        contatore_var = _tk.StringVar(value="")
        ctk.CTkLabel(dlg, textvariable=contatore_var, font=ctk.CTkFont(size=10),
                     text_color="gray").pack(pady=(0, 8))

        form = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        # NB: il form viene "packato" più in basso, DOPO i pulsanti, così la
        # barra pulsanti resta ancorata in fondo e sempre visibile anche su
        # schermi piccoli o con DPI alto; il form scorre se non ci sta.

        # ── Data / Ora live ──────────────────────────
        ctk.CTkLabel(form, text=T("addqso_data"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(form, text=T("addqso_ora"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=0, column=1, sticky="w", padx=(8, 0))
        e_data = ctk.CTkEntry(form, width=140)
        e_data.grid(row=1, column=0, sticky="w", pady=(0, 2))
        e_ora = ctk.CTkEntry(form, width=140)
        e_ora.grid(row=1, column=1, sticky="w", padx=(8, 0), pady=(0, 2))

        live_var = _tk.BooleanVar(value=True)
        ctk.CTkCheckBox(form, text=T("addqso_live"), variable=live_var,
                         font=ctk.CTkFont(size=9), checkbox_width=16, checkbox_height=16
                         ).grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 10))

        _clock_job = [None]
        def _tick():
            if live_var.get():
                now = _dt.datetime.now(_dt.timezone.utc)
                e_data.delete(0, 'end'); e_data.insert(0, now.strftime('%Y%m%d'))
                e_ora.delete(0, 'end'); e_ora.insert(0, now.strftime('%H%M%S'))
            _clock_job[0] = dlg.after(1000, _tick)
        _tick()
        def _on_close():
            if _clock_job[0]:
                dlg.after_cancel(_clock_job[0])
            dlg.destroy()
        dlg.protocol("WM_DELETE_WINDOW", _on_close)
        # Quando l'utente tocca a mano Data/Ora, ferma l'aggiornamento live
        # per non sovrascrivere quello che ha appena digitato.
        e_data.bind("<Key>", lambda e: live_var.set(False))
        e_ora.bind("<Key>", lambda e: live_var.set(False))

        # ── Nominativo / Country ──────────────────────
        ctk.CTkLabel(form, text=T("addqso_call"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=3, column=0, sticky="w")
        ctk.CTkLabel(form, text=T("addqso_country"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=3, column=1, sticky="w", padx=(8, 0))
        e_call = ctk.CTkEntry(form, width=140, placeholder_text=T("aq_ph_call"))
        e_call.grid(row=4, column=0, sticky="w", pady=(0, 10))
        e_country = ctk.CTkEntry(form, width=140, placeholder_text=T("aq_ph_country"))
        e_country.grid(row=4, column=1, sticky="w", padx=(8, 0), pady=(0, 10))

        _country_auto = [True]  # diventa False appena l'utente modifica Country a mano
        def _su_country_edit(*_a):
            _country_auto[0] = False
        e_country.bind("<Key>", _su_country_edit)

        def _su_cambio_call(*_a):
            if not _country_auto[0]:
                return  # l'utente ha già personalizzato il Country, non sovrascrivere
            call_attuale = e_call.get().strip()
            risultato = dxcc_da_nominativo(call_attuale) if call_attuale else None
            if risultato:
                country, _dxcc_code, _cont = risultato
                e_country.delete(0, 'end'); e_country.insert(0, country)
                _country_auto[0] = True  # resta "auto" finché l'utente non la tocca
        e_call.bind("<KeyRelease>", _su_cambio_call)
        e_call.bind("<FocusOut>", _su_cambio_call)

        # ── Banda / Frequenza ────────────────────────
        ctk.CTkLabel(form, text=T("addqso_banda"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=5, column=0, sticky="w")
        ctk.CTkLabel(form, text=T("addqso_freq"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=5, column=1, sticky="w", padx=(8, 0))
        var_banda = _tk.StringVar(value=self._addqso_ultimi['banda'])
        def _su_cambio_banda(*_a):
            nuova = self._freq_da_banda(var_banda.get())
            if nuova:
                e_freq.delete(0, 'end'); e_freq.insert(0, nuova)
        cb_banda = ctk.CTkOptionMenu(form, values=BANDE, variable=var_banda, width=140,
                                      command=_su_cambio_banda)
        cb_banda.grid(row=6, column=0, sticky="w", pady=(0, 10))
        e_freq = ctk.CTkEntry(form, width=140)
        e_freq.grid(row=6, column=1, sticky="w", padx=(8, 0), pady=(0, 10))
        freq_iniziale = self._addqso_ultimi['freq'] or self._freq_da_banda(self._addqso_ultimi['banda'])
        e_freq.insert(0, freq_iniziale)

        # ── Modo / Satellite ─────────────────────────
        ctk.CTkLabel(form, text=T("addqso_modo"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=7, column=0, sticky="w")
        ctk.CTkLabel(form, text=T("addqso_satellite"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=7, column=1, sticky="w", padx=(8, 0))
        var_modo = _tk.StringVar(value=self._addqso_ultimi['modo'])
        # Esponi i campi come attributi del dialog (per precompilazione da DX Cluster)
        dlg._aq_call    = e_call
        dlg._aq_freq    = e_freq
        dlg._aq_country = e_country
        dlg._aq_var_banda = var_banda
        dlg._aq_var_modo  = var_modo
        def _su_cambio_modo(*_a):
            tx, rx = RST_DEFAULT.get(var_modo.get().upper(), ('59', '59'))
            e_rst_tx.delete(0, 'end'); e_rst_tx.insert(0, tx)
            e_rst_rx.delete(0, 'end'); e_rst_rx.insert(0, rx)
        cb_modo = ctk.CTkOptionMenu(form, values=MODI, variable=var_modo, width=140,
                                     command=_su_cambio_modo)
        cb_modo.grid(row=8, column=0, sticky="w", pady=(0, 10))
        e_sat = ctk.CTkEntry(form, width=140, placeholder_text=T("aq_ph_sat"))
        e_sat.grid(row=8, column=1, sticky="w", padx=(8, 0), pady=(0, 10))
        e_sat.insert(0, self._addqso_ultimi['sat'])

        sat_status_var = _tk.StringVar(value="")
        lbl_sat_status = ctk.CTkLabel(form, textvariable=sat_status_var,
                                       font=ctk.CTkFont(size=9), justify="left",
                                       wraplength=140, anchor="w")
        lbl_sat_status.grid(row=9, column=1, sticky="nw", padx=(8, 0), pady=(0, 6))

        # Espone il campo Satellite per la precompilazione dal cruscotto tracking
        dlg._aq_sat = e_sat

        # ── Satellite: tratta RX (downlink) + SAT_MODE ───────────────
        # Sempre visibili sotto il campo Satellite: si auto-compilano
        # quando il satellite è noto nel database SAT_DB, e si azzerano
        # se il satellite non è riconosciuto. Per i QSO non satellitari
        # restano semplicemente vuoti (e non vengono scritti).
        var_banda_rx = _tk.StringVar(value='')
        ctk.CTkLabel(form, text=T("aq_sat_rx_hdr"), anchor="w",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="#4A90D9").grid(row=14, column=0, columnspan=2,
                                                sticky="w", pady=(6, 2))
        ctk.CTkLabel(form, text=T("aq_sat_rx_banda"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=15, column=0, sticky="w")
        ctk.CTkLabel(form, text=T("aq_sat_rx_freq"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=15, column=1, sticky="w", padx=(8, 0))
        cb_banda_rx = ctk.CTkOptionMenu(form, values=BANDE, variable=var_banda_rx, width=140)
        cb_banda_rx.grid(row=16, column=0, sticky="w", pady=(0, 6))
        e_freq_rx = ctk.CTkEntry(form, width=140, placeholder_text="—")
        e_freq_rx.grid(row=16, column=1, sticky="w", padx=(8, 0), pady=(0, 6))
        ctk.CTkLabel(form, text=T("aq_sat_mode"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=17, column=0, sticky="w")
        e_sat_mode = ctk.CTkEntry(form, width=140, placeholder_text="V/U")
        e_sat_mode.grid(row=18, column=0, sticky="w", pady=(0, 10))

        def _pulisci_rx():
            var_banda_rx.set('')
            e_freq_rx.delete(0, 'end')
            e_sat_mode.delete(0, 'end')

        def _riempi_rx(dati):
            var_banda_rx.set(dati['dn_band'])
            e_freq_rx.delete(0, 'end'); e_freq_rx.insert(0, dati['dn_freq'])
            e_sat_mode.delete(0, 'end'); e_sat_mode.insert(0, dati['mode'])

        def _verifica_satellite(*_a):
            sat_raw = e_sat.get().strip()
            if not sat_raw:
                sat_status_var.set("")
                lbl_sat_status.configure(text_color="gray")
                _pulisci_rx()
                return
            # Riconosce e normalizza varianti scritte senza spazi/trattini
            # per qualunque satellite noto (es. 'rs44' -> 'RS-44'), e le
            # diverse grafie testuali di QO-100 (ES'HAIL-2, QO100, ecc.).
            nome_canonico = self._normalizza_nome_satellite(sat_raw)
            sat_upper = nome_canonico or sat_raw.upper()
            if nome_canonico and nome_canonico != sat_raw:
                e_sat.delete(0, 'end'); e_sat.insert(0, nome_canonico)

            # Satellite con dati completi nel database: auto-compila la
            # coppia uplink/downlink e SAT_MODE, e mostra il pannello RX.
            dati = self._info_satellite(sat_upper)
            if dati:
                # TX (uplink): proponi solo se l'utente non l'ha già
                # personalizzato per questo stesso satellite in sessione.
                if self._addqso_ultimi.get('sat', '').upper() != sat_upper:
                    var_banda.set(dati['up_band'])
                    e_freq.delete(0, 'end'); e_freq.insert(0, dati['up_freq'])
                # RX (downlink) + SAT_MODE: sempre riallineati al satellite
                _riempi_rx(dati)
                sat_status_var.set(f"✓ {sat_upper} · {dati['mode']} · {dati['tipo']}")
                lbl_sat_status.configure(text_color=TH.OK_TEXT)
                return

            # Satellite noto solo come banda (vecchia tabella): niente RX.
            banda_tipica = self._banda_da_satellite(sat_upper)
            if banda_tipica:
                sat_status_var.set(f"✓ {sat_upper} riconosciuto ({banda_tipica})")
                lbl_sat_status.configure(text_color=TH.OK_TEXT)
                if self._addqso_ultimi.get('sat', '').upper() != sat_upper:
                    var_banda.set(banda_tipica)
                    freq_tipica = self._freq_da_banda(banda_tipica)
                    if freq_tipica:
                        e_freq.delete(0, 'end'); e_freq.insert(0, freq_tipica)
                _pulisci_rx()
            else:
                sat_status_var.set(T("addqso_sat_sconosciuto"))
                lbl_sat_status.configure(text_color=TH.WARN_TEXT)
                _pulisci_rx()
        e_sat.bind("<KeyRelease>", _verifica_satellite)
        e_sat.bind("<FocusOut>", _verifica_satellite)
        # Espone il verificatore per la precompilazione dal cruscotto tracking
        dlg._aq_verifica_sat = _verifica_satellite
        if e_sat.get().strip():
            _verifica_satellite()

        # ── RST TX/RX (auto da modo, modificabile) ───
        ctk.CTkLabel(form, text=T("addqso_rst_tx"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=10, column=0, sticky="w")
        ctk.CTkLabel(form, text=T("addqso_rst_rx"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=10, column=1, sticky="w", padx=(8, 0))
        e_rst_tx = ctk.CTkEntry(form, width=140)
        e_rst_tx.grid(row=11, column=0, sticky="w", pady=(0, 10))
        e_rst_rx = ctk.CTkEntry(form, width=140)
        e_rst_rx.grid(row=11, column=1, sticky="w", padx=(8, 0), pady=(0, 10))
        _su_cambio_modo()  # precompila subito all'apertura

        # ── Locator ──────────────────────────────────
        ctk.CTkLabel(form, text=T("addqso_locator"), anchor="w",
                     font=ctk.CTkFont(size=11)).grid(row=12, column=0, sticky="w")
        e_loc = ctk.CTkEntry(form, width=290, placeholder_text=T("aq_ph_loc"))
        e_loc.grid(row=13, column=0, columnspan=2, sticky="w", pady=(0, 14))

        n_inseriti = [0]

        def inserisci():
            call = e_call.get().strip().upper()
            if not call:
                messagebox.showwarning(T("attenzione"), T("addqso_err_call"), parent=dlg)
                return
            banda = var_banda.get().strip()
            if not banda:
                messagebox.showwarning(T("attenzione"), T("addqso_err_banda"), parent=dlg)
                return

            nuovo = adif_io.QSO({})
            nuovo['call'] = call
            nuovo['qso_date'] = e_data.get().strip()
            nuovo['time_on'] = e_ora.get().strip()
            nuovo['band'] = banda
            freq = e_freq.get().strip()
            if freq:
                nuovo['freq'] = freq
            nuovo['mode'] = var_modo.get().strip().upper()
            sat = e_sat.get().strip().upper()
            if sat:
                nuovo['sat_name'] = sat
                nuovo['prop_mode'] = 'SAT'
                # Tratta RX (downlink) + SAT_MODE, se compilati.
                brx   = var_banda_rx.get().strip()
                frx   = e_freq_rx.get().strip()
                smode = e_sat_mode.get().strip().upper()
                if brx:
                    nuovo['band_rx'] = brx
                if frx:
                    nuovo['freq_rx'] = frx
                if smode:
                    nuovo['sat_mode'] = smode
            loc = e_loc.get().strip().upper()
            if loc:
                nuovo['gridsquare'] = loc
            country = e_country.get().strip()
            if country:
                nuovo['country'] = country
                # Se il Country coincide con quanto dedotto dal prefisso,
                # arricchisce anche DXCC e Continente per coerenza con le
                # statistiche/pagina DXCC del PDF.
                risultato = dxcc_da_nominativo(call)
                if risultato and risultato[0] == country:
                    nuovo['dxcc'] = risultato[1]
                    nuovo['cont'] = risultato[2]
            nuovo['rst_sent'] = e_rst_tx.get().strip()
            nuovo['rst_rcvd'] = e_rst_rx.get().strip()
            profilo_call = self.entry_owner.get().strip().upper()
            if profilo_call:
                nuovo['station_callsign'] = profilo_call
                nuovo['operator'] = profilo_call
            profilo_grid = estrai_locator_da_testo(self.entry_details.get())
            if profilo_grid:
                nuovo['my_gridsquare'] = profilo_grid

            self.qsos_caricati.append(nuovo)
            self._log_modificato = True
            if self.qsos_filtrati:
                self.qsos_filtrati.append(nuovo)
            self._aggiorna_tree()

            # Memorizza i valori persistenti per il prossimo inserimento
            self._addqso_ultimi = {'banda': banda, 'freq': freq,
                                    'modo': var_modo.get().strip().upper(), 'sat': sat}

            # Svuota i campi specifici del QSO, mantiene banda/freq/modo/sat
            e_call.delete(0, 'end')
            e_country.delete(0, 'end')
            _country_auto[0] = True  # il prossimo Call può di nuovo auto-suggerire
            e_loc.delete(0, 'end')
            live_var.set(True)  # riprende l'orologio live per il prossimo QSO

            n_inseriti[0] += 1
            contatore_var.set(T("addqso_inseriti", n=n_inseriti[0]))
            e_call.focus_set()

        def _leggi_da_radio():
            """Legge frequenza e modo dalla radio via OmniRig e compila i campi."""
            rig = getattr(self, "_omnirig", None)
            if rig is None or not rig.disponibile():
                messagebox.showwarning(T("dxc_omnirig_no"),
                                       T("dxc_omnirig_assente"), parent=dlg)
                return
            hz = rig.get_freq()
            if not hz:
                # Mostra i valori grezzi per capire cosa vede OmniRig
                diag = ""
                try:
                    diag = rig.diagnostica()
                except Exception:
                    pass
                messagebox.showwarning(T("dxc_omnirig_no"),
                                       T("aq_radio_no_lettura") + "\n\n---\n" + diag,
                                       parent=dlg)
                return
            mhz = f"{hz/1_000_000:.6f}".rstrip('0').rstrip('.')
            e_freq.delete(0, 'end'); e_freq.insert(0, mhz)
            banda = self._banda_da_freq(hz/1_000_000)
            if banda:
                var_banda.set(banda)
            modo = rig.get_modo()
            if modo:
                # Se la radio è in USB-D (data), è quasi certamente un modo
                # digitale: usa il band plan per dedurre quale (FT8/RTTY...).
                if modo == "USB-D":
                    bp = modo_da_bandplan(hz)
                    modo = bp if bp and bp not in ("USB", "LSB", "CW", "AM", "FM") else "FT8"
                elif modo == "DIG":
                    modo = "FT8"
                elif modo in ("USB", "LSB"):
                    modo = "SSB"
                if modo in MODI:
                    var_modo.set(modo)

        # ── Barra inferiore FISSA: ancorata in fondo, sempre visibile
        #    anche se il form è più alto della finestra (schermi piccoli /
        #    DPI alto). Prima i pulsanti (side="bottom"), poi il form. ──
        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(side="bottom", fill="x", padx=24, pady=(6, 16))
        ctk.CTkButton(frame_btn, text=T("addqso_btn_inserisci"), command=inserisci,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                      height=40, font=ctk.CTkFont(size=13, weight="bold")
                      ).pack(side="left", expand=True, fill="x", padx=(0, 6))
        ctk.CTkButton(frame_btn, text=T("addqso_btn_chiudi"), command=_on_close,
                      fg_color="#718096", height=40).pack(side="left", expand=True, fill="x")

        _fr_radio = ctk.CTkFrame(dlg, fg_color="transparent")
        _fr_radio.pack(side="bottom", fill="x", padx=24, pady=(0, 4))
        ctk.CTkButton(_fr_radio, text=T("aq_leggi_radio"), command=_leggi_da_radio,
                      fg_color=TH.PRIMARY, hover_color=TH.PRIMARY_H,
                      height=32, font=ctk.CTkFont(size=12)).pack(fill="x")

        # Il form scorrevole riempie lo spazio centrale rimasto.
        form.pack(side="top", fill="both", expand=True, padx=20, pady=(0, 4))

        e_call.focus_set()
        dlg.bind("<Return>", lambda e: inserisci())

    def logga_qso_da_satellite(self, sat_name):
        """Apre (o riusa) la finestra Aggiungi QSO precompilando il
        satellite: nome e — se noto nel database SAT_DB — banda/freq TX,
        banda/freq RX e SAT_MODE (via _verifica_satellite). Pensato per il
        pulsante 'Logga QSO' del cruscotto di tracking satellitare."""
        if not sat_name:
            return
        canon = self._normalizza_nome_satellite(sat_name) or str(sat_name).strip().upper()
        dlg = getattr(self, "_aggiungi_qso_dlg", None)
        if dlg is None or not (hasattr(dlg, "winfo_exists") and dlg.winfo_exists()):
            self.apri_aggiungi_qso()
            dlg = getattr(self, "_aggiungi_qso_dlg", None)

        def _riempi():
            try:
                if hasattr(dlg, "_aq_sat"):
                    dlg._aq_sat.delete(0, 'end')
                    dlg._aq_sat.insert(0, canon)
                if hasattr(dlg, "_aq_verifica_sat"):
                    dlg._aq_verifica_sat()   # auto-compila TX/RX/SAT_MODE
                if hasattr(dlg, "_aq_call"):
                    dlg._aq_call.focus_set()
                try:
                    dlg.lift(); dlg.focus_force()
                except Exception:
                    pass
            except Exception:
                pass
        self.after(300, _riempi)

    def apri_calcolatore_distanza(self, locator_other_prefill=None):
        import tkinter as _tk

        dlg = ctk.CTkToplevel(self)
        dlg.title(T("dist_calc_title"))
        dlg.geometry("440x560")
        dlg.resizable(False, False)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        # Estrae il proprio locator dalle Note Operatore
        my_locator_guess = estrai_locator_da_testo(self.entry_details.get())

        ctk.CTkLabel(dlg, text="📡 " + T("dist_calc_title"),
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(16,12))

        form = ctk.CTkFrame(dlg, fg_color="transparent")
        form.pack(fill="x", padx=24)

        ctk.CTkLabel(form, text=T("dist_calc_my"), anchor="w",
                     font=ctk.CTkFont(size=12)).grid(row=0, column=0, sticky="w", pady=(0,2))
        e_my = ctk.CTkEntry(form, placeholder_text=T("dv_ph_myloc"), width=140)
        e_my.grid(row=1, column=0, sticky="w", padx=(0,8))
        if my_locator_guess:
            e_my.insert(0, my_locator_guess)

        # Pulsante inverti (la funzione calcola() viene assegnata più avanti)
        _calcola_ref = [lambda: None]

        def swap():
            a, b = e_my.get().strip(), e_other.get().strip()
            e_my.delete(0,'end'); e_my.insert(0, b)
            e_other.delete(0,'end'); e_other.insert(0, a)
            _calcola_ref[0]()

        ctk.CTkButton(form, text=T("dist_calc_swap"), command=swap,
                      width=80, height=28,
                      fg_color="#4A5568", hover_color="#2D3748",
                      font=ctk.CTkFont(size=11)).grid(row=1, column=1, padx=4)

        ctk.CTkLabel(form, text=T("dist_calc_other"), anchor="w",
                     font=ctk.CTkFont(size=12)).grid(row=0, column=2, sticky="w", pady=(0,2), padx=(8,0))
        e_other = ctk.CTkEntry(form, placeholder_text=T("dv_ph_jo65"), width=140)
        e_other.grid(row=1, column=2, sticky="w", padx=(8,0))
        if locator_other_prefill:
            e_other.insert(0, locator_other_prefill)

        # Risultati
        result_frame = ctk.CTkFrame(dlg, fg_color=("#EBF4FF","#101010"),
                                     border_width=1, border_color=("#2B6CB0","#242424"))
        result_frame.pack(fill="x", padx=24, pady=20)

        lbl_dist = ctk.CTkLabel(result_frame, text="—",
                                font=ctk.CTkFont(size=28, weight="bold"),
                                text_color=("#1A365D","#90CDF4"))
        lbl_dist.pack(pady=(16,2))
        lbl_dist_sub = ctk.CTkLabel(result_frame, text=T("dist_calc_dist"),
                                    font=ctk.CTkFont(size=11),
                                    text_color=("#4A5568","#94A3B8"))
        lbl_dist_sub.pack()

        lbl_bear = ctk.CTkLabel(result_frame, text="—",
                                font=ctk.CTkFont(size=22, weight="bold"),
                                text_color=("#2B6CB0","#63B3ED"))
        lbl_bear.pack(pady=(14,2))
        lbl_bear_sub = ctk.CTkLabel(result_frame, text=T("dist_calc_bearing"),
                                    font=ctk.CTkFont(size=11),
                                    text_color=("#4A5568","#94A3B8"))
        lbl_bear_sub.pack(pady=(0,16))

        # Mini bussola (Canvas)
        compass = _tk.Canvas(dlg, width=200, height=200, highlightthickness=0,
                              bg=self._canvas_bg())
        compass.pack(pady=(0,10))

        def draw_compass(bearing=None):
            compass.delete("all")
            cx, cy, r = 100, 100, 70
            bg = self._canvas_bg()
            compass.configure(bg=bg)
            fg = "#94A3B8"
            accent = "#3182CE"
            compass.create_oval(cx-r, cy-r, cx+r, cy+r, outline=fg, width=2)
            # Tacche cardinali
            for ang, lbl in [(0,"N"),(90,"E"),(180,"S"),(270,"W")]:
                rad = math.radians(ang - 90)
                x1 = cx + (r-8) * math.cos(rad)
                y1 = cy + (r-8) * math.sin(rad)
                x2 = cx + r * math.cos(rad)
                y2 = cy + r * math.sin(rad)
                compass.create_line(x1,y1,x2,y2, fill=fg, width=2)
                tx = cx + (r+16) * math.cos(rad)
                ty = cy + (r+16) * math.sin(rad)
                compass.create_text(tx, ty, text=lbl, fill=fg,
                                     font=("Segoe UI", 11, "bold"))
            if bearing is not None:
                rad = math.radians(bearing - 90)
                x2 = cx + (r-12) * math.cos(rad)
                y2 = cy + (r-12) * math.sin(rad)
                compass.create_line(cx, cy, x2, y2, fill=accent, width=3,
                                     arrow=_tk.LAST, arrowshape=(12,14,5))
            compass.create_oval(cx-3, cy-3, cx+3, cy+3, fill=fg, outline="")

        draw_compass()

        lbl_err = ctk.CTkLabel(dlg, text="", text_color=TH.DANGER,
                               font=ctk.CTkFont(size=11), wraplength=360)
        lbl_err.pack(pady=(0,6))

        def calcola():
            my = e_my.get().strip().upper()
            other = e_other.get().strip().upper()
            lbl_err.configure(text="")
            if not my and not other:
                lbl_dist.configure(text="—")
                lbl_bear.configure(text="—")
                draw_compass()
                return
            if not my or not other:
                lbl_err.configure(text=T("dist_calc_err"))
                return
            try:
                dist, bearing = distanza_bearing(my, other)
            except ValueError:
                lbl_dist.configure(text="—")
                lbl_bear.configure(text="—")
                draw_compass()
                lbl_err.configure(text=T("dist_calc_err"))
                return
            lbl_dist.configure(text=f"{dist:,.0f} km".replace(",", "."))
            compass_dir = bearing_to_compass(bearing)
            lbl_bear.configure(text=f"{bearing:.1f}°  ({compass_dir})")
            draw_compass(bearing)

        _calcola_ref[0] = calcola

        ctk.CTkButton(dlg, text=T("dist_calc_btn"), command=calcola,
                      height=38, fg_color=TH.PRIMARY, hover_color=TH.PRIMARY,
                      font=ctk.CTkFont(size=14, weight="bold")).pack(fill="x", padx=24, pady=(0,16))

        # Calcola subito se entrambi i campi sono già popolati
        if my_locator_guess and locator_other_prefill:
            calcola()

        e_my.bind("<Return>", lambda e: calcola())
        e_other.bind("<Return>", lambda e: calcola())

    def _canvas_bg(self):
        """Restituisce il colore di sfondo coerente col tema corrente per i Canvas."""
        is_dark = ctk.get_appearance_mode().lower() == "dark"
        return "#1a1a2e" if is_dark else "#F7FAFC"

    # ═══════════════════════════════════════════
    #  DOTTORE SAT — compatibilità satelliti/LoTW
    # ═══════════════════════════════════════════

    # Varianti note del nome QO-100 da normalizzare (LoTW richiede esattamente "QO-100")
    _QO100_VARIANTI = re.compile(r"^\s*(QO[\s\-]?100|ES[''`]?HAIL[\s\-]?2A?|ESHAILSAT[\s\-]?2)\s*(\(.*\))?\s*$", re.I)

    def _analizza_qso_sat(self, qso):
        """Analizza un singolo QSO e restituisce una lista di problemi rilevati,
        ognuno come dict: {testo, campo, valore_attuale, valore_proposto, sicuro}.
        'sicuro' indica se la correzione può essere pre-selezionata di default."""
        problemi = []
        sat_name = str(qso.get('sat_name', '')).strip()
        prop_mode = str(qso.get('prop_mode', '')).strip().upper()
        band = str(qso.get('band', '')).strip().upper()
        try:
            freq = float(qso.get('freq', '0') or 0)
        except (ValueError, TypeError):
            freq = 0.0

        is_qo100 = bool(sat_name) and self._QO100_VARIANTI.match(sat_name)
        sat_name_normalizzato = "QO-100" if is_qo100 else sat_name

        # 1) SAT_NAME presente ma PROP_MODE non è SAT
        if sat_name and prop_mode != "SAT":
            problemi.append({
                'testo': f"PROP_MODE manca o errato (attuale: '{prop_mode or '(vuoto)'}'). LoTW richiede PROP_MODE=SAT quando è presente SAT_NAME.",
                'campo': 'prop_mode', 'valore_attuale': prop_mode or '(vuoto)',
                'valore_proposto': 'SAT', 'sicuro': True,
            })

        # 2) PROP_MODE=SAT ma SAT_NAME vuoto
        if prop_mode == "SAT" and not sat_name:
            problemi.append({
                'testo': "PROP_MODE=SAT ma SAT_NAME è vuoto. LoTW rifiuterà questo QSO: "
                         "compila il campo Satellite nel pannello di modifica.",
                'campo': 'sat_name', 'valore_attuale': '(vuoto)',
                'valore_proposto': None, 'sicuro': False,
            })

        # 3) Nome QO-100 non nella grafia esatta richiesta da LoTW
        if is_qo100 and sat_name != "QO-100":
            problemi.append({
                'testo': f"Nome satellite '{sat_name}' non corrisponde alla grafia esatta "
                         f"richiesta da LoTW per QO-100 ('QO-100').",
                'campo': 'sat_name', 'valore_attuale': sat_name,
                'valore_proposto': 'QO-100', 'sicuro': True,
            })

        # 4) Banda per QO-100: LoTW preferisce 13CM (uplink/TX), non 3CM (downlink/RX)
        if is_qo100:
            if band == "13CM":
                pass  # OK
            elif 2400.0 <= freq <= 2410.0:
                # Frequenza coerente con 13cm uplink: fix sicuro
                problemi.append({
                    'testo': f"Banda '{band or '(vuota)'}' non coerente con la frequenza TX "
                             f"({freq} MHz, banda 13cm). LoTW richiede BAND=13CM per QO-100.",
                    'campo': 'band', 'valore_attuale': band or '(vuoto)',
                    'valore_proposto': '13CM', 'sicuro': True,
                })
            elif 10000.0 <= freq <= 10500.0:
                # Probabile frequenza RX (downlink, 3cm) loggata invece della TX
                problemi.append({
                    'testo': f"Frequenza {freq} MHz e banda '{band or '(vuota)'}' sembrano "
                             f"riferirsi al downlink (RX, 3cm). LoTW vuole la frequenza/banda "
                             f"TX (13cm, ~2400 MHz). Correggi manualmente FREQ e BAND nel "
                             f"pannello di modifica.",
                    'campo': 'band', 'valore_attuale': band or '(vuoto)',
                    'valore_proposto': None, 'sicuro': False,
                })
            elif band != "13CM":
                # Banda vuota o non riconosciuta, nessuna frequenza utile: default 13CM
                problemi.append({
                    'testo': f"Banda '{band or '(vuota)'}' non standard per QO-100. "
                             f"LoTW richiede BAND=13CM (uplink). Verifica che sia corretto "
                             f"prima di applicare.",
                    'campo': 'band', 'valore_attuale': band or '(vuoto)',
                    'valore_proposto': '13CM', 'sicuro': False,
                })

        return problemi

    def apri_dottore_sat(self):
        """Analizza il log alla ricerca di QSO satellite con campi non
        conformi ai requisiti LoTW (PROP_MODE, SAT_NAME, BAND per QO-100)
        e propone correzioni guidate."""
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return

        # Raccoglie i QSO con problemi
        risultati = []  # lista di (idx_in_caricati, qso, problemi)
        for idx, qso in enumerate(self.qsos_caricati):
            sat_name = str(qso.get('sat_name', '')).strip()
            prop_mode = str(qso.get('prop_mode', '')).strip().upper()
            if not sat_name and prop_mode != "SAT":
                continue  # non è un QSO satellite, salta
            problemi = self._analizza_qso_sat(qso)
            if problemi:
                risultati.append((idx, qso, problemi))

        # ── Finestra risultati ──────────────────────
        dlg = ctk.CTkToplevel(self)
        dlg.title("🛰️ Dottore SAT — Compatibilità LoTW")
        dlg.geometry("780x560")
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        header = ctk.CTkFrame(dlg, fg_color="transparent")
        header.pack(fill="x", padx=16, pady=(14,6))
        ctk.CTkLabel(header, text="🛰️ Dottore SAT — Verifica compatibilità LoTW",
                      font=ctk.CTkFont(size=15, weight="bold")).pack(anchor="w")

        n_sat = sum(1 for q in self.qsos_caricati
                    if str(q.get('sat_name','')).strip() or str(q.get('prop_mode','')).strip().upper()=="SAT")

        if not risultati:
            ctk.CTkLabel(header,
                text=f"Analizzati {n_sat} QSO satellite — nessun problema rilevato. ✅",
                font=ctk.CTkFont(size=12), text_color=TH.OK_TEXT).pack(anchor="w", pady=(6,0))
            ctk.CTkButton(dlg, text=T("cm_chiudi"), command=dlg.destroy,
                          height=34, width=100).pack(pady=20)
            return

        ctk.CTkLabel(header,
            text=f"Trovati {len(risultati)} QSO con possibili problemi su {n_sat} QSO satellite totali.\n"
                 f"Le correzioni con il segno ✓ sono sicure e pre-selezionate; quelle con ⚠ "
                 f"richiedono la tua verifica prima di applicarle.",
            font=ctk.CTkFont(size=11), text_color=("#4A5568","#94A3B8"),
            justify="left").pack(anchor="w", pady=(4,0))

        scroll = ctk.CTkScrollableFrame(dlg)
        scroll.pack(fill="both", expand=True, padx=16, pady=8)

        # checkbox_vars[(idx, campo)] = (BooleanVar, valore_proposto)
        checkbox_vars = {}
        # entry_vars[(idx, campo)] = CTkEntry, per i problemi senza correzione
        # automatica (valore_proposto is None): editing diretto nella finestra.
        entry_vars = {}
        idx_non_risolti = set()

        for idx, qso, problemi in risultati:
            call = str(qso.get('call','')).upper()
            data = str(qso.get('qso_date',''))
            if len(data) == 8:
                data = f"{data[6:8]}/{data[4:6]}/{data[0:4]}"
            sat = str(qso.get('sat_name','')).strip() or "—"

            card = ctk.CTkFrame(scroll, fg_color=("#EBF4FF","#161616"),
                                 border_width=1, border_color=("#CBD5E0","#2D3748"))
            card.pack(fill="x", pady=4)

            top = ctk.CTkFrame(card, fg_color="transparent")
            top.pack(fill="x", padx=10, pady=(8,2))
            ctk.CTkLabel(top, text=f"QSO #{idx+1} — {call}  {data}  —  SAT: {sat}",
                          font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w")

            for p in problemi:
                row = ctk.CTkFrame(card, fg_color="transparent")
                row.pack(fill="x", padx=10, pady=2)

                if p['valore_proposto'] is not None:
                    var = ctk.BooleanVar(value=p['sicuro'])
                    icona = "✓" if p['sicuro'] else "⚠"
                    colore = "#48BB78" if p['sicuro'] else "#F6AD55"
                    cb = ctk.CTkCheckBox(row, text="", variable=var, width=20,
                                          checkbox_width=18, checkbox_height=18)
                    cb.pack(side="left", padx=(0,6))
                    checkbox_vars[(idx, p['campo'])] = (var, p['valore_proposto'])
                    ctk.CTkLabel(row, text=icona, text_color=colore,
                                  font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(0,4))
                    testo_corr = (f"  →  {p['campo'].upper()}: '{p['valore_attuale']}' "
                                   f"⟹ '{p['valore_proposto']}'")
                    ctk.CTkLabel(row, text=p['testo'] + testo_corr,
                                  font=ctk.CTkFont(size=11),
                                  text_color=("#2D3748","#CBD5E0"),
                                  justify="left", wraplength=620, anchor="w").pack(side="left", fill="x")
                else:
                    # Nessuna correzione automatica disponibile: campo editabile
                    # direttamente qui, così non serve uscire dalla finestra.
                    idx_non_risolti.add(idx)
                    ctk.CTkLabel(row, text="⚠", text_color=TH.WARN_TEXT,
                                  font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(0,4))
                    ctk.CTkLabel(row, text=p['testo'],
                                  font=ctk.CTkFont(size=11),
                                  text_color=("#2D3748","#CBD5E0"),
                                  justify="left", wraplength=460, anchor="w").pack(side="left", fill="x")
                    entry_row = ctk.CTkFrame(card, fg_color="transparent")
                    entry_row.pack(fill="x", padx=(34,10), pady=(0,4))
                    ctk.CTkLabel(entry_row, text=f"{p['campo'].upper()}:",
                                  font=ctk.CTkFont(size=10, weight="bold"), width=70,
                                  anchor="w").pack(side="left")
                    e = ctk.CTkEntry(entry_row, width=200,
                                      placeholder_text=T("sat_placeholder_valore"))
                    e.pack(side="left", padx=(4,0))
                    entry_vars[(idx, p['campo'])] = e

        # ── Pulsanti azione ──────────────────────────
        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(fill="x", padx=16, pady=(0,14))

        def applica():
            n_applicate = 0
            for (idx, campo), (var, nuovo_val) in checkbox_vars.items():
                if var.get():
                    self.qsos_caricati[idx][campo] = nuovo_val
                    n_applicate += 1
            n_manuali = 0
            for (idx, campo), entry in entry_vars.items():
                valore = entry.get().strip()
                if valore:
                    self.qsos_caricati[idx][campo] = valore
                    n_manuali += 1
                    idx_non_risolti.discard(idx)
            if n_applicate or n_manuali:
                self._aggiorna_tree()
                msg = T("sat_msg_applicate", n=n_applicate)
                if n_manuali:
                    msg += T("sat_msg_manuali", n=n_manuali)
                msg += T("sat_msg_salva")
                messagebox.showinfo("Dottore SAT", msg)
            dlg.destroy()

        def esporta_non_risolti():
            if not idx_non_risolti:
                messagebox.showinfo("Dottore SAT", T("sat_export_nessuno"))
                return
            qsos_da_esportare = [self.qsos_caricati[i] for i in sorted(idx_non_risolti)]
            path = chiedi_cartella_output(dlg, "dottore_sat_da_correggere.adif",
                                           filepath_sorgente=self.filepath)
            if not path:
                return
            try:
                self._scrivi_adif(path, qsos_da_esportare)
                messagebox.showinfo("Dottore SAT",
                    T("sat_export_fatto", n=len(qsos_da_esportare), path=path))
            except Exception as ex:
                messagebox.showerror(T("errore"), T("sat_export_errore", err=ex))

        ctk.CTkButton(btns, text="✔ Applica correzioni selezionate", command=applica,
                      height=36, fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                      font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(0,8))
        if idx_non_risolti:
            ctk.CTkButton(btns, text=T("sat_btn_esporta"), command=esporta_non_risolti,
                          height=36, fg_color=TH.WARNING_H, hover_color=TH.WARNING_H,
                          font=ctk.CTkFont(size=12)).pack(side="left", padx=(0,8))
        ctk.CTkButton(btns, text=T("cm_annulla"), command=dlg.destroy,
                      height=36, width=100, fg_color="#4A5568",
                      hover_color="#2D3748").pack(side="left")

    # ── Imposta lingua ─────────────────────────
    def _set_lingua(self, lang):
        self.var_lingua.set(lang)
        imposta_lingua(lang)
        self._aggiorna_lingua()

    # ── Imposta tema ───────────────────────────
    def _set_tema(self, tema):
        self.var_tema.set(tema)
        ctk.set_appearance_mode(tema)
        self._aggiorna_colori_tree()

    def _on_main_focus(self, event=None):
        """Chiamato quando la finestra principale riceve il focus.
        Riapplica i colori del treeview nel caso siano stati alterati
        da un dialog etichette (o altro) ancora aperto in background."""
        if not hasattr(self, '_focus_after_id'):
            self._focus_after_id = None
        # Throttle: esegui al massimo una volta ogni 300ms
        if self._focus_after_id:
            self.after_cancel(self._focus_after_id)
        self._focus_after_id = self.after(300, self._aggiorna_colori_tree)

    # Tutte le colonne disponibili con label e larghezza default
    _COLONNE_DEF = [
        ("n",       "#",         38,  "center"),
        ("data",    "Data",      75,  "center"),
        ("utc",     "UTC",       52,  "center"),
        ("call",    "Callsign",  90,  "w"),
        ("nome",    "Nome Op.",  100, "w"),
        ("banda",   "Banda",     55,  "center"),
        ("sat",     "Satellite", 80,  "w"),
        ("prop",    "Prop.",     55,  "center"),
        ("modo",    "Mode",      55,  "center"),
        ("contest", "Contest",   80,  "w"),
        ("freq",    "Freq.",     65,  "center"),
        ("rst_s",   "RST TX",    52,  "center"),
        ("rst_r",   "RST RX",    52,  "center"),
        ("country", "Country",   150, "w"),
        ("state",   "Stato",     55,  "center"),
        ("locator", "Locator",   65,  "center"),
        ("lotw",    "LoTW",      42,  "center"),
        ("eqsl",    "eQSL",      42,  "center"),
        ("banda_rx","Banda RX",  60,  "center"),
        ("freq_rx", "Freq. RX",  70,  "center"),
        ("sat_mode","SAT Mode",  60,  "center"),
    ]
    # Colonne nascoste di default
    _COLONNE_HIDDEN_DEFAULT = {"sat", "prop", "contest", "state"}

    def _ordine_colonne(self):
        """Restituisce (ordine, nascoste) dal profilo o default."""
        default_ord = [c[0] for c in self._COLONNE_DEF]
        default_hid = set(self._COLONNE_HIDDEN_DEFAULT)
        try:
            profili = self._carica_profili()
            dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
            ord_s = dati.get('colonne_ordine')
            hid_s = dati.get('colonne_nascoste')
            if ord_s and isinstance(ord_s, list):
                tutti = {c[0] for c in self._COLONNE_DEF}
                validi   = [c for c in ord_s if c in tutti]
                mancanti = [c for c in default_ord if c not in validi]
                return validi + mancanti, set(hid_s or [])
        except Exception:
            pass
        return default_ord, default_hid

    def _applica_colonne(self):
        """Applica ordine e visibilità colonne al treeview."""
        ordine, nascoste = self._ordine_colonne()
        visibili = [c for c in ordine if c not in nascoste]
        try:
            self.tree.configure(displaycolumns=visibili)
        except Exception:
            pass

    def apri_dx_cluster(self):
        """Apre la finestra DX Cluster (singola istanza)."""
        if hasattr(self, '_dxc_win') and self._dxc_win and self._dxc_win.winfo_exists():
            self._dxc_win.lift(); self._dxc_win.focus_force()
            return
        self._dxc_win = DXClusterWindow(self, self)

    def apri_satelliti(self):
        """Apre la finestra di predizione passaggi satellitari (LEO)."""
        try:
            SatellitiDialog(self, self)
        except Exception as e:
            import traceback as _tb
            from tkinter import messagebox as _mb
            dettaglio = _tb.format_exc()
            try:
                _p = __import__("os").path.join(__import__("os").path.expanduser("~"),
                                                "adif_fzr_errore_satelliti.txt")
                open(_p, "w", encoding="utf-8").write(dettaglio)
            except Exception:
                _p = "(impossibile salvare)"
            _mb.showerror("Satelliti",
                "Errore aprendo la finestra:\n\n" + dettaglio[-1500:] +
                "\n\nSalvato anche in:\n" + _p)

    def apri_wsjtx(self):
        """Apre il listener WSJT-X Live (singola istanza)."""
        if hasattr(self, '_wsjtx_win') and self._wsjtx_win and self._wsjtx_win.winfo_exists():
            self._wsjtx_win.lift(); self._wsjtx_win.focus_force()
            return
        self._wsjtx_win = WSJTXListener(self, self)

    def apri_personalizza_colonne(self):
        """Dialog per personalizzare ordine e visibilità colonne della griglia."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("📋 Personalizza colonne griglia")
        dlg.geometry("400x520")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text="📋 Colonne della griglia principale",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(14,2))
        ctk.CTkLabel(dlg, text=T("dv_spunta_mostra"),
                     font=ctk.CTkFont(size=10), text_color="gray").pack(pady=(0,8))

        ordine_corrente, nascoste = self._ordine_colonne()
        label_map = {c[0]: c[1] for c in self._COLONNE_DEF}
        fissi = {"n"}  # colonna # sempre visibile

        scroll = ctk.CTkScrollableFrame(dlg, width=360, height=360)
        scroll.pack(fill="both", expand=True, padx=14, pady=(0,8))
        var_check = {}

        def _ridisegna():
            for w in scroll.winfo_children(): w.destroy()
            for col_id in ordine_corrente:
                lbl = label_map.get(col_id, col_id)
                row = ctk.CTkFrame(scroll, fg_color="transparent")
                row.pack(fill="x", pady=2)
                v = var_check.get(col_id, ctk.BooleanVar(value=col_id not in nascoste))
                var_check[col_id] = v
                cb = ctk.CTkCheckBox(row, text=lbl, variable=v,
                                     font=ctk.CTkFont(size=11), width=180)
                if col_id in fissi:
                    cb.configure(state="disabled")
                cb.pack(side="left")
                ctk.CTkButton(row, text="↑", width=28, height=24, fg_color="#4A5568",
                              command=lambda c=col_id: _sposta(c,-1)).pack(side="left", padx=2)
                ctk.CTkButton(row, text="↓", width=28, height=24, fg_color="#4A5568",
                              command=lambda c=col_id: _sposta(c,1)).pack(side="left", padx=2)

        def _sposta(col_id, delta):
            i = ordine_corrente.index(col_id)
            j = i + delta
            if 0 <= j < len(ordine_corrente):
                ordine_corrente[i], ordine_corrente[j] = ordine_corrente[j], ordine_corrente[i]
                _ridisegna()

        _ridisegna()

        def _salva():
            nascoste_nuove = {c for c in ordine_corrente
                              if c not in fissi and not var_check[c].get()}
            if self.profilo_attivo:
                try:
                    profili2 = self._carica_profili()
                    if self.profilo_attivo in profili2:
                        profili2[self.profilo_attivo]['colonne_ordine']   = ordine_corrente
                        profili2[self.profilo_attivo]['colonne_nascoste'] = list(nascoste_nuove)
                        with open(self.profili_path, 'w', encoding='utf-8') as f:
                            json.dump(profili2, f, ensure_ascii=False, indent=2)
                except Exception as ex:
                    messagebox.showerror(T("errore"), str(ex), parent=dlg)
                    return
            nascoste.clear(); nascoste.update(nascoste_nuove)
            self._applica_colonne()
            dlg.destroy()

        fr = ctk.CTkFrame(dlg, fg_color="transparent")
        fr.pack(fill="x", padx=14, pady=(0,14))
        ctk.CTkButton(fr, text="✔ Applica", command=_salva, height=34,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS
                      ).pack(side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(fr, text=T("cm_annulla"), command=dlg.destroy,
                      height=34, width=90, fg_color="#718096").pack(side="left")

    def _aggiorna_colori_tree(self):
        """Aggiorna i colori del Treeview in base al tema corrente."""
        import tkinter.ttk as _ttv
        is_dark = ctk.get_appearance_mode().lower() == "dark"
        if is_dark:
            bg_tree, bg_even, fg_tree, bg_field = "#12181F", "#1A222B", "#DCE3EC", "#12181F"
            head_bg, head_fg, head_act = "#1E2A38", "#9FB3C8", "#2B3B4E"
        else:
            bg_tree, bg_even, fg_tree, bg_field = "#FFFFFF", "#F1F5FA", "#1F2A37", "#FFFFFF"
            head_bg, head_fg, head_act = "#EDF2F7", "#3A4A61", "#DCE6F0"

        style = _ttv.Style()
        style.configure("QSO.Treeview",
                        background=bg_tree, foreground=fg_tree,
                        rowheight=28, fieldbackground=bg_field)
        style.configure("QSO.Treeview.Heading",
                        background=head_bg, foreground=head_fg)
        style.map("QSO.Treeview.Heading", background=[("active", head_act)])
        self.tree.tag_configure("even",    background=bg_even)
        self.tree.tag_configure("odd",     background=bg_tree)
        self.tree.tag_configure("ft8",     background="#12261A" if is_dark else "#E6F4EA", foreground="#4ADE80" if is_dark else "#166534")
        self.tree.tag_configure("ft4",     background="#12261A" if is_dark else "#E6F4EA", foreground="#4ADE80" if is_dark else "#166534")
        self.tree.tag_configure("cw",      background="#12202E" if is_dark else "#EBF4FF", foreground="#90CDF4" if is_dark else "#1A4480")
        self.tree.tag_configure("sat_row", background="#1E122E" if is_dark else "#FAF0FF", foreground="#D8B4FE" if is_dark else "#6B21A8")
        self.tree.tag_configure("eme",     background="#2A1212" if is_dark else "#FFF0F0", foreground="#FCA5A5" if is_dark else "#991B1B")

    # ── About ──────────────────────────────────
    def _apri_manuale(self):
        import webbrowser
        webbrowser.open("https://iw1fzr.it/manuale-adif-fzr/")

    def _about(self):
        from tkinter import messagebox as _mb
        _mb.showinfo(f"ADIF FZR {VERSIONE}",
            f"ADIF FZR {VERSIONE}  build {BUILD_DATE}\nProfessional Ham Radio Logbook Utility\n\n"
            "Autore: IW1FZR — Luca\nhttps://iw1fzr.it\n\n"
            "CustomTkinter + ReportLab + openpyxl")

    # ── Cambio lingua (compatibilità) ──────────
    def cambia_lingua(self, lang):
        self._set_lingua(lang)

    # ── Cambio tema (compatibilità) ────────────
    def _cambia_tema_compat_unused(self, tema):
        self._set_tema(tema)
        self._aggiorna_tree()

    def _freq_da_banda(self, banda):
        """Restituisce frequenza di default in MHz per una banda ADIF."""
        FREQ_BAND = {
            '2190m':'0.1365','630m':'0.475','560m':'0.501',
            '160m':'1.900','80m':'3.700','60m':'5.357',
            '40m':'7.100','30m':'10.130','20m':'14.225',
            '17m':'18.118','15m':'21.300','12m':'24.940',
            '10m':'28.500','6m':'51.000','4m':'70.200',
            '2m':'144.300','1.25m':'222.100','70cm':'432.200',
            '33cm':'902.100','23cm':'1296.200','13cm':'2320.200',
            '9cm':'3400.100','6cm':'5760.100','3cm':'10368.100',
            '1.25cm':'24048.100','6mm':'47088.100',
        }
        return FREQ_BAND.get(str(banda).lower().strip(), '')

    def _banda_da_freq(self, freq_mhz):
        """Deduce la banda ADIF dalla frequenza in MHz.
        Usata per compilare BAND quando è assente ma FREQ è presente."""
        try:
            f = float(str(freq_mhz).strip())
        except (ValueError, TypeError):
            return ''
        # Limiti di banda in MHz (da/a) → nome banda ADIF
        LIMITI = [
            (0.1357,  0.1378,  '2190m'),
            (0.472,   0.479,   '630m'),
            (0.500,   0.504,   '560m'),
            (1.800,   2.000,   '160m'),
            (3.500,   4.000,   '80m'),
            (5.330,   5.410,   '60m'),
            (7.000,   7.300,   '40m'),
            (10.100,  10.150,  '30m'),
            (14.000,  14.350,  '20m'),
            (18.068,  18.168,  '17m'),
            (21.000,  21.450,  '15m'),
            (24.890,  24.990,  '12m'),
            (28.000,  29.700,  '10m'),
            (50.000,  54.000,  '6m'),
            (70.000,  70.500,  '4m'),
            (144.000, 148.000, '2m'),
            (219.000, 225.000, '1.25m'),
            (420.000, 450.000, '70cm'),
            (902.000, 928.000, '33cm'),
            (1240.0,  1300.0,  '23cm'),
            (2300.0,  2450.0,  '13cm'),
            (3300.0,  3500.0,  '9cm'),
            (5650.0,  5925.0,  '6cm'),
            (10000.0, 10500.0, '3cm'),
            (24000.0, 24250.0, '1.25cm'),
            (47000.0, 47200.0, '6mm'),
        ]
        for f_min, f_max, banda in LIMITI:
            if f_min <= f <= f_max:
                return banda
        return ''

    _SAT_BAND = {
        'QO-100':  '13cm',   # transponder lineare geostazionario, uplink 13cm
        'ES-HAIL-2': '13cm',
        'RS-44':   '2m',     # FM/lineare LEO, uplink tipico 2m
        'VO-52':   '2m',
        'AO-7':    '2m',
        'AO-91':   '2m',
        'AO-92':   '2m',
        'SO-50':   '2m',
        'FO-29':   '2m',
        'PO-101':  '2m',
        'ISS':     '2m',
    }
    # Mappa "forma senza separatori" -> "forma canonica con trattino", per
    # riconoscere varianti scritte senza spazi/trattini (es. 'rs44' -> 'RS-44').
    _SAT_NOME_CANONICO = {
        re.sub(r'[\s\-]', '', k): k for k in _SAT_BAND
    }

    def _normalizza_nome_satellite(self, sat_name):
        """Restituisce la forma canonica (con trattino) di un nome
        satellite noto, anche se scritto senza spazi/trattini (es. 'rs44'
        -> 'RS-44', 'vo52' -> 'VO-52'). Per QO-100 gestisce anche le
        varianti testuali via _QO100_VARIANTI. Ritorna None se il nome
        non corrisponde a nessun satellite noto.
        Consulta prima il database frequenze SAT_DB (più ricco), poi la
        vecchia tabellina _SAT_BAND per retrocompatibilità."""
        sat_raw = str(sat_name).strip()
        if not sat_raw:
            return None
        # 1) database frequenze (RS-44, CAS-4A, JO-97, QO-100, ISS (ZARYA)…)
        canon = SATDB.normalizza(sat_raw)
        if canon:
            return canon
        # 2) fallback storico
        if self._QO100_VARIANTI.match(sat_raw):
            return "QO-100"
        chiave_compatta = re.sub(r'[\s\-]', '', sat_raw).upper()
        return self._SAT_NOME_CANONICO.get(chiave_compatta)

    def _info_satellite(self, sat_name):
        """Restituisce il dict up/down/mode/tipo del satellite dal database
        frequenze (SAT_DB), o None se non noto. Chiavi: up_band, up_freq,
        dn_band, dn_freq, mode, tipo, nome."""
        return SATDB.info(sat_name)

    def _banda_da_satellite(self, sat_name):
        """Restituisce la banda ADIF di uplink (TX) per un satellite noto,
        usata per correggere bande placeholder/non riconosciute (es. 'infm')
        sui QSO satellitari. Preferisce il database frequenze SAT_DB, con
        fallback alla vecchia tabella. Vuoto se il satellite non è in elenco."""
        b = SATDB.banda_uplink(sat_name)
        if b:
            return b
        return self._SAT_BAND.get(str(sat_name).upper().strip(), '')

    def _banda_non_valida(self, banda):
        """True se 'banda' non è una banda ADIF riconosciuta (placeholder
        come 'infm', vuota, o valore non presente nella tabella frequenze)."""
        b = str(banda).lower().strip()
        if not b or b == 'infm':
            return True
        FREQ_BAND_KEYS = {
            '2190m','630m','560m','160m','80m','60m','40m','30m','20m',
            '17m','15m','12m','10m','6m','4m','2m','1.25m','70cm',
            '33cm','23cm','13cm','9cm','6cm','3cm','1.25cm','6mm',
        }
        return b not in FREQ_BAND_KEYS

    def _correggi_banda_satellite(self, qsos):
        """Per i QSO satellitari (SAT_NAME presente o PROP_MODE=SAT) con una
        banda non riconosciuta (es. 'infm'), se il satellite è noto imposta
        la banda corretta e calcola la FREQ di default corrispondente.
        Ritorna il numero di QSO corretti."""
        n_corretti = 0
        for qso in qsos:
            keys_low = {k.lower(): k for k in qso.keys()}
            sat_name = qso.get(keys_low.get('sat_name', 'SAT_NAME'), '') if 'sat_name' in keys_low else ''
            prop_mode = qso.get(keys_low.get('prop_mode', 'PROP_MODE'), '') if 'prop_mode' in keys_low else ''
            if not sat_name and str(prop_mode).upper() != 'SAT':
                continue
            banda_attuale = qso.get(keys_low.get('band', 'BAND'), '') if 'band' in keys_low else ''
            if not self._banda_non_valida(banda_attuale):
                continue
            banda_corretta = self._banda_da_satellite(sat_name)
            if not banda_corretta:
                continue
            band_key = keys_low.get('band', 'band')
            qso[band_key] = banda_corretta.upper() if len(banda_corretta) <= 4 else banda_corretta
            freq_key = keys_low.get('freq', 'freq')
            freq_attuale = qso.get(freq_key, '') if freq_key in qso else ''
            if not str(freq_attuale).strip():
                qso[freq_key] = self._freq_da_banda(banda_corretta)
            n_corretti += 1
        return n_corretti

    def _normalizza_fase1_omogenizza(self, qsos):
        """Fase 1: uniforma i campi presenti su tutti i QSO della lista data.
        Per FREQ (assente dal singolo QSO, assente da TUTTO il log, oppure
        presente ma vuota/zero) calcola sempre la frequenza di default dalla
        banda. Ritorna (n_omogenea, tutti_campi, campi_critici_assenti):
        - n_omogenea: numero di campi scritti/omogenizzati
        - tutti_campi: insieme di tutti i campi presenti nel log (+ 'freq')
        - campi_critici_assenti: campi importanti (es. 'country') che non
          compaiono in NESSUN QSO del log, segnalati ma non riempiti
          automaticamente perché non c'è un modo sicuro di dedurli (es.
          COUNTRY richiederebbe una tabella prefissi->DXCC affidabile,
          non disponibile)."""
        ESCLUDI = {"adif_ver","programid","programversion","created_timestamp",
                   "app_l4ong_qso_confirmations","app_l4ong_qso_award_references"}
        CAMPI_CRITICI = {"country"}  # estendibile in futuro
        tutti_campi = set()
        for qso in qsos:
            tutti_campi.update(k.lower() for k in qso.keys() if k.lower() not in ESCLUDI)
        campi_critici_assenti = sorted(CAMPI_CRITICI - tutti_campi)
        n_omogenea = 0
        for qso in qsos:
            chiavi_qso = {k.lower() for k in qso.keys()}
            for campo in tutti_campi:
                if campo in ("freq", "band"):
                    continue  # gestiti sempre a parte, sotto
                if campo not in chiavi_qso:
                    qso[campo] = ""
                    n_omogenea += 1

            # FREQ: calcola dalla banda se assente/vuota/zero
            freq_key = next((k for k in qso.keys() if k.lower()=="freq"), None)
            band_key = next((k for k in qso.keys() if k.lower()=="band"), None)
            banda_val = str(qso.get(band_key, "")).strip() if band_key else ""
            freq_val  = str(qso.get(freq_key, "")).strip() if freq_key else ""

            if freq_key is None:
                if banda_val:
                    qso["freq"] = self._freq_da_banda(banda_val)
                    n_omogenea += 1
            else:
                if not freq_val or freq_val in ("0", "0.0", "0.000", "0.000000"):
                    if banda_val:
                        nuovo_val = self._freq_da_banda(banda_val)
                        if nuovo_val:
                            qso[freq_key] = nuovo_val
                            n_omogenea += 1

            # BAND: deduce dalla FREQ se BAND è assente o vuota
            # (es. QSO da LoTW o logger che scrive solo la frequenza)
            if band_key is None:
                if freq_val or (freq_key and str(qso.get(freq_key,'')).strip()):
                    f = freq_val or str(qso.get(freq_key,'')).strip()
                    banda_dedotta = self._banda_da_freq(f)
                    if banda_dedotta:
                        qso["band"] = banda_dedotta.upper()
                        n_omogenea += 1
            else:
                if not banda_val:
                    f = str(qso.get(freq_key, '')).strip() if freq_key else ''
                    if f:
                        banda_dedotta = self._banda_da_freq(f)
                        if banda_dedotta:
                            qso[band_key] = banda_dedotta.upper()
                            n_omogenea += 1

        tutti_campi.add("freq")
        tutti_campi.add("band")
        return n_omogenea, tutti_campi, campi_critici_assenti


    def _normalizza_fase2_dialogo(self, qsos, tutti_campi, n_omogenea, parent,
                                   on_done=None, n_sat_corretti=0, silenzioso=False,
                                   campi_critici_assenti=None):
        """Fase 2: mostra il dialogo dei valori di default per i campi
        completamente assenti dallo schema standard. parent è la finestra
        (Toplevel o self) su cui ancorare il dialogo; on_done è una callback
        opzionale invocata dopo l'applicazione/chiusura. n_sat_corretti è il
        numero di QSO satellitari con banda placeholder corretta a monte.
        Se silenzioso=True, non mostra il popup "Completato" quando non ci
        sono campi assenti da gestire (usato dal controllo automatico post-
        apertura, per non disturbare l'utente quando il log è già a posto).
        campi_critici_assenti è una lista di nomi di campi importanti (es.
        'country') che mancano da TUTTO il log e che NON vengono riempiti
        automaticamente: viene solo mostrato un avviso, perché dedurli in
        modo sicuro richiederebbe una tabella prefissi->DXCC affidabile non
        ancora disponibile."""
        from tkinter import messagebox
        sat_msg = f"Banda satellite corretta: {n_sat_corretti} QSO{chr(10)}" if n_sat_corretti else ""
        avviso_critici = ""
        if campi_critici_assenti:
            nomi = ", ".join(c.upper() for c in campi_critici_assenti)
            avviso_critici = (
                f"⚠ Attenzione: il campo {nomi} è assente da TUTTI i QSO del log.{chr(10)}"
                f"Non viene compilato automaticamente (richiederebbe una tabella "
                f"prefissi→DXCC). Verifica prima di stampare etichette/PDF.{chr(10)}"
            )
        profilo_call = self.entry_owner.get().strip().upper()
        profilo_grid = self.entry_details.get().strip()
        SCHEMA = [
            ("STATION_CALLSIGN",profilo_call,"Callsign TX"),
            ("OPERATOR",profilo_call,"Operatore"),("OWNER_CALLSIGN",profilo_call,"Owner"),
            ("MY_GRIDSQUARE",profilo_grid,"Locator"),("MY_NAME","","Nome"),
            ("MY_COUNTRY","","Paese"),("MY_CQ_ZONE","","CQ Zone"),
            ("MY_ITU_ZONE","","ITU Zone"),("MY_DXCC","","DXCC"),
            ("MY_CITY","","Citta"),("MY_STATE","","Provincia"),
            ("MY_POSTAL_CODE","","CAP"),("MY_LAT","","Lat"),("MY_LON","","Lon"),
            ("QSL_SENT","N","QSL inv"),("QSL_RCVD","N","QSL ric"),
            ("QSL_SENT_VIA","D","Via"),("QSL_RCVD_VIA","D","Via"),
            ("EQSL_QSL_SENT","N","eQSL inv"),("EQSL_QSL_RCVD","N","eQSL ric"),
            ("LOTW_QSL_SENT","N","LoTW inv"),("LOTW_QSL_RCVD","N","LoTW ric"),
            ("CLUBLOG_QSO_UPLOAD_STATUS","N","ClubLog"),
            ("HRDLOG_QSO_UPLOAD_STATUS","N","HRDLog"),
            ("QRZCOM_QSO_UPLOAD_STATUS","N","QRZ"),
            ("HAMQTH_QSO_UPLOAD_STATUS","N","HamQTH"),
            ("QSO_COMPLETE","Y","Completo"),("QSO_RANDOM","N","Casuale"),
            ("SWL","N","SWL"),("APP_L4ONG_CONTEST","N","Contest"),
            ("APP_L4ONG_SATELLITE_QSO","N","Satellite"),
        ]
        campi_assenti = [(c,d,desc) for c,d,desc in SCHEMA if c.lower() not in tutti_campi]
        if not campi_assenti:
            if silenzioso and not avviso_critici:
                if on_done:
                    on_done()
                return
            if hasattr(parent, 'lift'):
                parent.lift(); parent.focus_force()
            messagebox.showinfo(T("completato"),
                f"{avviso_critici}{sat_msg}" + T("completa_tutto_ok", n=n_omogenea, tot=len(qsos)))
            if on_done:
                on_done()
            return
        ndlg = ctk.CTkToplevel(parent)
        ndlg.title(T("completa_titolo"))
        ndlg.geometry("560x560")
        ndlg.minsize(480, 380)
        ndlg.resizable(True, True)
        ndlg.grab_set(); ndlg.lift(); ndlg.focus_force()

        # ── Pulsanti in fondo: creati e impacchettati SUBITO con side="bottom",
        # così hanno sempre spazio riservato e non vengono mai tagliati fuori
        # dalla finestra, indipendentemente da quanto cresce il contenuto sopra.
        frame_btn = ctk.CTkFrame(ndlg, fg_color="transparent")
        frame_btn.pack(side="bottom", fill="x", padx=15, pady=12)

        # ── Contenuto principale (titolo, avviso, spiegazione, lista campi) ──
        titolo_fase1 = T("completa_fase1_ok", n=n_omogenea)
        if n_sat_corretti:
            titolo_fase1 += T("completa_sat", n=n_sat_corretti)
        ctk.CTkLabel(ndlg, text=titolo_fase1, wraplength=500, justify="left",
                     font=ctk.CTkFont(size=11, weight="bold"), text_color=TH.OK_TEXT).pack(pady=(12,2), padx=20, anchor="w")
        if avviso_critici:
            ctk.CTkLabel(ndlg, text=avviso_critici.strip(),
                         font=ctk.CTkFont(size=10, weight="bold"), text_color=TH.WARNING_H,
                         justify="left", wraplength=500).pack(padx=20, pady=(0,6), anchor="w")
        ctk.CTkLabel(ndlg,
                     text=T("completa_spiega", n=len(campi_assenti)),
                     font=ctk.CTkFont(size=10), text_color="gray", justify="left").pack(padx=20, pady=(0,8), anchor="w")
        scroll_f = ctk.CTkScrollableFrame(ndlg)
        scroll_f.pack(fill="both", expand=True, padx=15, pady=4)
        vars_check = {}; vars_val = {}
        for campo, default, descr in campi_assenti:
            row = ctk.CTkFrame(scroll_f, fg_color="transparent")
            row.pack(fill="x", pady=2)
            var_c = ctk.BooleanVar(value=bool(default))
            ctk.CTkCheckBox(row, text="", variable=var_c, width=24).pack(side="left")
            ctk.CTkLabel(row, text=campo, width=190, anchor="w", font=ctk.CTkFont(size=10, weight="bold")).pack(side="left", padx=4)
            e = ctk.CTkEntry(row, width=190, placeholder_text=descr)
            e.insert(0, default); e.pack(side="left", padx=4)
            vars_check[campo] = var_c; vars_val[campo] = e
        def applica_default():
            n2 = 0
            for campo, _, _ in campi_assenti:
                if not vars_check[campo].get():
                    continue
                valore_base = vars_val[campo].get().strip()
                campo_low = campo.lower()
                for qso in qsos:
                    valore = valore_base
                    if campo_low == "freq" and not valore:
                        banda = next((v for k,v in qso.items() if k.lower()=="band"), "")
                        valore = self._freq_da_banda(banda)
                    if valore:
                        qso[campo_low] = valore
                        n2 += 1
            ndlg.destroy()
            if hasattr(parent, 'lift'):
                parent.lift(); parent.focus_force()
            messagebox.showinfo(
                T("completato"),
                f"{avviso_critici}{sat_msg}" + T("completa_finale", n1=n_omogenea, n2=n2, n3=len(qsos))
            )
            if on_done:
                on_done()
        def _solo_chiudi():
            ndlg.destroy()
            if hasattr(parent, 'lift'):
                parent.lift(); parent.focus_force()
            if on_done:
                on_done()
        ctk.CTkButton(frame_btn, text=T("completa_btn_applica"), command=applica_default,
                      fg_color=TH.SUCCESS_H, height=36).pack(side="left", expand=True, padx=(0,6), fill="x")
        ctk.CTkButton(frame_btn, text=T("completa_btn_chiudi"),
                      command=_solo_chiudi,
                      fg_color="#718096", height=36).pack(side="left", expand=True, fill="x")

    def apri_preview(self):
        """Apre finestra separata con preview completo del log ADIF."""
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("preview_vuota"))
            return
        import tkinter.ttk as ttv

        dlg = ctk.CTkToplevel(self)
        nome = os.path.basename(self.filepath) if self.filepath else "ADIF"
        dlg.title(f"{T('preview_titolo')} — {nome} ({len(self.qsos_caricati)} QSO)")
        dlg.geometry("1200x600")
        dlg.minsize(800, 400)
        dlg.lift()
        dlg.focus_force()
        dlg.attributes("-topmost", True)
        dlg.after(100, lambda: dlg.attributes("-topmost", False))

        # Stato modifica
        modificato = [False]

        def aggiorna_titolo():
            suf = " *" if modificato[0] else ""
            dlg.title(f"{T('preview_titolo')} — {nome} ({len(self.qsos_caricati)} QSO){suf}")

        def ricarica_tree():
            for iid in tree.get_children():
                tree.delete(iid)
            for idx, qso in enumerate(self.qsos_caricati):
                qso_low = {k.lower(): v for k, v in qso.items()}
                vals = tuple(str(qso_low.get(k, '')).strip() for k in cols)
                tree.insert("", "end", iid=str(idx), text=str(idx), values=vals,
                            tags=("even" if idx%2==0 else "odd",))
            tree.tag_configure("even", background="#0D0D0D")
            tree.tag_configure("odd",  background="#1A2132")
            lbl_count.configure(text=f"{len(self.qsos_caricati)} QSO — {nome}")

        def cancella_righe():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Attenzione", "Seleziona almeno una riga.")
                return
            if not messagebox.askyesno("Conferma", f"Cancellare {len(sel)} QSO?"):
                return
            idxs = sorted([int(tree.item(iid,'text')) for iid in sel], reverse=True)
            for i in idxs:
                del self.qsos_caricati[i]
            ricarica_tree()
            modificato[0] = True
            aggiorna_titolo()

        def modifica_cella():
            sel = tree.selection()
            if not sel or len(sel) > 1:
                messagebox.showwarning("Attenzione", "Seleziona esattamente una riga.")
                return
            iid = sel[0]
            idx = int(tree.item(iid, 'text'))
            qso = self.qsos_caricati[idx]
            edlg = ctk.CTkToplevel(dlg)
            edlg.title(f"Modifica QSO — {qso.get('call', qso.get('CALL','?'))}")
            edlg.geometry("500x580")
            edlg.grab_set()
            edlg.lift()
            edlg.focus_force()
            ctk.CTkLabel(edlg, text=f"Modifica campi QSO",
                         font=ctk.CTkFont(size=12, weight="bold")).pack(pady=10, padx=15)
            scroll_f = ctk.CTkScrollableFrame(edlg, height=420)
            scroll_f.pack(fill="both", expand=True, padx=15, pady=4)
            entries = {}
            qso_low = {k.lower(): v for k, v in qso.items()}
            for col in cols:
                val = qso_low.get(col, '')
                if not val:
                    continue
                row_f = ctk.CTkFrame(scroll_f, fg_color="transparent")
                row_f.pack(fill="x", pady=2)
                ctk.CTkLabel(row_f, text=col.upper().replace("_"," "),
                             width=130, anchor="e", font=ctk.CTkFont(size=10)
                             ).pack(side="left", padx=(0,8))
                e = ctk.CTkEntry(row_f, width=280)
                e.insert(0, str(val))
                e.pack(side="left")
                entries[col] = e
            def salva_mod():
                for col, entry in entries.items():
                    val = entry.get().strip()
                    for k in list(qso.keys()):
                        if k.lower() == col:
                            qso[k] = val
                            break
                    else:
                        qso[col] = val
                ricarica_tree()
                modificato[0] = True
                aggiorna_titolo()
                edlg.destroy()
                dlg.lift()
                dlg.focus_force()
            fbe = ctk.CTkFrame(edlg, fg_color="transparent")
            fbe.pack(fill="x", padx=15, pady=8)
            ctk.CTkButton(fbe, text=T("cm_salva"), command=salva_mod,
                          fg_color=TH.SUCCESS_H, height=34).pack(side="left", expand=True, padx=(0,6), fill="x")
            ctk.CTkButton(fbe, text=T("cm_annulla"),
                          command=lambda: [edlg.destroy(), dlg.lift(), dlg.focus_force()],
                          fg_color="#718096", height=34).pack(side="left", expand=True, fill="x")

        def duplica_riga():
            sel = tree.selection()
            if not sel or len(sel) > 1:
                messagebox.showwarning("Attenzione", "Seleziona esattamente una riga.")
                return
            idx = int(tree.item(sel[0], 'text'))
            import copy
            self.qsos_caricati.insert(idx + 1, copy.deepcopy(self.qsos_caricati[idx]))
            ricarica_tree()
            modificato[0] = True
            aggiorna_titolo()

        def modifica_colonna():
            """Imposta o sostituisce il valore di una colonna su tutti i QSO (o solo quelli selezionati)."""
            cdlg = ctk.CTkToplevel(dlg)
            cdlg.title("Modifica colonna")
            cdlg.geometry("420x280")
            cdlg.resizable(False, False)
            cdlg.grab_set()
            cdlg.lift()
            cdlg.focus_force()

            ctk.CTkLabel(cdlg, text=T("dv_mod_colonna"),
                         font=ctk.CTkFont(size=12, weight="bold")).pack(pady=12, padx=20)

            frame_c = ctk.CTkFrame(cdlg, fg_color="transparent")
            frame_c.pack(fill="x", padx=20, pady=4)
            ctk.CTkLabel(frame_c, text=T("dv_colonna_lbl"), width=80, anchor="e").pack(side="left", padx=(0,8))
            col_var = ctk.StringVar(value=cols[0] if cols else "")
            col_menu = ctk.CTkOptionMenu(frame_c, variable=col_var,
                                          values=[c.upper().replace("_"," ") for c in cols],
                                          width=250)
            col_menu.pack(side="left")

            frame_v = ctk.CTkFrame(cdlg, fg_color="transparent")
            frame_v.pack(fill="x", padx=20, pady=4)
            ctk.CTkLabel(frame_v, text=T("dv_valore_lbl"), width=80, anchor="e").pack(side="left", padx=(0,8))
            val_entry = ctk.CTkEntry(frame_v, width=250, placeholder_text=T("dv_nuovo_valore"))
            val_entry.pack(side="left")

            # Opzione: tutti i QSO o solo selezionati
            scope_var = ctk.StringVar(value="tutti")
            n_sel = len(tree.selection())
            frame_s = ctk.CTkFrame(cdlg, fg_color="transparent")
            frame_s.pack(fill="x", padx=20, pady=8)
            ctk.CTkRadioButton(frame_s, text=T("dv_tutti_qso"),
                               variable=scope_var, value="tutti").pack(side="left", padx=8)
            rb_sel = ctk.CTkRadioButton(frame_s, text=f"Solo selezionati ({n_sel})",
                               variable=scope_var, value="sel")
            rb_sel.pack(side="left", padx=8)
            if n_sel == 0:
                rb_sel.configure(state="disabled")

            def applica_col():
                col_idx = col_menu._values.index(col_var.get()) if col_var.get() in col_menu._values else 0
                col_key = cols[col_idx]
                nuovo_val = val_entry.get()
                if scope_var.get() == "sel" and tree.selection():
                    idxs = [int(tree.item(iid,'text')) for iid in tree.selection()]
                else:
                    idxs = list(range(len(self.qsos_caricati)))
                for i in idxs:
                    qso = self.qsos_caricati[i]
                    # Aggiorna chiave esistente (upper o lower) o crea nuova
                    aggiornato = False
                    for k in list(qso.keys()):
                        if k.lower() == col_key:
                            qso[k] = nuovo_val
                            aggiornato = True
                            break
                    if not aggiornato:
                        qso[col_key] = nuovo_val
                ricarica_tree()
                modificato[0] = True
                aggiorna_titolo()
                cdlg.destroy()
                dlg.lift()
                dlg.focus_force()
                messagebox.showinfo("Fatto", f"Colonna {col_key.upper()} aggiornata su {len(idxs)} QSO.")

            frame_btn = ctk.CTkFrame(cdlg, fg_color="transparent")
            frame_btn.pack(fill="x", padx=20, pady=10)
            ctk.CTkButton(frame_btn, text=T("editor_applica"), command=applica_col,
                          fg_color=TH.SUCCESS_H, height=34).pack(side="left", expand=True, padx=(0,6), fill="x")
            ctk.CTkButton(frame_btn, text=T("cm_annulla"),
                          command=lambda: [cdlg.destroy(), dlg.lift(), dlg.focus_force()],
                          fg_color="#718096", height=34).pack(side="left", expand=True, fill="x")

        def normalizza_adif():
            n_sat_corretti = self._correggi_banda_satellite(self.qsos_caricati)
            n_omogenea, tutti_campi, campi_critici_assenti = self._normalizza_fase1_omogenizza(self.qsos_caricati)
            # Reset filtri anche dalla finestra editor — stessa logica di _ep_normalizza
            self.qsos_filtrati = list(self.qsos_caricati)
            if hasattr(self, 'entry_search'):
                self.entry_search.delete(0, 'end')
            self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
            ricarica_tree(); modificato[0] = True; aggiorna_titolo()
            def _dopo():
                ricarica_tree(); modificato[0] = True; aggiorna_titolo()
            self._normalizza_fase2_dialogo(self.qsos_caricati, tutti_campi, n_omogenea,
                                            parent=dlg, on_done=_dopo,
                                            n_sat_corretti=n_sat_corretti,
                                            campi_critici_assenti=campi_critici_assenti)

        def importa_cabrillo_editor(parent_dlg, cols):

            """Importa un file CBR e aggiunge i QSO al log in memoria."""
            path = filedialog.askopenfilename(
                parent=parent_dlg,
                title=T("dv_importa_cbr_file"),
                filetypes=[("Cabrillo","*.cbr *.log *.txt"),("All files","*.*")])
            if not path:
                return
            # Usa il parser CBR della UnisciDialog
            try:
                tmp = UnisciDialog.__new__(UnisciDialog)
                tmp.files = []
                qsos_cbr = tmp._leggi_cabrillo(path)
                if not qsos_cbr:
                    messagebox.showwarning("Attenzione", "Nessun QSO trovato nel file Cabrillo.")
                    return
                self.qsos_caricati.extend(qsos_cbr)
                ricarica_tree()
                modificato[0] = True
                aggiorna_titolo()
                parent_dlg.lift()
                parent_dlg.focus_force()
                messagebox.showinfo("Importato",
                    f"{len(qsos_cbr)} QSO importati dal file Cabrillo.{chr(10)}{os.path.basename(path)}")
            except Exception as ex:
                messagebox.showerror("Errore", f"Impossibile leggere il file:{chr(10)}{ex}")

        def esporta_cabrillo():
            sel = tree.selection()
            qsos_exp = [self.qsos_caricati[int(tree.item(iid,"text"))] for iid in sel] if sel else list(self.qsos_caricati)
            cdlg = ctk.CTkToplevel(dlg)
            cdlg.title("Esporta Cabrillo")
            cdlg.geometry("460x420"); cdlg.resizable(False, False)
            cdlg.grab_set(); cdlg.lift(); cdlg.focus_force()
            ctk.CTkLabel(cdlg, text=T("dv_esporta_cbr"),
                         font=ctk.CTkFont(size=12, weight="bold")).pack(pady=10, padx=20)
            CONTEST_LIST = ["CQ-WW-CW","CQ-WW-SSB","CQ-WPX-CW","CQ-WPX-SSB",
                "IARU-HF","CQ-VHF","ARRL-DX-CW","ARRL-DX-SSB","ARRL-10","ARRL-160",
                "ARRL-SS-CW","ARRL-SS-SSB","EU-HF","ARI-HF","ARI-DX","ITALIA-40-80","GENERIC","OTHER"]
            scroll_c = ctk.CTkScrollableFrame(cdlg, height=260)
            scroll_c.pack(fill="x", padx=15, pady=4)
            def add_field(lbl, default="", ph=""):
                row = ctk.CTkFrame(scroll_c, fg_color="transparent"); row.pack(fill="x", pady=3)
                ctk.CTkLabel(row, text=lbl, width=130, anchor="e", font=ctk.CTkFont(size=10)).pack(side="left", padx=(0,8))
                e = ctk.CTkEntry(row, width=240, placeholder_text=ph); e.insert(0, default); e.pack(side="left")
                return e
            profilo_call = self.entry_owner.get().strip().upper()
            f_call   = add_field("Callsign:", profilo_call)
            ctk.CTkLabel(scroll_c, text=T("cm_contest"), anchor="e", width=130, font=ctk.CTkFont(size=10)).pack(anchor="w", padx=(15,0), pady=(4,0))
            contest_var = ctk.StringVar(value="CQ-WW-SSB")
            ctk.CTkOptionMenu(scroll_c, variable=contest_var, values=CONTEST_LIST, width=240).pack(padx=15, pady=2)
            f_cat_op   = add_field("Category-Op:", "SINGLE-OP")
            f_cat_band = add_field("Category-Band:", "ALL")
            f_cat_mode = add_field("Category-Mode:", "SSB")
            f_cat_pwr  = add_field("Category-Power:", "HIGH")
            f_ops      = add_field("Operators:", profilo_call)
            f_club     = add_field("Club:", "", "Nome club (opz.)")
            f_loc      = add_field("Location:", "", "es. ITA / EU")
            f_score    = add_field("Claimed-Score:", "0")
            def genera_cbr():
                call = f_call.get().strip().upper()
                if not call: messagebox.showwarning("Attenzione","Inserire il callsign."); return
                nome_base = os.path.splitext(os.path.basename(self.filepath))[0] if self.filepath else "contest"
                save_path = filedialog.asksaveasfilename(parent=cdlg, title=T("dv_salva_cbr"),
                    defaultextension=".cbr", filetypes=[("Cabrillo","*.cbr *.log"),("All","*.*")],
                    initialfile=f"{nome_base}_{contest_var.get()}.cbr")
                if not save_path: return
                nl = chr(10)
                try:
                    with open(save_path, "w", encoding="ascii", errors="replace") as fw:
                        fw.write(f"START-OF-LOG: 3.0{nl}CONTEST: {contest_var.get()}{nl}CALLSIGN: {call}{nl}")
                        fw.write(f"CATEGORY-OPERATOR: {f_cat_op.get().strip()}{nl}")
                        fw.write(f"CATEGORY-BAND: {f_cat_band.get().strip()}{nl}")
                        fw.write(f"CATEGORY-MODE: {f_cat_mode.get().strip()}{nl}")
                        fw.write(f"CATEGORY-POWER: {f_cat_pwr.get().strip()}{nl}")
                        if f_ops.get().strip(): fw.write(f"OPERATORS: {f_ops.get().strip()}{nl}")
                        if f_club.get().strip(): fw.write(f"CLUB: {f_club.get().strip()}{nl}")
                        if f_loc.get().strip(): fw.write(f"LOCATION: {f_loc.get().strip()}{nl}")
                        fw.write(f"CLAIMED-SCORE: {f_score.get().strip() or '0'}{nl}CREATED-BY: ADIF FZR 2.5{nl}{nl}")
                        FREQ_MAP = {"160m":"1800","80m":"3500","60m":"5357","40m":"7000",
                            "30m":"10100","20m":"14000","17m":"18068","15m":"21000",
                            "12m":"24890","10m":"28000","6m":"50000","2m":"144000",
                            "70cm":"432000","23cm":"1296000"}
                        for qso in qsos_exp:
                            q = {k.lower():v for k,v in qso.items()}
                            try: freq_khz = str(int(float(q.get("freq","0"))*1000))
                            except: freq_khz = FREQ_MAP.get(q.get("band","").lower(),"14000")
                            modo = q.get("mode","SSB").upper()
                            if modo in ("FT8","FT4","WSPR","JT65","JT9","MFSK"): modo="DG"
                            elif modo in ("PSK31","PSK63","RTTY"): modo="RY"
                            data = q.get("qso_date","")
                            if len(data)==8: data=f"{data[:4]}-{data[4:6]}-{data[6:]}"
                            ora = q.get("time_on","0000")[:4]
                            fw.write(f"QSO: {freq_khz:>6} {modo:<2} {data} {ora} {q.get('station_callsign',call):<13} {q.get('rst_sent','59')[:3]} {q.get('stx','001'):<6} {q.get('call','').upper():<13} {q.get('rst_rcvd','59')[:3]} {q.get('srx','001')}{nl}")
                        fw.write(f"END-OF-LOG:{nl}")
                    cdlg.destroy(); dlg.lift(); dlg.focus_force()
                    messagebox.showinfo("Esportato", f"{len(qsos_exp)} QSO in Cabrillo:{nl}{os.path.basename(save_path)}")
                except Exception as ex:
                    messagebox.showerror("Errore", f"Impossibile salvare:{nl}{ex}")
            frame_btn = ctk.CTkFrame(cdlg, fg_color="transparent")
            frame_btn.pack(fill="x", padx=15, pady=8)
            ctk.CTkButton(frame_btn, text=T("dv_genera_cbr"), command=genera_cbr,
                          fg_color=TH.SUCCESS_H, height=34).pack(side="left", expand=True, padx=(0,6), fill="x")
            ctk.CTkButton(frame_btn, text=T("cm_annulla"),
                          command=lambda: [cdlg.destroy(), dlg.lift(), dlg.focus_force()],
                          fg_color="#718096", height=34).pack(side="left", expand=True, fill="x")

        def sposta_su():
            sel = tree.selection()
            if not sel or len(sel) > 1:
                return
            idx = int(tree.item(sel[0], 'text'))
            if idx == 0:
                return
            self.qsos_caricati[idx], self.qsos_caricati[idx-1] =                 self.qsos_caricati[idx-1], self.qsos_caricati[idx]
            ricarica_tree()
            new_iid = str(idx-1)
            tree.selection_set(new_iid)
            tree.see(new_iid)
            modificato[0] = True

        def sposta_giu():
            sel = tree.selection()
            if not sel or len(sel) > 1:
                return
            idx = int(tree.item(sel[0], 'text'))
            if idx >= len(self.qsos_caricati) - 1:
                return
            self.qsos_caricati[idx], self.qsos_caricati[idx+1] =                 self.qsos_caricati[idx+1], self.qsos_caricati[idx]
            ricarica_tree()
            new_iid = str(idx+1)
            tree.selection_set(new_iid)
            tree.see(new_iid)
            modificato[0] = True

        def esporta_selezionati():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Attenzione", "Seleziona almeno una riga.")
                return
            qsos_exp = [self.qsos_caricati[int(tree.item(iid,'text'))] for iid in sel]
            nome_base = os.path.splitext(os.path.basename(self.filepath))[0] if self.filepath else "export"
            save_path = filedialog.asksaveasfilename(
                parent=dlg,
                title=f"Esporta {len(qsos_exp)} QSO selezionati",
                defaultextension=".adif",
                filetypes=[("ADIF files","*.adif"),("All files","*.*")],
                initialfile=f"{nome_base}_sel{len(qsos_exp)}.adif")
            if not save_path:
                return
            try:
                CAMPI_H = {"adif_ver","programid","programversion","created_timestamp"}
                nl = chr(10)
                with open(save_path, "w", encoding="utf-8") as fw:
                    fw.write("<ADIF_VER:5>3.1.4" + nl)
                    fw.write(f"<PROGRAMID:{len(PROGRAMID_ADIF)}>{PROGRAMID_ADIF}" + nl)
                    fw.write(f"<PROGRAMVERSION:{len(VERSIONE)}>{VERSIONE}" + nl)
                    fw.write("<EOH>" + nl + nl)
                    for qso in qsos_exp:
                        for k, v in qso.items():
                            if k.lower() in CAMPI_H or not str(v).strip():
                                continue
                            fw.write(f"<{k.upper()}:{len(str(v))}>{v} ")
                        fw.write("<EOR>" + nl)
                dlg.lift()
                dlg.focus_force()
                messagebox.showinfo("Esportato",
                    f"{len(qsos_exp)} QSO salvati in:" + nl + os.path.basename(save_path))
            except Exception as ex:
                messagebox.showerror("Errore", f"Impossibile salvare:" + nl + str(ex))

        def salva_modifiche():
            if not modificato[0]:
                return
            self.qsos_filtrati = list(self.qsos_caricati)
            self._aggiorna_preview()
            messagebox.showinfo("Applicate",
                f"Modifiche applicate al log in memoria.\n{len(self.qsos_caricati)} QSO totali.\n\nUsa Salva ADIF per salvare su disco.")
            modificato[0] = False
            aggiorna_titolo()

        def chiudi():
            if modificato[0]:
                r = messagebox.askyesnocancel("Modifiche non salvate",
                    "Hai modifiche non applicate.\nApplicarle al log prima di chiudere?")
                if r is None:
                    return
                if r:
                    salva_modifiche()
            dlg.destroy()

        # Toolbar
        frame_tb = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_tb.pack(fill="x", padx=10, pady=6)
        lbl_count = ctk.CTkLabel(frame_tb,
                     text=f"{len(self.qsos_caricati)} QSO — {nome}",
                     font=ctk.CTkFont(size=12, weight="bold"))
        lbl_count.pack(side="left", padx=10)
        _b = ctk.CTkButton(frame_tb, text=T("editor_modifica"), command=modifica_cella,
                      width=90, height=28, fg_color=TH.PRIMARY); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_modifica"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_cancella"), command=cancella_righe,
                      width=90, height=28, fg_color=TH.WARNING_H); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_cancella"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_duplica"), command=duplica_riga,
                      width=90, height=28, fg_color="#4A5568"); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_duplica"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_colonna"), command=modifica_colonna,
                      width=90, height=28, fg_color="#2C7A7B"); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_colonna"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_normalizza"), command=normalizza_adif,
                      width=100, height=28, fg_color="#4A5568"); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_normalizza"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_su"), command=sposta_su,
                      width=55, height=28, fg_color=TH.SUCCESS_H); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_su"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_giu"), command=sposta_giu,
                      width=55, height=28, fg_color=TH.SUCCESS_H); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_giu"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_applica"), command=salva_modifiche,
                      width=80, height=28, fg_color="#4A5568"); _b.pack(side="left", padx=10)
        _tip(_b, T("tip_applica"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_esporta"), command=esporta_selezionati,
                      width=100, height=28, fg_color=TH.SUCCESS_H); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_esporta_sel"))
        _b = ctk.CTkButton(frame_tb, text=T("editor_cbr"), command=esporta_cabrillo,
                      width=100, height=28, fg_color="#4A5568"); _b.pack(side="left", padx=2)
        _tip(_b, T("tip_cbr"))
        _b = ctk.CTkButton(frame_tb, text=T("dv_import_cbr"), command=lambda: importa_cabrillo_editor(dlg, cols),
                      width=90, height=28, fg_color=TH.PRIMARY); _b.pack(side="left", padx=2)
        _b = ctk.CTkButton(frame_tb, text=T("editor_chiudi"), command=chiudi,
                      width=80, height=28, fg_color="#718096"); _b.pack(side="right", padx=10)
        _tip(_b, T("tip_chiudi"))

        # Stile treeview
        st = ttv.Style()
        st.theme_use("default")
        st.configure("Prv.Treeview", background="#0D0D0D", foreground="#E2E8F0",
                     rowheight=22, fieldbackground="#0D0D0D", font=("Arial", 9))
        st.configure("Prv.Treeview.Heading", background="#1A365D",
                     foreground="white", font=("Arial", 9, "bold"))
        st.map("Prv.Treeview", background=[("selected","#2B6CB0")],
               foreground=[("selected","white")])

        # Treeview con scroll V e H
        frame_t = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_t.pack(fill="both", expand=True, padx=10, pady=(0,10))

        # Scrollbar orizzontale in FONDO, verticale a destra
        sb_h = ttv.Scrollbar(frame_t, orient="horizontal")
        sb_v = ttv.Scrollbar(frame_t, orient="vertical")
        tree = ttv.Treeview(frame_t, show="headings", style="Prv.Treeview",
                            yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
        sb_v.configure(command=tree.yview)
        sb_h.configure(command=tree.xview)

        sb_h.pack(side="bottom", fill="x")
        sb_v.pack(side="right",  fill="y")
        tree.pack(side="left", fill="both", expand=True)

        # Colonne — priorità ai campi noti usati nella conversione
        ESCLUDI = {'adif_ver','programid','programversion','created_timestamp'}
        PRINCIPALI = [
            # Identificazione QSO
            'call','qso_date','time_on','time_off',
            # Radio
            'band','freq','mode','submode',
            # Segnale
            'rst_sent','rst_rcvd',
            # Posizione
            'country','gridsquare','cqz','ituz','cont',
            # Stazione
            'station_callsign','operator','my_gridsquare',
            # QSL / conferme
            'qsl_sent','qsl_rcvd','qslsdate',
            'lotw_qsl_sent','lotw_qsl_rcvd','lotw_qslsdate','lotw_qslrdate',
            'eqsl_qsl_sent','eqsl_qsl_rcvd',
            # Satellite
            'sat_name','sat_mode','prop_mode',
            # Extra
            'name','comment','notes','tx_pwr','distance',
        ]
        chiavi_set = set()
        for q in self.qsos_caricati:
            chiavi_set.update(k.lower() for k in q.keys() if k.lower() not in ESCLUDI)
        cols = [k for k in PRINCIPALI if k in chiavi_set]
        cols += sorted(k for k in chiavi_set if k not in PRINCIPALI and k not in cols)

        # Larghezze colonne
        LARGHE = {
            'call':80,'qso_date':80,'time_on':60,'time_off':60,
            'band':50,'freq':80,'mode':60,'submode':60,
            'rst_sent':55,'rst_rcvd':55,
            'country':110,'gridsquare':80,'cqz':40,'ituz':40,'cont':40,
            'station_callsign':100,'operator':80,'my_gridsquare':80,
            'qsl_sent':60,'qsl_rcvd':60,'qslsdate':80,
            'lotw_qsl_sent':70,'lotw_qsl_rcvd':80,'lotw_qslsdate':90,'lotw_qslrdate':90,
            'eqsl_qsl_sent':70,'eqsl_qsl_rcvd':80,
            'sat_name':80,'sat_mode':70,'prop_mode':70,
            'name':90,'comment':140,'notes':140,'tx_pwr':60,'distance':70,
        }

        tree["columns"] = cols
        for col in cols:
            tree.heading(col, text=col.upper().replace('_',' '),
                         command=lambda c=col: self._sort_preview(tree, c, cols))
            w = LARGHE.get(col, max(len(col)*8, 60))
            tree.column(col, width=w, minwidth=36, anchor="center")

        # Popola righe
        for idx, qso in enumerate(self.qsos_caricati):
            qso_low = {k.lower(): v for k, v in qso.items()}
            vals = tuple(str(qso_low.get(k, '')).strip() for k in cols)
            tree.insert("", "end", iid=str(idx), text=str(idx), values=vals,
                        tags=("even" if idx%2==0 else "odd",))
        tree.tag_configure("even", background="#0D0D0D")
        tree.tag_configure("odd",  background="#1A2132")

    def _sort_preview(self, tree, col, cols):
        """Ordina il treeview per colonna cliccata."""
        col_idx = cols.index(col)
        items = [(tree.set(iid, col), iid) for iid in tree.get_children()]
        items.sort(key=lambda x: x[0].lower())
        for i, (_, iid) in enumerate(items):
            tree.move(iid, '', i)
            tree.item(iid, tags=("even" if i%2==0 else "odd",))
        tree.tag_configure("even", background="#0D0D0D")
        tree.tag_configure("odd",  background="#1A2132")

    def _aggiorna_preview(self):
        pass  # placeholder — preview ora in finestra separata

    def _crea_menubar(self):
        """Costruisce (o ricostruisce) l'intera barra dei menù usando T()
        per ogni etichetta, così l'intero menù è bilingue e viene
        rigenerato da capo ad ogni cambio lingua da _aggiorna_lingua()."""
        import tkinter as _tk
        menubar = _tk.Menu(self)
        self.configure(menu=menubar)

        # File
        m_file = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_file"), menu=m_file)
        m_file.add_command(label=T("menu_apri_adif"),      command=self.sfoglia_file,    accelerator="Ctrl+O")
        m_file.add_command(label=T("menu_importa_adif_smart"), command=self.importa_adif_intelligente)
        m_file.add_command(label=T("menu_unisci_adif"),    command=self.apri_unisci)
        m_file.add_command(label=T("menu_importa_cbr"),    command=self.importa_cbr)
        m_file.add_command(label=T("menu_importa_hrd"), command=self.importa_hrd)
        m_file.add_command(label=T("menu_importa_log4om"), command=self.importa_log4om)
        m_file.add_separator()
        m_file.add_command(label=T("menu_salva_adif"),     command=self.salva_adif,      accelerator="Ctrl+S")
        m_file.add_separator()
        m_file.add_command(label=T("menu_file_recenti"),   command=self.apri_storico)
        m_file.add_separator()
        m_file.add_checkbutton(
            label=T("menu_controllo_apertura"),
            variable=self.var_controllo_post_apertura,
            command=self._salva_impostazioni_apertura
        )
        m_file.add_command(label=T("menu_deduci_country"), command=self.deduci_country_da_nominativo)
        m_file.add_separator()
        m_file.add_command(label=T("menu_esci"),            command=self.destroy,         accelerator="Alt+F4")
        self.bind_all("<Control-o>", lambda e: self.sfoglia_file())
        self.bind_all("<Control-s>", lambda e: self.salva_adif())

        # Modifica
        m_edit = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_modifica"), menu=m_edit)
        m_edit.add_command(label=T("menu_aggiungi_qso"),   command=self.apri_aggiungi_qso, accelerator="Ctrl+N")
        m_edit.add_separator()
        m_edit.add_command(label=T("menu_editor_log"),     command=self.apri_preview)
        m_edit.add_command(label=T("menu_dupe_check"),     command=self.apri_duplicati)
        m_edit.add_separator()
        m_edit.add_command(label=T("menu_filtri_qso"),     command=self.apri_filtri)
        m_edit.add_command(label=T("menu_colonne_pdf"),    command=self._apri_dialog_colonne)
        self.bind_all("<Control-n>", lambda e: self.apri_aggiungi_qso())

        # Esporta
        m_exp = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_esporta"), menu=m_exp)
        m_exp.add_command(label=T("menu_genera_pdf"),      command=self.processa_e_salva, accelerator="Ctrl+P")
        m_exp.add_command(label=T("menu_esporta_csv"),     command=self.esporta_csv)
        m_exp.add_command(label=T("menu_esporta_excel"),   command=self.esporta_excel)
        m_exp.add_command(label=T("menu_esporta_html"),    command=self.esporta_html)
        m_exp.add_separator()
        m_exp.add_command(label=T("menu_esporta_cbr"),     command=self.esporta_cabrillo_standalone)
        self.bind_all("<Control-p>", lambda e: self.processa_e_salva())
        self.bind_all("<F1>", lambda e: self._apri_manuale())

        # QSL
        m_qsl = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_qsl"), menu=m_qsl)
        m_qsl.add_command(label=T("menu_stampa_qsl"),        command=self.apri_qsl_card)
        m_qsl.add_command(label=T("menu_qsl_designer"),    command=self.apri_qsl_card_designer)
        m_qsl.add_separator()
        m_qsl.add_command(label=T("menu_cerca_qsl_mgr"),
                          command=self._hint_cerca_manager)
        m_qsl.add_command(label=T("menu_info_hamqth"),
                          command=self._hint_hamqth)
        m_qsl.add_separator()
        qsl_upload_sub = _tk.Menu(m_qsl, tearoff=0)
        m_qsl.add_cascade(label=T("menu_carica_su"), menu=qsl_upload_sub)
        qsl_upload_sub.add_command(label=T("menu_cloudlog_upload"), command=self.apri_cloudlog_upload)
        qsl_upload_sub.add_command(label=T("menu_clublog_upload"),  command=self.apri_clublog_upload)
        qsl_upload_sub.add_command(label=T("menu_lotw_upload"),     command=self.apri_lotw_upload)
        qsl_upload_sub.add_command(label=T("menu_eqsl_upload"),     command=self.apri_eqsl_upload)
        qsl_upload_sub.add_command(label=T("menu_qo100_upload"),    command=self.apri_qo100_upload)
        qsl_download_sub = _tk.Menu(m_qsl, tearoff=0)
        m_qsl.add_cascade(label=T("menu_scarica_da"), menu=qsl_download_sub)
        qsl_download_sub.add_command(label=T("menu_lotw_download"), command=self.apri_lotw_download)
        qsl_download_sub.add_command(label=T("menu_eqsl_download"), command=self.apri_eqsl_download)

        # Visualizza
        m_view = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_visualizza"), menu=m_view)
        m_view.add_command(label=T("menu_grafici"),        command=self.apri_grafici)
        m_view.add_command(label=T("menu_dx_cluster"),      command=self.apri_dx_cluster)
        m_view.add_command(label=T("menu_wsjtx"),           command=self.apri_wsjtx)
        m_view.add_command(label="Passaggi satelliti",   command=self.apri_satelliti)
        m_view.add_separator()
        m_view.add_checkbutton(
            label=T("menu_colora_righe"),
            variable=self.var_colora_righe,
            command=self._aggiorna_tree)
        m_view.add_separator()
        tema_sub = _tk.Menu(m_view, tearoff=0)
        m_view.add_cascade(label=T("menu_tema"), menu=tema_sub)
        for t in ["System", "Light", "Dark"]:
            tema_sub.add_command(label=t, command=lambda x=t: self._set_tema(x))
        lingua_sub = _tk.Menu(m_view, tearoff=0)
        m_view.add_cascade(label=T("menu_lingua"), menu=lingua_sub)
        lingua_sub.add_command(label=T("menu_italiano"), command=lambda: self._set_lingua("IT"))
        lingua_sub.add_command(label=T("menu_english"),  command=lambda: self._set_lingua("EN"))

        # Strumenti
        m_tools = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_strumenti"), menu=m_tools)

        # ── Sottomenu: Aspetto PDF ──
        tools_pdf_sub = _tk.Menu(m_tools, tearoff=0)
        m_tools.add_cascade(label=T("menu_aspetto_pdf"), menu=tools_pdf_sub)
        tools_pdf_sub.add_command(label=T("menu_colori_pdf"),    command=self.apri_colori)
        tools_pdf_sub.add_command(label=T("menu_colori_html"),   command=self.apri_colori_html)
        tools_pdf_sub.add_command(label=T("menu_opzioni_reg_pdf"), command=self.apri_opzioni_registro_pdf)
        tools_pdf_sub.add_command(label=T("menu_formato_pdf"),   command=self._apri_dialog_formato)

        # ── Sottomenu: Calcolatori ──
        tools_calc_sub = _tk.Menu(m_tools, tearoff=0)
        m_tools.add_cascade(label=T("menu_calcolatori"), menu=tools_calc_sub)
        tools_calc_sub.add_command(label=T("menu_dist_calc"),     command=self.apri_calcolatore_distanza)
        tools_calc_sub.add_command(label=T("menu_dottore_sat"),   command=self.apri_dottore_sat)

        m_tools.add_separator()
        m_tools.add_command(label=T("menu_preferenze"),
                            command=self.apri_preferenze)
        m_tools.add_command(label=T("menu_tb_custom"),
                            command=self.apri_personalizza_toolbar)
        m_tools.add_command(label=T("menu_pers_colonne"),
                            command=self.apri_personalizza_colonne)
        m_tools.add_command(label=T("menu_radio_omnirig"),
                            command=self.apri_impostazioni_radio)
        m_tools.add_command(label=T("menu_radio_display"),
                            command=self.apri_display_radio)
        m_tools.add_separator()

        # ── Sottomenu: Profili operatore ──
        tools_prof_sub = _tk.Menu(m_tools, tearoff=0)
        m_tools.add_cascade(label=T("menu_profili_op"), menu=tools_prof_sub)
        tools_prof_sub.add_command(label=T("menu_profili"),       command=self.apri_gestione_profili)
        tools_prof_sub.add_command(label=T("menu_salva_profilo"), command=self.salva_profilo)

        # Aiuto
        m_help = _tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label=T("menu_cascade_aiuto"), menu=m_help)
        m_help.add_command(label=T("menu_manuale"), command=self._apri_manuale)
        m_help.add_separator()
        m_help.add_command(label=T("menu_about"), command=self._about)

    def _aggiorna_lingua(self):
        self.title(f"{APP_TITOLO}  ·  build {BUILD_DATE}  —  IW1FZR")
        self._crea_menubar()  # ricostruisce l'intera barra menu nella lingua corrente

        # Aggiorna i widget della sidebar principale (label e pulsanti tradotti)
        for _w, _chiave, _extra in getattr(self, '_sidebar_lang_refs', []):
            try:
                _w.configure(text=T(_chiave))
            except Exception:
                pass
        # Placeholder degli entry (owner, details, ricerca)
        try:
            self.entry_owner.configure(placeholder_text=T("sb_owner_ph"))
            self.entry_details.configure(placeholder_text=T("sb_details_ph"))
            self.entry_search.configure(placeholder_text=T("sb_cerca_ph"))
        except Exception:
            pass
        # Filtri sidebar: aggiorna "Tutte"/"Tutti" se sono al valore di default
        try:
            _tutti_set = {"tutte","tutti","all"}
            if hasattr(self,'_fil_banda_var') and self._fil_banda_var.get().lower() in _tutti_set:
                self._fil_banda_var.set(T("filtri_tutte"))
            if hasattr(self,'_fil_modo_var') and self._fil_modo_var.get().lower() in _tutti_set:
                self._fil_modo_var.set(T("filtri_tutti"))
            if hasattr(self,'_fil_sat_var') and self._fil_sat_var.get().lower() in _tutti_set:
                self._fil_sat_var.set(T("filtri_tutti"))
            self._aggiorna_sidebar_filtri()
        except Exception:
            pass

        # Widget refs originali (sidebar + labels)
        _emoji = {
            'colori_pdf': '🎨 ', 'filtri_qso': '⚗ ',
            'grafici': '📊 ', 'storico': '🕐 ', 'salva_profilo': '💾 ',
            'duplicati': '🔍 ', 'preview_adif': '🔎 ',
            'qsl_card': '📮  ', 'importa_cbr': '📥 ',
            'apri_adif': '📂 ', 'unisci': '🔗 ', 'salva_adif': '💾 ',
            'ep_applica': '', 'ep_duplica': '', 'ep_su': '', 'ep_giu': '',
            'ep_elimina': '', 'ep_colonna': '', 'ep_normalizza': '',
            'ep_modifica': '✏  ', 'dist_calc_from_qso': '',
        }
        _uppercase = {'qsl_card'}
        for key, widget in self._widget_refs.items():
            trad = TRADUZIONI.get(key)
            if not trad:
                continue
            try:
                prefix = _emoji.get(key, '')
                testo = T(key).upper() if key in _uppercase else T(key)
                widget.configure(text=prefix + testo)
            except Exception:
                pass

        # Toolbar btn1
        try:
            for btn, key, emoji in self._tb1_refs:
                btn.configure(text=emoji + T(key))
        except Exception:
            pass

        # Toolbar btn2
        try:
            for btn, key, emoji in self._tb2_refs:
                btn.configure(text=emoji + T(key))
        except Exception:
            pass

        # Status / filtri
        if not self.qsos_filtrati or len(self.qsos_filtrati) == len(self.qsos_caricati):
            self.lbl_filtri.configure(text=T("nessun_filtro"))
        if not self.filepath:
            self.lbl_status.configure(text=T("nessun_file"))

        # Pannello modifica — label campi
        try:
            field_labels_map = {
                "band": T("col_banda"), "mode": T("col_modo"), "name": T("nome_op")
            }
            for tag, lbl in self._ep_field_labels.items():
                if tag in field_labels_map:
                    lbl.configure(text=field_labels_map[tag])
        except Exception:
            pass

        # Pannello modifica inline
        try:
            if self._ep_idx is None:
                self._ep_info.configure(text=T("ep_seleziona"))
        except Exception:
            pass

        # Sezioni sidebar
        try:
            for lbl, key in self._sidebar_section_refs:
                lbl.configure(text=T(key))
        except Exception:
            pass

    # ── Cambio tema ───────────────────────────
    def cambia_tema(self, valore):
        self._set_tema(valore)
        self._aggiorna_tree()

    # ── Dialoghi ──────────────────────────────
    def apri_filtri(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        dlg = FiltriDialog(self, self.qsos_caricati)
        self.wait_window(dlg)
        if dlg.risultato is not None:
            self.qsos_filtrati = dlg.risultato
            n = len(self.qsos_filtrati)
            tot = len(self.qsos_caricati)
            # Il filtro avanzato sostituisce eventuali filtri rapidi/ricerca attivi
            if hasattr(self, 'entry_search'):
                self.entry_search.delete(0, 'end')
            if hasattr(self, 'sb_filt'):
                self.sb_filt.configure(text="")
            if n == tot:
                self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
            else:
                self.lbl_filtri.configure(text=T("filtro_attivo", n=n, tot=tot), text_color=TH.OK_TEXT)
            self._aggiorna_tree()

    def apri_colori(self):
        dlg = ColoriDialog(self, self.colori_pdf)
        self.wait_window(dlg)
        if dlg.risultato is not None:
            self.colori_pdf = dlg.risultato

    def apri_opzioni_registro_pdf(self):
        dlg = OpzioniRegistroPDFDialog(
            self,
            self.campi_disponibili,
            self.ordine_campi_pdf,
            self.checkboxes,
            self.width_pdf,
            self.titolo_pdf_custom,
            self.font_size_pdf,
        )
        self.wait_window(dlg)
        if dlg.risultato is not None:
            self.ordine_campi_pdf, self.width_pdf, self.titolo_pdf_custom, self.font_size_pdf = dlg.risultato

    def apri_colori_html(self):
        dlg = ColoriHtmlDialog(self, self.colori_html)
        self.wait_window(dlg)
        if dlg.risultato is not None:
            self.colori_html = dlg.risultato


    # ── Grafici attività ──────────────────────
    def _esporta_html_legacy_unused(self):
        """Genera una pagina HTML con tabella QSO e filtri multipli."""
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_nessun_file"))
            return
        qsos = self._qsos_attivi()
        stazione = self.entry_owner.get().strip().upper() or "LOG"
        nome_def = os.path.splitext(os.path.basename(self.filepath))[0] + "_log.html" if self.filepath else "log.html"
        save_path = filedialog.asksaveasfilename(
            title=T("dv_esporta_html"), defaultextension=".html",
            filetypes=[("HTML files","*.html"),("All files","*.*")],
            initialfile=nome_def)
        if not save_path:
            return
        bande = sorted(set(str(q.get('band','?')).upper() for q in qsos if q.get('band')))
        modi  = sorted(set(str(q.get('mode','?')).upper() for q in qsos if q.get('mode')))
        import json as _json
        rows = []
        for q in qsos:
            data = str(q.get('qso_date',''))
            if len(data)==8: data = f"{data[6:8]}/{data[4:6]}/{data[0:4]}"
            ora = str(q.get('time_on',''))[:4]
            if len(ora)==4: ora = f"{ora[:2]}:{ora[2:]}"
            lotw = str(q.get('lotw_qsl_rcvd','')).upper().strip()
            eqsl = str(q.get('eqsl_qsl_rcvd','')).upper().strip()
            rows.append({
                'call':    str(q.get('call','')).upper(),
                'date':    data, 'time': ora,
                'band':    str(q.get('band','')).upper(),
                'mode':    str(q.get('mode','')).upper(),
                'rst_s':   str(q.get('rst_sent','')),
                'rst_r':   str(q.get('rst_rcvd','')),
                'country': str(q.get('country','')),
                'grid':    str(q.get('gridsquare','')).upper(),
                'lotw':    'Y' if lotw in ('Y','V') else 'N',
                'eqsl':    'Y' if eqsl == 'Y' else 'N',
            })
        rows_json  = _json.dumps(rows, ensure_ascii=False).replace("</", "<\\/")
        bande_opts = ''.join(f'<option value="{b}">{b}</option>' for b in bande)
        modi_opts  = ''.join(f'<option value="{m}">{m}</option>' for m in modi)
        nl = chr(10)
        h = []
        a = h.append
        a('<!DOCTYPE html><html lang="it"><head>')
        a('<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">')
        a(f'<title>Log {stazione} - ADIF FZR 2.5</title>')
        a('<style>')
        a('*{box-sizing:border-box;margin:0;padding:0}')
        a('body{font-family:Arial,sans-serif;background:#0D1117;color:#E2E8F0;font-size:.93em}')
        a('.header{background:linear-gradient(135deg,#0D1F35,#1A365D,#2B6CB0);padding:28px 30px 22px;margin-bottom:20px}')
        a('.header h1{color:#fff;font-size:1.8em;margin-bottom:4px}')
        a('.header p{color:#90CDF4;font-size:.9em}')
        a('.container{max-width:1200px;margin:0 auto;padding:0 16px 40px}')
        a('.filters{background:#141414;border:1px solid #2D3748;border-radius:10px;padding:16px 20px;margin-bottom:16px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}')
        a('.fg{display:flex;flex-direction:column;gap:4px}')
        a('.fg label{font-size:.78em;color:#90CDF4;font-weight:700;letter-spacing:1px;text-transform:uppercase}')
        a('.fg input,.fg select{background:#1A365D;border:1px solid #2B6CB0;border-radius:6px;color:#E2E8F0;padding:6px 10px;font-size:.88em;outline:none;min-width:110px}')
        a('.fg input:focus,.fg select:focus{border-color:#4299E1}')
        a('.btn-r{background:#718096;border:none;color:#fff;padding:7px 16px;border-radius:6px;cursor:pointer;font-size:.88em;margin-top:18px}')
        a('.btn-r:hover{background:#4A5568}')
        a('.stats{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:14px}')
        a('.sc{background:#141414;border:1px solid #2D3748;border-radius:8px;padding:10px 18px;text-align:center;min-width:90px}')
        a('.sc .n{font-size:1.5em;font-weight:700;color:#4299E1}')
        a('.sc .l{font-size:.75em;color:#718096}')
        a('.tw{overflow-x:auto;border-radius:10px;border:1px solid #2D3748}')
        a('table{width:100%;border-collapse:collapse;font-size:.88em}')
        a('thead th{background:#1A365D;color:#90CDF4;padding:10px 12px;text-align:left;font-weight:700;white-space:nowrap;cursor:pointer;user-select:none}')
        a('thead th:hover{background:#2B6CB0;color:#fff}')
        a('thead th.sa::after{content:" \25B2"}thead th.sd::after{content:" \25BC"}')
        a('tbody tr{border-bottom:1px solid #141414;transition:background .15s}')
        a('tbody tr:hover{background:#141414}')
        a('tbody tr:nth-child(even){background:#111827}')
        a('tbody tr:nth-child(even):hover{background:#141414}')
        a('td{padding:8px 12px;white-space:nowrap}')
        a('.by{display:inline-block;padding:1px 8px;border-radius:10px;font-size:.8em;font-weight:700;background:#276749;color:#9AE6B4}')
        a('.bn{display:inline-block;padding:1px 8px;border-radius:10px;font-size:.8em;font-weight:700;background:#2D3748;color:#718096}')
        a('.nr{text-align:center;padding:40px;color:#718096;font-size:1.1em}')
        a('.footer{text-align:center;padding:20px;color:#4A5568;font-size:.8em;margin-top:20px}')
        a('</style></head><body>')
        a(f'<div class="header"><div style="max-width:1200px;margin:0 auto">')
        a(f'<h1>Log {stazione}</h1><p>ADIF FZR 2.5 &nbsp;&middot;&nbsp; {len(qsos)} QSO totali</p>')
        a('</div></div>')
        a('<div class="container">')
        a('<div class="filters">')
        a('<div class="fg"><label>Callsign</label><input type="text" id="fc" placeholder="es. DL2..." oninput="af()"></div>')
        a('<div class="fg"><label>Banda</label><select id="fb" onchange="af()"><option value="">Tutte</option>' + bande_opts + '</select></div>')
        a('<div class="fg"><label>Modo</label><select id="fm" onchange="af()"><option value="">Tutti</option>' + modi_opts + '</select></div>')
        a('<div class="fg"><label>Da data</label><input type="text" id="fd1" placeholder="gg/mm/aaaa" oninput="af()"></div>')
        a('<div class="fg"><label>A data</label><input type="text" id="fd2" placeholder="gg/mm/aaaa" oninput="af()"></div>')
        a('<div class="fg"><label>Country</label><input type="text" id="fco" placeholder="es. Italy..." oninput="af()"></div>')
        a('<div class="fg"><label>LoTW</label><select id="fl" onchange="af()"><option value="">Tutti</option><option value="Y">Confermato</option><option value="N">Non conf.</option></select></div>')
        a('<button class="btn-r" onclick="rf()">Reset</button>')
        a('</div>')
        a('<div class="stats">')
        a('<div class="sc"><div class="n" id="st">0</div><div class="l">QSO</div></div>')
        a('<div class="sc"><div class="n" id="sl">0</div><div class="l">LoTW</div></div>')
        a('<div class="sc"><div class="n" id="se">0</div><div class="l">eQSL</div></div>')
        a('<div class="sc"><div class="n" id="sd">0</div><div class="l">DXCC</div></div>')
        a('</div>')
        a('<div class="tw"><table id="t"><thead><tr>')
        for i,col in enumerate(['Data','UTC','Callsign','Banda','Modo','RST TX','RST RX','Country','Locator','LoTW','eQSL']):
            a(f'<th onclick="st2({i})">{col}</th>')
        a('</tr></thead><tbody id="tb"></tbody></table>')
        a('<div class="nr" id="nr" style="display:none">Nessun QSO trovato</div>')
        a('</div></div>')
        a(f'<div class="footer">ADIF FZR 2.5 &middot; {stazione} &middot; {len(qsos)} QSO</div>')
        a('<script>')
        a('const D=' + rows_json + ';')
        a('let sc=0,sd2=1,fi=[...D];')
        a('function pd(s){if(!s)return 0;const p=s.split("/");return p.length===3?new Date(p[2],p[1]-1,p[0]).getTime():0}')
        a('function af(){')
        a('  const fc=document.getElementById("fc").value.toUpperCase();')
        a('  const fb=document.getElementById("fb").value;')
        a('  const fm=document.getElementById("fm").value;')
        a('  const fd1=document.getElementById("fd1").value;')
        a('  const fd2=document.getElementById("fd2").value;')
        a('  const fco=document.getElementById("fco").value.toUpperCase();')
        a('  const fl=document.getElementById("fl").value;')
        a('  const t1=pd(fd1),t2=pd(fd2);')
        a('  fi=D.filter(r=>{')
        a('    if(fc&&!r.call.includes(fc))return false;')
        a('    if(fb&&r.band!==fb)return false;')
        a('    if(fm&&r.mode!==fm)return false;')
        a('    if(fco&&!r.country.toUpperCase().includes(fco))return false;')
        a('    if(fl&&r.lotw!==fl)return false;')
        a('    if(t1){const td=pd(r.date);if(td<t1)return false;}')
        a('    if(t2){const td=pd(r.date);if(td>t2)return false;}')
        a('    return true;')
        a('  });rt();')
        a('}')
        a('function rf(){["fc","fd1","fd2","fco"].forEach(id=>document.getElementById(id).value="");["fb","fm","fl"].forEach(id=>document.getElementById(id).value="");fi=[...D];rt();}')
        a('function st2(c){const ths=document.querySelectorAll("thead th");ths.forEach(th=>th.classList.remove("sa","sd"));if(sc===c)sd2*=-1;else{sc=c;sd2=1;}ths[c].classList.add(sd2===1?"sa":"sd");const ks=["date","time","call","band","mode","rst_s","rst_r","country","grid","lotw","eqsl"];fi.sort((a,b)=>(a[ks[c]]>b[ks[c]]?1:-1)*sd2);rt();}')
        a('function bdg(v){return v==="Y"?"<span class=\"by\">Y</span>":"<span class=\"bn\">-</span>";}')
        a('function rt(){')
        a('  const tb=document.getElementById("tb");')
        a('  const nr=document.getElementById("nr");')
        a('  if(!fi.length){tb.innerHTML="";nr.style.display="block";}')
        a('  else{nr.style.display="none";tb.innerHTML=fi.map(r=>`<tr><td>${r.date}</td><td>${r.time}</td><td><strong style="color:#90CDF4">${r.call}</strong></td><td>${r.band}</td><td>${r.mode_vis}</td><td>${r.rst_s}</td><td>${r.rst_r}</td><td>${r.country}</td><td>${r.grid}</td><td>${bdg(r.lotw)}</td><td>${bdg(r.eqsl)}</td></tr>`).join("");}')
        a('  document.getElementById("st").textContent=fi.length;')
        a('  document.getElementById("sl").textContent=fi.filter(r=>r.lotw==="Y").length;')
        a('  document.getElementById("se").textContent=fi.filter(r=>r.eqsl==="Y").length;')
        a('  document.getElementById("sd").textContent=new Set(fi.map(r=>r.country).filter(Boolean)).size;')
        a('}')
        a('rt();')
        a('</script></body></html>')
        html = nl.join(h)
        try:
            with open(save_path, 'w', encoding='utf-8') as fw:
                fw.write(html)
            messagebox.showinfo(T("successo"),
                "HTML esportato:" + nl + os.path.basename(save_path) + nl + str(len(qsos)) + " QSO")
            os.startfile(os.path.abspath(save_path))
        except Exception as ex:
            messagebox.showerror(T("errore"), "Errore:" + nl + str(ex))

    def importa_adif_intelligente(self):
        """Import ADIF universale con anteprima, statistiche e opzioni.
        Funziona con qualsiasi logger che esporta ADIF: MSHV, N3FJP,
        DXKeeper, Logger32, WSJT-X, QRZ, LoTW, ecc."""
        path = filedialog.askopenfilename(
            title=T("imp_smart_sel"),
            filetypes=[("File ADIF","*.adi *.adif"),("Tutti i file","*.*")])
        if not path:
            return

        try:
            # Il parser ADIF vive in UnisciDialog: crea istanza senza __init__
            _tmp = UnisciDialog.__new__(UnisciDialog)
            qsos = _tmp._leggi_adif_puro(path)
        except Exception as ex:
            messagebox.showerror(T("imp_errore_lettura"), T("imp_file_nonvalido", ex=ex))
            return

        if not qsos:
            messagebox.showwarning(T("imp_no_qso"),
                T("imp_no_qso_validi"))
            return

        # ── Calcola statistiche ──
        import os as _os
        from collections import Counter
        date = [str(q.get('qso_date','')).strip() for q in qsos if q.get('qso_date')]
        date = [d for d in date if len(d) == 8]
        data_min = min(date) if date else "?"
        data_max = max(date) if date else "?"
        def _fmtd(d):
            return f"{d[6:8]}/{d[4:6]}/{d[0:4]}" if len(d)==8 else "?"

        bande = Counter(str(q.get('band','')).lower() for q in qsos if q.get('band'))
        modi  = Counter(str(q.get('mode','')).upper() for q in qsos if q.get('mode'))
        calls = Counter(str(q.get('call','')).upper() for q in qsos if q.get('call'))
        con_grid = sum(1 for q in qsos if q.get('gridsquare'))
        con_lotw = sum(1 for q in qsos if str(q.get('lotw_qsl_rcvd','')).upper()=='Y')

        # Rileva il programma di origine (header ADIF)
        origine = "?"
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                testa = f.read(2000)
            mo = re.search(r'<programid:\d+>([^<]+)', testa, re.IGNORECASE)
            if mo:
                origine = mo.group(1).strip()
        except Exception:
            pass

        # ── Dialog anteprima ──
        dlg = ctk.CTkToplevel(self)
        dlg.title(T("imp_smart_titolo"))
        dlg.geometry("560x640")
        dlg.transient(self)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()
        dlg.after(200, lambda: (dlg.lift(), dlg.focus_force()))

        ctk.CTkLabel(dlg, text=T("imp_smart_head"),
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(14,2))
        ctk.CTkLabel(dlg, text=_os.path.basename(path),
                     font=ctk.CTkFont(size=10), text_color="gray").pack()

        # Card statistiche
        card = ctk.CTkFrame(dlg)
        card.pack(fill="x", padx=20, pady=12)

        def _riga(parent, etichetta, valore, grassetto=False):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", padx=14, pady=2)
            ctk.CTkLabel(r, text=etichetta, anchor="w",
                         font=ctk.CTkFont(size=11)).pack(side="left")
            ctk.CTkLabel(r, text=valore, anchor="e",
                         font=ctk.CTkFont(size=12,
                             weight="bold" if grassetto else "normal")).pack(side="right")

        _riga(card, T("imp_qso_totali"), str(len(qsos)), True)
        _riga(card, T("imp_periodo"), f"{_fmtd(data_min)} → {_fmtd(data_max)}")
        _riga(card, T("imp_stazioni"), str(len(calls)))
        _riga(card, T("imp_con_loc"), f"{con_grid} ({100*con_grid//len(qsos)}%)")
        _riga(card, T("imp_lotw_conf"), str(con_lotw))
        if origine != "?":
            _riga(card, T("imp_origine"), origine)

        # Bande e modi
        top_bande = ", ".join(f"{b}({n})" for b,n in bande.most_common(6))
        top_modi  = ", ".join(f"{m}({n})" for m,n in modi.most_common(6))
        ctk.CTkLabel(dlg, text=T("imp_bande"), anchor="w",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(fill="x", padx=24, pady=(6,0))
        ctk.CTkLabel(dlg, text=top_bande or "—", anchor="w", wraplength=500,
                     font=ctk.CTkFont(size=11), text_color="gray").pack(fill="x", padx=24)
        ctk.CTkLabel(dlg, text=T("imp_modi"), anchor="w",
                     font=ctk.CTkFont(size=11, weight="bold")).pack(fill="x", padx=24, pady=(6,0))
        ctk.CTkLabel(dlg, text=top_modi or "—", anchor="w", wraplength=500,
                     font=ctk.CTkFont(size=11), text_color="gray").pack(fill="x", padx=24)

        # ── Opzioni import ──
        ctk.CTkLabel(dlg, text=T("imp_modalita"),
                     font=ctk.CTkFont(size=12, weight="bold")).pack(pady=(14,4))

        modo_import = ctk.StringVar(value="aggiungi" if self.qsos_caricati else "sostituisci")
        frame_opt = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_opt.pack(fill="x", padx=30)
        if self.qsos_caricati:
            ctk.CTkRadioButton(frame_opt,
                text=T("imp_aggiungi_a", n=len(self.qsos_caricati)),
                variable=modo_import, value="aggiungi",
                font=ctk.CTkFont(size=11)).pack(anchor="w", pady=3)
        ctk.CTkRadioButton(frame_opt, text=T("imp_sostituisci"),
                variable=modo_import, value="sostituisci",
                font=ctk.CTkFont(size=11)).pack(anchor="w", pady=3)

        salta_dup = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(dlg, text=T("imp_salta_dup"),
                        variable=salta_dup,
                        font=ctk.CTkFont(size=11)).pack(pady=(10,4))

        lbl_esito = ctk.CTkLabel(dlg, text="", font=ctk.CTkFont(size=11),
                                  text_color=TH.OK_TEXT)
        lbl_esito.pack(pady=2)

        def _esegui():
            nuovi = list(qsos)
            saltati = 0
            if salta_dup.get():
                # Costruisci indice dei QSO esistenti se in modalità aggiungi
                esistenti = set()
                base = self.qsos_caricati if modo_import.get()=="aggiungi" else []
                for q in base:
                    esistenti.add(self._chiave_dup(q))
                filtrati = []
                visti = set(esistenti)
                for q in nuovi:
                    k = self._chiave_dup(q)
                    if k in visti:
                        saltati += 1
                        continue
                    visti.add(k)
                    filtrati.append(q)
                nuovi = filtrati

            if modo_import.get() == "aggiungi":
                self.qsos_caricati.extend(nuovi)
            else:
                self.qsos_caricati = nuovi

            self.qsos_caricati.sort(key=lambda q: (
                str(q.get('qso_date','')).strip(),
                str(q.get('time_on','')).strip().zfill(6)))
            self.qsos_filtrati = list(self.qsos_caricati)
            self._aggiorna_tree()
            dlg.destroy()
            msg = T("imp_importati_n", n=len(nuovi))
            if saltati:
                msg += "\n" + T("imp_dup_saltati", n=saltati)
            msg += "\n\n" + T("imp_totale_log", n=len(self.qsos_caricati))
            msg += "\n" + T("imp_salva_ricorda")
            messagebox.showinfo(T("imp_completato"), msg)

        fr = ctk.CTkFrame(dlg, fg_color="transparent")
        fr.pack(fill="x", padx=30, pady=16)
        ctk.CTkButton(fr, text=T("imp_importa"), command=_esegui, height=38,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS
                      ).pack(side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(fr, text=T("imp_annulla"), command=dlg.destroy,
                      height=38, width=100, fg_color="#718096").pack(side="left")

    def _chiave_dup(self, q):
        """Chiave per rilevare duplicati: call+banda+modo+tempo arrotondato a 2 min."""
        call = str(q.get('call','')).upper().strip()
        band = str(q.get('band','')).lower().strip()
        mode = str(q.get('mode','')).upper().strip()
        # Timestamp in secondi (autonomo, senza dipendenze esterne)
        data = str(q.get('qso_date','')).strip()
        ora  = str(q.get('time_on','')).strip().ljust(6,'0')[:6]
        blocco = 0
        if len(data) == 8:
            try:
                from datetime import datetime as _dt
                sec = int(_dt.strptime(data + ora, "%Y%m%d%H%M%S").timestamp())
                blocco = sec // 120   # blocchi di 2 minuti
            except Exception:
                blocco = 0
        return (call, band, mode, blocco)

    def importa_log4om(self):
        """Importa QSO dal database SQLite di Log4OM (v2).
        Log4OM usa una tabella 'log' con colonne in nomi ADIF standard."""
        import sqlite3

        path = filedialog.askopenfilename(
            title=T("imp_l4o_sel"),
            filetypes=[("Log4OM SQLite","*.sqlite *.db *.sqlite3"),
                       ("Tutti i file","*.*")])
        if not path:
            return

        try:
            con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            con.row_factory = sqlite3.Row
            cur = con.cursor()

            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tabelle = [r[0] for r in cur.fetchall()]

            # Cerca la tabella dei QSO: quella con più righe che ha una
            # colonna 'call' (con vari nomi possibili). Log4OM usa 'log'
            # ma alcune versioni usano nomi diversi.
            def _trova_col_call(cols_lower):
                for cand in ('call', 'callsign', 'col_call', 'dxcall', 'dx_call'):
                    if cand in cols_lower:
                        return cand
                return None

            tab = None
            col_call_name = None
            diag = []
            best_count = -1
            for t in tabelle:
                try:
                    cur.execute(f"SELECT * FROM [{t}] LIMIT 1")
                    cols = [d[0] for d in cur.description]
                    cols_lower = [c.lower() for c in cols]
                    cc = _trova_col_call(cols_lower)
                    # Conta righe
                    try:
                        cur.execute(f"SELECT COUNT(*) FROM [{t}]")
                        n = cur.fetchone()[0]
                    except Exception:
                        n = 0
                    diag.append(f"  {t}: {n} righe, {len(cols)} colonne" +
                                (f" [call={cc}]" if cc else " [no call]"))
                    if cc and n > best_count:
                        best_count = n
                        tab = t
                        col_call_name = cc
                except Exception:
                    continue

            if not tab:
                con.close()
                messagebox.showerror(T("imp_l4o_no_tab"),
                    T("imp_l4o_no_tab_msg", diag="\n".join(diag[:15])))
                return

            cur.execute(f"SELECT * FROM [{tab}] LIMIT 1")
            col_names = [d[0] for d in cur.description]

            # Log4OM usa già nomi ADIF (call, qso_date, band, mode...).
            # Costruisce la mappa identità per le colonne ADIF note.
            # Mappa esplicita colonne Log4OM → campi ADIF.
            # Log4OM usa 'callsign', 'qsodate' (timestamp), 'startdate'/'timeon' ecc.
            mappa_l4o = {
                'callsign': 'call',
                'qsodate': 'qso_date',       # spesso timestamp completo → estrai data
                'startdate': 'qso_date',
                'timeon': 'time_on',
                'starttime': 'time_on',
                'timeoff': 'time_off',
                'endtime': 'time_off',
                'band': 'band', 'bandrx': 'band_rx',
                'freq': 'freq', 'frequency': 'freq', 'freqrx': 'freq_rx',
                'mode': 'mode', 'submode': 'submode',
                'rstsent': 'rst_sent', 'rst_sent': 'rst_sent', 'rstsnt': 'rst_sent',
                'rstrcvd': 'rst_rcvd', 'rst_rcvd': 'rst_rcvd', 'rstrcv': 'rst_rcvd',
                'name': 'name', 'qth': 'qth',
                'gridsquare': 'gridsquare', 'grid': 'gridsquare', 'locator': 'gridsquare',
                'country': 'country', 'dxcc': 'dxcc',
                'cqz': 'cqz', 'cqzone': 'cqz', 'ituz': 'ituz', 'ituzone': 'ituz',
                'cont': 'cont', 'continent': 'cont', 'iota': 'iota',
                'sota_ref': 'sota_ref', 'sotaref': 'sota_ref', 'sota': 'sota_ref',
                'pota_ref': 'pota_ref', 'potaref': 'pota_ref', 'pota': 'pota_ref',
                'state': 'state', 'cnty': 'cnty', 'county': 'cnty',
                'comment': 'comment', 'comments': 'comment', 'notes': 'comment',
                'qslsent': 'qsl_sent', 'qsl_sent': 'qsl_sent',
                'qslrcvd': 'qsl_rcvd', 'qsl_rcvd': 'qsl_rcvd',
                'qslsdate': 'qslsdate', 'qslrdate': 'qslrdate', 'qslvia': 'qsl_via', 'qsl_via':'qsl_via',
                'lotwqslsent': 'lotw_qsl_sent', 'lotw_qsl_sent': 'lotw_qsl_sent',
                'lotwqslrcvd': 'lotw_qsl_rcvd', 'lotw_qsl_rcvd': 'lotw_qsl_rcvd',
                'lotwqslsdate': 'lotw_qslsdate', 'lotwqslrdate': 'lotw_qslrdate',
                'eqslqslsent': 'eqsl_qsl_sent', 'eqsl_qsl_sent': 'eqsl_qsl_sent',
                'eqslqslrcvd': 'eqsl_qsl_rcvd', 'eqsl_qsl_rcvd': 'eqsl_qsl_rcvd',
                'eqslqslsdate': 'eqsl_qslsdate', 'eqslqslrdate': 'eqsl_qslrdate',
                'satname': 'sat_name', 'sat_name': 'sat_name',
                'propmode': 'prop_mode', 'prop_mode': 'prop_mode',
                'txpwr': 'tx_pwr', 'tx_pwr': 'tx_pwr', 'power': 'tx_pwr', 'rxpwr':'tx_pwr',
                'operator': 'operator', 'mygridsquare': 'my_gridsquare', 'my_gridsquare':'my_gridsquare',
                'stationcallsign': 'station_callsign', 'station_callsign': 'station_callsign',
                'contestid': 'contest_id', 'contest_id':'contest_id',
                'srx': 'srx', 'stx': 'stx', 'srxstring': 'srx_string', 'stxstring': 'stx_string',
                'email': 'email', 'address': 'address', 'age': 'age',
                'sig': 'sig', 'siginfo': 'sig_info', 'sig_info':'sig_info',
                'arrlsect': 'arrl_sect', 'arrl_sect': 'arrl_sect',
            }

            cur.execute(f"SELECT * FROM [{tab}]")
            righe = cur.fetchall()
            con.close()

            def _norm_key(col):
                """Normalizza il nome colonna: minuscolo, senza underscore/prefissi."""
                k = col.lower()
                for pfx in ('col_', 'app_log4om_', 'log4om_', 'l4o_'):
                    if k.startswith(pfx):
                        k = k[len(pfx):]
                return k

            qsos = []
            for row in righe:
                q = {}
                for col in col_names:
                    key_norm = _norm_key(col)
                    adif_key = mappa_l4o.get(key_norm)
                    if not adif_key:
                        continue
                    val = row[col]
                    if val is None or str(val).strip() == "":
                        continue
                    v = str(val).strip()

                    # qsodate di Log4OM è spesso un timestamp "2026-03-10 08:15:00"
                    # o ISO "2026-03-10T08:15:00": estrai sia data che ora.
                    if adif_key == 'qso_date':
                        # Estrai ora se presente nello stesso campo
                        mt = re.search(r'[T ](\d{2}):(\d{2})(?::(\d{2}))?', v)
                        if mt and 'time_on' not in q:
                            q['time_on'] = mt.group(1)+mt.group(2)+(mt.group(3) or '')
                        v = re.sub(r'[^0-9]', '', v.split('T')[0].split(' ')[0])[:8]
                        if len(v) != 8:
                            continue
                        q['qso_date'] = v
                        continue

                    if adif_key in ('qslsdate','qslrdate','lotw_qslsdate',
                                    'lotw_qslrdate','eqsl_qslsdate','eqsl_qslrdate'):
                        v = re.sub(r'[^0-9]', '', v.split('T')[0].split(' ')[0])[:8]
                        if len(v) != 8:
                            continue
                    if adif_key in ('time_on','time_off'):
                        v = re.sub(r'[^0-9]', '', v)[:6]
                    if adif_key in ('band','band_rx'):
                        v = v.lower()
                    # Non sovrascrivere time_on già estratto da qsodate
                    if adif_key == 'time_on' and q.get('time_on'):
                        continue
                    q[adif_key] = v
                if q.get('call'):
                    qsos.append(q)

            if not qsos:
                # Diagnostica: mostra le colonne trovate per capire il mapping
                messagebox.showwarning(T("imp_l4o_no_qso"),
                    T("imp_l4o_no_qso_msg", tab=tab, n=len(righe), cols=", ".join(col_names[:30]) + ("…" if len(col_names) > 30 else "")))
                return

            if self.qsos_caricati:
                risp = messagebox.askyesnocancel(T("imp_l4o_titolo"),
                    T("imp_l4o_scelta", n=len(qsos), tot=len(self.qsos_caricati)))
                if risp is None:
                    return
                if risp:
                    self.qsos_caricati.extend(qsos)
                else:
                    self.qsos_caricati = qsos
            else:
                self.qsos_caricati = qsos

            self.qsos_caricati.sort(key=lambda q: (
                str(q.get('qso_date','')).strip(),
                str(q.get('time_on','')).strip().zfill(6)))
            self.qsos_filtrati = list(self.qsos_caricati)
            self._aggiorna_tree()
            messagebox.showinfo(T("imp_completato"),
                T("imp_l4o_ok", n=len(qsos), tot=len(self.qsos_caricati)))

        except sqlite3.DatabaseError:
            messagebox.showerror(T("imp_db_errore"),
                T("imp_l4o_db_err"))
        except Exception as ex:
            messagebox.showerror(T("imp_errore"), T("imp_err_durante", ex=ex))

    def importa_hrd(self):
        """Importa QSO direttamente dal database SQLite di Ham Radio Deluxe.
        HRD v6.x usa SQLite con tabella TABLE_HRD_CONTACTS_V01 e colonne col_*."""
        import sqlite3

        path = filedialog.askopenfilename(
            title=T("imp_hrd_sel"),
            filetypes=[("HRD Logbook","*.hrdsql *.hrd *.db *.sqlite *.db3 *.sqlite3"),
                       ("HRD SQLite (.hrdsql)","*.hrdsql"),
                       ("Tutti i file","*.*")])
        if not path:
            return

        try:
            con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
            con.row_factory = sqlite3.Row
            cur = con.cursor()

            # Trova la tabella dei contatti (nome può variare leggermente)
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tabelle = [r[0] for r in cur.fetchall()]
            tab_contatti = None
            for t in tabelle:
                if 'CONTACTS' in t.upper() or 'HRD_CONTACTS' in t.upper():
                    tab_contatti = t
                    break
            if not tab_contatti:
                con.close()
                messagebox.showerror(T("imp_hrd_db_nonrec"),
                    T("imp_hrd_no_tab", tab=", ".join(tabelle[:8])))
                return

            # Legge le colonne disponibili
            cur.execute(f"SELECT * FROM [{tab_contatti}] LIMIT 1")
            col_names = [d[0] for d in cur.description]

            # Mappa colonne HRD (col_*) → campi ADIF
            mappa = {
                'col_call': 'call', 'col_qso_date': 'qso_date', 'col_time_on': 'time_on',
                'col_time_off': 'time_off', 'col_band': 'band', 'col_band_rx': 'band_rx',
                'col_freq': 'freq', 'col_freq_rx': 'freq_rx', 'col_mode': 'mode',
                'col_submode': 'submode', 'col_rst_sent': 'rst_sent', 'col_rst_rcvd': 'rst_rcvd',
                'col_name': 'name', 'col_qth': 'qth', 'col_gridsquare': 'gridsquare',
                'col_country': 'country', 'col_dxcc': 'dxcc', 'col_cqz': 'cqz',
                'col_ituz': 'ituz', 'col_cont': 'cont', 'col_iota': 'iota',
                'col_sota_ref': 'sota_ref', 'col_pota_ref': 'pota_ref',
                'col_state': 'state', 'col_cnty': 'cnty', 'col_comment': 'comment',
                'col_qsl_sent': 'qsl_sent', 'col_qsl_rcvd': 'qsl_rcvd',
                'col_qsl_sent_date': 'qslsdate', 'col_qsl_rcvd_date': 'qslrdate',
                'col_qsl_via': 'qsl_via',
                'col_lotw_qsl_sent': 'lotw_qsl_sent', 'col_lotw_qsl_rcvd': 'lotw_qsl_rcvd',
                'col_lotw_qslsdate': 'lotw_qslsdate', 'col_lotw_qslrdate': 'lotw_qslrdate',
                'col_eqsl_qsl_sent': 'eqsl_qsl_sent', 'col_eqsl_qsl_rcvd': 'eqsl_qsl_rcvd',
                'col_eqsl_qslsdate': 'eqsl_qslsdate', 'col_eqsl_qslrdate': 'eqsl_qslrdate',
                'col_sat_name': 'sat_name', 'col_prop_mode': 'prop_mode',
                'col_tx_pwr': 'tx_pwr', 'col_operator': 'operator',
                'col_my_gridsquare': 'my_gridsquare', 'col_station_callsign': 'station_callsign',
                'col_contest_id': 'contest_id', 'col_srx': 'srx', 'col_stx': 'stx',
            }

            cur.execute(f"SELECT * FROM [{tab_contatti}]")
            righe = cur.fetchall()
            con.close()

            qsos = []
            for row in righe:
                q = {}
                for col in col_names:
                    val = row[col]
                    if val is None or str(val).strip() == "":
                        continue
                    adif_key = mappa.get(col.lower())
                    if not adif_key:
                        continue
                    v = str(val).strip()
                    # Normalizza data (HRD usa spesso YYYY-MM-DD o timestamp)
                    if adif_key in ('qso_date','qslsdate','qslrdate',
                                    'lotw_qslsdate','lotw_qslrdate',
                                    'eqsl_qslsdate','eqsl_qslrdate'):
                        v = re.sub(r'[^0-9]', '', v)[:8]
                        if len(v) != 8:
                            continue
                    # Normalizza ora (HRD HH:MM:SS → HHMMSS)
                    if adif_key in ('time_on','time_off'):
                        v = re.sub(r'[^0-9]', '', v)[:6]
                    # Band in minuscolo
                    if adif_key in ('band','band_rx'):
                        v = v.lower()
                    q[adif_key] = v
                if q.get('call'):
                    qsos.append(q)

            if not qsos:
                messagebox.showwarning(T("imp_no_qso"),
                    T("imp_hrd_letto"))
                return

            # Chiede se sostituire o aggiungere al log corrente
            if self.qsos_caricati:
                risp = messagebox.askyesnocancel(T("imp_hrd_titolo"),
                    T("imp_hrd_scelta", n=len(qsos), tot=len(self.qsos_caricati)))
                if risp is None:
                    return
                if risp:
                    self.qsos_caricati.extend(qsos)
                else:
                    self.qsos_caricati = qsos
            else:
                self.qsos_caricati = qsos

            # Ordina cronologicamente (HRD restituisce per chiave primaria, non per data)
            def _chiave_ord(q):
                return (str(q.get('qso_date','')).strip(),
                        str(q.get('time_on','')).strip().zfill(6))
            self.qsos_caricati.sort(key=_chiave_ord)

            self.qsos_filtrati = list(self.qsos_caricati)
            self._aggiorna_tree()
            messagebox.showinfo(T("imp_completato"),
                T("imp_hrd_ok", n=len(qsos), tot=len(self.qsos_caricati)))

        except sqlite3.DatabaseError:
            messagebox.showerror(T("imp_db_errore"),
                T("imp_hrd_db_err"))
        except Exception as ex:
            messagebox.showerror(T("imp_errore"), T("imp_err_durante", ex=ex))

    def importa_cbr(self):
        """Importa un file Cabrillo creando un nuovo log ADIF in memoria."""
        path = filedialog.askopenfilename(
            title=T("dv_importa_cbr_file"),
            filetypes=[("Cabrillo","*.cbr *.log *.txt"),("All files","*.*")])
        if not path:
            return
        try:
            # Usa il parser della UnisciDialog
            tmp = UnisciDialog.__new__(UnisciDialog)
            qsos_cbr = tmp._leggi_cabrillo(path)
            if not qsos_cbr:
                messagebox.showwarning(T("attenzione"),
                    "Nessun QSO trovato nel file Cabrillo.")
                return
            # Se c'è già un log chiedere cosa fare
            if self.qsos_caricati:
                r = messagebox.askyesnocancel(
                    "Importa CBR",
                    f"Trovati {len(qsos_cbr)} QSO nel file CBR." + chr(10) +
                    chr(10) +
                    "SI = aggiungi al log corrente" + chr(10) +
                    "NO = crea nuovo log (sostituisce quello corrente)" + chr(10) +
                    "ANNULLA = annulla")
                if r is None:
                    return
                if r:  # SI — aggiungi
                    self.qsos_caricati.extend(qsos_cbr)
                else:  # NO — sostituisci
                    self.qsos_caricati = qsos_cbr
                    self.qsos_filtrati = list(qsos_cbr)
                    self.filepath = path
            else:
                # Nessun log — crea nuovo
                self.qsos_caricati = qsos_cbr
                self.qsos_filtrati = list(qsos_cbr)
                self.filepath = path

            # Aggiorna interfaccia
            nome = os.path.basename(path)
            n = len(self.qsos_caricati)
            self.lbl_status.configure(
                text=f"{nome} — {n} QSO (da CBR)",
                text_color="#ED8936")
            self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
            messagebox.showinfo("Importato",
                f"{len(qsos_cbr)} QSO importati dal file Cabrillo." + chr(10) +
                f"Totale log: {len(self.qsos_caricati)} QSO" + chr(10) + chr(10) +
                "Usa 'Salva ADIF' per salvare il log su disco.")
        except Exception as ex:
            messagebox.showerror(T("errore"),
                f"Impossibile leggere il file Cabrillo:" + chr(10) + str(ex))

    def salva_adif(self):
        """Salva il log corrente come nuovo file ADIF aggiornato.
        Se è attivo un filtro (incluso il filtro rapido Ctrl+Click), salva
        solo i QSO visibili/filtrati, chiedendo conferma per evitare
        salvataggi parziali accidentali. Funziona anche senza un file di
        origine caricato (es. QSO inseriti solo con "Aggiungi QSO")."""
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("qsl_no_adif"))
            return

        qsos = self._qsos_attivi()
        tot = len(self.qsos_caricati)
        n = len(qsos)

        if n != tot:
            if not messagebox.askyesno(
                "Filtro attivo",
                f"È attivo un filtro: verranno salvati solo {n} QSO su {tot} totali.\n\n"
                f"Vuoi procedere e salvare solo i QSO filtrati?\n\n"
                f"(Per salvare l'intero log, premi 'No' e rimuovi prima il filtro "
                f"con il pulsante '✕ Reset filtro')"
            ):
                return

        if self.filepath:
            base = os.path.splitext(os.path.basename(self.filepath))[0]
            nome_def = base + "_QSL.adif"
        else:
            # Nessun file di origine (es. log iniziato da zero con
            # "Aggiungi QSO"): propone un nome generico con la data odierna.
            nome_def = "ADIF_FZR_" + datetime.now().strftime("%Y%m%d") + ".adif"

        save_path = self._chiedi_cartella_output(nome_def)
        if save_path is None:
            return  # dialogo chiuso con X
        if save_path == "":
            # Usa dialogo standard
            save_path = filedialog.asksaveasfilename(
                title=T("dv_salva_adif_agg"),
                defaultextension=".adif",
                filetypes=[("ADIF files", "*.adif"), ("All files", "*.*")],
                initialfile=nome_def
            )
        if not save_path:
            return
        try:
            self._scrivi_adif(save_path, qsos)
            self._log_modificato = False   # salvato: niente più modifiche pendenti
            messagebox.showinfo(T("successo"), T("salva_adif_ok", f=os.path.basename(save_path)))
        except Exception as ex:
            messagebox.showerror(T("errore"), f"{T('salva_adif_err')}{ex}")

    def _chiedi_cartella_output(self, nome_file):
        """Chiede se usare cartella dedicata, poi restituisce il path completo."""
        return chiedi_cartella_output(self, nome_file, self.filepath)

    def _scrivi_adif(self, path, qsos):
        """Scrive una lista di QSO in formato ADIF."""
        # Campi header da escludere dai record QSO
        CAMPI_HEADER = {'adif_ver', 'programid', 'programversion', 'created_timestamp'}
        with open(path, "w", encoding="utf-8") as f:
            f.write("<ADIF_VER:5>3.1.4\n")
            f.write(f"<PROGRAMID:{len(PROGRAMID_ADIF)}>{PROGRAMID_ADIF}\n")
            f.write(f"<PROGRAMVERSION:{len(VERSIONE)}>{VERSIONE}\n")
            f.write("<EOH>\n\n")
            for qso in qsos:
                for k, v in qso.items():
                    if k.lower() in CAMPI_HEADER:
                        continue
                    if v and str(v).strip():
                        tag = k.upper()
                        val = str(v)
                        f.write(f"<{tag}:{len(val)}>{val} ")
                f.write("<EOR>\n")

    def apri_duplicati(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        DuplicatiDialog(self, self)

    def deduci_country_da_nominativo(self):
        """Per i QSO con COUNTRY assente o vuoto, deduce Country, DXCC e
        Continente dal prefisso del nominativo (CALL), usando la tabella
        DXCC_PREFIX_TABLE. Non sovrascrive mai un COUNTRY già presente.
        Opera sui QSO attualmente visibili (rispettando eventuali filtri)."""
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        qsos = self._qsos_attivi()
        n_compilati = 0
        n_non_risolti = 0
        prefissi_non_risolti = set()
        for qso in qsos:
            keys_low = {k.lower(): k for k in qso.keys()}
            country_key = keys_low.get('country', 'COUNTRY')
            country_val = str(qso.get(country_key, '')).strip() if 'country' in keys_low else ''
            if country_val:
                continue  # già presente, non tocco
            call = qso.get(keys_low.get('call', 'CALL'), '') if 'call' in keys_low else ''
            if not call:
                continue
            risultato = dxcc_da_nominativo(call)
            if risultato is None:
                n_non_risolti += 1
                pfx = str(call).strip().upper().split('/')[0][:3]
                prefissi_non_risolti.add(pfx)
                continue
            country, dxcc_code, cont = risultato
            qso[country_key] = country
            dxcc_key = keys_low.get('dxcc', 'dxcc')
            qso[dxcc_key] = dxcc_code
            cont_key = keys_low.get('cont', 'cont')
            qso[cont_key] = cont
            n_compilati += 1
        self._aggiorna_tree()
        msg = T("deduci_risultato", n=n_compilati)
        if n_non_risolti:
            esempio = ", ".join(sorted(prefissi_non_risolti)[:8])
            etc = "…" if len(prefissi_non_risolti) > 8 else ""
            msg += T("deduci_non_risolti", n=n_non_risolti, esempio=esempio, etc=etc)
        messagebox.showinfo(T("deduci_titolo"), msg)

    # ── Formati etichetta QSL personalizzati ──
    def _carica_formati_qsl(self):
        """Carica i formati etichetta custom dal file JSON e li restituisce come dict."""
        try:
            if os.path.exists(self.formati_qsl_path):
                with open(self.formati_qsl_path, 'r', encoding='utf-8') as f:
                    dati = json.load(f)
                # dati = {nome: [lw, lh, cols, rows, ml, mt, gh, gv, psize]}
                return {k: tuple(v) for k, v in dati.items()}
        except Exception:
            pass
        return {}

    def _salva_formati_qsl(self, formati_custom):
        """Salva i formati etichetta custom su file JSON."""
        try:
            with open(self.formati_qsl_path, 'w', encoding='utf-8') as f:
                json.dump({k: list(v) for k, v in formati_custom.items()},
                          f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def apri_qsl_card(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        # Impedisce finestre duplicate — porta in primo piano quella esistente
        if hasattr(self, '_qsl_master_dlg') and self._qsl_master_dlg and \
                self._qsl_master_dlg.winfo_exists():
            self._qsl_master_dlg.lift()
            self._qsl_master_dlg.focus_force()
            return
        stazione = self.entry_owner.get().strip().upper()
        formati_custom = self._carica_formati_qsl()
        dlg = QSLMasterDialog(self, self._qsos_attivi(), stazione, self.colori_pdf)
        self._qsl_master_dlg = dlg
        # Inietta i formati custom salvati nel dialog
        if formati_custom:
            dlg.FORMATI = {**formati_custom, **dlg.FORMATI}
            voci = list(dlg.FORMATI.keys()) + ["▶ Formato personalizzato…"]
            try:
                dlg._menu_formato.configure(values=voci)
            except Exception:
                pass
        # Callback per salvare quando il dialog aggiunge un formato custom
        dlg._app_salva_formati = self._salva_formati_qsl

    def apri_qsl_card_designer(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        stazione = self.entry_owner.get().strip().upper()
        QSLCardDesignerDialog(self, self._qsos_attivi(), stazione, self.colori_pdf)


    # ── Cloudlog upload ────────────────────────
    def apri_cloudlog_upload(self):
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        cl_url = dati.get('cloudlog_url', '').strip()
        cl_key = dati.get('cloudlog_api_key', '').strip()
        cl_station = dati.get('cloudlog_station', '').strip()
        if not (cl_url and cl_key and cl_station):
            messagebox.showwarning(T("attenzione"), T("cl_no_config"))
            return
        CloudlogUploadDialog(self, self._qsos_attivi(), self._qsos_da_selezione(),
                              cl_url, cl_key, cl_station)

    # ── Clublog upload ─────────────────────────
    def apri_clublog_upload(self):
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        cb_email = dati.get('clublog_email', '').strip()
        cb_password = dati.get('clublog_password', '').strip()
        cb_api = dati.get('clublog_api_key', '').strip()  # opzionale: alcuni client (es. HRD) caricano anche senza
        cb_callsign = dati.get('callsign', '').strip() or self.entry_owner.get().strip()
        if not (cb_email and cb_password and cb_callsign):
            messagebox.showwarning(T("attenzione"), T("cb_no_config"))
            return
        ClublogUploadDialog(self, self._qsos_attivi(), self._qsos_da_selezione(),
                             cb_email, cb_password, cb_callsign, cb_api)

    # ── LoTW upload ────────────────────────────
    def apri_lotw_upload(self):
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        tqsl_path = dati.get('lotw_tqsl_path', '').strip()
        station_location = dati.get('lotw_station_location', '').strip()
        if not (tqsl_path and station_location):
            messagebox.showwarning(T("attenzione"), T("lw_no_config"))
            return
        LotwUploadDialog(self, self._qsos_attivi(), self._qsos_da_selezione(),
                          tqsl_path, station_location)

    # ── eQSL upload ────────────────────────────
    def apri_eqsl_upload(self):
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        eq_user = dati.get('eqsl_username', '').strip()
        eq_pass = dati.get('eqsl_password', '').strip()
        eq_qth = dati.get('eqsl_qth_nickname', '').strip()
        if not (eq_user and eq_pass):
            messagebox.showwarning(T("attenzione"), T("eq_no_config"))
            return
        EqslUploadDialog(self, self._qsos_attivi(), self._qsos_da_selezione(),
                          eq_user, eq_pass, eq_qth)

    # ── QO-100 DX Club upload ──────────────────
    def apri_qo100_upload(self):
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_carica_prima"))
            return
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        api_key  = dati.get('qo100_api_key', '').strip()
        my_grid  = dati.get('qo100_my_grid', '') or dati.get('locator', '')
        callsign = dati.get('callsign', '').strip() or self.entry_owner.get().strip()
        if not api_key:
            messagebox.showwarning(T("attenzione"), T("qo100_no_config"))
            return
        QO100UploadDialog(self, self._qsos_attivi(), api_key, callsign, my_grid)

    # ── LoTW download ──────────────────────────
    def apri_lotw_download(self):
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        lw_user = dati.get('lotw_username', '').strip()
        lw_pass = dati.get('lotw_password', '').strip()
        if not (lw_user and lw_pass):
            messagebox.showwarning(T("attenzione"), T("lwd_no_config"))
            return
        default_call = dati.get('callsign', '').strip() or self.entry_owner.get().strip()
        LotwDownloadDialog(self, lw_user, lw_pass, default_call)

    # ── eQSL download ──────────────────────────
    def apri_eqsl_download(self):
        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
        eq_user = dati.get('eqsl_username', '').strip()
        eq_pass = dati.get('eqsl_password', '').strip()
        eq_qth  = dati.get('eqsl_qth_nickname', '').strip()
        if not (eq_user and eq_pass):
            messagebox.showwarning(T("attenzione"), T("eqd_no_config"))
            return
        EqslDownloadDialog(self, eq_user, eq_pass, eq_qth)

    def apri_grafici(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        GraficiDialog(self, self._qsos_attivi(), self.colori_pdf)
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        GraficiDialog(self, self._qsos_attivi(), self.colori_pdf)

    # ── Storico file ──────────────────────────
    def _aggiungi_storico(self, path):
        path = os.path.abspath(path)
        if path in self.storico_files:
            self.storico_files.remove(path)
        self.storico_files.insert(0, path)
        self.storico_files = self.storico_files[:10]  # max 10 file
        try:
            with open(self.storico_path, 'w', encoding='utf-8') as f:
                json.dump(self.storico_files, f, ensure_ascii=False)
        except Exception:
            pass

    def _carica_impostazioni_apertura(self):
        """Carica da disco lo stato salvato del flag 'Controlla campi
        principali dopo l'apertura', mostrato come checkbox nella finestra
        Apri ADIF. Disattivo di default se non è mai stato salvato."""
        try:
            if os.path.exists(self.impostazioni_apertura_path):
                with open(self.impostazioni_apertura_path, 'r', encoding='utf-8') as f:
                    dati = json.load(f)
                self.var_controllo_post_apertura.set(bool(dati.get('controllo_post_apertura', False)))
        except Exception:
            pass

    def _salva_impostazioni_apertura(self):
        """Salva su disco lo stato corrente del flag 'Controlla campi
        principali dopo l'apertura', così viene ricordato ai prossimi avvii."""
        try:
            with open(self.impostazioni_apertura_path, 'w', encoding='utf-8') as f:
                json.dump({'controllo_post_apertura': self.var_controllo_post_apertura.get()}, f)
        except Exception:
            pass

    def apri_storico(self):
        try:
            if os.path.exists(self.storico_path):
                with open(self.storico_path, 'r', encoding='utf-8') as f:
                    self.storico_files = json.load(f)
        except Exception:
            self.storico_files = []

        # Rimuovi file non più esistenti
        self.storico_files = [p for p in self.storico_files if os.path.exists(p)]

        if not self.storico_files:
            messagebox.showinfo(T("storico"), T("ok_nessun_rec"))
            return

        dlg = ctk.CTkToplevel(self)
        dlg.title(T("file_recenti"))
        dlg.geometry("520x360")
        dlg.resizable(False, False)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text=T("file_recenti"),
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=10)

        frame_lista = ctk.CTkScrollableFrame(dlg, height=220)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=5)

        def carica_da_storico(path):
            dlg.destroy()
            self.filepath = path
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    testo = f.read()
                testo = self._fix_adif(testo)
                qsos = self._leggi_adif_sicuro(testo)
                self.qsos_caricati = sorted(qsos, key=lambda x: (x.get('qso_date', ''), x.get('time_on', '')))
                self.qsos_filtrati = list(self.qsos_caricati)
                nome = os.path.basename(path)
                self.lbl_status.configure(text=T("caricato_status", nome=nome, count=len(self.qsos_caricati)), text_color=TH.PRIMARY)
                self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
                self._aggiungi_storico(path)
            except Exception as ex:
                messagebox.showerror("Errore", f"{T('impossibile_aprire')}{ex}")

        for i, path in enumerate(self.storico_files):
            nome = os.path.basename(path)
            cartella = os.path.dirname(path)
            fr = ctk.CTkFrame(frame_lista, fg_color="#E2E8F0" if i%2==0 else "transparent", corner_radius=4)
            fr.pack(fill="x", pady=2)
            ctk.CTkLabel(fr, text=f"  {nome}", font=ctk.CTkFont(weight="bold"),
                         anchor="w").pack(side="left", padx=8, pady=4)
            ctk.CTkLabel(fr, text=cartella, font=ctk.CTkFont(size=10),
                         text_color="gray", anchor="w").pack(side="left", padx=4)
            ctk.CTkButton(fr, text=T("apri"), width=60, height=26,
                          fg_color=TH.PRIMARY,
                          command=lambda p=path: carica_da_storico(p)).pack(side="right", padx=8, pady=4)

        ctk.CTkButton(dlg, text=T("chiudi"), command=dlg.destroy,
                      fg_color="#718096", width=100).pack(pady=8)

    # ── Profilo ───────────────────────────────
    def _dati_stazione_correnti(self):
        """Ritorna dict con i dati stazione dai campi UI."""
        return {
            'callsign': self.entry_owner.get().strip().upper(),
            'locator':  self.entry_details.get().strip(),
            'nome_op':  '',
            'qth':      '',
            'cq_zone':  '',
            'itu_zone': '',
        }

    def _applica_profilo(self, profilo):
        """Applica i dati di un profilo all'interfaccia."""
        self.entry_owner.delete(0, 'end')
        self.entry_owner.insert(0, profilo.get('callsign',''))
        self.entry_details.delete(0, 'end')
        self.entry_details.insert(0, profilo.get('locator',''))
        self.profilo_attivo = profilo.get('nome','')
        self.btn_profili.configure(text=self.profilo_attivo or '—')
        # Lingua
        if profilo.get('lingua') and hasattr(self, '_set_lingua'):
            self._set_lingua(profilo['lingua'])
        # Tema
        if profilo.get('tema') and profilo.get('pref_ricorda_tema', True):
            tema = profilo['tema']
            if tema != ctk.get_appearance_mode():
                ctk.set_appearance_mode(tema)
        # Colora righe
        if hasattr(self, 'var_colora_righe'):
            self.var_colora_righe.set(bool(profilo.get('pref_colora_righe', True)))
            if hasattr(self, 'qsos_caricati') and self.qsos_caricati:
                self._aggiorna_tree()

    def _carica_profili(self):
        """Carica il file profili multipli."""
        if not os.path.exists(self.profili_path):
            return {}
        try:
            with open(self.profili_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def _salva_profili(self, profili):
        """Salva il file profili multipli."""
        try:
            with open(self.profili_path, 'w', encoding='utf-8') as f:
                json.dump(profili, f, ensure_ascii=False, indent=2)
        except Exception as ex:
            messagebox.showerror("Errore", f"Impossibile salvare profili:\n{ex}")

    def _dialog_profilo(self, parent, titolo, dati=None):
        """Dialogo per creare o modificare un profilo. Ritorna dict o None."""
        dlg = ctk.CTkToplevel(parent)
        dlg.title(titolo)
        dlg.geometry("440x560")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text=titolo,
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(14,8), padx=20)

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=4, pady=(0,4))

        CAMPI = [
            ("nome",     T("profilo_nome"),     "es. IW1FZR Home"),
            ("callsign", T("profilo_callsign"), "es. IW1FZR"),
            ("locator",  T("profilo_locator"),  "es. JN45bj"),
            ("nome_op",  T("profilo_nome_op"),  "es. Luca"),
            ("qth",      T("profilo_qth"),      "es. Cavaglià (BI)"),
            ("cq_zone",  T("profilo_cq"),       "es. 15"),
            ("itu_zone", T("profilo_itu"),       "es. 28"),
            ("cloudlog_url",     T("profilo_cl_url"),     "es. https://log.iw1fzr.it"),
            ("cloudlog_api_key", T("profilo_cl_key"),     "API key (read/write)"),
            ("cloudlog_station", T("profilo_cl_station"), "es. 1"),
            ("clublog_email",    T("profilo_cb_email"),   "es. me@example.com"),
            ("clublog_password", T("profilo_cb_pass"),    "App Password Clublog"),
            ("clublog_api_key",  T("profilo_cb_api"),     "API key Clublog"),
            ("lotw_tqsl_path",   T("profilo_lw_path"),    r"es. C:\Program Files\TrustedQSL\tqsl.exe"),
            ("lotw_station_location", T("profilo_lw_loc"),"es. Home"),
            ("lotw_username", T("profilo_lw_user"), "es. IW1FZR"),
            ("lotw_password", T("profilo_lw_pass"), "Password LoTW"),
            ("eqsl_username", T("profilo_eq_user"), "es. IW1FZR"),
            ("eqsl_password", T("profilo_eq_pass"), "Password eQSL"),
            ("eqsl_qth_nickname", T("profilo_eq_qth"), "es. HOME (lascia vuoto se unico QTH)"),
            ("hamqth_username", T("profilo_hqth_user"), "es. IW1FZR"),
            ("hamqth_password", T("profilo_hqth_pass"), "Password HamQTH"),
            ("qo100_api_key",   T("profilo_qo100_key"), "API Key da qo100dx.club/profile"),
            ("qo100_my_grid",   T("profilo_qo100_grid"), "es. JN45bj"),
        ]
        entries = {}
        for key, lbl, ph in CAMPI:
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", padx=16, pady=3)
            ctk.CTkLabel(row, text=lbl, width=140, anchor="e",
                         font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,8))
            show_char = "*" if key in ("clublog_password", "eqsl_password", "lotw_password", "hamqth_password") else None
            e = ctk.CTkEntry(row, width=200, placeholder_text=ph, show=show_char)
            if dati and dati.get(key):
                e.insert(0, dati[key])
            e.pack(side="left")
            entries[key] = e
            if key == "lotw_tqsl_path":
                def _sfoglia_tqsl(ent=e):
                    p = filedialog.askopenfilename(
                        title=T("dv_sel_tqsl"),
                        filetypes=[("Eseguibile", "*.exe"), ("Tutti i file", "*.*")])
                    if p:
                        ent.delete(0, 'end'); ent.insert(0, p)
                ctk.CTkButton(row, text=T("lw_sfoglia"), width=70, height=26,
                              command=_sfoglia_tqsl, fg_color="#4A5568").pack(side="left", padx=(6,0))

        result = [None]
        def salva():
            nome = entries["nome"].get().strip()
            call = entries["callsign"].get().strip().upper()
            if not nome or not call:
                messagebox.showwarning("Attenzione", "Nome profilo e Callsign sono obbligatori.")
                return
            result[0] = {k: entries[k].get().strip() for k in entries}
            result[0]['callsign'] = result[0]['callsign'].upper()
            dlg.destroy()

        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=20, pady=10)
        ctk.CTkButton(frame_btn, text=T("conferma_btn"), command=salva,
                      fg_color=TH.SUCCESS_H, height=34).pack(side="left", expand=True, padx=(0,6), fill="x")
        ctk.CTkButton(frame_btn, text=T("cm_annulla"), command=dlg.destroy,
                      fg_color="#718096", height=34).pack(side="left", expand=True, fill="x")
        dlg.wait_window()
        return result[0]

    def _crea_filtri_inline(self, parent):
        """Crea un pannello filtri compatto direttamente nella sidebar."""
        def _lbl(txt):
            ctk.CTkLabel(parent, text=txt, font=ctk.CTkFont(size=10),
                         text_color="gray", anchor="w").pack(fill="x", padx=4, pady=(4,0))

        # Data da / a
        _lbl("Data dal:")
        self._fil_data_da = ctk.CTkEntry(parent, placeholder_text="YYYYMMDD",
                                          font=ctk.CTkFont(size=10), height=26)
        self._fil_data_da.pack(fill="x", padx=4, pady=(0,2))
        _lbl("Data al:")
        self._fil_data_al = ctk.CTkEntry(parent, placeholder_text="YYYYMMDD",
                                          font=ctk.CTkFont(size=10), height=26)
        self._fil_data_al.pack(fill="x", padx=4, pady=(0,2))

        # Callsign
        _lbl("Callsign (contiene):")
        self._fil_call = ctk.CTkEntry(parent, font=ctk.CTkFont(size=10), height=26)
        self._fil_call.pack(fill="x", padx=4, pady=(0,2))

        # Banda
        _lbl("Banda:")
        self._fil_banda = ctk.CTkEntry(parent, placeholder_text=T("dv_ph_20m"),
                                        font=ctk.CTkFont(size=10), height=26)
        self._fil_banda.pack(fill="x", padx=4, pady=(0,2))

        # Modo
        _lbl("Modo:")
        self._fil_modo = ctk.CTkEntry(parent, placeholder_text=T("dv_ph_ft8"),
                                       font=ctk.CTkFont(size=10), height=26)
        self._fil_modo.pack(fill="x", padx=4, pady=(0,4))

        frame_btns = ctk.CTkFrame(parent, fg_color="transparent")
        frame_btns.pack(fill="x", padx=4, pady=(0,4))
        ctk.CTkButton(frame_btns, text=T("cm_applica"), command=self._applica_filtri_inline,
                      height=26, fg_color=TH.PRIMARY, font=ctk.CTkFont(size=10)
                      ).pack(side="left", expand=True, fill="x", padx=(0,4))
        ctk.CTkButton(frame_btns, text=T("cm_reset"), command=self._reset_filtri_inline,
                      height=26, fg_color="#4A5568", font=ctk.CTkFont(size=10), width=60
                      ).pack(side="left")

    def _applica_filtri_sidebar(self):
        """Applica i filtri della sidebar al log."""
        import datetime as _dt
        def _parse(s):
            for fmt in ("%d/%m/%Y", "%Y%m%d"):
                try: return _dt.datetime.strptime(s.strip(), fmt)
                except: pass
            return None

        da    = _parse(self._fil_data_da.get()) if hasattr(self,'_fil_data_da') else None
        a     = _parse(self._fil_data_al.get()) if hasattr(self,'_fil_data_al') else None
        call  = self._fil_call.get().strip().upper() if hasattr(self,'_fil_call') else ""
        banda = self._fil_banda_var.get() if hasattr(self,'_fil_banda_var') else "Tutte"
        modo  = self._fil_modo_var.get()  if hasattr(self,'_fil_modo_var')  else "Tutti"
        sat   = self._fil_sat_var.get()   if hasattr(self,'_fil_sat_var')   else "Tutti"
        # "Tutte"/"Tutti"/"All" = nessun filtro (robusto al cambio lingua)
        _tutti = {"tutte", "tutti", "all"}

        filtrati = []
        for q in self.qsos_caricati:
            d_raw = str(q.get('qso_date','')).strip()
            d_qso = None
            if len(d_raw)==8:
                try: d_qso = _dt.datetime.strptime(d_raw,"%Y%m%d")
                except: pass
            if da and d_qso and d_qso < da: continue
            if a  and d_qso and d_qso > a:  continue
            if call and call not in str(q.get('call','')).upper(): continue
            if banda.lower() not in _tutti and str(q.get('band','')).upper().strip() != banda: continue
            if modo.lower()  not in _tutti and str(q.get('mode','')).upper().strip() != modo:  continue
            if sat.lower()   not in _tutti and sat not in str(q.get('sat_name','')).upper():   continue
            filtrati.append(q)

        self.qsos_filtrati = filtrati
        self._aggiorna_tree()

    def _reset_filtri_sidebar(self):
        """Resetta i filtri della sidebar."""
        for attr in ('_fil_data_da','_fil_data_al','_fil_call'):
            if hasattr(self, attr): getattr(self, attr).delete(0,'end')
        for var, chiave in (('_fil_banda_var','filtri_tutte'),('_fil_modo_var','filtri_tutti'),('_fil_sat_var','filtri_tutti')):
            if hasattr(self, var): getattr(self, var).set(T(chiave))
        self._azzera_filtro()

    def _applica_filtri_inline(self):

        """Applica i filtri dal pannello sidebar come filtro rapido."""
        # Usa la stessa logica del dialog filtri avanzati
        filtri = {}
        if hasattr(self, '_fil_data_da') and self._fil_data_da.get().strip():
            filtri['data_da'] = self._fil_data_da.get().strip()
        if hasattr(self, '_fil_data_al') and self._fil_data_al.get().strip():
            filtri['data_al'] = self._fil_data_al.get().strip()
        if hasattr(self, '_fil_call') and self._fil_call.get().strip():
            filtri['call'] = self._fil_call.get().strip().upper()
        if hasattr(self, '_fil_banda') and self._fil_banda.get().strip():
            filtri['band'] = self._fil_banda.get().strip().upper()
        if hasattr(self, '_fil_modo') and self._fil_modo.get().strip():
            filtri['mode'] = self._fil_modo.get().strip().upper()

        qsos = self.qsos_caricati
        for k, v in filtri.items():
            if k == 'data_da':
                qsos = [q for q in qsos if str(q.get('qso_date','')).strip() >= v]
            elif k == 'data_al':
                qsos = [q for q in qsos if str(q.get('qso_date','')).strip() <= v]
            elif k == 'call':
                qsos = [q for q in qsos if v in str(q.get('call','')).upper()]
            elif k == 'band':
                qsos = [q for q in qsos if str(q.get('band','')).upper().strip() == v]
            elif k == 'mode':
                qsos = [q for q in qsos if v in str(q.get('mode','')).upper()]

        self.qsos_filtrati = qsos
        self._aggiorna_tree()
        self._aggiorna_statusbar()

    def _reset_filtri_inline(self):
        """Resetta tutti i filtri del pannello sidebar."""
        for attr in ('_fil_data_da','_fil_data_al','_fil_call','_fil_banda','_fil_modo'):
            if hasattr(self, attr):
                getattr(self, attr).delete(0, 'end')
        self.qsos_filtrati = list(self.qsos_caricati)
        self._aggiorna_tree()
        self._aggiorna_statusbar()

    def _ricostruisci_toolbar1_wrap(self):
        """Ricostruisce la sezione riga-1 nella WrapToolbar."""
        try:
            profili = self._carica_profili()
            dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
            disattivati = set(dati.get('toolbar1_disattivati', []))
        except Exception:
            disattivati = set()
        self._tb1_refs = []
        for tb_id in self._ordine_toolbar1():
            if tb_id not in self._tb1_disponibili or tb_id in disattivati:
                continue
            text, cmd, tip, color, emoji = self._tb1_disponibili[tb_id]
            # Testo coerente con la lingua corrente (emoji + traduzione)
            self._tb_btn1_factory(self._tb_wrap, emoji + T(tb_id), cmd, tip, color, tb_id, emoji)

    def _ricostruisci_toolbar2_wrap(self):
        """Ricostruisce la sezione riga-2 nella WrapToolbar."""
        try:
            profili = self._carica_profili()
            dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
            disattivati = set(dati.get('toolbar2_disattivati', []))
        except Exception:
            disattivati = set()
        self._tb2_refs = []
        for tb_id in self._ordine_toolbar2():
            if tb_id not in self._tb2_disponibili or tb_id in disattivati:
                continue
            text, cmd, tip, color, emoji = self._tb2_disponibili[tb_id]
            # Testo coerente con la lingua corrente (emoji + traduzione)
            self._tb_btn2_factory(self._tb_wrap, emoji + T(tb_id), cmd, tip, color, tb_id, emoji)

    def _ricostruisci_toolbar1(self):
        """Compatibilità: ricostruisce l'intera WrapToolbar."""
        if hasattr(self, '_tb_wrap'):
            self._tb_wrap.clear()
            self._ricostruisci_toolbar1_wrap()
            s = ctk.CTkLabel(self._tb_wrap, text="", width=1, height=32,
                             fg_color=("#B0BBC8","#2A2A2A"))
            self._tb_wrap.add(s, is_sep=True)
            self._ricostruisci_toolbar2_wrap()

    def _ricostruisci_toolbar2(self):
        self._ricostruisci_toolbar1()

    def _applica_ordine_toolbar1(self, visibili):
        self._ricostruisci_toolbar1()

    def _applica_ordine_toolbar2(self, visibili):
        self._ricostruisci_toolbar1()

    def apri_personalizza_toolbar(self):
        """Finestra per personalizzare quali pulsanti mostrare nelle toolbar
        riga 1 e riga 2, e in che ordine (frecce ↑↓)."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("🎛 Personalizza toolbar")
        dlg.geometry("420x520")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text="🎛 Personalizza toolbar",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(16,4))
        ctk.CTkLabel(dlg,
                     text=T("dv_spunta_pulsanti"),
                     font=ctk.CTkFont(size=10), text_color="gray").pack(pady=(0,10))

        tabs = ctk.CTkTabview(dlg, width=380, height=370)
        tabs.pack(fill="both", expand=True, padx=16, pady=(0,10))
        tabs.add("Riga 1 — File")
        tabs.add("Riga 2 — Esporta/QSL")

        def _costruisci_tab(tab, disponibili, ordine_iniziale, disattivati_iniziali):
            scroll = ctk.CTkScrollableFrame(tab, width=340, height=310)
            scroll.pack(fill="both", expand=True)
            ordine_corrente = list(ordine_iniziale)
            attivi = set(ordine_corrente) - set(disattivati_iniziali)
            var_check = {}

            def _ridisegna():
                for w in scroll.winfo_children():
                    w.destroy()
                for tb_id in ordine_corrente:
                    if tb_id not in disponibili:
                        continue
                    text = disponibili[tb_id][0]
                    row = ctk.CTkFrame(scroll, fg_color="transparent")
                    row.pack(fill="x", pady=2)
                    v = var_check.get(tb_id, ctk.BooleanVar(value=tb_id in attivi))
                    var_check[tb_id] = v
                    ctk.CTkCheckBox(row, text=text, variable=v,
                                    font=ctk.CTkFont(size=11), width=200).pack(side="left")
                    ctk.CTkButton(row, text="↑", width=28, height=24,
                                  fg_color="#4A5568",
                                  command=lambda t=tb_id: _sposta(t,-1)).pack(side="left", padx=2)
                    ctk.CTkButton(row, text="↓", width=28, height=24,
                                  fg_color="#4A5568",
                                  command=lambda t=tb_id: _sposta(t,1)).pack(side="left", padx=2)

            def _sposta(tb_id, delta):
                i = ordine_corrente.index(tb_id)
                j = i + delta
                if 0 <= j < len(ordine_corrente):
                    ordine_corrente[i], ordine_corrente[j] = ordine_corrente[j], ordine_corrente[i]
                    _ridisegna()

            _ridisegna()
            return ordine_corrente, var_check

        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}

        ord1, chk1 = _costruisci_tab(
            tabs.tab("Riga 1 — File"), self._tb1_disponibili,
            self._ordine_toolbar1(), dati.get('toolbar1_disattivati', []))
        ord2, chk2 = _costruisci_tab(
            tabs.tab("Riga 2 — Esporta/QSL"), self._tb2_disponibili,
            self._ordine_toolbar2(), dati.get('toolbar2_disattivati', []))

        def _salva():
            attivi1 = [t for t in ord1 if chk1[t].get()]
            disatt1 = [t for t in ord1 if not chk1[t].get()]
            attivi2 = [t for t in ord2 if chk2[t].get()]
            disatt2 = [t for t in ord2 if not chk2[t].get()]

            if self.profilo_attivo:
                try:
                    profili2 = self._carica_profili()
                    if self.profilo_attivo in profili2:
                        profili2[self.profilo_attivo]['toolbar1_ordine'] = ord1
                        profili2[self.profilo_attivo]['toolbar1_disattivati'] = disatt1
                        profili2[self.profilo_attivo]['toolbar2_ordine'] = ord2
                        profili2[self.profilo_attivo]['toolbar2_disattivati'] = disatt2
                        with open(self.profili_path, 'w', encoding='utf-8') as f:
                            json.dump(profili2, f, ensure_ascii=False, indent=2)
                except Exception as ex:
                    messagebox.showerror(T("errore"), str(ex), parent=dlg)
                    return
            else:
                messagebox.showinfo("Nessun profilo attivo",
                    "La disposizione verrà applicata solo a questa sessione\n"
                    "(crea/seleziona un profilo per renderla permanente).",
                    parent=dlg)

            self._applica_ordine_toolbar1(attivi1)
            self._applica_ordine_toolbar2(attivi2)
            dlg.destroy()

        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=16, pady=(0,16))
        ctk.CTkButton(frame_btn, text="✔ Applica", command=_salva, height=34,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS).pack(side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(frame_btn, text=T("cm_annulla"), command=dlg.destroy,
                      height=34, width=90, fg_color="#718096").pack(side="left")


    def _ordine_toolbar1(self):
        """Restituisce l'ordine/visibilità dei pulsanti tb1 dal profilo."""
        default = ["apri_adif","unisci","importa_cbr","salva_adif",
                   "aggiungi_qso","filtri_qso","duplicati","deduci_country"]
        try:
            profili = self._carica_profili()
            dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
            ordine = dati.get('toolbar1_ordine')
            if ordine and isinstance(ordine, list):
                validi = [i for i in ordine if i in self._tb1_disponibili]
                mancanti = [i for i in default if i not in validi]
                return validi + mancanti
        except Exception:
            pass
        return default

    def _ordine_toolbar2(self):
        """Restituisce l'ordine/visibilità dei pulsanti tb2 dal profilo,
        o l'ordine di default se non personalizzato."""
        default = ["genera_pdf","esporta_csv","esporta_html","esporta_excel",
                   "qsl_card","grafici","qsl_designer","storico","dx_cluster"]
        try:
            profili = self._carica_profili()
            dati = profili.get(self.profilo_attivo, {}) if self.profilo_attivo else {}
            ordine = dati.get('toolbar2_ordine')
            if ordine and isinstance(ordine, list):
                # Filtra eventuali id non più validi, aggiunge i nuovi mancanti in coda
                validi = [i for i in ordine if i in self._tb2_disponibili]
                mancanti = [i for i in default if i not in validi]
                return validi + mancanti
        except Exception:
            pass
        return default


    def apri_display_radio(self):
        """Finestra compatta sempre-in-primo-piano che mostra QRG, modo e banda
        letti dalla radio in tempo reale (polling ~1s), stile Log4OM. Include un
        pulsante per loggare al volo la frequenza corrente in Aggiungi QSO."""
        rig = getattr(self, "_omnirig", None)
        if rig is None or not rig.disponibile():
            messagebox.showwarning(T("dxc_omnirig_no"),
                                   T("dxc_omnirig_assente"), parent=self)
            return
        # Se già aperta, portala in primo piano invece di duplicarla
        if getattr(self, "_display_radio_win", None) is not None:
            try:
                if self._display_radio_win.winfo_exists():
                    self._display_radio_win.lift()
                    self._display_radio_win.focus_force()
                    return
            except Exception:
                pass

        is_dark = ctk.get_appearance_mode().lower() == "dark"
        bg = "#0D0D0D" if is_dark else "#12233B"
        col_qrg = "#4ADE80"     # verde acceso per la frequenza
        col_mode = "#60A5FA"    # blu per il modo
        col_band = "#FBBF24"    # ambra per la banda
        col_dim = "#64748B"

        win = ctk.CTkToplevel(self)
        self._display_radio_win = win
        win.title(T("rdisp_titolo"))
        win.geometry("440x250")
        win.resizable(False, False)
        win.configure(fg_color=bg)
        win.attributes("-topmost", True)
        win.lift(); win.focus_force()
        win.after(200, lambda: (win.lift(), win.focus_force()))

        # Stato connessione (pallino + testo)
        top = ctk.CTkFrame(win, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(12, 0))
        lbl_stato = ctk.CTkLabel(top, text="●", font=ctk.CTkFont(size=14),
                                 text_color=col_dim)
        lbl_stato.pack(side="left")
        lbl_stato_txt = ctk.CTkLabel(top, text=T("rdisp_attesa"),
                                     font=ctk.CTkFont(size=11), text_color=col_dim)
        lbl_stato_txt.pack(side="left", padx=(4, 0))

        # Due VFO affiancati (A = principale/RX, B = secondario/TX)
        vfo_frame = ctk.CTkFrame(win, fg_color="transparent")
        vfo_frame.pack(fill="x", padx=12, pady=(8, 0))

        col_bg = "#0A1A2E" if is_dark else "#0E2038"
        def _crea_vfo(parent, etichetta, colore_freq):
            box = ctk.CTkFrame(parent, fg_color=col_bg, corner_radius=8)
            box.pack(side="left", expand=True, fill="both", padx=4)
            ctk.CTkLabel(box, text=etichetta, font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=col_dim).pack(pady=(8, 0))
            lbl_f = ctk.CTkLabel(box, text="—.————",
                                 font=ctk.CTkFont(size=24, weight="bold"),
                                 text_color=colore_freq)
            lbl_f.pack(pady=(2, 0))
            riga_mb = ctk.CTkFrame(box, fg_color="transparent")
            riga_mb.pack(pady=(0, 8))
            lbl_m = ctk.CTkLabel(riga_mb, text="—",
                                 font=ctk.CTkFont(size=14, weight="bold"),
                                 text_color=col_mode)
            lbl_m.pack(side="left", padx=(0, 8))
            lbl_bd = ctk.CTkLabel(riga_mb, text="—",
                                  font=ctk.CTkFont(size=14, weight="bold"),
                                  text_color=col_band)
            lbl_bd.pack(side="left")
            return lbl_f, lbl_m, lbl_bd

        lbl_qrg_a, lbl_mode_a, lbl_band_a = _crea_vfo(vfo_frame, T("rdisp_vfoa"), col_qrg)
        lbl_qrg_b, lbl_mode_b, lbl_band_b = _crea_vfo(vfo_frame, T("rdisp_vfob"), "#F472B6")

        # Pulsante logga al volo (usa VFO-A, quello del downlink/RX)
        btn = ctk.CTkButton(win, text=T("rdisp_logga"),
                            fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                            height=34, font=ctk.CTkFont(size=13, weight="bold"),
                            command=lambda: self._logga_da_display())
        btn.pack(fill="x", padx=16, pady=(12, 12))

        # Loop di aggiornamento
        stato = {"job": None, "attivo": True}

        def _fmt(hz):
            return f"{hz/1_000_000:.5f}" if hz else "—.————"

        def _aggiorna():
            if not stato["attivo"]:
                return
            try:
                if not win.winfo_exists():
                    stato["attivo"] = False
                    return
            except Exception:
                stato["attivo"] = False
                return
            fa = fb = None
            ma = mb = None
            try:
                fa, fb = rig.get_freq_ab()
                ma, mb = rig.get_modo_ab()
            except Exception:
                pass
            if fa or fb:
                # VFO-A
                lbl_qrg_a.configure(text=_fmt(fa))
                ba = self._banda_da_freq(fa/1_000_000) if fa else None
                lbl_band_a.configure(text=ba if ba else "—")
                lbl_mode_a.configure(text=ma if ma else "—")
                # VFO-B
                lbl_qrg_b.configure(text=_fmt(fb))
                bb = self._banda_da_freq(fb/1_000_000) if fb else None
                lbl_band_b.configure(text=bb if bb else "—")
                lbl_mode_b.configure(text=mb if mb else ("—" if fb else "—"))
                lbl_stato.configure(text_color=col_qrg)
                lbl_stato_txt.configure(text=T("rdisp_connesso"), text_color=col_qrg)
                # Per il pulsante logga usa VFO-A (downlink/RX)
                self._display_ultimo = {"hz": fa, "modo": ma,
                                        "banda": ba, "hz_b": fb, "modo_b": mb}
            else:
                lbl_stato.configure(text_color=TH.DANGER)
                lbl_stato_txt.configure(text=T("rdisp_no_radio"), text_color=TH.DANGER)
            stato["job"] = win.after(1000, _aggiorna)

        def _on_close():
            stato["attivo"] = False
            if stato["job"]:
                try: win.after_cancel(stato["job"])
                except Exception: pass
            self._display_radio_win = None
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", _on_close)

        self._display_ultimo = {}
        _aggiorna()

    def _logga_da_display(self):
        """Apre Aggiungi QSO precompilando frequenza/banda/modo dall'ultima
        lettura del display radio."""
        d = getattr(self, "_display_ultimo", None)
        if not d or not d.get("hz"):
            return
        # Apri (o riusa) la finestra Aggiungi QSO
        dlg = getattr(self, "_aggiungi_qso_dlg", None)
        if dlg is None or not (hasattr(dlg, "winfo_exists") and dlg.winfo_exists()):
            self.apri_aggiungi_qso()
            dlg = getattr(self, "_aggiungi_qso_dlg", None)

        def _riempi():
            try:
                d2 = self._display_ultimo
                mhz = f"{d2['hz']/1_000_000:.6f}".rstrip('0').rstrip('.')
                if hasattr(dlg, "_aq_freq"):
                    dlg._aq_freq.delete(0, 'end'); dlg._aq_freq.insert(0, mhz)
                if hasattr(dlg, "_aq_var_banda") and d2.get("banda"):
                    dlg._aq_var_banda.set(d2["banda"])
                modo = d2.get("modo")
                if modo and hasattr(dlg, "_aq_var_modo"):
                    if modo == "USB-D":
                        bp = modo_da_bandplan(d2['hz'])
                        modo = bp if bp and bp not in ("USB","LSB","CW","AM","FM") else "FT8"
                    elif modo in ("USB", "LSB"):
                        modo = "SSB"
                    _modi_validi = ('SSB','CW','FT8','FT4','FT2','RTTY','PSK31','MFSK',
                                    'JT65','JT9','WSPR','AM','FM','SSTV','PKT')
                    if modo in _modi_validi:
                        dlg._aq_var_modo.set(modo)
                if hasattr(dlg, "_aq_call"):
                    dlg._aq_call.focus_set()
            except Exception:
                pass
        self.after(300, _riempi)

    def apri_impostazioni_radio(self):
        """Finestra dedicata alle impostazioni OmniRig: avvio automatico e
        percorsi personalizzati di OmniRig.exe e della cartella Rigs."""
        import tkinter as _tk
        dati = {}
        try:
            dati = self._carica_profili().get(self.profilo_attivo, {})
        except Exception:
            pass

        dlg = ctk.CTkToplevel(self)
        dlg.title(T("radio_titolo"))
        dlg.geometry("560x420")
        dlg.transient(self); dlg.lift(); dlg.focus_force()
        dlg.after(200, lambda: (dlg.lift(), dlg.focus_force()))

        ctk.CTkLabel(dlg, text=T("radio_head"),
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(16,4), padx=20)
        ctk.CTkLabel(dlg, text=T("radio_sub"), font=ctk.CTkFont(size=11),
                     text_color="gray", wraplength=500,
                     justify="left").pack(pady=(0,10), padx=20)

        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20)

        # ── Sorgente radio: OmniRig o SDR Console ────────────────
        ctk.CTkLabel(frame, text=T("radio_sorgente_lbl"), anchor="w",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(fill="x", pady=(4,2))
        var_backend = ctk.StringVar(value=dati.get('radio_backend', 'omnirig'))
        fr_src = ctk.CTkFrame(frame, fg_color="transparent")
        fr_src.pack(fill="x", pady=(0,6))
        ctk.CTkRadioButton(fr_src, text="OmniRig", variable=var_backend,
                           value="omnirig", font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,16))
        ctk.CTkRadioButton(fr_src, text="SDR Console", variable=var_backend,
                           value="sdrconsole", font=ctk.CTkFont(size=11)).pack(side="left")
        # Porta COM per SDR Console
        fr_sdr = ctk.CTkFrame(frame, fg_color="transparent")
        fr_sdr.pack(fill="x", pady=(0,10))
        ctk.CTkLabel(fr_sdr, text=T("radio_sdr_porta"), anchor="w",
                     font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,6))
        e_sdrport = ctk.CTkEntry(fr_sdr, width=90, placeholder_text="COM11")
        e_sdrport.pack(side="left")
        if dati.get('sdr_porta'):
            e_sdrport.insert(0, dati['sdr_porta'])
        ctk.CTkLabel(fr_sdr, text=T("radio_sdr_nota"), anchor="w",
                     font=ctk.CTkFont(size=9), text_color="gray").pack(side="left", padx=(8,0))

        # Avvio automatico
        var_auto = ctk.BooleanVar(value=bool(dati.get('omnirig_avvio_auto', True)))
        ctk.CTkCheckBox(frame, text=T("radio_avvio_auto"), variable=var_auto,
                        font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(6,12))

        # Percorso exe
        ctk.CTkLabel(frame, text=T("radio_exe_lbl"), anchor="w",
                     font=ctk.CTkFont(size=11)).pack(fill="x")
        fr_exe = ctk.CTkFrame(frame, fg_color="transparent")
        fr_exe.pack(fill="x", pady=(2,10))
        e_exe = ctk.CTkEntry(fr_exe, placeholder_text=T("radio_exe_ph"))
        e_exe.pack(side="left", fill="x", expand=True, padx=(0,6))
        if dati.get('omnirig_exe_path'):
            e_exe.insert(0, dati['omnirig_exe_path'])

        def _sfoglia_exe():
            p = filedialog.askopenfilename(
                title=T("radio_sfoglia_exe"),
                filetypes=[("OmniRig.exe", "*.exe"), ("Tutti i file", "*.*")])
            if p:
                e_exe.delete(0, 'end'); e_exe.insert(0, p)
        ctk.CTkButton(fr_exe, text=T("radio_sfoglia"), width=90,
                      command=_sfoglia_exe).pack(side="left")

        # Percorso cartella Rigs
        ctk.CTkLabel(frame, text=T("radio_rigs_lbl"), anchor="w",
                     font=ctk.CTkFont(size=11)).pack(fill="x")
        fr_rigs = ctk.CTkFrame(frame, fg_color="transparent")
        fr_rigs.pack(fill="x", pady=(2,10))
        e_rigs = ctk.CTkEntry(fr_rigs, placeholder_text=T("radio_rigs_ph"))
        e_rigs.pack(side="left", fill="x", expand=True, padx=(0,6))
        if dati.get('omnirig_rigs_path'):
            e_rigs.insert(0, dati['omnirig_rigs_path'])

        def _sfoglia_rigs():
            p = filedialog.askdirectory(title=T("radio_sfoglia_rigs"))
            if p:
                e_rigs.delete(0, 'end'); e_rigs.insert(0, p)
        ctk.CTkButton(fr_rigs, text=T("radio_sfoglia"), width=90,
                      command=_sfoglia_rigs).pack(side="left")

        # Stato / test
        lbl_stato = ctk.CTkLabel(frame, text="", font=ctk.CTkFont(size=11),
                                 wraplength=500, justify="left")
        lbl_stato.pack(anchor="w", pady=(4,0))

        def _applica_percorsi():
            """Applica i percorsi all'oggetto OmniRig live."""
            self._omnirig.exe_path = e_exe.get().strip()
            self._omnirig.rigs_path = e_rigs.get().strip()

        def _test():
            _applica_percorsi()
            try:
                info = self._omnirig.diagnostica_ini()
                # Aggiunge la diagnostica dello split (per capire perché non risponde)
                try:
                    info += "\n\n── SPLIT ──\n" + self._omnirig.diagnostica_split()
                except Exception:
                    pass
                lbl_stato.configure(text=info, text_color=("#1A365D", "#90CDF4"))
            except Exception as ex:
                lbl_stato.configure(text=f"Errore: {ex}", text_color=TH.DANGER)

        def _avvia_ora():
            _applica_percorsi()
            ok = self._omnirig._avvia_omnirig_exe()
            lbl_stato.configure(
                text=T("radio_avviato") if ok else T("radio_non_trovato"),
                text_color="#38A169" if ok else "#E53E3E")

        def _salva():
            _applica_percorsi()
            # Applica il flag avvio auto all'istanza OmniRig reale
            self._omnirig_reale.avvio_auto = bool(var_auto.get())
            # Applica la scelta di sorgente radio (OmniRig o SDR Console)
            backend = var_backend.get()
            porta_sdr = e_sdrport.get().strip() or "COM11"
            self._imposta_backend_radio(backend, porta_sdr)
            try:
                profili = self._carica_profili()
                if self.profilo_attivo in profili:
                    profili[self.profilo_attivo]['omnirig_avvio_auto'] = var_auto.get()
                    profili[self.profilo_attivo]['omnirig_exe_path'] = e_exe.get().strip()
                    profili[self.profilo_attivo]['omnirig_rigs_path'] = e_rigs.get().strip()
                    profili[self.profilo_attivo]['radio_backend'] = backend
                    profili[self.profilo_attivo]['sdr_porta'] = porta_sdr
                    with open(self.profili_path, 'w', encoding='utf-8') as f:
                        json.dump(profili, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            dlg.destroy()

        def _monitor_split():
            """Finestra che mostra .Split, FreqA, FreqB in tempo reale, così
            l'utente vede quale valore cambia quando attiva/disattiva lo split
            dalla radio."""
            rig = self._omnirig
            if not rig.disponibile():
                lbl_stato.configure(text="OmniRig non disponibile.", text_color=TH.DANGER)
                return
            mon = ctk.CTkToplevel(dlg)
            mon.title("Monitor Split")
            mon.geometry("360x200")
            mon.attributes("-topmost", True)
            mon.after(200, lambda: (mon.lift(), mon.focus_force()))
            ctk.CTkLabel(mon, text="Attiva/disattiva lo SPLIT dalla radio\ne guarda quale valore cambia:",
                         font=ctk.CTkFont(size=12)).pack(pady=(12, 6))
            lbl_val = ctk.CTkLabel(mon, text="—",
                                   font=ctk.CTkFont(size=15, weight="bold", family="Consolas"),
                                   justify="left", text_color=TH.OK_TEXT)
            lbl_val.pack(pady=8)
            job = {"id": None, "on": True}
            def _upd():
                if not job["on"]:
                    return
                try:
                    if not mon.winfo_exists():
                        job["on"] = False; return
                except Exception:
                    job["on"] = False; return
                try:
                    sv = rig.leggi_valore_split()
                    fa, fb = rig.get_freq_ab()
                    dsplit = ""
                    if fa and fb:
                        dsplit = f"Δ = {(fb-fa)} Hz"
                    txt = (f".Split = {sv}  (hex {sv:#x})\n"
                           f"FreqA = {fa}\nFreqB = {fb}\n{dsplit}")
                    lbl_val.configure(text=txt)
                except Exception as ex:
                    lbl_val.configure(text=f"Errore: {ex}")
                job["id"] = mon.after(500, _upd)
            def _close():
                job["on"] = False
                if job["id"]:
                    try: mon.after_cancel(job["id"])
                    except Exception: pass
                mon.destroy()
            mon.protocol("WM_DELETE_WINDOW", _close)
            _upd()

        fr_test = ctk.CTkFrame(dlg, fg_color="transparent")
        fr_test.pack(fill="x", padx=20, pady=(8,4))
        ctk.CTkButton(fr_test, text=T("radio_test"), command=_test,
                      fg_color="#4A5568", width=90).pack(side="left", padx=(0,4))
        ctk.CTkButton(fr_test, text="📊 Monitor Split", command=_monitor_split,
                      fg_color="#4A5568", width=120).pack(side="left", padx=(0,4))
        ctk.CTkButton(fr_test, text=T("radio_avvia_ora"), command=_avvia_ora,
                      fg_color=TH.PRIMARY, width=110).pack(side="left")

        fr_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        fr_btn.pack(fill="x", padx=20, pady=(4,16))
        ctk.CTkButton(fr_btn, text=T("radio_salva"), command=_salva,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                      height=38).pack(side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(fr_btn, text=T("radio_annulla"), command=dlg.destroy,
                      fg_color="#718096", height=38).pack(side="left", expand=True, fill="x")

    def apri_preferenze(self):
        """Finestra Preferenze — personalizzazione comportamento all'avvio,
        salvata nel profilo operatore attivo."""
        if not self.profilo_attivo:
            messagebox.showwarning(T("attenzione"),
                "Seleziona o crea un profilo operatore prima di impostare le preferenze.")
            return

        profili = self._carica_profili()
        dati = profili.get(self.profilo_attivo, {})

        dlg = ctk.CTkToplevel(self)
        dlg.title(T("pref_titolo"))
        dlg.geometry("460x500")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text=T("pref_head", prof=self.profilo_attivo),
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(16,4), padx=20)
        ctk.CTkLabel(dlg, text=T("pref_sub"),
                     font=ctk.CTkFont(size=10), text_color="gray").pack(pady=(0,12))

        frame = ctk.CTkFrame(dlg, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=24)

        var_ultimo_log = ctk.BooleanVar(value=bool(dati.get('pref_apri_ultimo_log', False)))
        ctk.CTkCheckBox(frame, text=T("pref_apri_ultimo"),
                        variable=var_ultimo_log,
                        font=ctk.CTkFont(size=11)).pack(anchor="w", pady=6)

        ultimo_path = dati.get('ultimo_log_path', '')
        ctk.CTkLabel(frame,
                     text=f"Ultimo log: {os.path.basename(ultimo_path) if ultimo_path else '(nessuno ancora)'}",
                     font=ctk.CTkFont(size=9), text_color=TH.LINK).pack(anchor="w", padx=(24,0), pady=(0,12))

        var_controllo = ctk.BooleanVar(value=bool(
            self.var_controllo_post_apertura.get() if hasattr(self,'var_controllo_post_apertura') else False))
        ctk.CTkCheckBox(frame, text=T("pref_controlla"),
                        variable=var_controllo,
                        font=ctk.CTkFont(size=11)).pack(anchor="w", pady=6)

        var_default_profilo = ctk.BooleanVar(value=bool(dati.get('default', False)))
        ctk.CTkCheckBox(frame, text=T("pref_predefinito"),
                        variable=var_default_profilo,
                        font=ctk.CTkFont(size=11)).pack(anchor="w", pady=6)

        var_ricorda_tema = ctk.BooleanVar(value=bool(dati.get('pref_ricorda_tema', True)))
        ctk.CTkCheckBox(frame, text=T("pref_ricorda_tema"),
                        variable=var_ricorda_tema,
                        font=ctk.CTkFont(size=11)).pack(anchor="w", pady=6)

        var_colora_righe = ctk.BooleanVar(value=self.var_colora_righe.get())
        ctk.CTkCheckBox(frame, text=T("pref_colora_righe"),
                        variable=var_colora_righe,
                        font=ctk.CTkFont(size=11)).pack(anchor="w", pady=6)

        # ── Lingua e Tema ──
        ctk.CTkLabel(frame, text="──────────────────────────────",
                     font=ctk.CTkFont(size=9), text_color="gray").pack(anchor="w", pady=(8,2))

        frame_lingua = ctk.CTkFrame(frame, fg_color="transparent")
        frame_lingua.pack(anchor="w", pady=4, fill="x")
        ctk.CTkLabel(frame_lingua, text=T("pref_lingua"), width=80,
                     font=ctk.CTkFont(size=11)).pack(side="left")
        var_lingua = ctk.StringVar(value=dati.get('lingua', 'IT'))
        ctk.CTkRadioButton(frame_lingua, text=T("menu_italiano"), variable=var_lingua, value="IT",
                           font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,12))
        ctk.CTkRadioButton(frame_lingua, text=T("menu_english"), variable=var_lingua, value="EN",
                           font=ctk.CTkFont(size=11)).pack(side="left")

        frame_tema = ctk.CTkFrame(frame, fg_color="transparent")
        frame_tema.pack(anchor="w", pady=4, fill="x")
        ctk.CTkLabel(frame_tema, text=T("pref_tema"), width=80,
                     font=ctk.CTkFont(size=11)).pack(side="left")
        tema_corrente = ctk.get_appearance_mode()
        var_tema = ctk.StringVar(value=tema_corrente)
        ctk.CTkRadioButton(frame_tema, text=T("pref_dark"), variable=var_tema, value="Dark",
                           font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,12))
        ctk.CTkRadioButton(frame_tema, text=T("pref_light"), variable=var_tema, value="Light",
                           font=ctk.CTkFont(size=11)).pack(side="left")

        def _salva():
            profili2 = self._carica_profili()
            if self.profilo_attivo not in profili2:
                dlg.destroy(); return
            profili2[self.profilo_attivo]['pref_apri_ultimo_log'] = var_ultimo_log.get()
            profili2[self.profilo_attivo]['pref_ricorda_tema'] = var_ricorda_tema.get()
            profili2[self.profilo_attivo]['pref_colora_righe'] = var_colora_righe.get()
            profili2[self.profilo_attivo]['lingua'] = var_lingua.get()
            profili2[self.profilo_attivo]['tema'] = var_tema.get()
            # Applica subito
            self.var_colora_righe.set(var_colora_righe.get())
            self._aggiorna_tree()
            if var_lingua.get() != self.var_lingua.get():
                self._set_lingua(var_lingua.get())
            if var_tema.get() != ctk.get_appearance_mode():
                ctk.set_appearance_mode(var_tema.get())
                self._aggiorna_colori_tree()
            if var_default_profilo.get():
                for n in profili2:
                    profili2[n]['default'] = (n == self.profilo_attivo)
            try:
                with open(self.profili_path, 'w', encoding='utf-8') as f:
                    json.dump(profili2, f, ensure_ascii=False, indent=2)
            except Exception as ex:
                messagebox.showerror(T("errore"), str(ex), parent=dlg)
                return
            if hasattr(self, 'var_controllo_post_apertura'):
                self.var_controllo_post_apertura.set(var_controllo.get())
            dlg.destroy()
            messagebox.showinfo("Preferenze salvate",
                "Le preferenze sono state salvate per questo profilo.", parent=self)

        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=24, pady=(16,16))
        ctk.CTkButton(frame_btn, text=T("pref_salva"), command=_salva, height=34,
                      fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS).pack(
                      side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(frame_btn, text=T("cm_annulla"), command=dlg.destroy,
                      height=34, width=100, fg_color="#718096").pack(side="left")

    def apri_gestione_profili(self):
        """Finestra di gestione profili multipli."""
        profili = self._carica_profili()

        dlg = ctk.CTkToplevel(self)
        dlg.title(T("profili"))
        dlg.geometry("520x480")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text=T("profili"),
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(14,4), padx=20)
        ctk.CTkLabel(dlg, text=T("profilo_attivo") + f" {self.profilo_attivo or '—'}",
                     font=ctk.CTkFont(size=10), text_color=TH.LINK).pack(padx=20, pady=(0,8))

        scroll = ctk.CTkScrollableFrame(dlg, height=260)
        scroll.pack(fill="both", expand=True, padx=15, pady=4)

        def aggiorna_lista():
            for w in scroll.winfo_children():
                w.destroy()
            profili_now = self._carica_profili()
            if not profili_now:
                ctk.CTkLabel(scroll, text=T("pref_no_profili"),
                             text_color="gray").pack(pady=20)
                return
            for nome, dati in profili_now.items():
                row = ctk.CTkFrame(scroll, border_width=1,
                                   border_color=("#4299E1" if nome==self.profilo_attivo else "#2D3748"))
                row.pack(fill="x", pady=3)
                lbl = ctk.CTkLabel(row,
                    text=f"  {dati.get('callsign','')}  |  {nome}  |  {dati.get('locator','')}  |  {dati.get('qth','')}",
                    anchor="w", font=ctk.CTkFont(size=11))
                lbl.pack(side="left", expand=True, padx=8, pady=6)
                ctk.CTkButton(row, text=T("cm_usa"), width=50, height=26,
                              fg_color=TH.SUCCESS_H,
                              command=lambda n=nome, d=dati: [self._imposta_default(n), self._applica_profilo({**d,'nome':n}), aggiorna_lista(), None]
                              ).pack(side="right", padx=2, pady=4)
                ctk.CTkButton(row, text=T("cm_mod"), width=50, height=26,
                              fg_color=TH.PRIMARY,
                              command=lambda n=nome, d=dati: modifica(n, d)
                              ).pack(side="right", padx=2, pady=4)
                ctk.CTkButton(row, text=T("cm_del"), width=50, height=26,
                              fg_color=TH.WARNING_H,
                              command=lambda n=nome: elimina(n)
                              ).pack(side="right", padx=2, pady=4)

        def nuovo():
            dati = self._dialog_profilo(dlg, T("profilo_nuovo"))
            if dati:
                p = self._carica_profili()
                p[dati['nome']] = dati
                self._salva_profili(p)
                aggiorna_lista()

        def modifica(nome, dati):
            nuovo_dati = self._dialog_profilo(dlg, T("profilo_modifica"), dati)
            if nuovo_dati:
                p = self._carica_profili()
                if nuovo_dati['nome'] != nome:
                    del p[nome]
                p[nuovo_dati['nome']] = nuovo_dati
                self._salva_profili(p)
                aggiorna_lista()

        def elimina(nome):
            if messagebox.askyesno("Elimina", f"Eliminare il profilo '{nome}'?"):
                p = self._carica_profili()
                if nome in p: del p[nome]
                self._salva_profili(p)
                if self.profilo_attivo == nome:
                    self.profilo_attivo = None
                    self.btn_profili.configure(text='—')
                aggiorna_lista()

        aggiorna_lista()

        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=15, pady=10)
        ctk.CTkButton(frame_btn, text="+ " + T("profilo_nuovo"), command=nuovo,
                      fg_color=TH.SUCCESS_H, height=34).pack(side="left", expand=True, fill="x", padx=(0, 4))
        ctk.CTkButton(frame_btn, text="📥 " + T("profilo_importa"),
                      command=lambda: (self._importa_profili(dlg), aggiorna_lista()),
                      fg_color=TH.PRIMARY, height=34).pack(side="left", expand=True, fill="x", padx=4)
        ctk.CTkButton(frame_btn, text="📤 " + T("profilo_esporta"),
                      command=lambda: self._esporta_profili(dlg),
                      fg_color="#4A5568", height=34).pack(side="left", expand=True, fill="x", padx=(4, 0))

    def _profilo_e_sensibile(self, key):
        """True se la chiave di profilo contiene una credenziale (password/API key)."""
        k = str(key).lower()
        return ("password" in k) or ("api_key" in k) or ("apikey" in k) or k.endswith("_key")

    def _esporta_profili(self, parent=None):
        """Esporta tutti i profili in un file JSON (backup e travaso tra PC).
        Chiede se includere le credenziali (password/API key)."""
        parent = parent or self
        profili = self._carica_profili()
        if not profili:
            messagebox.showinfo("Esporta profili",
                                "Non c'è nessun profilo da esportare.", parent=parent)
            return
        incl = messagebox.askyesnocancel(
            "Esporta profili",
            "Includere password e API key nel file?\n\n"
            "• Sì  → backup completo: riutilizzabile su un altro PC senza reinserire le credenziali.\n"
            "• No  → solo i dati operatore, senza password.\n\n"
            "⚠ Con 'Sì' il file contiene le password in chiaro: conservalo in un posto sicuro.",
            parent=parent)
        if incl is None:
            return
        import datetime as _dt
        default = f"adif_fzr_profili_{_dt.datetime.now():%Y%m%d}.json"
        path = filedialog.asksaveasfilename(
            title="Esporta profili", defaultextension=".json", initialfile=default,
            filetypes=[("File JSON", "*.json"), ("Tutti i file", "*.*")], parent=parent)
        if not path:
            return
        dati = {}
        for nome, prof in profili.items():
            if incl:
                dati[nome] = dict(prof)
            else:
                dati[nome] = {k: v for k, v in prof.items()
                              if not self._profilo_e_sensibile(k)}
        pacchetto = {
            "_adif_fzr_export": True,
            "tipo": "profili",
            "versione": VERSIONE,
            "data": _dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "con_credenziali": bool(incl),
            "profili": dati,
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(pacchetto, f, ensure_ascii=False, indent=2)
        except Exception as ex:
            messagebox.showerror("Errore", f"Esportazione fallita:\n{ex}", parent=parent)
            return
        messagebox.showinfo(
            "Esporta profili",
            f"Esportati {len(dati)} profili in:\n{path}"
            + ("" if incl else "\n\n(senza password / API key)"),
            parent=parent)

    def _importa_profili(self, parent=None):
        """Importa profili da un file JSON esportato (o da un file
        .adif_fzr_profili.json grezzo). Gestisce i doppioni di nome.
        Ritorna il numero di profili importati."""
        parent = parent or self
        path = filedialog.askopenfilename(
            title="Importa profili",
            filetypes=[("File JSON", "*.json"), ("Tutti i file", "*.*")], parent=parent)
        if not path:
            return 0
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
        except Exception as ex:
            messagebox.showerror("Errore", f"File non leggibile:\n{ex}", parent=parent)
            return 0
        # Accetta sia il formato esportato ({..., "profili": {...}}) sia un
        # file profili grezzo ({nome: {...}, ...}).
        if isinstance(raw, dict) and isinstance(raw.get("profili"), dict):
            in_profili = raw["profili"]
        elif isinstance(raw, dict) and raw and all(isinstance(v, dict) for v in raw.values()):
            in_profili = raw
        else:
            messagebox.showerror(
                "Errore", "Il file non contiene profili in un formato riconosciuto.",
                parent=parent)
            return 0
        if not in_profili:
            messagebox.showinfo("Importa profili", "Il file non contiene profili.",
                                parent=parent)
            return 0
        esistenti = self._carica_profili()
        conflitti = [n for n in in_profili if n in esistenti]
        sovrascrivi = True
        if conflitti:
            r = messagebox.askyesnocancel(
                "Importa profili",
                f"{len(conflitti)} profili hanno un nome già presente "
                f"({', '.join(conflitti[:5])}{'…' if len(conflitti) > 5 else ''}).\n\n"
                "• Sì  → sovrascrivi gli esistenti con la versione importata\n"
                "• No  → salta i doppioni, importa solo i profili nuovi\n"
                "• Annulla → non importare nulla",
                parent=parent)
            if r is None:
                return 0
            sovrascrivi = bool(r)
        n_imp = 0
        for nome, prof in in_profili.items():
            if nome in esistenti and not sovrascrivi:
                continue
            prof = dict(prof)
            prof.setdefault("nome", nome)
            esistenti[nome] = prof
            n_imp += 1
        self._salva_profili(esistenti)
        messagebox.showinfo("Importa profili", f"Importati {n_imp} profili.", parent=parent)
        return n_imp

    def _chiudi_app(self):
        """Chiusura pulita: chiede di salvare le modifiche non salvate, poi
        chiude OmniRig se è stato avviato da noi (per liberare la porta COM)."""
        # 1) Modifiche non salvate → chiedi cosa fare
        if getattr(self, '_log_modificato', False) and self.qsos_caricati:
            risp = messagebox.askyesnocancel(
                T("chiudi_titolo"), T("chiudi_msg"), parent=self)
            if risp is None:
                return          # Annulla: non chiudere
            if risp is True:    # Sì: salva prima di uscire
                try:
                    self.salva_adif()
                    # Se dopo il salvataggio è ancora modificato (salvataggio
                    # annullato dall'utente), non chiudere.
                    if getattr(self, '_log_modificato', False):
                        return
                except Exception:
                    if not messagebox.askyesno(T("chiudi_titolo"),
                                               T("chiudi_salva_fallito"), parent=self):
                        return
            # risp False = No: esci senza salvare

        # 2) Chiudi OmniRig se l'abbiamo avviato noi (libera la porta COM)
        try:
            rig = getattr(self, '_omnirig', None)
            if rig is not None and getattr(rig, '_avviato_da_noi', False):
                rig.chiudi_omnirig()
        except Exception:
            pass

        # Ferma il polling della barra radio
        try:
            if getattr(self, '_radiobar_job', None):
                self.after_cancel(self._radiobar_job)
        except Exception:
            pass

        # 3) Chiudi l'app
        try:
            self.destroy()
        except Exception:
            import sys
            sys.exit(0)

    def _controlla_primo_avvio(self):
        """Al secondo avvio: applica automaticamente il profilo attivo.
        Protetto: nessun errore qui deve impedire l'avvio del programma."""
        try:
            self._applica_profilo_avvio()
        except Exception:
            # Qualsiasi problema (profilo corrotto, path morto…) non deve bloccare
            pass

    def _applica_profilo_avvio(self):
        if os.path.exists(self.profili_path):
            profili = self._carica_profili()
            if profili:
                nome = next((n for n,d in profili.items() if d.get('default')), None)
                if not nome:
                    nome = next(iter(profili.keys()))
                dati = profili[nome]
                self.profilo_attivo = nome
                # Applica le impostazioni OmniRig del profilo reale (qui il
                # profilo è noto; nell'__init__ profilo_attivo era ancora None).
                try:
                    self._omnirig_reale.avvio_auto = bool(dati.get('omnirig_avvio_auto', True))
                    self._omnirig_reale.exe_path = dati.get('omnirig_exe_path', '') or ''
                    self._omnirig_reale.rigs_path = dati.get('omnirig_rigs_path', '') or ''
                    # Applica la sorgente radio salvata (OmniRig o SDR Console)
                    self._imposta_backend_radio(
                        dati.get('radio_backend', 'omnirig'),
                        dati.get('sdr_porta', 'COM11'))
                    # Avvio automatico OmniRig: solo se il backend è OmniRig.
                    if (self._radio_backend == "omnirig"
                            and self._omnirig_reale.avvio_auto):
                        import threading
                        threading.Thread(target=self._omnirig_reale._avvia_omnirig_exe,
                                         daemon=True).start()
                except Exception:
                    pass
                self.btn_profili.configure(text=nome)
                self.entry_owner.delete(0, 'end')
                self.entry_owner.insert(0, dati.get('callsign',''))
                self.entry_details.delete(0, 'end')
                self.entry_details.insert(0, dati.get('locator',''))

                # Lingua
                if dati.get('lingua'):
                    self._set_lingua(dati['lingua'])

                # Tema (se preferenza abilitata)
                if dati.get('pref_ricorda_tema', True) and dati.get('tema'):
                    if dati['tema'] != ctk.get_appearance_mode():
                        ctk.set_appearance_mode(dati['tema'])
                        self.after(100, self._aggiorna_colori_tree)

                # Colorazione righe
                if hasattr(self, 'var_colora_righe'):
                    self.var_colora_righe.set(bool(dati.get('pref_colora_righe', True)))

                # Toolbar personalizzata
                if hasattr(self, '_ricostruisci_toolbar1'):
                    self._ricostruisci_toolbar1()
                if hasattr(self, '_ricostruisci_toolbar2'):
                    self._ricostruisci_toolbar2()

                # Riapri ultimo log (protetto: chiavetta rimossa, path di rete morto…)
                if dati.get('pref_apri_ultimo_log'):
                    ultimo = dati.get('ultimo_log_path', '')
                    def _try_apri_ultimo(p=ultimo):
                        try:
                            if p and os.path.exists(p):
                                self._carica_adif_da_path(p)
                            elif p:
                                # File non più raggiungibile — avvisa senza bloccare
                                self.lbl_status.configure(
                                    text=f"⚠ Ultimo log non trovato: {os.path.basename(p)}",
                                    text_color=TH.WARN_TEXT)
                        except Exception:
                            # Non deve mai impedire l'avvio del programma
                            try:
                                self.lbl_status.configure(
                                    text="⚠ Impossibile aprire l'ultimo log",
                                    text_color=TH.WARN_TEXT)
                            except Exception:
                                pass
                    if ultimo:
                        self.after(300, _try_apri_ultimo)
            return
        # Migrazione profilo vecchio — apre wizard precompilato
        old_call = ''
        old_loc  = ''
        if os.path.exists(self.profilo_path):
            try:
                with open(self.profilo_path, 'r', encoding='utf-8') as f:
                    old = json.load(f)
                old_call = old.get('callsign','')
                old_loc  = old.get('note_op','')
            except Exception:
                pass
        self.after(800, lambda: self._wizard_primo_avvio(old_call, old_loc))

    def _imposta_default(self, nome):
        """Marca un profilo come default (selezionato al prossimo avvio)."""
        profili = self._carica_profili()
        for n in profili:
            profili[n]['default'] = (n == nome)
        self._salva_profili(profili)

    def _wizard_primo_avvio(self, prefill_call="", prefill_loc=""):
        """Popup wizard al primo avvio."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(f"ADIF FZR {VERSIONE} — Benvenuto")
        dlg.geometry("420x400")
        dlg.resizable(False, False)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()
        ctk.CTkLabel(dlg, text=T("dv_benvenuto"),
                     font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(18,4), padx=20)
        ctk.CTkLabel(dlg, text=T("dv_crea_primo"),
                     font=ctk.CTkFont(size=11), text_color="gray").pack(padx=20, pady=(0,12))

        CAMPI = [
            ("nome",     "Nome profilo",    "es. IW1FZR Home"),
            ("callsign", "Callsign",        "es. IW1FZR"),
            ("locator",  "Locator",         "es. JN45bj"),
            ("nome_op",  "Nome operatore",  "es. Luca"),
            ("qth",      "QTH",             "es. Cavaglià (BI)"),
            ("cq_zone",  "CQ Zone",         "es. 15"),
            ("itu_zone", "ITU Zone",        "es. 28"),
        ]
        entries = {}
        for key, lbl, ph in CAMPI:
            row = ctk.CTkFrame(dlg, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=3)
            ctk.CTkLabel(row, text=lbl, width=120, anchor="e",
                         font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,8))
            e = ctk.CTkEntry(row, width=200, placeholder_text=ph)
            if key == "callsign" and prefill_call:
                e.insert(0, prefill_call)
            elif key == "locator" and prefill_loc:
                e.insert(0, prefill_loc)
            e.pack(side="left")
            entries[key] = e

        def crea():
            nome = entries["nome"].get().strip()
            call = entries["callsign"].get().strip().upper()
            if not nome or not call:
                messagebox.showwarning("Attenzione", "Nome profilo e Callsign sono obbligatori.")
                return
            dati = {k: entries[k].get().strip() for k in entries}
            dati['callsign'] = dati['callsign'].upper()
            dati['default'] = True
            profili = {nome: dati}
            self._salva_profili(profili)
            self._applica_profilo({**dati, 'nome': nome})
            dlg.destroy()

        frame_wbtn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_wbtn.pack(padx=20, pady=14, fill="x")
        ctk.CTkButton(frame_wbtn, text=T("dv_crea_inizia"), command=crea,
                      fg_color=TH.SUCCESS_H, height=38,
                      font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", expand=True, padx=(0,6), fill="x")
        ctk.CTkButton(frame_wbtn, text=T("dv_salta"), command=dlg.destroy,
                      fg_color="#718096", height=38).pack(side="left", width=80)

    def salva_profilo(self):
        profilo = {
            'callsign': self.entry_owner.get().strip(),
            'lingua': self.var_lingua.get(),
            'note_op': self.entry_details.get().strip(),
            'tema': self.var_tema.get(),
            'colori_pdf': self.colori_pdf,
            'colori_html': self.colori_html,
            'ordine_campi_pdf': self.ordine_campi_pdf,
            'width_pdf': self.width_pdf,
            'titolo_pdf_custom': self.titolo_pdf_custom,
            'font_size_pdf': self.font_size_pdf,
            'campi': {tag: var.get() for tag, var in self.checkboxes.items()},
            'dxcc_page': self.var_dxcc_page.get(),
            'formato_pdf': self.var_formato_pdf.get(),
            'window_state': self.state(),
        }
        try:
            with open(self.profilo_path, 'w', encoding='utf-8') as f:
                json.dump(profilo, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("Profilo", T("ok_profilo_new"))
        except Exception as ex:
            messagebox.showerror("Errore", f"{T('err_profilo')}{ex}")

    def carica_profilo(self):
        if not os.path.exists(self.profilo_path):
            return
        try:
            with open(self.profilo_path, 'r', encoding='utf-8') as f:
                p = json.load(f)
            # Se i profili multipli esistono, non sovrascrivere callsign/locator
            # (verranno impostati da _controlla_primo_avvio)
            if not os.path.exists(self.profili_path):
                if p.get('callsign'):
                    self.entry_owner.insert(0, p['callsign'])
                if p.get('note_op'):
                    self.entry_details.insert(0, p['note_op'])
            if p.get('tema'):
                self.var_tema.set(p['tema'])
                ctk.set_appearance_mode(p['tema'])
            if p.get('colori_pdf'):
                self.colori_pdf.update(p['colori_pdf'])
            if p.get('colori_html'):
                self.colori_html.update(p['colori_html'])
            if p.get('ordine_campi_pdf'):
                self.ordine_campi_pdf = p['ordine_campi_pdf']
            if p.get('width_pdf'):
                self.width_pdf = p['width_pdf']
            if 'titolo_pdf_custom' in p:
                self.titolo_pdf_custom = p['titolo_pdf_custom']
            if 'font_size_pdf' in p:
                self.font_size_pdf = int(p['font_size_pdf'])
            if p.get('campi'):
                for tag, val in p['campi'].items():
                    if tag in self.checkboxes:
                        self.checkboxes[tag].set(val)
            if 'dxcc_page' in p:
                self.var_dxcc_page.set(p['dxcc_page'])
            if p.get('formato_pdf'):
                self.var_formato_pdf.set(p['formato_pdf'])
            if p.get('lingua'):
                self.var_lingua.set(p['lingua'])
                imposta_lingua(p['lingua'])
                self._aggiorna_lingua()
            # Ripristina stato finestra — sempre zoomed se era zoomed
            ws = p.get('window_state', 'zoomed')
            if ws == 'zoomed':
                self.after(10, lambda: self.state('zoomed'))
        except Exception:
            pass  # profilo corrotto, ignora silenziosamente

    def apri_unisci(self):
        dlg = UnisciDialog(self)
        self.wait_window(dlg)
        if dlg.risultato_qsos is not None:
            self.qsos_caricati = dlg.risultato_qsos
            self.qsos_filtrati = list(self.qsos_caricati)
            nome = dlg.risultato_nome or "log_unito"
            self.filepath = nome + ".adif"  # path virtuale per i nomi default export
            self.lbl_status.configure(
                text=T("log_unito_status", nome=nome, count=len(self.qsos_caricati)),
                text_color="#9F7AEA"
            )
            self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")

    # ── Caricamento file ──────────────────────
    def sfoglia_file(self):
        """Apre direttamente la finestra di selezione file di sistema. Se la
        voce di menu File 'Controlla campi principali dopo l'apertura' è
        spuntata, dopo il caricamento esegue il controllo automatico dei
        campi principali (FREQ e altri campi assenti dallo schema)."""
        path = filedialog.askopenfilename(
            title=T("dv_sel_log_adif"),
            filetypes=[("ADIF files", "*.adi *.adif"), ("All files", "*.*")]
        )
        if not path:
            return
        self._carica_adif_da_path(path)

    def _carica_adif_da_path(self, path):
        """Carica effettivamente il file ADIF indicato e, se il flag è
        attivo, esegue il controllo automatico dei campi principali."""
        self.filepath = path
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                testo = f.read()
            testo = self._fix_adif(testo)
            qsos, _ = adif_io.read_from_string(testo)
            self.qsos_caricati = sorted(qsos, key=lambda x: (x.get('qso_date', ''), x.get('time_on', '')))
            self.qsos_filtrati = list(self.qsos_caricati)
            self._log_modificato = False   # file appena aperto: nessuna modifica
            nome = os.path.basename(path)
            self.lbl_status.configure(text=T("caricato_status", nome=nome, count=len(self.qsos_caricati)), text_color=TH.PRIMARY)
            self.lbl_filtri.configure(text=T("nessun_filtro"), text_color="gray")
            self._aggiungi_storico(path)
            self._aggiorna_tree()
            self._salva_ultimo_log_path(path)
            if self.var_controllo_post_apertura.get():
                self._controlla_campi_post_apertura()
        except Exception as ex:
            messagebox.showerror("Errore", f"{T('errore_caricamento')}{ex}")

    def _salva_ultimo_log_path(self, path):
        """Salva il percorso dell'ultimo log aperto nel profilo attivo
        (usato dalla preferenza 'Apri ultimo log all'avvio')."""
        try:
            if not self.profilo_attivo:
                return
            profili = self._carica_profili()
            if self.profilo_attivo in profili:
                profili[self.profilo_attivo]['ultimo_log_path'] = path
                with open(self.profili_path, 'w', encoding='utf-8') as f:
                    json.dump(profili, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _controlla_campi_post_apertura(self):
        """Controllo automatico dopo l'apertura di un ADIF: omogenizza i
        campi (calcolando la FREQ di default dalla banda dove possibile) e,
        se restano campi dello schema standard completamente assenti dal
        log, mostra il dialogo per impostarne i valori di default — la
        stessa logica della normalizzazione, ma senza toccare uppercase di
        CALL/BAND/MODE né correggere bande satellite (quelle restano
        riservate al pulsante "Normalizza" esplicito)."""
        if not self.qsos_caricati:
            return
        n_omogenea, tutti_campi, campi_critici_assenti = self._normalizza_fase1_omogenizza(self.qsos_caricati)
        self._aggiorna_tree()  # mostra subito le FREQ calcolate dalla Fase 1
        def _dopo():
            self._aggiorna_tree()
        self._normalizza_fase2_dialogo(self.qsos_caricati, tutti_campi, n_omogenea,
                                        parent=self, on_done=_dopo, silenzioso=True,
                                        campi_critici_assenti=campi_critici_assenti)

    def _pulisci_qsos(self, qsos):
        """Rimuove i campi header che adif_io inserisce nel primo QSO."""
        CAMPI_HEADER = {'adif_ver', 'programid', 'programversion',
                        'created_timestamp', 'userdef', 'appdef'}
        return [
            {k.lower(): v for k, v in qso.items() if k.lower() not in CAMPI_HEADER}
            for qso in qsos
        ]

    def _fix_adif(self, testo):
        up = testo.upper()
        if "<EOH" in up and "<EOH>" not in up:
            testo = testo.replace("<EOH", "<EOH>").replace("<eoh", "<EOH>")
        if "<EOR" in up and "<EOR>" not in up:
            testo = testo.replace("<EOR", "<EOR>").replace("<eor", "<EOR>")
        if "<EOH>" not in testo.upper():
            testo = "<EOH>\n" + testo
        return testo

    def _leggi_adif_sicuro(self, testo):
        """Legge ADIF gestendo duplicati e altri errori del file sorgente."""
        import re
        CAMPI_HEADER = {'adif_ver','programid','programversion',
                        'created_timestamp','userdef','appdef'}
        try:
            qsos, _ = adif_io.read_from_string(testo)
            return [{k: v for k, v in q.items()
                     if k.lower() not in CAMPI_HEADER} for q in qsos]
        except Exception:
            pass

        # Fallback: parser manuale tollerante ai duplicati
        qsos = []
        eoh_pos = testo.upper().find('<EOH>')
        body = testo[eoh_pos+5:] if eoh_pos >= 0 else testo
        record_pat = re.compile(r'<([^:>]+)(?::(\d+)(?::[^>]*)?)?>([^<]*)', re.IGNORECASE)
        records = re.split(r'<EOR>', body, flags=re.IGNORECASE)
        for rec in records:
            rec = rec.strip()
            if not rec:
                continue
            qso = {}
            for m in record_pat.finditer(rec):
                tag = m.group(1).upper()
                length = int(m.group(2)) if m.group(2) else None
                value = m.group(3)
                if length is not None:
                    value = value[:length]
                value = value.strip()
                if tag.lower() not in CAMPI_HEADER and value:
                    qso[tag.lower()] = value
            if qso.get('call') or qso.get('CALL'):
                # ── Normalizzazione campi LoTW da download LoTW / HRD ──
                # LoTW e HRD esportano QSL ricevuta via LoTW come QSL_RCVD=Y
                # + QSLRDATE, insieme a campi APP_LOTW_* che identificano la fonte.
                # Mappiamo ai campi ADIF standard LOTW_QSL_RCVD / LOTW_QSLRDATE.
                _ha_app_lotw = any(k.startswith('app_lotw') for k in qso)
                if _ha_app_lotw:
                    if not qso.get('lotw_qsl_rcvd') and qso.get('qsl_rcvd', '').upper() == 'Y':
                        qso['lotw_qsl_rcvd'] = 'Y'
                    if not qso.get('lotw_qslrdate') and qso.get('qslrdate', '').strip():
                        qso['lotw_qslrdate'] = qso['qslrdate'].strip()
                # HRD esporta LOTW_QSL_RCVD=V (Verified) invece dello standard Y
                if qso.get('lotw_qsl_rcvd', '').upper() == 'V':
                    qso['lotw_qsl_rcvd'] = 'Y'
                qsos.append(qso)
        return qsos

    # ── QSO da usare (con filtri) ─────────────
    def _qsos_attivi(self):
        return self.qsos_filtrati if self.qsos_filtrati else self.qsos_caricati

    # ── Formattazione valore cella ────────────
    def _country_con_stato(self, qso):
        """Ritorna il country del QSO, con lo stato USA tra parentesi se il
        country è USA e il campo state è compilato (es. 'USA (CA)').
        Centralizza la logica usata in griglia, PDF e HTML."""
        country = str(qso.get('country', '')).strip()
        v_up = country.upper()
        if v_up in ("USA", "UNITED STATES", "UNITED STATES OF AMERICA"):
            stato = str(qso.get('state', '')).strip().upper()
            return f"USA ({stato})" if stato else "USA"
        return country

    def _formatta_valore(self, tag, qso, per_pdf=True):
        valore = qso.get(tag.lower(), '')
        sat_name = str(qso.get('sat_name', '')).strip().upper()

        # Quando per_pdf=True i valori vengono inseriti dentro Paragraph ReportLab
        # che usa un mini-parser XML. I caratteri & < > nei valori grezzi (es. campo
        # NAME="W&W", COMMENT="<QRP>") spaccano il parser con "unclosed tags".
        # _xe() esegue l'escape SOLO quando il valore viene usato come testo libero
        # (non dentro tag ReportLab che costruiamo noi stessi).
        def _xe(s):
            if not per_pdf:
                return str(s)
            return (str(s)
                    .replace('&', '&amp;')
                    .replace('<', '&lt;')
                    .replace('>', '&gt;'))

        if tag == 'QSO_DATE' and len(valore) == 8:
            return f"{valore[6:8]}/{valore[4:6]}/{valore[0:4]}"
        elif tag == 'TIME_ON' and len(valore) >= 4:
            return f"{valore[0:2]}:{valore[2:4]}"
        elif tag == 'NAME':
            return _xe(str(valore)[:15])
        elif tag in ['CALL', 'MODE', 'GRIDSQUARE', 'BAND', 'FREQ']:
            v = _xe(str(valore).upper())
            if tag in ['BAND', 'FREQ'] and sat_name:
                sat_safe = _xe(sat_name)
                return f"{v} ({sat_safe})" if not per_pdf else f"{v}<br/><font color='#2B6CB0'><b>({sat_safe})</b></font>"
            if tag == 'MODE':
                submode = _xe(str(qso.get('submode', '')).strip().upper())
                v_raw = str(valore).upper()  # versione non escaped per confronto
                sub_raw = str(qso.get('submode', '')).strip().upper()
                if sub_raw and sub_raw != v_raw:
                    return f"{v} ({submode})" if not per_pdf else f"{v} <font color='#2B6CB0'><b>({submode})</b></font>"
            return v
        elif tag == 'COUNTRY':
            v_raw = str(valore).upper()
            if v_raw in ["USA", "UNITED STATES", "UNITED STATES OF AMERICA"]:
                stato = _xe(str(qso.get('state', '')).strip().upper())
                return f"USA ({stato})" if stato else "USA"
            return _xe(v_raw)
        elif tag == 'LOTW_QSL_RCVD':
            rcvd = str(qso.get('lotw_qsl_rcvd', '')).upper().strip()
            date = str(qso.get('lotw_qslrdate', '')).strip()
            confermato = rcvd == 'Y' or (not rcvd and date and date != '00000000')
            if per_pdf:
                return "<font color='green'><b>V</b></font>" if confermato else "<font color='gray'>-</font>"
            return "Y" if confermato else "-"
        elif tag == 'EQSL_QSL_RCVD':
            rcvd = str(qso.get('eqsl_qsl_rcvd', '')).upper().strip()
            date = str(qso.get('eqsl_qslrdate', '')).strip()
            confermato = rcvd == 'Y' or (not rcvd and date and date != '00000000')
            if per_pdf:
                return "<font color='green'><b>V</b></font>" if confermato else "<font color='gray'>-</font>"
            return "Y" if confermato else "-"
        return _xe(str(valore))

    # ── Statistiche ───────────────────────────
    def calcola_statistiche(self, qsos):
        dxcc_unici = {}
        bande_uniche = set()
        modi_unici = set()
        conteggio_modi = {}
        conteggio_continenti = {}
        conteggio_bande = {}
        lotw = eqsl = 0

        for qso in qsos:
            country = str(qso.get('country', '')).strip().upper()
            band = str(qso.get('band', '')).strip().upper()
            mode = str(qso.get('mode', '')).strip().upper()
            cont = str(qso.get('cont', '')).strip().upper()

            if country:
                if country not in dxcc_unici:
                    dxcc_unici[country] = {'qso': 0, 'bande': set(), 'modi': set(),
                                            'lotw': 0, 'eqsl': 0, 'continente': ''}
                dxcc_unici[country]['qso'] += 1
                if band:
                    dxcc_unici[country]['bande'].add(band)
                if mode:
                    dxcc_unici[country]['modi'].add(mode)

            if band:
                bande_uniche.add(band)
                conteggio_bande[band] = conteggio_bande.get(band, 0) + 1
            if mode:
                modi_unici.add(mode)
                conteggio_modi[mode] = conteggio_modi.get(mode, 0) + 1

            if not cont or cont not in ['EU', 'NA', 'SA', 'AS', 'AF', 'OC']:
                cont = MAPPA_CONTINENTI_DXCC.get(country, 'Altri')
            if country:
                dxcc_unici[country]['continente'] = cont
            conteggio_continenti[cont] = conteggio_continenti.get(cont, 0) + 1

            lotw_r = str(qso.get('lotw_qsl_rcvd', '')).upper().strip()
            lotw_d = str(qso.get('lotw_qslrdate', '')).strip()
            if lotw_r in ('Y','V') or (not lotw_r and lotw_d and lotw_d != '00000000'):
                lotw += 1
                if country:
                    dxcc_unici[country]['lotw'] += 1

            eqsl_r = str(qso.get('eqsl_qsl_rcvd', '')).upper().strip()
            eqsl_d = str(qso.get('eqsl_qslrdate', '')).strip()
            if eqsl_r == 'Y' or (not eqsl_r and eqsl_d and eqsl_d != '00000000'):
                eqsl += 1
                if country:
                    dxcc_unici[country]['eqsl'] += 1

        return {
            'totale': len(qsos),
            'dxcc': len(dxcc_unici),
            'dxcc_dettaglio': dxcc_unici,
            'bande': len(bande_uniche),
            'modi_distinti': len(modi_unici),
            'ripartizione_modi': conteggio_modi,
            'ripartizione_continenti': conteggio_continenti,
            'ripartizione_bande': conteggio_bande,
            'lotw_confermati': lotw,
            'eqsl_confermati': eqsl,
        }

    # ── Grafici torta ─────────────────────────
    def crea_grafico_torta(self, diz, totale, x, y, raggio):
        dati_ord = sorted(diz.items(), key=lambda i: i[1], reverse=True)
        labels, valori, altri = [], [], 0
        for idx, (k, v) in enumerate(dati_ord):
            if idx < 5:
                labels.append(k)
                valori.append(v)
            else:
                altri += v
        if altri:
            labels.append("Altri")
            valori.append(altri)

        leg_testi = [f"{l} ({valori[i]/totale*100:.1f}%)" for i, l in enumerate(labels)]

        c1 = self.colori_pdf['primario']
        c2 = self.colori_pdf['secondario']
        palette = [
            colors.HexColor(c1), colors.HexColor(c2),
            colors.HexColor("#4299E1"), colors.HexColor("#48BB78"),
            colors.HexColor("#ED8936"), colors.HexColor("#A0AEC0")
        ]

        pie = Pie()
        pie.x, pie.y, pie.width, pie.height = x, y, raggio * 2, raggio * 2
        pie.data = valori
        pie.labels = [f"{v/totale*100:.0f}%" for v in valori]
        pie.sideLabels = 1
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = colors.white
        pie.slices.fontName = "Helvetica-Bold"
        pie.slices.fontSize = 9
        for idx, col in enumerate(palette[:len(valori)]):
            pie.slices[idx].fillColor = col

        leg = Legend()
        leg.x = x + raggio * 2 + 35
        leg.y = y + raggio * 2 - 10
        leg.dxTextSpace = 8
        leg.dy = 6
        leg.dx = 10
        leg.deltay = 14
        leg.alignment = 'right'
        leg.fontName = 'Helvetica'
        leg.fontSize = 10
        leg.colorNamePairs = [(palette[i], leg_testi[i]) for i in range(len(valori))]
        return pie, leg

    def genera_grafico_reportlab(self, stats):
        d = Drawing(740, 450)
        totale = stats['totale']
        c1 = self.colori_pdf['primario']

        d.add(String(10, 432, "Distribuzione per Continente", fontName="Helvetica-Bold", fontSize=12, fillColor=colors.HexColor(c1)))
        pie_c, leg_c = self.crea_grafico_torta(stats['ripartizione_continenti'], totale, 10, 260, 75)
        d.add(pie_c); d.add(leg_c)

        d.add(String(390, 432, "Distribuzione per Modo Operativo", fontName="Helvetica-Bold", fontSize=12, fillColor=colors.HexColor(c1)))
        pie_m, leg_m = self.crea_grafico_torta(stats['ripartizione_modi'], totale, 390, 260, 75)
        d.add(pie_m); d.add(leg_m)

        d.add(String(10, 187, "Distribuzione per Banda", fontName="Helvetica-Bold", fontSize=12, fillColor=colors.HexColor(c1)))
        pie_b, leg_b = self.crea_grafico_torta(stats['ripartizione_bande'], totale, 10, 15, 75)
        d.add(pie_b); d.add(leg_b)

        d.add(String(390, 187, "Stato Conferme Digitali", fontName="Helvetica-Bold", fontSize=12, fillColor=colors.HexColor(c1)))
        c2 = self.colori_pdf['secondario']
        pct_l = stats['lotw_confermati'] / totale * 100 if totale else 0
        pct_e = stats['eqsl_confermati'] / totale * 100 if totale else 0

        d.add(String(390, 135, f"LoTW: {stats['lotw_confermati']} / {totale} QSO ({pct_l:.1f}%)", fontName="Helvetica", fontSize=10, fillColor=colors.HexColor("#2D3748")))
        d.add(Rect(390, 115, 250, 14, fillColor=colors.HexColor("#E2E8F0"), strokeColor=None))
        d.add(Rect(390, 115, max(1, int(250 * pct_l / 100)), 14, fillColor=colors.HexColor("#48BB78"), strokeColor=None))

        d.add(String(390, 75, f"eQSL: {stats['eqsl_confermati']} / {totale} QSO ({pct_e:.1f}%)", fontName="Helvetica", fontSize=10, fillColor=colors.HexColor("#2D3748")))
        d.add(Rect(390, 55, 250, 14, fillColor=colors.HexColor("#E2E8F0"), strokeColor=None))
        d.add(Rect(390, 55, max(1, int(250 * pct_e / 100)), 14, fillColor=colors.HexColor(c2), strokeColor=None))
        return d

    # ── Pagina DXCC ───────────────────────────
    def genera_pagina_dxcc(self, stats, styles):
        c1 = self.colori_pdf['primario']
        c2 = self.colori_pdf['secondario']
        cp = self.colori_pdf['riga_pari']

        style_h = ParagraphStyle('DxccH', parent=styles['Normal'], fontSize=10, leading=12,
                                  textColor=colors.whitesmoke, fontName="Helvetica-Bold", alignment=1)
        style_c = ParagraphStyle('DxccC', parent=styles['Normal'], fontSize=8, leading=10, alignment=1)
        style_title = ParagraphStyle('DxccTitle', fontName="Helvetica-Bold", fontSize=16,
                                      textColor=colors.HexColor(c1), leading=20)

        elementi = []
        elementi.append(Paragraph(T("dxcc_riepilogo"), style_title))
        elementi.append(Spacer(1, 10))

        # Raggruppa per continente nell'ordine standard
        per_cont = {}
        for paese, info in sorted(stats['dxcc_dettaglio'].items()):
            cont = info.get('continente', 'Altri') or 'Altri'
            per_cont.setdefault(cont, []).append((paese, info))

        intestazione = [Paragraph(t, style_h) for t in ["Country / DXCC", "Cont.", "QSO", "Bande", "Modi", "LoTW", "eQSL"]]

        for cont in CONTINENTS_ORDER:
            if cont not in per_cont:
                continue
            paesi = per_cont[cont]
            # Riga separatore continente
            style_sep = ParagraphStyle('Sep', fontName="Helvetica-Bold", fontSize=9,
                                        textColor=colors.HexColor(c2), alignment=0)
            lbl = f"── {cont} ({len(paesi)} entità) ──"
            elementi.append(Paragraph(lbl, style_sep))
            elementi.append(Spacer(1, 4))

            righe = [intestazione]
            for idx, (paese, info) in enumerate(paesi):
                bande_str = ", ".join(sorted(info['bande']))
                modi_str = ", ".join(sorted(info['modi']))
                lotw_str = "<font color='green'><b>V</b></font>" if info['lotw'] else "<font color='gray'>-</font>"
                eqsl_str = "<font color='green'><b>V</b></font>" if info['eqsl'] else "<font color='gray'>-</font>"
                righe.append([
                    Paragraph(paese, style_c),
                    Paragraph(cont, style_c),
                    Paragraph(str(info['qso']), style_c),
                    Paragraph(bande_str, style_c),
                    Paragraph(modi_str, style_c),
                    Paragraph(lotw_str, style_c),
                    Paragraph(eqsl_str, style_c),
                ])

            col_w = [200, 40, 40, 120, 120, 40, 40]
            t = Table(righe, colWidths=col_w, repeatRows=1)
            ts = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(c1)),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ])
            for i in range(1, len(righe)):
                if i % 2 == 0:
                    ts.add('BACKGROUND', (0, i), (-1, i), colors.HexColor(cp))
            t.setStyle(ts)
            elementi.append(t)
            elementi.append(Spacer(1, 12))

        return elementi

    # ── PDF ───────────────────────────────────
    def _crea_progress_dialog(self, titolo):
        """Crea una finestra progress bar modale."""
        dlg = ctk.CTkToplevel(self)
        dlg.title(titolo)
        dlg.geometry("420x130")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.attributes("-topmost", True)
        dlg.protocol("WM_DELETE_WINDOW", lambda: None)  # blocca chiusura

        self._progress_lbl = ctk.CTkLabel(dlg, text=T("pdf_generazione"),
                                           font=ctk.CTkFont(size=12))
        self._progress_lbl.pack(pady=(18, 8), padx=20)

        self._progress_bar = ctk.CTkProgressBar(dlg, width=360, height=18)
        self._progress_bar.pack(padx=20, pady=4)
        self._progress_bar.set(0)

        self._progress_pct = ctk.CTkLabel(dlg, text="0%",
                                           font=ctk.CTkFont(size=10), text_color="gray")
        self._progress_pct.pack(pady=(4, 0))

        dlg.update()
        return dlg

    def _aggiorna_progress(self, dlg, valore, testo):
        """Aggiorna progress bar — valore tra 0.0 e 1.0."""
        if dlg and dlg.winfo_exists():
            self._progress_bar.set(valore)
            self._progress_lbl.configure(text=testo)
            self._progress_pct.configure(text=f"{int(valore*100)}%")
            dlg.update()

    def processa_e_salva(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_nessun_file"))
            return
        # Mostra dialog opzioni PDF prima di generare
        self._apri_dialog_genera_pdf()

    def _apri_dialog_genera_pdf(self):
        """Dialog opzioni PDF (colonne, formato, ecc.) prima della generazione."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("📄 " + T("genera_pdf"))
        dlg.geometry("540x600")
        dlg.resizable(False, True)
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text="📄 " + T("genera_pdf"),
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(14,2), padx=16)
        ctk.CTkLabel(dlg, text=T("sez_colonne"),
                     font=ctk.CTkFont(size=10), text_color="gray").pack(pady=(0,6))

        # Colonne
        frame_col = ctk.CTkScrollableFrame(dlg, width=490, height=200)
        frame_col.pack(fill="x", padx=14, pady=(0,8))
        grid_cb = ctk.CTkFrame(frame_col, fg_color="transparent")
        grid_cb.pack(fill="x")
        for idx, (tag, info) in enumerate(self.campi_disponibili.items()):
            row, col = idx // 2, idx % 2
            cb = ctk.CTkCheckBox(grid_cb, text=info['nome'], variable=self.checkboxes[tag],
                                 font=ctk.CTkFont(size=11))
            cb.grid(row=row, column=col, padx=6, pady=2, sticky="w")
            self._checkbox_widgets[tag] = cb

        ctk.CTkFrame(dlg, height=1, fg_color="#4A5568").pack(fill="x", padx=14, pady=6)

        # Opzioni
        frame_opt = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_opt.pack(fill="x", padx=14, pady=(0,6))

        self._widget_refs['dxcc_page'] = ctk.CTkCheckBox(
            frame_opt, text=T("dxcc_page"), variable=self.var_dxcc_page,
            font=ctk.CTkFont(size=11))
        self._widget_refs['dxcc_page'].pack(anchor="w", pady=3)

        frame_fmt = ctk.CTkFrame(frame_opt, fg_color="transparent")
        frame_fmt.pack(anchor="w", pady=3, fill="x")
        ctk.CTkLabel(frame_fmt, text=T("formato_pagina"),
                     font=ctk.CTkFont(size=11), width=130).pack(side="left")
        ctk.CTkOptionMenu(frame_fmt, variable=self.var_formato_pdf,
                          values=["A4", "US Letter"], width=120).pack(side="left")

        frame_extra = ctk.CTkFrame(frame_opt, fg_color="transparent")
        frame_extra.pack(anchor="w", pady=3, fill="x")
        ctk.CTkButton(frame_extra, text="🎨 " + T("colori_pdf"),
                      command=self.apri_colori, fg_color="#4A5568",
                      height=28, width=160).pack(side="left", padx=(0,6))
        ctk.CTkButton(frame_extra, text="📋 Opzioni registro PDF",
                      command=self.apri_opzioni_registro_pdf, fg_color="#4A5568",
                      height=28, width=180).pack(side="left")

        # Pulsanti
        frame_btn = ctk.CTkFrame(dlg, fg_color="transparent")
        frame_btn.pack(fill="x", padx=14, pady=(8,14))

        def _genera():
            dlg.destroy()
            self._esegui_genera_pdf()

        ctk.CTkButton(frame_btn, text="📄 Genera PDF", command=_genera,
                      height=38, fg_color=TH.PRIMARY_H, hover_color="#2A4365",
                      font=ctk.CTkFont(size=13, weight="bold")).pack(
                      side="left", expand=True, fill="x", padx=(0,6))
        ctk.CTkButton(frame_btn, text=T("cm_annulla"), command=dlg.destroy,
                      height=38, width=100, fg_color="#718096").pack(side="left")

    def _esegui_genera_pdf(self):
        campi_scelti = {}
        for tag in self.ordine_campi_pdf:
            if tag in self.campi_disponibili and self.checkboxes[tag].get():
                info = dict(self.campi_disponibili[tag])
                if tag in self.width_pdf:
                    info['width_base'] = self.width_pdf[tag]
                campi_scelti[tag] = info
        if not campi_scelti:
            messagebox.showwarning("Attenzione", T("warn_nessun_campo"))
            return

        # Controllo larghezza colonne
        spazio_max = 756 if self.var_formato_pdf.get() == "US Letter" else 792
        tot_assegnato = sum(info['width_base'] for info in campi_scelti.values())
        if tot_assegnato > spazio_max * 1.05:
            messagebox.showwarning(
                "Troppi campi selezionati",
                f"Le colonne selezionate occupano {tot_assegnato} pt ma lo spazio disponibile "
                f"è {spazio_max} pt (formato {self.var_formato_pdf.get()}, orizzontale).\n\n"
                f"Il PDF risulterà con testo tagliato o illeggibile.\n\n"
                f"Suggerimento: deseleziona alcuni campi dalla sidebar — ad esempio "
                f"Frequenza, Nome, Locator — e riprova.\n"
                f"Si consigliano al massimo 8-9 colonne."
            )
            return

        base_nome = (os.path.splitext(os.path.basename(self.filepath))[0]
                     if self.filepath else "ADIF_FZR_" + datetime.now().strftime("%Y%m%d"))

        stazione = self.entry_owner.get().strip().upper()
        dettagli = self.entry_details.get().strip()
        qsos = self._qsos_attivi()
        stats = self.calcola_statistiche(qsos)

        # Raccoglie gli anni presenti nel log
        anni_disponibili = sorted(set(
            str(q.get('qso_date', ''))[:4]
            for q in qsos
            if len(str(q.get('qso_date', ''))) == 8 and str(q.get('qso_date', ''))[:4].isdigit()
        ), reverse=True)

        multi_anno = len(anni_disponibili) > 1
        large_log  = len(qsos) > 10000

        anno_filtro = None  # None = tutto il log

        # Mostra il dialog di scelta anno se multi-anno OPPURE log grande
        if multi_anno or large_log:
            _scelta = [None]   # 'tutto' | 'anno' | None (annulla)
            _anno_scelto = [anni_disponibili[0] if anni_disponibili else '']

            dlg_anno = ctk.CTkToplevel(self)
            dlg_anno.title("Opzioni PDF")
            dlg_anno.geometry("460x300" if multi_anno else "460x220")
            dlg_anno.resizable(False, False)
            dlg_anno.grab_set(); dlg_anno.lift(); dlg_anno.focus_force()

            if large_log:
                ctk.CTkLabel(dlg_anno,
                    text=f"⚠  {len(qsos):,} QSO — log molto grande",
                    font=ctk.CTkFont(size=13, weight="bold"),
                    text_color=TH.WARN_TEXT).pack(pady=(16, 2), padx=20)
                ctk.CTkLabel(dlg_anno,
                    text=T("dv_gen_minuti"),
                    font=ctk.CTkFont(size=10), text_color="gray").pack(pady=(0, 10))
            else:
                ctk.CTkLabel(dlg_anno,
                    text=f"📅  Log con QSO di {len(anni_disponibili)} anni diversi",
                    font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(16, 10), padx=20)

            def _scegli_tutto():
                _scelta[0] = 'tutto'; dlg_anno.destroy()
            ctk.CTkButton(dlg_anno,
                text=f"📄 Stampa tutto ({len(qsos):,} QSO)",
                command=_scegli_tutto, height=38,
                fg_color=TH.PRIMARY, hover_color=TH.PRIMARY_H,
                font=ctk.CTkFont(size=11, weight="bold")).pack(fill="x", padx=20, pady=(0, 8))

            if multi_anno:
                frame_anno = ctk.CTkFrame(dlg_anno, fg_color="transparent")
                frame_anno.pack(fill="x", padx=20, pady=(0, 8))
                var_anno = ctk.StringVar(value=_anno_scelto[0])
                n_anno_init = sum(1 for q in qsos if str(q.get('qso_date',''))[:4] == var_anno.get())
                lbl_n_anno = ctk.CTkLabel(frame_anno, text=f"({n_anno_init} QSO)",
                                           font=ctk.CTkFont(size=10), text_color="gray", width=70)
                def _aggiorna_n_anno(anno):
                    n = sum(1 for q in qsos if str(q.get('qso_date',''))[:4] == anno)
                    lbl_n_anno.configure(text=f"({n} QSO)")
                    _anno_scelto[0] = anno
                ctk.CTkOptionMenu(frame_anno, values=anni_disponibili,
                                   variable=var_anno, width=90,
                                   command=_aggiorna_n_anno).pack(side="left", padx=(0,6))
                lbl_n_anno.pack(side="left", padx=(0,8))
                def _scegli_anno():
                    _scelta[0] = 'anno'; _anno_scelto[0] = var_anno.get(); dlg_anno.destroy()
                ctk.CTkButton(frame_anno,
                    text="📅 Solo l'anno selezionato",
                    command=_scegli_anno, height=38,
                    fg_color=TH.SUCCESS_H, hover_color=TH.SUCCESS,
                    font=ctk.CTkFont(size=11)).pack(side="left", expand=True, fill="x")

            ctk.CTkButton(dlg_anno, text=T("cm_annulla"),
                fg_color="#718096", height=30,
                command=dlg_anno.destroy).pack(pady=(0, 14), padx=20, fill="x")

            dlg_anno.wait_window()

            if _scelta[0] is None:
                return
            elif _scelta[0] == 'anno':
                anno_filtro = _anno_scelto[0]
                qsos = [q for q in qsos if str(q.get('qso_date', ''))[:4] == anno_filtro]
                stats = self.calcola_statistiche(qsos)

        # Ora chiedi dove salvare — con il nome già contenente l'anno se filtrato
        suffisso = f"_{anno_filtro}" if anno_filtro else ""
        nome_def = f"{base_nome}{suffisso}_registro.pdf"
        save_path = filedialog.asksaveasfilename(
            title=T("dv_salva_pdf") + (f" — {anno_filtro}" if anno_filtro else ""),
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=nome_def)
        if not save_path:
            return

        # Apri progress dialog
        prog = self._crea_progress_dialog(T("pdf_generazione"))
        self._aggiorna_progress(prog, 0.05, T("pdf_step_copertina"))

        c1 = self.colori_pdf['primario']
        c2 = self.colori_pdf['secondario']
        cp = self.colori_pdf['riga_pari']

        styles = getSampleStyleSheet()
        fs = getattr(self, 'font_size_pdf', 7)
        style_cell = ParagraphStyle('Cella', parent=styles['Normal'], fontSize=fs, leading=fs+2, alignment=1)
        style_header_pdf = ParagraphStyle('Head', parent=styles['Normal'], fontSize=10, leading=12,
                                           textColor=colors.whitesmoke, fontName="Helvetica-Bold", alignment=1)

        try:
            # Copertina
            style_cover_title = ParagraphStyle('CT', fontName="Helvetica-Bold", fontSize=26, leading=32,
                                                textColor=colors.HexColor(c1), alignment=1)
            style_cover_sub = ParagraphStyle('CS', fontName="Helvetica", fontSize=13, leading=16,
                                              textColor=colors.HexColor("#4A5568"), alignment=1)
            style_cover_sec = ParagraphStyle('CSc', fontName="Helvetica-Bold", fontSize=15, leading=19,
                                              textColor=colors.HexColor(c2), alignment=1, spaceAfter=8)
            style_stat_val = ParagraphStyle('SV', fontName="Helvetica-Bold", fontSize=18, leading=22,
                                             textColor=colors.HexColor(c1), alignment=1)
            style_grande = ParagraphStyle('SG', fontName="Helvetica-Bold", fontSize=80, leading=90,
                                           textColor=colors.HexColor(c1), alpha=0.07, alignment=1)

            elementi = []
            self._aggiorna_progress(prog, 0.10, T("pdf_step_copertina"))
            elementi.append(Spacer(1, 10))
            elementi.append(Paragraph(T("copertina_titolo"), style_cover_title))
            elementi.append(Spacer(1, 5))
            info_op = T("copertina_sub")
            if stazione:
                info_op += f" - CALLSIGN: {stazione}"
            if dettagli:
                info_op += f" ({dettagli})"
            elementi.append(Paragraph(info_op, style_cover_sub))
            n_filtrati = len(qsos)
            n_totali = len(self.qsos_caricati)
            note_filtro = f" [Filtro: {n_filtrati}/{n_totali} QSO]" if n_filtrati != n_totali else ""
            sorgente = os.path.basename(self.filepath) if self.filepath else "Log corrente"
            elementi.append(Paragraph(f"File sorgente: {sorgente}{note_filtro}", style_cover_sub))

            elementi.append(Spacer(1, 20))
            # Usa il titolo personalizzato se impostato, altrimenti il callsign
            titolo_pdf = self.titolo_pdf_custom if self.titolo_pdf_custom else (stazione if stazione else "RADIO LOG")
            elementi.append(Paragraph(titolo_pdf, style_grande))
            elementi.append(Spacer(1, 25))
            elementi.append(Paragraph(T("stat_generali"), style_cover_sec))

            h_stats = [Paragraph(T(k), style_header_pdf) for k in ["stat_qso_totali","stat_dxcc","stat_bande","stat_modi"]]
            v_stats = [Paragraph(str(stats[k]), style_stat_val) for k in ['totale', 'dxcc', 'bande', 'modi_distinti']]
            t_stats = Table([h_stats, v_stats], colWidths=[150, 150, 150, 150], hAlign='CENTER')
            t_stats.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(c1)),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E0")),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor(cp)),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elementi.append(t_stats)
            elementi.append(Spacer(1, 15))
            elementi.append(self.genera_grafico_reportlab(stats))
            elementi.append(PageBreak())

            # Pagina DXCC (opzionale)
            if self.var_dxcc_page.get() and stats['dxcc_dettaglio']:
                elementi.extend(self.genera_pagina_dxcc(stats, styles))
                elementi.append(PageBreak())

            # Tabella QSO
            style_main_title = ParagraphStyle('MT', fontName="Helvetica-Bold", fontSize=16,
                                               leading=20, textColor=colors.HexColor(c1))
            titolo_log = T("log_stazione")
            if stazione:
                titolo_log += f" - {stazione}"
            elementi.append(Paragraph(titolo_log, style_main_title))
            elementi.append(Spacer(1, 10))

            self._aggiorna_progress(prog, 0.45, T("pdf_step_tabella"))
            data_table = [[Paragraph(info['nome'], style_header_pdf) for info in campi_scelti.values()]]
            for qso in qsos:
                riga = [Paragraph(self._formatta_valore(tag, qso, per_pdf=True), style_cell)
                        for tag in campi_scelti.keys()]
                data_table.append(riga)

            larg = [info['width_base'] for info in campi_scelti.values()]
            tot_assegnato = sum(larg)
            spazio_max = 756 if self.var_formato_pdf.get() == "US Letter" else 792
            if tot_assegnato < spazio_max:
                delta = (spazio_max - tot_assegnato) / tot_assegnato
                larg = [w + w * delta for w in larg]

            t = Table(data_table, colWidths=larg, repeatRows=1)
            ts = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(c1)),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 6),
            ])
            for i in range(1, len(data_table)):
                if i % 2 == 0:
                    ts.add('BACKGROUND', (0, i), (-1, i), colors.HexColor(cp))
            t.setStyle(ts)
            elementi.append(t)

            _psize = LETTER if self.var_formato_pdf.get() == "US Letter" else A4
            self._aggiorna_progress(prog, 0.80, T("pdf_step_salvataggio"))
            doc = SimpleDocTemplate(save_path, pagesize=landscape(_psize),
                                    rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=50)

            # Stima pagine per progress bar (circa 40-50 QSO per pagina orizzontale)
            n_qso_pdf = len(qsos)
            _n_pagine_stimate = max(3, 2 + n_qso_pdf // 45)

            # Contatore pagine condiviso tra thread e main thread
            _pagine_counter = [0]
            _build_done = [False]
            _build_error = [None]

            def canvas_builder(*args, **kwargs):
                c = ElegantNumberedCanvas(*args, **kwargs)
                c.stazione_call = stazione
                c.dettagli_op = dettagli
                c.colore_primario = c1
                # Sovrascrivi showPage per intercettare ogni nuova pagina
                _orig_showPage = c.showPage.__func__ if hasattr(c.showPage, '__func__') else None
                def _showPage_counted():
                    _pagine_counter[0] += 1
                    ElegantNumberedCanvas.showPage(c)
                c.showPage = _showPage_counted
                return c

            import threading as _threading

            def _build_thread():
                try:
                    doc.build(elementi, canvasmaker=canvas_builder)
                except Exception as ex:
                    _build_error[0] = ex
                finally:
                    _build_done[0] = True

            t_build = _threading.Thread(target=_build_thread, daemon=True)
            t_build.start()

            # Polling sul main thread: aggiorna progress bar page by page
            # senza bloccare la UI
            def _poll_progress():
                if not _build_done[0]:
                    pag = _pagine_counter[0]
                    pct = min(0.98, 0.80 + 0.18 * (pag / _n_pagine_stimate))
                    if prog.winfo_exists():
                        self._aggiorna_progress(
                            prog, pct,
                            f"{T('pdf_step_salvataggio')} (pag. {pag}…)")
                        prog.after(400, _poll_progress)
                else:
                    if _build_error[0]:
                        try:
                            prog.destroy()
                        except Exception:
                            pass
                        messagebox.showerror(
                            "Errore", f"Errore durante la generazione del PDF:\n{_build_error[0]}")
                    else:
                        self._aggiorna_progress(prog, 1.0, T("pdf_completato"))
                        prog.after(600, prog.destroy)
                        if messagebox.askyesno("Successo", f"PDF generato!\n\nVuoi aprirlo?"):
                            percorso = os.path.abspath(save_path)
                            if os.path.exists(percorso):
                                os.startfile(percorso)

            prog.after(400, _poll_progress)
            # Non bloccare il main thread — il flusso normale termina qui,
            # il resto avviene nei callback di _poll_progress
            return

        except Exception as ex:
            messagebox.showerror("Errore", f"Errore durante la generazione del PDF:\n{ex}")

    # ── Esporta CSV ───────────────────────────
    def esporta_csv(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        campi_scelti = {}
        for tag in self.ordine_campi_pdf:
            if tag in self.campi_disponibili and self.checkboxes[tag].get():
                info = dict(self.campi_disponibili[tag])
                if tag in self.width_pdf:
                    info['width_base'] = self.width_pdf[tag]
                campi_scelti[tag] = info
        if not campi_scelti:
            messagebox.showwarning("Attenzione", "Seleziona almeno un campo.")
            return

        base_nome = (os.path.splitext(os.path.basename(self.filepath))[0]
                     if self.filepath else "ADIF_FZR_" + datetime.now().strftime("%Y%m%d"))
        nome_def = base_nome + "_log.csv"
        save_path = filedialog.asksaveasfilename(title=T("dv_salva_csv"), defaultextension=".csv",
                                                  filetypes=[("CSV files", "*.csv")], initialfile=nome_def)
        if not save_path:
            return

        qsos = self._qsos_attivi()
        try:
            with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                intestazione = [info['nome'] for info in campi_scelti.values()]
                # Aggiungi colonne extra utili
                intestazione += ["SAT_NAME", "CONT"]
                writer.writerow(intestazione)
                for qso in qsos:
                    riga = [self._formatta_valore(tag, qso, per_pdf=False) for tag in campi_scelti.keys()]
                    riga.append(str(qso.get('sat_name', '')).strip())
                    country = str(qso.get('country', '')).strip().upper()
                    cont = str(qso.get('cont', '')).strip().upper()
                    if not cont or cont not in ['EU', 'NA', 'SA', 'AS', 'AF', 'OC']:
                        cont = MAPPA_CONTINENTI_DXCC.get(country, '')
                    riga.append(cont)
                    writer.writerow(riga)

            messagebox.showinfo("Successo", f"CSV esportato:\n{os.path.basename(save_path)}\n({len(qsos)} righe)")
        except Exception as ex:
            messagebox.showerror("Errore", f"Errore durante l'esportazione CSV:\n{ex}")

    # ── Esporta Excel ─────────────────────────
    def esporta_html(self):
        if not self.qsos_caricati:
            messagebox.showwarning(T("attenzione"), T("warn_nessun_file"))
            return
        qsos = self._qsos_attivi()
        stazione = self.entry_owner.get().strip().upper() or "LOG"
        nome_def = (os.path.splitext(os.path.basename(self.filepath))[0] + "_log.html"
                    if self.filepath else "log.html")
        save_path = filedialog.asksaveasfilename(
            title=T("dv_esporta_html"), defaultextension=".html",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
            initialfile=nome_def)
        if not save_path:
            return
        bande = sorted(set(str(q.get("band", "?")).upper() for q in qsos if q.get("band")))
        modi  = sorted(set(str(q.get("mode", "?")).upper() for q in qsos if q.get("mode")))
        import json as _json
        import html as _html
        rows = []
        for q in qsos:
            data = str(q.get("qso_date", ""))
            if len(data) == 8:
                data = data[6:8] + "/" + data[4:6] + "/" + data[0:4]
            ora = str(q.get("time_on", ""))[:4]
            if len(ora) == 4:
                ora = ora[:2] + ":" + ora[2:]
            lotw = str(q.get("lotw_qsl_rcvd", "")).upper().strip()
            eqsl = str(q.get("eqsl_qsl_rcvd", "")).upper().strip()
            modo_puro = str(q.get("mode", "")).upper()
            submode = str(q.get("submode", "")).upper().strip()
            modo_vis = f"{modo_puro} ({submode})" if submode and submode != modo_puro else modo_puro
            rows.append({
                "call":    str(q.get("call", "")).upper(),
                "date":    data,
                "time":    ora,
                "band":    str(q.get("band", "")).upper(),
                "mode":    modo_puro,
                "mode_vis": modo_vis,
                "rst_s":   str(q.get("rst_sent", "")),
                "rst_r":   str(q.get("rst_rcvd", "")),
                "country": self._country_con_stato(q),
                "grid":    str(q.get("gridsquare", "")).upper(),
                "lotw":    "Y" if lotw in ("Y", "V") else "N",
                "eqsl":    "Y" if eqsl == "Y" else "N",
            })
        rows_json  = _json.dumps(rows, ensure_ascii=False).replace("</", "<\\/")
        stazione_safe = _html.escape(stazione, quote=True)
        bande_opts = "".join(
            '<option value="' + _html.escape(b, quote=True) + '">' + _html.escape(b) + "</option>"
            for b in bande
        )
        modi_opts  = "".join(
            '<option value="' + _html.escape(m, quote=True) + '">' + _html.escape(m) + "</option>"
            for m in modi
        )
        try:
            html = self._build_html(stazione, len(qsos), rows_json, bande_opts, modi_opts, self.colori_html)
            with open(save_path, "w", encoding="utf-8") as fw:
                fw.write(html)
            messagebox.showinfo(T("successo"),
                "HTML esportato:" + chr(10) + os.path.basename(save_path) +
                chr(10) + str(len(qsos)) + " QSO")
            try:
                os.startfile(os.path.abspath(save_path))
            except Exception:
                pass
        except Exception as ex:
            messagebox.showerror(T("errore"), "Errore:" + chr(10) + str(ex))

    def _build_html(self, stazione, n_qso, rows_json, bande_opts, modi_opts, colori=None):
        import html as _html
        stazione_safe = _html.escape(str(stazione), quote=True)

        # Colori personalizzati o default
        _c = colori or {}
        C1  = _c.get('primario',   '#1A365D')
        C2  = _c.get('secondario', '#2B6CB0')
        BG_D = _c.get('bg_scuro',  '#0D1117')
        BG_L = _c.get('bg_chiaro', '#F7FAFC')

        # Colori derivati automaticamente
        def _lighten(hex_c, pct=40):
            """Schiarisce un colore hex di pct/255 su ogni canale."""
            h = hex_c.lstrip('#')
            r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
            r = min(255, r + pct); g = min(255, g + pct); b = min(255, b + pct)
            return f"#{r:02X}{g:02X}{b:02X}"

        C1_hover = _lighten(C1, 30)
        C2_hover = _lighten(C2, 30)
        h = []
        a = h.append
        css = (
            "*{box-sizing:border-box;margin:0;padding:0}"
            ":root{"
            f"--c1:{C1};--c2:{C2};--c1h:{C1_hover};--c2h:{C2_hover};"
            f"--bg:{BG_D};--bg2:#111827;--fg:#E2E8F0;--fg2:#90CDF4;"
            "--brd:#2D3748;--fil:#141414;--even:#111827;--sc:#141414}"
            "body.light{"
            f"--bg:{BG_L};--bg2:#FFFFFF;--fg:#1A202C;--fg2:#2D3748;"
            "--brd:#CBD5E0;--fil:#EBF4FF;--even:#F7FAFC;--sc:#FFFFFF}"
            "body{font-family:Arial,sans-serif;background:var(--bg);color:var(--fg);font-size:.93em;transition:background .2s,color .2s}"
            ".hdr{background:linear-gradient(135deg,var(--c1),var(--c2));"
            "padding:24px 30px;margin-bottom:20px;display:flex;align-items:center;justify-content:space-between}"
            ".hdr h1{color:#fff;font-size:1.8em;margin-bottom:4px}"
            ".hdr p{color:rgba(255,255,255,.75);font-size:.9em}"
            ".hdr-left{flex:1}"
            ".theme-btn{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);"
            "color:#fff;padding:6px 14px;border-radius:20px;cursor:pointer;font-size:.9em;"
            "transition:background .2s}"
            ".theme-btn:hover{background:rgba(255,255,255,.25)}"
            ".wrap{max-width:1200px;margin:0 auto;padding:0 16px 40px}"
            ".fil{background:var(--fil);border:1px solid var(--brd);border-radius:10px;"
            "padding:14px 18px;margin-bottom:14px;display:flex;flex-wrap:wrap;gap:10px;align-items:flex-end}"
            ".fg{display:flex;flex-direction:column;gap:3px}"
            ".fg label{font-size:.75em;color:var(--c2);font-weight:700;"
            "letter-spacing:1px;text-transform:uppercase}"
            ".fg input,.fg select{background:var(--bg);border:1px solid var(--c2);"
            "border-radius:6px;color:var(--fg);padding:5px 9px;font-size:.87em;"
            "outline:none;min-width:100px}"
            ".fg input:focus,.fg select:focus{border-color:var(--c2h)}"
            ".br{background:#718096;border:none;color:#fff;padding:6px 14px;"
            "border-radius:6px;cursor:pointer;font-size:.87em;margin-top:16px}"
            ".br:hover{background:#4A5568}"
            ".stats{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px}"
            ".sc{background:var(--sc);border:1px solid var(--brd);border-radius:8px;"
            "padding:8px 16px;text-align:center;min-width:80px}"
            f".sc .n{{font-size:1.4em;font-weight:700;color:{C2}}}"
            ".sc .l{font-size:.72em;color:#718096}"
            ".tw{overflow-x:auto;border-radius:10px;border:1px solid var(--brd)}"
            "table{width:100%;border-collapse:collapse;font-size:.87em}"
            f"thead th{{background:{C1};color:#fff;padding:9px 11px;"
            "text-align:left;font-weight:700;white-space:nowrap;cursor:pointer;user-select:none}"
            f"thead th:hover{{background:{C2};color:#fff}}"
            "thead th.sa::after{content:' \u25b2'}thead th.sd::after{content:' \u25bc'}"
            "tbody tr{border-bottom:1px solid var(--brd)}"
            "tbody tr:hover{background:var(--fil)}"
            "tbody tr:nth-child(even){background:var(--even)}"
            "tbody tr:nth-child(even):hover{background:var(--fil)}"
            "td{padding:7px 11px;white-space:nowrap}"
            ".by{display:inline-block;padding:1px 7px;border-radius:9px;"
            "font-size:.79em;font-weight:700;background:#276749;color:#9AE6B4}"
            ".bn{display:inline-block;padding:1px 7px;border-radius:9px;"
            "font-size:.79em;font-weight:700;background:#2D3748;color:#718096}"
            ".nr{text-align:center;padding:36px;color:#718096}"
            ".ftr{text-align:center;padding:18px;color:#718096;"
            "font-size:.78em;margin-top:18px}"
        )
        a("<!DOCTYPE html><html lang='it'>")
        a("<head><meta charset='UTF-8'>")
        a("<meta name='viewport' content='width=device-width,initial-scale=1'>")
        a("<title>Log " + stazione_safe + " - ADIF FZR 2.5</title>")
        a("<style>" + css + "</style></head><body>")
        a("<div class='hdr'><div class='hdr-left'>")
        a("<h1>&#128251; Log " + stazione_safe + "</h1>")
        a("<p>ADIF FZR 2.5 &nbsp;&middot;&nbsp; " + str(n_qso) + " QSO totali</p>")
        a("</div>")
        a("<button class='theme-btn' onclick='toggleTheme()' id='tbtn'>&#9790; Tema chiaro</button>")
        a("</div><div class='wrap'>")
        a("<div class='fil'>")
        a("<div class='fg'><label>Callsign</label>"
          "<input id='fc' placeholder='es. DL2...' oninput='af()'></div>")
        a("<div class='fg'><label>Banda</label>"
          "<select id='fb' onchange='af()'><option value=''>Tutte</option>"
          + bande_opts + "</select></div>")
        a("<div class='fg'><label>Modo</label>"
          "<select id='fm' onchange='af()'><option value=''>Tutti</option>"
          + modi_opts + "</select></div>")
        a("<div class='fg'><label>Da data</label>"
          "<input id='fd1' placeholder='gg/mm/aaaa' oninput='af()'></div>")
        a("<div class='fg'><label>A data</label>"
          "<input id='fd2' placeholder='gg/mm/aaaa' oninput='af()'></div>")
        a("<div class='fg'><label>Country</label>"
          "<input id='fco' placeholder='es. Italy...' oninput='af()'></div>")
        a("<div class='fg'><label>LoTW</label>"
          "<select id='fl' onchange='af()'><option value=''>Tutti</option>"
          "<option value='Y'>Confermato</option>"
          "<option value='N'>Non conf.</option></select></div>")
        a("<button class='br' onclick='rf()'>Reset</button></div>")
        a("<div class='stats'>")
        a("<div class='sc'><div class='n' id='st'>0</div><div class='l'>QSO</div></div>")
        a("<div class='sc'><div class='n' id='sl'>0</div><div class='l'>LoTW &#10003;</div></div>")
        a("<div class='sc'><div class='n' id='se'>0</div><div class='l'>eQSL &#10003;</div></div>")
        a("<div class='sc'><div class='n' id='sd'>0</div><div class='l'>DXCC</div></div>")
        a("</div>")
        a("<div class='tw'><table><thead><tr>")
        a("<th onclick='s(0)'>Data</th><th onclick='s(1)'>UTC</th>")
        a("<th onclick='s(2)'>Callsign</th><th onclick='s(3)'>Banda</th>")
        a("<th onclick='s(4)'>Modo</th><th onclick='s(5)'>RST TX</th>")
        a("<th onclick='s(6)'>RST RX</th><th onclick='s(7)'>Country</th>")
        a("<th onclick='s(8)'>Locator</th><th onclick='s(9)'>LoTW</th>")
        a("<th onclick='s(10)'>eQSL</th>")
        a("</tr></thead><tbody id='lb'></tbody></table>")
        a("<div class='nr' id='nr' style='display:none'>Nessun QSO trovato</div></div></div>")
        a("<div class='ftr'>ADIF FZR 2.5 &middot; " + stazione_safe +
          " &middot; " + str(n_qso) + " QSO</div>")
        # JavaScript
        js = (
            "const D=" + rows_json + ";"
            "let sc=0,sd=1,f=[...D];"
            "function pd(s){"
            "if(!s)return 0;"
            "const p=s.split('/');"
            "return p.length===3?new Date(p[2],p[1]-1,p[0]).getTime():0;}"
            "function af(){"
            "const fc=document.getElementById('fc').value.toUpperCase(),"
            "fb=document.getElementById('fb').value,"
            "fm=document.getElementById('fm').value,"
            "fd1=document.getElementById('fd1').value,"
            "fd2=document.getElementById('fd2').value,"
            "fco=document.getElementById('fco').value.toUpperCase(),"
            "fl=document.getElementById('fl').value,"
            "t1=pd(fd1),t2=pd(fd2);"
            "f=D.filter(r=>{"
            "if(fc&&!r.call.includes(fc))return false;"
            "if(fb&&r.band!==fb)return false;"
            "if(fm&&r.mode!==fm)return false;"
            "if(fco&&!r.country.toUpperCase().includes(fco))return false;"
            "if(fl&&r.lotw!==fl)return false;"
            "if(t1){const td=pd(r.date);if(td<t1)return false;}"
            "if(t2){const td=pd(r.date);if(td>t2)return false;}"
            "return true;});rt();}"
            "function rf(){"
            "['fc','fd1','fd2','fco'].forEach(id=>document.getElementById(id).value='');"
            "['fb','fm','fl'].forEach(id=>document.getElementById(id).value='');"
            "f=[...D];rt();}"
            "function s(c){"
            "const ths=document.querySelectorAll('thead th');"
            "ths.forEach(t=>t.classList.remove('sa','sd'));"
            "if(sc===c)sd*=-1;else{sc=c;sd=1;}"
            "ths[c].classList.add(sd===1?'sa':'sd');"
            "const k=['date','time','call','band','mode','rst_s','rst_r','country','grid','lotw','eqsl'];"
            "f.sort((a,b)=>(a[k[c]]>b[k[c]]?1:-1)*sd);rt();}"
            "function b(v){"
            "return v==='Y'?"
            "'<span class=\"by\">&#10003;</span>'"
            ":'<span class=\"bn\">&ndash;</span>';}"
            "function e(v){"
            "return String(v??'').replace(/[&<>\"']/g,c=>"
            "({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',\"'\":'&#39;'}[c]));}"
            "function rt(){"
            "const tb=document.getElementById('lb'),"
            "nr=document.getElementById('nr');"
            "if(!f.length){tb.innerHTML='';nr.style.display='block';}"
            "else{"
            "nr.style.display='none';"
            "tb.innerHTML=f.map(r=>"
            "'<tr>'"
            "+'<td>'+e(r.date)+'</td>'"
            "+'<td>'+e(r.time)+'</td>'"
            "+'<td><strong style=\"color:#90CDF4\">'+e(r.call)+'</strong></td>'"
            "+'<td>'+e(r.band)+'</td>'"
            "+'<td>'+e(r.mode_vis||r.mode)+'</td>'"
            "+'<td>'+e(r.rst_s)+'</td>'"
            "+'<td>'+e(r.rst_r)+'</td>'"
            "+'<td>'+e(r.country)+'</td>'"
            "+'<td>'+e(r.grid)+'</td>'"
            "+'<td>'+b(r.lotw)+'</td>'"
            "+'<td>'+b(r.eqsl)+'</td>'"
            "+'</tr>').join('');}"
            "document.getElementById('st').textContent=f.length;"
            "document.getElementById('sl').textContent=f.filter(r=>r.lotw==='Y').length;"
            "document.getElementById('se').textContent=f.filter(r=>r.eqsl==='Y').length;"
            "document.getElementById('sd').textContent="
            "new Set(f.map(r=>r.country).filter(Boolean)).size;}"
            "rt();"
        )
        toggle_js = (
            "function toggleTheme(){"
            "var b=document.body,btn=document.getElementById('tbtn');"
            "if(b.classList.toggle('light')){"
            "btn.innerHTML='&#9728; Tema scuro';localStorage.setItem('adfzr-theme','light');}"
            "else{btn.innerHTML='&#9790; Tema chiaro';localStorage.setItem('adfzr-theme','dark');}}"
            "(function(){"
            "var t=localStorage.getItem('adfzr-theme');"
            "if(t==='light'){document.body.classList.add('light');"
            "var b=document.getElementById('tbtn');"
            "if(b)b.innerHTML='&#9728; Tema scuro';}})();"
        )
        a("<script>" + js + toggle_js + "</script></body></html>")
        return chr(10).join(h)


    def esporta_excel(self):
        if not self.qsos_caricati:
            messagebox.showwarning("Attenzione", T("warn_carica_prima"))
            return
        campi_scelti = {}
        for tag in self.ordine_campi_pdf:
            if tag in self.campi_disponibili and self.checkboxes[tag].get():
                info = dict(self.campi_disponibili[tag])
                if tag in self.width_pdf:
                    info['width_base'] = self.width_pdf[tag]
                campi_scelti[tag] = info
        if not campi_scelti:
            messagebox.showwarning("Attenzione", "Seleziona almeno un campo.")
            return

        base_nome = (os.path.splitext(os.path.basename(self.filepath))[0]
                     if self.filepath else "ADIF_FZR_" + datetime.now().strftime("%Y%m%d"))
        nome_def = base_nome + "_log.xlsx"
        save_path = filedialog.asksaveasfilename(title=T("dv_salva_excel"), defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx")], initialfile=nome_def)
        if not save_path:
            return

        qsos = self._qsos_attivi()
        stazione = self.entry_owner.get().strip().upper()
        stats = self.calcola_statistiche(qsos)

        try:
            wb = openpyxl.Workbook()

            # ── Foglio LOG ─────────────────────────────────────────────
            ws_log = wb.active
            ws_log.title = "Log QSO"

            c1_hex = self.colori_pdf['primario'].lstrip('#')
            c2_hex = self.colori_pdf['secondario'].lstrip('#')
            cp_hex = self.colori_pdf['riga_pari'].lstrip('#')

            fill_header = PatternFill("solid", fgColor=c1_hex)
            fill_pari = PatternFill("solid", fgColor=cp_hex)
            font_header = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            font_norm = Font(name="Arial", size=9)
            align_c = Alignment(horizontal="center", vertical="center")
            align_l = Alignment(horizontal="left", vertical="center")
            bordo = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            )

            intestazione_log = [info['nome'] for info in campi_scelti.values()] + ["SAT", "Continente"]
            ws_log.append(intestazione_log)

            col_widths = {i + 1: len(h) + 4 for i, h in enumerate(intestazione_log)}

            for col_idx, cell in enumerate(ws_log[1], 1):
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = align_c
                cell.border = bordo

            for r_idx, qso in enumerate(qsos, 2):
                riga = [self._formatta_valore(tag, qso, per_pdf=False) for tag in campi_scelti.keys()]
                riga.append(str(qso.get('sat_name', '')).strip())
                country = str(qso.get('country', '')).strip().upper()
                cont = str(qso.get('cont', '')).strip().upper()
                if not cont or cont not in ['EU', 'NA', 'SA', 'AS', 'AF', 'OC']:
                    cont = MAPPA_CONTINENTI_DXCC.get(country, '')
                riga.append(cont)
                ws_log.append(riga)

                fill_riga = fill_pari if r_idx % 2 == 0 else None
                for col_idx, cell in enumerate(ws_log[r_idx], 1):
                    if fill_riga:
                        cell.fill = fill_riga
                    cell.font = font_norm
                    cell.alignment = align_c
                    cell.border = bordo
                    val_len = len(str(cell.value or '')) + 2
                    if val_len > col_widths.get(col_idx, 0):
                        col_widths[col_idx] = val_len

            for col_idx, width in col_widths.items():
                ws_log.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(width, 40)
            ws_log.row_dimensions[1].height = 20
            ws_log.freeze_panes = "A2"

            # ── Foglio STATISTICHE ─────────────────────────────────────
            ws_st = wb.create_sheet("Statistiche")
            ws_st.column_dimensions['A'].width = 28
            ws_st.column_dimensions['B'].width = 18

            def scrivi_sezione(ws, titolo, dati, riga_start):
                cell_t = ws.cell(row=riga_start, column=1, value=titolo)
                cell_t.font = Font(bold=True, color=c1_hex, size=12, name="Arial")
                riga_start += 1
                cell_h1 = ws.cell(row=riga_start, column=1, value="Voce")
                cell_h2 = ws.cell(row=riga_start, column=2, value="Valore")
                for c in [cell_h1, cell_h2]:
                    c.fill = fill_header
                    c.font = font_header
                    c.alignment = align_c
                    c.border = bordo
                riga_start += 1
                for idx, (k, v) in enumerate(dati):
                    r = riga_start + idx
                    ca = ws.cell(row=r, column=1, value=k)
                    cb = ws.cell(row=r, column=2, value=v)
                    fill_r = fill_pari if idx % 2 == 0 else None
                    for c in [ca, cb]:
                        if fill_r:
                            c.fill = fill_r
                        c.font = font_norm
                        c.alignment = align_l
                        c.border = bordo
                return riga_start + len(dati) + 2

            r = scrivi_sezione(ws_st, "Riepilogo Generale", [
                ("QSO Totali", stats['totale']),
                ("DXCC Unici", stats['dxcc']),
                ("Bande Utilizzate", stats['bande']),
                ("Modi Distinti", stats['modi_distinti']),
                ("Conferme LoTW", stats['lotw_confermati']),
                ("Conferme eQSL", stats['eqsl_confermati']),
            ], 1)

            r = scrivi_sezione(ws_st, "QSO per Modo", sorted(stats['ripartizione_modi'].items(), key=lambda x: x[1], reverse=True), r)
            r = scrivi_sezione(ws_st, "QSO per Banda", sorted(stats['ripartizione_bande'].items(), key=lambda x: x[1], reverse=True), r)
            scrivi_sezione(ws_st, "QSO per Continente", sorted(stats['ripartizione_continenti'].items(), key=lambda x: x[1], reverse=True), r)

            # ── Foglio DXCC ────────────────────────────────────────────
            ws_dxcc = wb.create_sheet("DXCC Lavorati")
            ws_dxcc.column_dimensions['A'].width = 30
            for col, w in [('B', 10), ('C', 10), ('D', 25), ('E', 25), ('F', 10), ('G', 10)]:
                ws_dxcc.column_dimensions[col].width = w

            hdrs = ["Country / DXCC", "Continente", "QSO", "Bande", "Modi", "LoTW", "eQSL"]
            ws_dxcc.append(hdrs)
            for col_idx, cell in enumerate(ws_dxcc[1], 1):
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = align_c
                cell.border = bordo

            riga_idx = 2
            for cont in CONTINENTS_ORDER:
                paesi_cont = [(p, i) for p, i in stats['dxcc_dettaglio'].items() if i.get('continente', 'Altri') == cont]
                if not paesi_cont:
                    continue
                for idx, (paese, info) in enumerate(sorted(paesi_cont)):
                    riga_d = [
                        paese, cont, info['qso'],
                        ", ".join(sorted(info['bande'])),
                        ", ".join(sorted(info['modi'])),
                        "Y" if info['lotw'] else "-",
                        "Y" if info['eqsl'] else "-",
                    ]
                    ws_dxcc.append(riga_d)
                    fill_r = fill_pari if idx % 2 == 0 else None
                    for col_idx, cell in enumerate(ws_dxcc[riga_idx], 1):
                        if fill_r:
                            cell.fill = fill_r
                        cell.font = font_norm
                        cell.alignment = align_c if col_idx > 1 else align_l
                        cell.border = bordo
                    riga_idx += 1

            ws_dxcc.freeze_panes = "A2"

            wb.save(save_path)
            messagebox.showinfo("Successo", f"Excel esportato!\n{os.path.basename(save_path)}\n3 fogli: Log QSO, Statistiche, DXCC Lavorati")

        except Exception as ex:
            messagebox.showerror("Errore", f"Errore durante l'esportazione Excel:\n{ex}")


if __name__ == "__main__":
    app = ADIFtoPDFApp()
    _APP_REF = app   # riferimento globale usato da _ripristina_tema_ttk
    app.mainloop()
